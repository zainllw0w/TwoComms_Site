import json
from decimal import Decimal, InvalidOperation
from datetime import datetime, time, timedelta
from typing import Any

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Sum
from django.db.models.functions import Coalesce
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.views.decorators.http import require_POST

from orders.models import WholesaleInvoice

from .models import Shop, ShopCommunication, ShopInventoryMovement, ShopPhone, ShopShipment
from .views import get_manager_bot_username, get_reminders, get_user_stats, has_report_today, user_is_management


def _decimal_or_none(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("грн", "").replace("₴", "").replace(" ", "").replace("\u00a0", "")
    text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _int_or_zero(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _guess_category(title: str) -> str:
    t = (title or "").lower()
    if any(k in t for k in ("худі", "hoodie", "худи", "флис", "fleece")):
        return "hoodie"
    if any(k in t for k in ("футбол", "t-shirt", "tee", "футб", "tshirt")):
        return "tshirt"
    return "other"


def _invoice_summary_from_wholesale_invoice(inv: WholesaleInvoice) -> dict[str, Any]:
    items = []
    company_data = {}
    order_details = inv.order_details or {}
    if isinstance(order_details, dict):
        company_data = order_details.get("company_data") or {}
        raw_items = order_details.get("order_items") or []
    else:
        raw_items = []

    total_amount = _decimal_or_none(inv.total_amount) or Decimal("0")
    category_totals: dict[str, dict[str, Any]] = {}

    for raw in raw_items:
        product = (raw or {}).get("product", {}) or {}
        title = str(product.get("title") or "").strip() or "—"
        size = str((raw or {}).get("size") or "").strip()
        color = str((raw or {}).get("color") or "").strip()
        qty = _int_or_zero((raw or {}).get("quantity"))
        price = _decimal_or_none((raw or {}).get("price")) or Decimal("0")
        line_total = _decimal_or_none((raw or {}).get("total"))
        if line_total is None:
            line_total = (price * Decimal(qty)) if qty else Decimal("0")

        category = (product.get("type") or "").strip() or _guess_category(title)
        items.append(
            {
                "title": title,
                "size": size,
                "color": color,
                "quantity": qty,
                "price": str(price),
                "total": str(line_total),
                "category": category,
            }
        )

        bucket = category_totals.setdefault(category, {"qty": 0, "amount": Decimal("0")})
        bucket["qty"] += qty
        bucket["amount"] += line_total

    serialized_category_totals = {
        k: {"qty": v["qty"], "amount": str(v["amount"])} for k, v in category_totals.items()
    }

    return {
        "source": "system",
        "invoice_id": inv.id,
        "invoice_number": inv.invoice_number,
        "company_name": inv.company_name,
        "company_data": company_data,
        "items": items,
        "total_amount": str(total_amount),
        "category_totals": serialized_category_totals,
        "created_at": timezone.localtime(inv.created_at).isoformat() if inv.created_at else "",
    }


def _parse_uploaded_invoice_xlsx(uploaded_file) -> dict[str, Any]:
    try:
        import openpyxl
    except Exception:
        return {"source": "upload", "items": [], "total_amount": "0", "category_totals": {}}

    try:
        uploaded_file.seek(0)
    except Exception:
        pass
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    try:
        uploaded_file.seek(0)
    except Exception:
        pass
    ws = wb.active

    def norm(v: Any) -> str:
        s = str(v or "").strip().lower()
        for ch in ("\n", "\r", "\t"):
            s = s.replace(ch, " ")
        s = " ".join(s.split())
        return s

    header_row_idx = None
    header_cells = []
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(120, ws.max_row), values_only=True), start=1):
        row_norm = [norm(c) for c in row]
        if not any(row_norm):
            continue
        # heuristic: look for товар/назва + кількість + сума
        joined = " | ".join(row_norm)
        if (("назва" in joined or "наим" in joined or "товар" in joined) and
                ("кільк" in joined or "колич" in joined) and
                ("сума" in joined or "сумма" in joined or "итог" in joined)):
            header_row_idx = idx
            header_cells = row_norm
            break

    col_map: dict[str, int] = {}
    if header_row_idx:
        for i, cell in enumerate(header_cells):
            if not cell:
                continue
            if "назва" in cell or "наим" in cell or "товар" in cell:
                col_map["title"] = i
            elif "розм" in cell or "разм" in cell:
                col_map["size"] = i
            elif "колір" in cell or "цвет" in cell:
                col_map["color"] = i
            elif "кільк" in cell or "колич" in cell:
                col_map["qty"] = i
            elif "ціна" in cell or "цена" in cell:
                col_map["price"] = i
            elif "сума" in cell or "сумма" in cell or "итог" in cell:
                col_map["total"] = i

    items: list[dict[str, Any]] = []
    category_totals: dict[str, dict[str, Any]] = {}
    total_amount = Decimal("0")

    if not header_row_idx or "title" not in col_map:
        return {"source": "upload", "items": [], "total_amount": "0", "category_totals": {}}

    blank_streak = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
        title = str(row[col_map["title"]] or "").strip() if col_map["title"] < len(row) else ""
        if not title:
            blank_streak += 1
            if blank_streak >= 8:
                break
            continue
        blank_streak = 0

        size = str(row[col_map.get("size", -1)] or "").strip() if col_map.get("size", -1) >= 0 else ""
        color = str(row[col_map.get("color", -1)] or "").strip() if col_map.get("color", -1) >= 0 else ""
        qty = _int_or_zero(row[col_map.get("qty", -1)]) if col_map.get("qty", -1) >= 0 else 0
        price = _decimal_or_none(row[col_map.get("price", -1)]) if col_map.get("price", -1) >= 0 else None
        line_total = _decimal_or_none(row[col_map.get("total", -1)]) if col_map.get("total", -1) >= 0 else None
        if line_total is None:
            line_total = (price or Decimal("0")) * Decimal(qty or 0)
        if price is None:
            price = Decimal("0")

        category = _guess_category(title)
        items.append(
            {
                "title": title,
                "size": size,
                "color": color,
                "quantity": qty,
                "price": str(price),
                "total": str(line_total),
                "category": category,
            }
        )
        total_amount += line_total
        bucket = category_totals.setdefault(category, {"qty": 0, "amount": Decimal("0")})
        bucket["qty"] += qty
        bucket["amount"] += line_total

    serialized_category_totals = {
        k: {"qty": v["qty"], "amount": str(v["amount"])} for k, v in category_totals.items()
    }

    return {"source": "upload", "items": items, "total_amount": str(total_amount), "category_totals": serialized_category_totals}


def _shop_accessible_to_user(shop: Shop, user) -> bool:
    if user.is_staff or user.is_superuser:
        return True
    return shop.created_by_id == user.id


def _rebuild_receipts_for_shipment(shipment: ShopShipment) -> None:
    # Drop previous receipts for this shipment to keep idempotency
    ShopInventoryMovement.objects.filter(shipment=shipment, kind=ShopInventoryMovement.Kind.RECEIPT).delete()

    summary = shipment.invoice_summary or {}
    items = summary.get("items") if isinstance(summary, dict) else None
    if not isinstance(items, list):
        return

    moves = []
    for it in items:
        if not isinstance(it, dict):
            continue
        title = str(it.get("title") or "").strip()
        if not title:
            continue
        qty = _int_or_zero(it.get("quantity"))
        if qty <= 0:
            continue
        moves.append(
            ShopInventoryMovement(
                shop=shipment.shop,
                shipment=shipment,
                kind=ShopInventoryMovement.Kind.RECEIPT,
                product_name=title,
                category=str(it.get("category") or "").strip(),
                size=str(it.get("size") or "").strip(),
                color=str(it.get("color") or "").strip(),
                delta_qty=qty,
                created_by=shipment.created_by,
                note=f"Надходження по ТТН {shipment.ttn_number}",
            )
        )
    if moves:
        ShopInventoryMovement.objects.bulk_create(moves)


@login_required(login_url="management_login")
def shops(request):
    if not user_is_management(request.user):
        return redirect("management_login")

    qs = Shop.objects.all() if request.user.is_staff else Shop.objects.filter(created_by=request.user)
    qs = qs.annotate(
        total_amount=Coalesce(
            Sum("shipments__invoice_total_amount"),
            Decimal("0"),
            output_field=models.DecimalField(max_digits=12, decimal_places=2),
        )
    )
    qs = qs.prefetch_related("phones", "shipments__wholesale_invoice")
    qs = qs.order_by("-created_at")

    paginator = Paginator(qs, 11)  # 12 cards total incl. "+", so 11 shops per page
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    now_dt = timezone.localtime(timezone.now())
    now_date = now_dt.date()

    shops_payload: list[dict[str, Any]] = []
    for shop in page_obj.object_list:
        primary_phone = None
        phones = []
        for p in shop.phones.all():
            phones.append(
                {
                    "id": p.id,
                    "role": p.role,
                    "role_other": p.role_other,
                    "phone": p.phone,
                    "is_primary": bool(p.is_primary),
                    "sort_order": p.sort_order,
                }
            )
            if p.is_primary and not primary_phone:
                primary_phone = p.phone
        if not primary_phone and phones:
            primary_phone = phones[0]["phone"]

        shipments = []
        for s in shop.shipments.all():
            invoice_kind = "none"
            invoice_title = ""
            invoice_download_url = ""
            invoice_total = s.invoice_total_amount
            if s.wholesale_invoice_id:
                invoice_kind = "system"
                inv_num = ""
                try:
                    inv_num = s.wholesale_invoice.invoice_number if s.wholesale_invoice else ""
                except Exception:
                    inv_num = ""
                invoice_title = f"#{inv_num}" if inv_num else "Накладна"
                invoice_download_url = reverse("management_invoices_download", args=[s.wholesale_invoice_id])
            elif s.uploaded_invoice_file:
                invoice_kind = "upload"
                invoice_title = s.uploaded_invoice_file.name.split("/")[-1]
                invoice_download_url = reverse("management_shop_shipment_invoice_download", args=[s.id])

            shipments.append(
                {
                    "id": s.id,
                    "ttn_number": s.ttn_number,
                    "shipped_at": s.shipped_at.isoformat() if s.shipped_at else "",
                    "invoice_kind": invoice_kind,
                    "wholesale_invoice_id": s.wholesale_invoice_id,
                    "invoice_title": invoice_title,
                    "invoice_total_amount": str(invoice_total) if invoice_total is not None else "",
                    "invoice_download_url": invoice_download_url,
                }
            )

        timer = None
        if shop.shop_type == Shop.ShopType.TEST and shop.test_connected_at:
            end_date = shop.test_connected_at + timedelta(days=int(shop.test_period_days or 14))
            end_dt = timezone.make_aware(datetime.combine(end_date, time.min), timezone.get_current_timezone())
            remaining = end_dt - now_dt
            remaining_seconds = int(remaining.total_seconds())
            if remaining_seconds <= 0:
                timer = {"status": "expired", "label": "Тест завершено", "seconds": remaining_seconds}
            elif remaining_seconds <= 86400:
                hours = max(1, remaining_seconds // 3600)
                timer = {"status": "urgent", "label": f"Залишилось {hours} год", "seconds": remaining_seconds}
            else:
                days = remaining_seconds // 86400
                timer = {"status": "active", "label": f"Залишилось {days} дн", "seconds": remaining_seconds}

        shops_payload.append(
            {
                "id": shop.id,
                "name": shop.name,
                "photo_url": shop.photo.url if shop.photo else "",
                "owner_full_name": shop.owner_full_name,
                "shop_type": shop.shop_type,
                "registration_place": shop.registration_place,
                "is_physical": bool(shop.is_physical),
                "city": shop.city,
                "address": shop.address,
                "website_url": shop.website_url,
                "instagram_url": shop.instagram_url,
                "prom_url": shop.prom_url,
                "other_sales_channel": shop.other_sales_channel,
                "test_product_id": shop.test_product_id,
                "test_package": shop.test_package or {},
                "test_contract_name": shop.test_contract_file.name.split("/")[-1] if shop.test_contract_file else "",
                "test_contract_download_url": reverse("management_shop_contract_download", args=[shop.id]) if shop.test_contract_file else "",
                "test_connected_at": shop.test_connected_at.isoformat() if shop.test_connected_at else "",
                "test_period_days": int(shop.test_period_days or 14),
                "next_contact_at": timezone.localtime(shop.next_contact_at).isoformat() if shop.next_contact_at else "",
                "notes": shop.notes,
                "primary_phone": primary_phone or "",
                "phones": phones,
                "shipments": shipments,
                "total_amount": str(getattr(shop, "total_amount", "") or ""),
                "timer": timer,
                "created_by": (shop.created_by.get_full_name() or shop.created_by.username) if shop.created_by else "",
                "managed_by": (shop.managed_by.get_full_name() or shop.managed_by.username) if shop.managed_by else "",
                "can_delete": bool(request.user.is_staff or request.user.is_superuser),
            }
        )

    stats = get_user_stats(request.user)
    report_sent_today = has_report_today(request.user)
    reminders = get_reminders(request.user, stats=stats, report_sent=report_sent_today)
    bot_username = get_manager_bot_username()

    try:
        from storefront.models import Product

        test_products = list(
            Product.objects.all()
            .select_related("category")
            .order_by("category__name", "title")
            .values("id", "title", "category__name")[:600]
        )
    except Exception:
        test_products = []

    return render(
        request,
        "management/shops.html",
        {
            "page_obj": page_obj,
            "shops_payload": shops_payload,
            "test_products": test_products,
            "now_date": now_date.isoformat(),
            "reminders": reminders,
            "manager_bot_username": bot_username,
        },
    )


@login_required(login_url="management_login")
@require_POST
def shops_save_api(request):
    if not user_is_management(request.user):
        return JsonResponse({"ok": False}, status=403)

    phones_raw = (request.POST.get("phones_json") or "").strip()
    try:
        phones_data = json.loads(phones_raw) if phones_raw else []
    except Exception:
        phones_data = []
    if not isinstance(phones_data, list):
        phones_data = []

    shipments_raw = (request.POST.get("shipments_json") or "").strip()
    try:
        shipments_data = json.loads(shipments_raw) if shipments_raw else []
    except Exception:
        shipments_data = []
    if not isinstance(shipments_data, list):
        shipments_data = []

    if not any(isinstance(p, dict) and str(p.get("phone") or "").strip() for p in phones_data):
        return JsonResponse({"ok": False, "error": "Додайте мінімум 1 номер телефону"}, status=400)
    if not any(
        isinstance(s, dict)
        and str(s.get("ttn_number") or "").strip()
        and parse_date(str(s.get("shipped_at") or "").strip())
        for s in shipments_data
    ):
        return JsonResponse({"ok": False, "error": "Додайте мінімум 1 ТТН"}, status=400)

    shop_id = (request.POST.get("shop_id") or "").strip()
    is_update = bool(shop_id)

    if is_update:
        shop = get_object_or_404(Shop, id=shop_id)
        if not _shop_accessible_to_user(shop, request.user):
            return JsonResponse({"ok": False}, status=403)
    else:
        shop = Shop(created_by=request.user, managed_by=request.user)

    name = (request.POST.get("name") or "").strip()
    if not name:
        return JsonResponse({"ok": False, "error": "Вкажіть назву магазину"}, status=400)

    shop.name = name
    shop.owner_full_name = (request.POST.get("owner_full_name") or "").strip()
    shop.registration_place = (request.POST.get("registration_place") or "").strip()
    shop.is_physical = (request.POST.get("is_physical") or "").strip() in {"1", "true", "on", "yes"}
    shop.city = (request.POST.get("city") or "").strip()
    shop.address = (request.POST.get("address") or "").strip()

    shop.website_url = (request.POST.get("website_url") or "").strip()
    shop.instagram_url = (request.POST.get("instagram_url") or "").strip()
    shop.prom_url = (request.POST.get("prom_url") or "").strip()
    shop.other_sales_channel = (request.POST.get("other_sales_channel") or "").strip()

    shop.shop_type = (request.POST.get("shop_type") or Shop.ShopType.FULL).strip()
    if shop.shop_type not in {Shop.ShopType.FULL, Shop.ShopType.TEST}:
        shop.shop_type = Shop.ShopType.FULL

    test_product_id = (request.POST.get("test_product_id") or "").strip()
    shop.test_product_id = int(test_product_id) if test_product_id.isdigit() else None

    test_connected_at = parse_date((request.POST.get("test_connected_at") or "").strip()) if shop.shop_type == Shop.ShopType.TEST else None
    shop.test_connected_at = test_connected_at

    try:
        shop.test_period_days = int((request.POST.get("test_period_days") or "14").strip() or 14)
    except Exception:
        shop.test_period_days = 14

    test_package_raw = (request.POST.get("test_package_json") or "").strip()
    if test_package_raw:
        try:
            shop.test_package = json.loads(test_package_raw)
        except Exception:
            shop.test_package = {}

    shop.notes = (request.POST.get("notes") or "").strip()

    if request.FILES.get("photo"):
        shop.photo = request.FILES["photo"]

    if shop.shop_type == Shop.ShopType.TEST and request.FILES.get("test_contract_file"):
        shop.test_contract_file = request.FILES["test_contract_file"]

    shop.save()

    # Phones (replace)
    ShopPhone.objects.filter(shop=shop).delete()
    phone_objs = []
    has_primary = any(bool(p.get("is_primary")) for p in phones_data if isinstance(p, dict))
    for idx, p in enumerate(phones_data):
        if not isinstance(p, dict):
            continue
        phone = str(p.get("phone") or "").strip()
        if not phone:
            continue
        role = str(p.get("role") or ShopPhone.Role.OWNER).strip()
        if role not in {c for c, _ in ShopPhone.Role.choices}:
            role = ShopPhone.Role.OTHER
        phone_objs.append(
            ShopPhone(
                shop=shop,
                role=role,
                role_other=str(p.get("role_other") or "").strip(),
                phone=phone,
                is_primary=bool(p.get("is_primary")) if has_primary else idx == 0,
                sort_order=_int_or_zero(p.get("sort_order")) or idx,
            )
        )
    ShopPhone.objects.bulk_create(phone_objs)

    # Shipments (upsert)
    existing_ids = set(shop.shipments.values_list("id", flat=True))
    seen_ids: set[int] = set()

    for idx, s in enumerate(shipments_data):
        if not isinstance(s, dict):
            continue
        sid = s.get("id")
        ttn_number = str(s.get("ttn_number") or "").strip()
        shipped_at = parse_date(str(s.get("shipped_at") or "").strip())
        if not ttn_number or not shipped_at:
            continue

        shipment = None
        if sid:
            try:
                sid_int = int(sid)
            except Exception:
                sid_int = None
            if sid_int and sid_int in existing_ids:
                shipment = ShopShipment.objects.filter(id=sid_int, shop=shop).first()
                seen_ids.add(sid_int)

        if shipment is None:
            shipment = ShopShipment(shop=shop, created_by=request.user)

        shipment.ttn_number = ttn_number
        shipment.shipped_at = shipped_at

        invoice_kind = "none" if shop.shop_type == Shop.ShopType.TEST else str(s.get("invoice_kind") or "none").strip()
        if invoice_kind == "system":
            wid = s.get("wholesale_invoice_id")
            shipment.uploaded_invoice_file = None
            shipment.invoice_summary = {}
            shipment.invoice_total_amount = None
            if wid:
                try:
                    wid_int = int(wid)
                except Exception:
                    wid_int = None
                inv = WholesaleInvoice.objects.filter(id=wid_int).first() if wid_int else None
                if inv and (request.user.is_staff or inv.created_by_id == request.user.id):
                    shipment.wholesale_invoice = inv
                    summary = _invoice_summary_from_wholesale_invoice(inv)
                    shipment.invoice_summary = summary
                    shipment.invoice_total_amount = _decimal_or_none(summary.get("total_amount")) or _decimal_or_none(inv.total_amount)
                else:
                    shipment.wholesale_invoice = None
        elif invoice_kind == "upload":
            shipment.wholesale_invoice = None
            file_field = str(s.get("file_field") or "").strip()
            if file_field and request.FILES.get(file_field):
                uploaded = request.FILES[file_field]
                summary = _parse_uploaded_invoice_xlsx(uploaded)
                shipment.invoice_summary = summary
                shipment.invoice_total_amount = _decimal_or_none(summary.get("total_amount"))
                try:
                    uploaded.seek(0)
                except Exception:
                    pass
                shipment.uploaded_invoice_file = uploaded
        else:
            shipment.wholesale_invoice = None
            shipment.uploaded_invoice_file = None
            shipment.invoice_summary = {}
            shipment.invoice_total_amount = None

        shipment.save()

        if shipment.invoice_summary:
            _rebuild_receipts_for_shipment(shipment)

    if shop.shop_type == Shop.ShopType.TEST and not shop.test_connected_at:
        shipped_dates = [d for d in shop.shipments.values_list("shipped_at", flat=True) if d]
        if shipped_dates:
            shop.test_connected_at = min(shipped_dates)
            shop.save(update_fields=["test_connected_at"])

    # NOTE: deletions are not applied implicitly to avoid accidental data loss.

    return JsonResponse({"ok": True, "shop_id": shop.id})


@login_required(login_url="management_login")
def shops_detail_api(request, shop_id: int):
    if not user_is_management(request.user):
        return JsonResponse({"ok": False}, status=403)

    shop = get_object_or_404(Shop, id=shop_id)
    if not _shop_accessible_to_user(shop, request.user):
        return JsonResponse({"ok": False}, status=403)

    comms = list(
        ShopCommunication.objects.filter(shop=shop)
        .order_by("-contacted_at")[:50]
        .values("id", "contacted_at", "contact_person", "phone", "note")
    )
    for c in comms:
        try:
            c["contacted_at"] = timezone.localtime(c["contacted_at"]).isoformat()
        except Exception:
            pass

    # Inventory aggregation
    from django.db.models import Q
    from django.db.models.functions import Abs

    inv_rows = (
        ShopInventoryMovement.objects.filter(shop=shop)
        .values("product_name", "category", "size", "color")
        .annotate(
            available=Coalesce(Sum("delta_qty"), 0),
            sold=Coalesce(Sum(Abs("delta_qty"), filter=Q(kind=ShopInventoryMovement.Kind.SALE)), 0),
            received=Coalesce(Sum("delta_qty", filter=Q(kind=ShopInventoryMovement.Kind.RECEIPT)), 0),
            adjusted=Coalesce(Sum("delta_qty", filter=Q(kind=ShopInventoryMovement.Kind.ADJUST)), 0),
        )
        .order_by("category", "product_name", "size", "color")
    )
    inventory = []
    for row in inv_rows:
        inventory.append(
            {
                "product_name": row.get("product_name") or "",
                "category": row.get("category") or "",
                "size": row.get("size") or "",
                "color": row.get("color") or "",
                "available": int(row.get("available") or 0),
                "sold": int(row.get("sold") or 0),
                "received": int(row.get("received") or 0),
                "adjusted": int(row.get("adjusted") or 0),
            }
        )

    shipments = []
    for s in shop.shipments.select_related("wholesale_invoice").order_by("-shipped_at", "-id"):
        invoice_kind = "none"
        invoice_title = ""
        invoice_download_url = ""
        if s.wholesale_invoice_id:
            invoice_kind = "system"
            inv_num = ""
            try:
                inv_num = s.wholesale_invoice.invoice_number if s.wholesale_invoice else ""
            except Exception:
                inv_num = ""
            invoice_title = f"#{inv_num}" if inv_num else "Накладна"
            invoice_download_url = reverse("management_invoices_download", args=[s.wholesale_invoice_id])
        elif s.uploaded_invoice_file:
            invoice_kind = "upload"
            invoice_title = s.uploaded_invoice_file.name.split("/")[-1]
            invoice_download_url = reverse("management_shop_shipment_invoice_download", args=[s.id])
        shipments.append(
            {
                "id": s.id,
                "ttn_number": s.ttn_number,
                "shipped_at": s.shipped_at.isoformat() if s.shipped_at else "",
                "invoice_kind": invoice_kind,
                "invoice_title": invoice_title,
                "invoice_total_amount": str(s.invoice_total_amount) if s.invoice_total_amount is not None else "",
                "invoice_summary": s.invoice_summary or {},
                "invoice_download_url": invoice_download_url,
            }
        )

    return JsonResponse(
        {
            "ok": True,
            "shop": {
                "id": shop.id,
                "name": shop.name,
                "shop_type": shop.shop_type,
                "owner_full_name": shop.owner_full_name,
                "test_contract_name": shop.test_contract_file.name.split("/")[-1] if shop.test_contract_file else "",
                "test_contract_download_url": reverse("management_shop_contract_download", args=[shop.id]) if shop.test_contract_file else "",
                "phones": list(shop.phones.order_by("sort_order", "id").values("id", "role", "role_other", "phone", "is_primary")),
                "next_contact_at": timezone.localtime(shop.next_contact_at).isoformat() if shop.next_contact_at else "",
            },
            "shipments": shipments,
            "communications": comms,
            "inventory": inventory,
        }
    )


@login_required(login_url="management_login")
@require_POST
def shops_add_contact_api(request):
    if not user_is_management(request.user):
        return JsonResponse({"ok": False}, status=403)
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        payload = {}

    shop_id = payload.get("shop_id")
    shop = get_object_or_404(Shop, id=shop_id)
    if not _shop_accessible_to_user(shop, request.user):
        return JsonResponse({"ok": False}, status=403)

    contacted_at = parse_datetime(str(payload.get("contacted_at") or "").strip())
    if contacted_at is None:
        contacted_at = timezone.localtime(timezone.now())
    if timezone.is_naive(contacted_at):
        contacted_at = timezone.make_aware(contacted_at, timezone.get_current_timezone())

    comm = ShopCommunication.objects.create(
        shop=shop,
        contacted_at=contacted_at,
        contact_person=str(payload.get("contact_person") or "").strip(),
        phone=str(payload.get("phone") or "").strip(),
        note=str(payload.get("note") or "").strip(),
        created_by=request.user,
    )

    return JsonResponse(
        {
            "ok": True,
            "communication": {
                "id": comm.id,
                "contacted_at": timezone.localtime(comm.contacted_at).isoformat(),
                "contact_person": comm.contact_person,
                "phone": comm.phone,
                "note": comm.note,
            },
        }
    )


@login_required(login_url="management_login")
@require_POST
def shops_set_next_contact_api(request):
    if not user_is_management(request.user):
        return JsonResponse({"ok": False}, status=403)
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        payload = {}

    shop_id = payload.get("shop_id")
    shop = get_object_or_404(Shop, id=shop_id)
    if not _shop_accessible_to_user(shop, request.user):
        return JsonResponse({"ok": False}, status=403)

    dt_str = str(payload.get("next_contact_at") or "").strip()
    if not dt_str:
        shop.next_contact_at = None
        shop.save(update_fields=["next_contact_at"])
        return JsonResponse({"ok": True, "next_contact_at": ""})

    next_dt = parse_datetime(dt_str)
    if next_dt is None:
        return JsonResponse({"ok": False, "error": "Невірний формат дати"}, status=400)
    if timezone.is_naive(next_dt):
        next_dt = timezone.make_aware(next_dt, timezone.get_current_timezone())

    shop.next_contact_at = next_dt
    shop.save(update_fields=["next_contact_at"])

    return JsonResponse({"ok": True, "next_contact_at": timezone.localtime(next_dt).isoformat()})


@login_required(login_url="management_login")
@require_POST
def shops_inventory_move_api(request):
    if not user_is_management(request.user):
        return JsonResponse({"ok": False}, status=403)
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        payload = {}

    shop_id = payload.get("shop_id")
    shop = get_object_or_404(Shop, id=shop_id)
    if not _shop_accessible_to_user(shop, request.user):
        return JsonResponse({"ok": False}, status=403)

    kind = str(payload.get("kind") or "").strip()
    if kind not in {k for k, _ in ShopInventoryMovement.Kind.choices}:
        return JsonResponse({"ok": False, "error": "Невірний тип операції"}, status=400)

    product_name = str(payload.get("product_name") or "").strip()
    if not product_name:
        return JsonResponse({"ok": False, "error": "Вкажіть товар"}, status=400)

    lines = payload.get("lines")
    if isinstance(lines, list):
        moves = []
        note = str(payload.get("note") or "").strip()
        category = str(payload.get("category") or "").strip()
        color = str(payload.get("color") or "").strip()
        for line in lines:
            if not isinstance(line, dict):
                continue
            line_product_name = str(line.get("product_name") or product_name).strip()
            if not line_product_name:
                continue
            line_category = str(line.get("category") or category).strip()
            line_color = str(line.get("color") or color).strip()
            line_size = str(line.get("size") or "").strip()

            delta = 0
            if kind == ShopInventoryMovement.Kind.SALE:
                qty = _int_or_zero(line.get("qty"))
                if qty <= 0:
                    continue
                delta = -qty
            elif kind == ShopInventoryMovement.Kind.ADJUST:
                delta = _int_or_zero(line.get("delta_qty"))
                if delta == 0:
                    continue
            else:
                qty = _int_or_zero(line.get("qty"))
                if qty <= 0:
                    continue
                delta = qty

            moves.append(
                ShopInventoryMovement(
                    shop=shop,
                    kind=kind,
                    product_name=line_product_name,
                    category=line_category,
                    size=line_size,
                    color=line_color,
                    delta_qty=delta,
                    note=note,
                    created_by=request.user,
                )
            )
        if not moves:
            return JsonResponse({"ok": False, "error": "Немає позицій для збереження"}, status=400)
        ShopInventoryMovement.objects.bulk_create(moves)
        return JsonResponse({"ok": True, "count": len(moves)})

    delta = 0
    if kind == ShopInventoryMovement.Kind.SALE:
        qty = _int_or_zero(payload.get("qty"))
        if qty <= 0:
            return JsonResponse({"ok": False, "error": "Кількість має бути > 0"}, status=400)
        delta = -qty
    elif kind == ShopInventoryMovement.Kind.ADJUST:
        delta = _int_or_zero(payload.get("delta_qty"))
        if delta == 0:
            return JsonResponse({"ok": False, "error": "Коригування має бути ≠ 0"}, status=400)
    else:
        qty = _int_or_zero(payload.get("qty"))
        if qty <= 0:
            return JsonResponse({"ok": False, "error": "Кількість має бути > 0"}, status=400)
        delta = qty

    ShopInventoryMovement.objects.create(
        shop=shop,
        kind=kind,
        product_name=product_name,
        category=str(payload.get("category") or "").strip(),
        size=str(payload.get("size") or "").strip(),
        color=str(payload.get("color") or "").strip(),
        delta_qty=delta,
        note=str(payload.get("note") or "").strip(),
        created_by=request.user,
    )

    return JsonResponse({"ok": True})


@login_required(login_url="management_login")
@require_POST
def shops_delete_api(request, shop_id: int):
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({"ok": False}, status=403)
    if not user_is_management(request.user):
        return JsonResponse({"ok": False}, status=403)
    shop = get_object_or_404(Shop, id=shop_id)
    shop.delete()
    return JsonResponse({"ok": True})


@login_required(login_url="management_login")
def shop_shipment_invoice_download(request, shipment_id: int):
    if not user_is_management(request.user):
        return redirect("management_login")
    shipment = get_object_or_404(ShopShipment, id=shipment_id)
    if not _shop_accessible_to_user(shipment.shop, request.user):
        return HttpResponse("Доступ заборонено", status=403)
    if not shipment.uploaded_invoice_file:
        return HttpResponse("Файл не знайдено", status=404)
    filename = shipment.uploaded_invoice_file.name.split("/")[-1]
    return FileResponse(shipment.uploaded_invoice_file.open("rb"), as_attachment=True, filename=filename)


@login_required(login_url="management_login")
def shop_contract_download(request, shop_id: int):
    if not user_is_management(request.user):
        return redirect("management_login")
    shop = get_object_or_404(Shop, id=shop_id)
    if not _shop_accessible_to_user(shop, request.user):
        return HttpResponse("Доступ заборонено", status=403)
    if not shop.test_contract_file:
        return HttpResponse("Файл не знайдено", status=404)
    filename = shop.test_contract_file.name.split("/")[-1]
    return FileResponse(shop.test_contract_file.open("rb"), as_attachment=True, filename=filename)
