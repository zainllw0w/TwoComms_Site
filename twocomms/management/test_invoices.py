import json
import os
import tempfile
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import Client as DjangoClient
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from management.shop_views import _invoice_summary_from_wholesale_invoice
from orders.models import WholesaleInvoice


def _find_row_by_value(ws, value):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell.row
    return None


@override_settings(ROOT_URLCONF="twocomms.urls_management")
class InvoiceItemNormalizationTests(TestCase):
    def test_normalize_invoice_items_recalculates_auto_and_manual_prices(self):
        from management.invoice_service import normalize_management_invoice_payload

        company_data = {
            "companyName": "ТОВ Тест",
            "contactPhone": "+380000000000",
            "deliveryAddress": "Київ, вул. Тестова, 1",
        }
        order_items = [
            {
                "product": {"id": "1", "type": "tshirt", "title": "Футболка 1"},
                "size": "M",
                "color": "Чорний",
                "quantity": 6,
                "pricing_mode": "auto",
                "extra_description": "з кашкорсе",
            },
            {
                "product": {"id": "2", "type": "tshirt", "title": "Футболка 2"},
                "size": "L",
                "color": "Білий",
                "quantity": 2,
                "pricing_mode": "manual",
                "manual_price": "610",
            },
            {
                "product": {"id": "3", "type": "hoodie", "title": "Худі 1 [фліс]"},
                "size": "XL",
                "color": "Чорний",
                "quantity": 8,
                "pricing_mode": "auto",
            },
        ]

        normalized = normalize_management_invoice_payload(
            company_data=company_data,
            order_items=order_items,
        )

        self.assertEqual(normalized["totals"]["total_tshirts"], 8)
        self.assertEqual(normalized["totals"]["total_hoodies"], 8)
        self.assertEqual(normalized["items"][0]["display_title"], "Футболка 1 [з кашкорсе]")
        self.assertEqual(normalized["items"][0]["base_unit_price"], Decimal("540.00"))
        self.assertEqual(normalized["items"][0]["unit_price"], Decimal("540.00"))
        self.assertEqual(normalized["items"][0]["line_total"], Decimal("3240.00"))
        self.assertEqual(normalized["items"][1]["pricing_mode"], "manual")
        self.assertEqual(normalized["items"][1]["base_unit_price"], Decimal("540.00"))
        self.assertEqual(normalized["items"][1]["unit_price"], Decimal("610.00"))
        self.assertEqual(normalized["items"][1]["line_total"], Decimal("1220.00"))
        self.assertEqual(normalized["items"][2]["base_unit_price"], Decimal("1300.00"))
        self.assertEqual(normalized["items"][2]["line_total"], Decimal("10400.00"))
        self.assertEqual(normalized["totals"]["total_amount"], Decimal("14860.00"))
        self.assertTrue(normalized["pricing"]["has_manual_prices"])
        self.assertEqual(normalized["pricing"]["policy"], "custom_manual")
        self.assertEqual(normalized["items"][0]["base_unit_price"], Decimal("540.00"))
        self.assertEqual(normalized["items"][2]["base_unit_price"], Decimal("1300.00"))

    def test_manual_price_disables_tier_discount_recalculation_for_auto_items(self):
        from management.invoice_service import normalize_management_invoice_payload

        company_data = {
            "companyName": "ТОВ Тест",
            "contactPhone": "+380000000000",
            "deliveryAddress": "Київ, вул. Тестова, 1",
        }
        order_items = [
            {
                "product": {"id": "1", "type": "tshirt", "title": "Футболка 1"},
                "size": "M",
                "color": "Чорний",
                "quantity": 20,
                "pricing_mode": "auto",
            },
            {
                "product": {"id": "2", "type": "tshirt", "title": "Футболка 2"},
                "size": "L",
                "color": "Білий",
                "quantity": 1,
                "pricing_mode": "manual",
                "manual_price": "610",
            },
        ]

        normalized = normalize_management_invoice_payload(
            company_data=company_data,
            order_items=order_items,
        )

        self.assertTrue(normalized["pricing"]["has_manual_prices"])
        self.assertFalse(normalized["pricing"]["discounts_enabled"])
        self.assertEqual(normalized["items"][0]["base_unit_price"], Decimal("540.00"))
        self.assertEqual(normalized["items"][0]["unit_price"], Decimal("540.00"))
        self.assertEqual(normalized["items"][0]["line_total"], Decimal("10800.00"))
        self.assertEqual(normalized["items"][1]["unit_price"], Decimal("610.00"))
        self.assertEqual(normalized["totals"]["total_amount"], Decimal("11410.00"))

    def test_normalize_invoice_items_expands_full_size_run_multiplier(self):
        from management.invoice_service import normalize_management_invoice_payload

        company_data = {
            "companyName": "ТОВ Тест",
            "contactPhone": "+380000000000",
            "deliveryAddress": "Київ, вул. Тестова, 1",
        }
        order_items = [
            {
                "product": {"id": "1", "type": "tshirt", "title": "Футболка 1"},
                "size": "all",
                "run_multiplier": 3,
                "quantity": 4,
                "pricing_mode": "auto",
            },
        ]

        normalized = normalize_management_invoice_payload(
            company_data=company_data,
            order_items=order_items,
        )

        self.assertEqual(normalized["totals"]["total_tshirts"], 12)
        self.assertEqual(normalized["items"][0]["size"], "Всі ростовки (S-XL) ×3")
        self.assertEqual(normalized["items"][0]["run_multiplier"], 3)
        self.assertEqual(normalized["items"][0]["quantity"], 12)
        self.assertEqual(normalized["items"][0]["base_unit_price"], Decimal("540.00"))
        self.assertEqual(normalized["items"][0]["line_total"], Decimal("6480.00"))


@override_settings(
    ROOT_URLCONF="twocomms.urls_management",
    ALLOWED_HOSTS=["testserver", "management.twocomms.shop", "localhost", "127.0.0.1"],
)
class ManagementInvoiceApiTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username="manager", password="x", is_staff=True)
        self.client_http = DjangoClient()
        self.client_http.force_login(self.user)

    def test_invoices_page_renders_for_management_user(self):
        response = self.client_http.get(
            reverse("management_invoices"),
            HTTP_HOST="management.twocomms.shop",
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Сформувати накладну")
        self.assertContains(response, "Ручна ціна за 1 шт.")
        self.assertContains(response, "Множник ростовки")
        self.assertContains(response, "Включити 2XL у ростовку")
        self.assertContains(response, '<option value="2XL">2XL</option>', html=True)

    def test_generate_invoice_uses_normalized_rows_and_writes_excel(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            with override_settings(MEDIA_ROOT=tmpdir):
                payload = {
                    "companyData": {
                        "companyName": "ТОВ Тест",
                        "companyNumber": "12345678",
                        "contactPhone": "+380000000000",
                        "deliveryAddress": "Київ, вул. Тестова, 1",
                        "storeLink": "https://example.com/store",
                    },
                    "orderItems": [
                        {
                            "product": {"id": "1", "type": "tshirt", "title": "Футболка 1"},
                            "size": "M",
                            "color": "Чорний",
                            "quantity": 6,
                            "pricing_mode": "auto",
                            "extra_description": "з кашкорсе",
                        },
                        {
                            "product": {"id": "2", "type": "tshirt", "title": "Футболка 2"},
                            "size": "L",
                            "color": "Білий",
                            "quantity": 2,
                            "pricing_mode": "manual",
                            "manual_price": "610",
                        },
                    ],
                }

                response = self.client_http.post(
                    reverse("management_invoices_generate_api"),
                    data=json.dumps(payload),
                    content_type="application/json",
                    HTTP_HOST="management.twocomms.shop",
                    secure=True,
                )

                self.assertEqual(response.status_code, 200)
                data = response.json()
                self.assertTrue(data["ok"])

                invoice = WholesaleInvoice.objects.get(id=data["invoice"]["id"])
                self.assertEqual(invoice.total_tshirts, 8)
                self.assertEqual(invoice.total_hoodies, 0)
                self.assertEqual(invoice.total_amount, Decimal("4460.00"))
                self.assertEqual(invoice.order_details["pricing"]["policy"], "custom_manual")
                self.assertFalse(invoice.order_details["pricing"]["discounts_enabled"])

                saved_items = invoice.order_details["order_items"]
                self.assertEqual(saved_items[0]["display_title"], "Футболка 1 [з кашкорсе]")
                self.assertEqual(saved_items[0]["unit_price"], "540.00")
                self.assertEqual(saved_items[1]["pricing_mode"], "manual")
                self.assertEqual(saved_items[1]["unit_price"], "610.00")

                workbook_path = Path(invoice.file_path)
                self.assertTrue(workbook_path.exists())

                wb = load_workbook(workbook_path)
                ws = wb.active
                header_row = _find_row_by_value(ws, "Назва товару")
                self.assertIsNotNone(header_row)
                first_item_row = header_row + 1
                second_item_row = header_row + 2
                policy_row = _find_row_by_value(ws, "Ручна ціна присутня в замовленні: автоматичні знижки за кількість вимкнені.")
                link_row = _find_row_by_value(ws, "Посилання на магазин")

                self.assertEqual(ws.freeze_panes, f"A{first_item_row}")
                self.assertEqual(ws.cell(row=first_item_row, column=2).value, "Футболка 1 [з кашкорсе]")
                self.assertTrue(ws.cell(row=first_item_row, column=2).alignment.wrapText)
                self.assertEqual(ws.cell(row=first_item_row, column=6).value, 540)
                self.assertEqual(ws.cell(row=second_item_row, column=7).value, 1220)
                self.assertIsNotNone(policy_row)
                self.assertIn("₴", ws.cell(row=first_item_row, column=6).number_format)
                self.assertIsNotNone(link_row)
                self.assertEqual(ws[f"B{link_row}"].hyperlink.target, "https://example.com/store")

    def test_generate_invoice_serializes_full_size_run_multiplier(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            with override_settings(MEDIA_ROOT=tmpdir):
                payload = {
                    "companyData": {
                        "companyName": "ТОВ Тест",
                        "companyNumber": "12345678",
                        "contactPhone": "+380000000000",
                        "deliveryAddress": "Київ, вул. Тестова, 1",
                    },
                    "orderItems": [
                        {
                            "product": {"id": "1", "type": "tshirt", "title": "Футболка 1"},
                            "size": "all",
                            "runMultiplier": 2,
                            "include2xlInRun": True,
                            "quantity": 4,
                            "pricing_mode": "auto",
                        },
                    ],
                }

                response = self.client_http.post(
                    reverse("management_invoices_generate_api"),
                    data=json.dumps(payload),
                    content_type="application/json",
                    HTTP_HOST="management.twocomms.shop",
                    secure=True,
                )

                self.assertEqual(response.status_code, 200)
                data = response.json()
                self.assertTrue(data["ok"])

                invoice = WholesaleInvoice.objects.get(id=data["invoice"]["id"])
                self.assertEqual(invoice.total_tshirts, 10)

                saved_item = invoice.order_details["order_items"][0]
                self.assertEqual(saved_item["size"], "Всі ростовки (S-2XL) ×2")
                self.assertEqual(saved_item["run_multiplier"], 2)
                self.assertTrue(saved_item["include_2xl"])
                self.assertEqual(saved_item["quantity"], 10)

                wb = load_workbook(Path(invoice.file_path))
                ws = wb.active
                item_row = _find_row_by_value(ws, "Футболка 1")
                self.assertIsNotNone(item_row)
                self.assertEqual(ws.cell(row=item_row, column=3).value, "Всі ростовки (S-2XL) ×2")
                self.assertEqual(ws.cell(row=item_row, column=5).value, 10)

    def test_generate_invoice_returns_validation_error_for_malformed_order_item(self):
        payload = {
            "companyData": {
                "companyName": "ТОВ Тест",
                "contactPhone": "+380000000000",
                "deliveryAddress": "Київ, вул. Тестова, 1",
            },
            "orderItems": ["broken-item"],
        }

        response = self.client_http.post(
            reverse("management_invoices_generate_api"),
            data=json.dumps(payload),
            content_type="application/json",
            HTTP_HOST="management.twocomms.shop",
            secure=True,
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()["ok"])

    @patch.dict(os.environ, {"MANAGEMENT_TG_ADMIN_CHAT_ID": "111,222", "MANAGEMENT_TG_BOT_TOKEN": "bot-token"}, clear=False)
    @patch("management.views._tg_send_message")
    def test_submit_invoice_for_review_sends_to_all_admin_chats(self, send_message_mock):
        send_message_mock.side_effect = [
            {"message_id": 501},
            {"message_id": 777},
        ]
        invoice = WholesaleInvoice.objects.create(
            invoice_number="INV-MULTI-1",
            company_name="ТОВ Тест",
            company_number="123",
            contact_phone="+380000000000",
            delivery_address="Київ, вул. Тестова, 1",
            store_link="",
            total_tshirts=2,
            total_hoodies=0,
            total_amount=Decimal("1080.00"),
            status="draft",
            created_by=self.user,
            review_status="draft",
            order_details={"company_data": {"companyName": "ТОВ Тест"}, "order_items": []},
        )

        response = self.client_http.post(
            reverse("management_invoices_submit_for_review_api", args=[invoice.id]),
            HTTP_HOST="management.twocomms.shop",
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        invoice.refresh_from_db()
        self.assertEqual(send_message_mock.call_count, 2)
        self.assertEqual(send_message_mock.call_args_list[0].args[1], 111)
        self.assertEqual(send_message_mock.call_args_list[1].args[1], 222)
        self.assertEqual(invoice.admin_tg_chat_id, 111)
        self.assertEqual(invoice.admin_tg_message_id, 501)
        self.assertEqual(
            invoice.order_details["admin_tg_messages"],
            [
                {"chat_id": 111, "message_id": 501},
                {"chat_id": 222, "message_id": 777},
            ],
        )

    @patch.dict(os.environ, {"MANAGEMENT_TG_BOT_TOKEN": "bot-token"}, clear=False)
    @patch("management.views._tg_edit_message")
    @patch("management.views._tg_edit_caption")
    def test_update_invoice_status_syncs_all_admin_messages(self, edit_caption_mock, edit_message_mock):
        edit_caption_mock.return_value = {"message_id": 1}
        invoice = WholesaleInvoice.objects.create(
            invoice_number="INV-MULTI-2",
            company_name="ТОВ Тест",
            company_number="123",
            contact_phone="+380000000000",
            delivery_address="Київ, вул. Тестова, 1",
            store_link="",
            total_tshirts=2,
            total_hoodies=0,
            total_amount=Decimal("1080.00"),
            status="pending",
            created_by=self.user,
            review_status="approved",
            is_approved=True,
            admin_tg_chat_id=111,
            admin_tg_message_id=501,
            order_details={
                "company_data": {"companyName": "ТОВ Тест"},
                "order_items": [],
                "admin_tg_messages": [
                    {"chat_id": 111, "message_id": 501},
                    {"chat_id": 222, "message_id": 777},
                ],
            },
        )

        from management.views import _try_update_admin_invoice_message

        _try_update_admin_invoice_message(invoice, final=True)

        self.assertEqual(edit_caption_mock.call_count, 2)
        seen_pairs = {(call.args[1], call.args[2]) for call in edit_caption_mock.call_args_list}
        self.assertEqual(seen_pairs, {(111, 501), (222, 777)})
        self.assertEqual(edit_message_mock.call_count, 0)


class InvoiceSummaryTests(TestCase):
    def test_shop_invoice_summary_prefers_normalized_fields(self):
        invoice = WholesaleInvoice(
            invoice_number="INV-1",
            company_name="ТОВ Тест",
            total_tshirts=8,
            total_hoodies=0,
            total_amount=Decimal("4460.00"),
            order_details={
                "company_data": {"companyName": "ТОВ Тест"},
                "order_items": [
                    {
                        "product": {"title": "Футболка 1", "type": "tshirt"},
                        "display_title": "Футболка 1 [з кашкорсе]",
                        "size": "M",
                        "color": "Чорний",
                        "quantity": 6,
                        "unit_price": "540.00",
                        "line_total": "3240.00",
                    }
                ],
            },
        )

        summary = _invoice_summary_from_wholesale_invoice(invoice)

        self.assertEqual(summary["items"][0]["title"], "Футболка 1 [з кашкорсе]")
        self.assertEqual(summary["items"][0]["price"], "540.00")
        self.assertEqual(summary["items"][0]["total"], "3240.00")
