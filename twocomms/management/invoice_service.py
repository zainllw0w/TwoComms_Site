from __future__ import annotations

from copy import deepcopy
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_TWOPLACES = Decimal("0.01")
_FULL_SIZE_RUN_BASE_QUANTITY = 4

_WHOLESALE_PRICE_CONTEXT = {
    "tshirt": {
        "drop": 570,
        "wholesale": [540, 520, 500, 490, 480],
        "ranges": ["8–15 шт.", "16–31 шт.", "32–63 шт.", "64–99 шт.", "100+ шт."],
    },
    "hoodie": {
        "drop": 1350,
        "wholesale": [1300, 1250, 1200, 1175, 1150],
        "ranges": ["8–15 шт.", "16–31 шт.", "32–63 шт.", "64–99 шт.", "100+ шт."],
    },
}


class InvoicePayloadError(ValueError):
    pass


def get_management_wholesale_price_context() -> dict[str, dict[str, Any]]:
    return deepcopy(_WHOLESALE_PRICE_CONTEXT)


def _require_mapping(value: Any, error_message: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise InvoicePayloadError(error_message)
    return value


def _quantize_money(value: Decimal) -> Decimal:
    return value.quantize(_TWOPLACES, rounding=ROUND_HALF_UP)


def _parse_money(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return _quantize_money(value)
    if isinstance(value, int):
        return _quantize_money(Decimal(value))
    if isinstance(value, float):
        return _quantize_money(Decimal(str(value)))

    text = str(value).strip()
    if not text:
        return None
    text = (
        text.replace("грн", "")
        .replace("₴", "")
        .replace("\u00a0", "")
        .replace(" ", "")
        .replace(",", ".")
    )
    try:
        return _quantize_money(Decimal(text))
    except InvalidOperation:
        return None


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def _parse_quantity(value: Any) -> int:
    try:
        quantity = int(value)
    except (TypeError, ValueError):
        raise InvoicePayloadError("Некоректна кількість товару") from None
    if quantity <= 0:
        raise InvoicePayloadError("Кількість товару повинна бути більшою за нуль")
    return quantity


def _parse_run_multiplier(value: Any) -> int:
    if value in (None, ""):
        return 1
    try:
        multiplier = int(value)
    except (TypeError, ValueError):
        raise InvoicePayloadError("Некоректний множник ростовки") from None
    if multiplier <= 0:
        raise InvoicePayloadError("Множник ростовки повинен бути більшим за нуль")
    return multiplier


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _normalize_size_label(raw_size: Any, run_multiplier: int, include_2xl: bool) -> tuple[str, bool, int, bool]:
    cleaned_size = _clean_text(raw_size)
    normalized_size = cleaned_size.lower().replace("×", "x")
    is_full_size_run = normalized_size == "all" or ("s-xl" in normalized_size and "ростов" in normalized_size)
    if "s-2xl" in normalized_size and "ростов" in normalized_size:
        is_full_size_run = True
        include_2xl = True

    if not is_full_size_run:
        return cleaned_size, False, 1, False

    if run_multiplier == 1:
        multiplier_match = re.search(r"[x×]\s*(\d+)", normalized_size)
        if multiplier_match:
            run_multiplier = _parse_run_multiplier(multiplier_match.group(1))

    label = "Всі ростовки (S-2XL)" if include_2xl else "Всі ростовки (S-XL)"
    if run_multiplier > 1:
        label = f"{label} ×{run_multiplier}"
    return label, True, run_multiplier, include_2xl


def _guess_product_type(product_type: str, title: str) -> str:
    cleaned_type = _clean_text(product_type).lower()
    if cleaned_type in {"tshirt", "hoodie"}:
        return cleaned_type

    title_lower = _clean_text(title).lower()
    if any(token in title_lower for token in ("худі", "hoodie", "фліс", "флис", "fleece")):
        return "hoodie"
    if any(token in title_lower for token in ("футбол", "t-shirt", "tshirt", "tee")):
        return "tshirt"
    raise InvoicePayloadError("Не вдалося визначити тип товару для накладної")


def _wholesale_price_for_quantity(product_type: str, quantity: int) -> Decimal:
    prices = _WHOLESALE_PRICE_CONTEXT[product_type]["wholesale"]
    if quantity >= 100:
        return Decimal(str(prices[4]))
    if quantity >= 64:
        return Decimal(str(prices[3]))
    if quantity >= 32:
        return Decimal(str(prices[2]))
    if quantity >= 16:
        return Decimal(str(prices[1]))
    return Decimal(str(prices[0]))


def _entry_wholesale_price(product_type: str) -> Decimal:
    return Decimal(str(_WHOLESALE_PRICE_CONTEXT[product_type]["wholesale"][0]))


def _compose_display_title(title: str, extra_description: str) -> str:
    if not extra_description:
        return title
    return f"{title} [{extra_description}]"


def normalize_management_invoice_payload(*, company_data: dict[str, Any], order_items: list[dict[str, Any]]) -> dict[str, Any]:
    company_data = _require_mapping(company_data or {}, "Некоректні дані компанії")
    if not isinstance(order_items, list):
        raise InvoicePayloadError("Некоректний список товарів для накладної")

    normalized_company = {
        "companyName": _clean_text(company_data.get("companyName")),
        "companyNumber": _clean_text(company_data.get("companyNumber")),
        "contactPhone": _clean_text(company_data.get("contactPhone")),
        "deliveryAddress": _clean_text(company_data.get("deliveryAddress")),
        "storeLink": _clean_text(company_data.get("storeLink")),
    }

    required_fields = [
        ("companyName", "назву компанії"),
        ("contactPhone", "номер телефону"),
        ("deliveryAddress", "адресу доставки"),
    ]
    missing = [label for key, label in required_fields if not normalized_company[key]]
    if missing:
        raise InvoicePayloadError(f"Заповніть: {', '.join(missing)}")

    if not order_items:
        raise InvoicePayloadError("Немає товарів для накладної")

    prepared_items: list[dict[str, Any]] = []
    totals_by_type = {"tshirt": 0, "hoodie": 0}
    has_manual_prices = False

    for index, raw_item in enumerate(order_items, start=1):
        raw_item = _require_mapping(raw_item, f"Некоректні дані товару в позиції {index}")
        product = _require_mapping(raw_item.get("product") or {}, f"Некоректні дані товару в позиції {index}")
        raw_title = _clean_text(product.get("title"))
        if not raw_title:
            raise InvoicePayloadError("У позиції відсутня назва товару")

        product_type = _guess_product_type(product.get("type"), raw_title)
        run_multiplier = _parse_run_multiplier(raw_item.get("run_multiplier") or raw_item.get("runMultiplier"))
        include_2xl = _parse_bool(raw_item.get("include_2xl") or raw_item.get("include2xlInRun"))
        size_label, is_full_size_run, run_multiplier, include_2xl = _normalize_size_label(
            raw_item.get("size"),
            run_multiplier,
            include_2xl,
        )
        fit_label = _clean_text(raw_item.get("fit") or raw_item.get("fit_label") or raw_item.get("fitOptionLabel"))
        quantity = _parse_quantity(raw_item.get("quantity"))
        if is_full_size_run:
            quantity = (_FULL_SIZE_RUN_BASE_QUANTITY + (1 if include_2xl else 0)) * run_multiplier
        totals_by_type[product_type] += quantity

        pricing_mode = _clean_text(raw_item.get("pricing_mode") or raw_item.get("pricingMode")).lower()
        pricing_mode = "manual" if pricing_mode == "manual" else "auto"

        manual_price = _parse_money(raw_item.get("manual_price") or raw_item.get("manualPrice"))
        if pricing_mode == "manual":
            if manual_price is None or manual_price <= 0:
                raise InvoicePayloadError("Вкажіть коректну ручну ціну для всіх ручних позицій")
            has_manual_prices = True
        extra_description = _clean_text(
            raw_item.get("extra_description") or raw_item.get("extraDescription")
        )

        prepared_items.append(
            {
                "product": {
                    "id": product.get("id"),
                    "type": product_type,
                    "title": raw_title,
                    "image": product.get("image") or product.get("main_image") or "",
                },
                "size": size_label,
                "fit": fit_label,
                "color": _clean_text((raw_item or {}).get("color")) or "Чорний",
                "quantity": quantity,
                "run_multiplier": run_multiplier,
                "include_2xl": include_2xl,
                "pricing_mode": pricing_mode,
                "manual_price": manual_price,
                "extra_description": extra_description,
            }
        )

    normalized_items: list[dict[str, Any]] = []
    total_amount = Decimal("0.00")
    discounts_enabled = not has_manual_prices

    for item in prepared_items:
        product_type = item["product"]["type"]
        if discounts_enabled:
            base_unit_price = _quantize_money(_wholesale_price_for_quantity(product_type, totals_by_type[product_type]))
        else:
            base_unit_price = _quantize_money(_entry_wholesale_price(product_type))
        use_manual_price = item["pricing_mode"] == "manual" and item["manual_price"] is not None and item["manual_price"] > 0
        unit_price = _quantize_money(item["manual_price"] if use_manual_price else base_unit_price)
        line_total = _quantize_money(unit_price * Decimal(item["quantity"]))
        display_title = _compose_display_title(item["product"]["title"], item["extra_description"])

        normalized_items.append(
            {
                "product": dict(item["product"]),
                "size": item["size"],
                "fit": item.get("fit", ""),
                "color": item["color"],
                "quantity": item["quantity"],
                "run_multiplier": item["run_multiplier"],
                "include_2xl": item["include_2xl"],
                "pricing_mode": "manual" if use_manual_price else "auto",
                "extra_description": item["extra_description"],
                "display_title": display_title,
                "title": display_title,
                "base_unit_price": base_unit_price,
                "unit_price": unit_price,
                "line_total": line_total,
                "price": unit_price,
                "total": line_total,
            }
        )
        total_amount += line_total

    return {
        "company_data": normalized_company,
        "items": normalized_items,
        "pricing": {
            "policy": "custom_manual" if has_manual_prices else "tier_wholesale",
            "has_manual_prices": has_manual_prices,
            "discounts_enabled": discounts_enabled,
            "note": (
                "Ручна ціна присутня в замовленні: автоматичні знижки за кількість вимкнені."
                if has_manual_prices
                else "Автоматичні оптові tier-знижки застосовані за загальною кількістю."
            ),
        },
        "totals": {
            "total_tshirts": totals_by_type["tshirt"],
            "total_hoodies": totals_by_type["hoodie"],
            "total_amount": _quantize_money(total_amount),
        },
    }


def serialize_management_invoice_payload(normalized_payload: dict[str, Any]) -> dict[str, Any]:
    def _serialize_money(value: Decimal) -> str:
        return f"{_quantize_money(value):.2f}"

    serialized_items = []
    for item in normalized_payload["items"]:
        serialized_items.append(
            {
                "product": dict(item["product"]),
                "size": item["size"],
                "fit": item.get("fit", ""),
                "color": item["color"],
                "quantity": item["quantity"],
                "run_multiplier": item.get("run_multiplier", 1),
                "include_2xl": item.get("include_2xl", False),
                "pricing_mode": item["pricing_mode"],
                "extra_description": item["extra_description"],
                "display_title": item["display_title"],
                "title": item["title"],
                "base_unit_price": _serialize_money(item["base_unit_price"]),
                "unit_price": _serialize_money(item["unit_price"]),
                "line_total": _serialize_money(item["line_total"]),
                "price": _serialize_money(item["price"]),
                "total": _serialize_money(item["total"]),
            }
        )

    totals = normalized_payload["totals"]
    pricing = normalized_payload.get("pricing") or {}
    return {
        "company_data": dict(normalized_payload["company_data"]),
        "order_items": serialized_items,
        "pricing": {
            "policy": pricing.get("policy") or "tier_wholesale",
            "has_manual_prices": bool(pricing.get("has_manual_prices")),
            "discounts_enabled": bool(pricing.get("discounts_enabled", True)),
            "note": pricing.get("note") or "",
        },
        "totals": {
            "total_tshirts": totals["total_tshirts"],
            "total_hoodies": totals["total_hoodies"],
            "total_amount": _serialize_money(totals["total_amount"]),
        },
    }


def build_management_invoice_workbook(
    *,
    company_data: dict[str, Any],
    normalized_items: list[dict[str, Any]],
    pricing: dict[str, Any],
    totals: dict[str, Any],
    invoice_number: str,
    created_at_label: str,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Оптова накладна"

    title_fill = PatternFill("solid", fgColor="E8F4FD")
    section_fill = PatternFill("solid", fgColor="F5F9FF")
    company_fill = PatternFill("solid", fgColor="DCEEFF")
    header_fill = PatternFill("solid", fgColor="366092")
    zebra_fill = PatternFill("solid", fgColor="F8FBFE")
    auto_note_fill = PatternFill("solid", fgColor="EAF7EE")
    manual_note_fill = PatternFill("solid", fgColor="FFF1DE")
    summary_fill = PatternFill("solid", fgColor="E8F4FD")
    link_fill = PatternFill("solid", fgColor="EEF7FF")
    thin = Side(style="thin", color="C9D2DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(name="Arial", size=16, bold=True, color="366092")
    section_font = Font(name="Arial", size=11, bold=True, color="244061")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10, color="1F1F1F")
    emphasis_font = Font(name="Arial", size=10, bold=True, color="1F1F1F")
    company_font = Font(name="Arial", size=11, bold=True, color="366092")
    link_font = Font(name="Arial", size=10, underline="single", color="0563C1")

    center = Alignment(horizontal="center", vertical="center", wrapText=True)
    left = Alignment(horizontal="left", vertical="center", wrapText=True)
    right = Alignment(horizontal="right", vertical="center", wrapText=True)

    def _style_row(row_index: int, start_column: int = 1, end_column: int = 7, *, fill=None, font=None, alignment=None):
        for column in range(start_column, end_column + 1):
            cell = ws.cell(row=row_index, column=column)
            cell.border = border
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment

    def _merge_row(row_index: int, start_column: int, end_column: int, value, *, fill=None, font=None, alignment=None):
        start_letter = get_column_letter(start_column)
        end_letter = get_column_letter(end_column)
        ws.merge_cells(f"{start_letter}{row_index}:{end_letter}{row_index}")
        cell = ws.cell(row=row_index, column=start_column, value=value)
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font
        if alignment is not None:
            cell.alignment = alignment
        _style_row(row_index, start_column, end_column, fill=fill, font=font, alignment=alignment)
        return cell

    _merge_row(1, 1, 7, "ОПТОВА НАКЛАДНА", fill=title_fill, font=title_font, alignment=center)
    ws.row_dimensions[1].height = 26

    row = 3
    _merge_row(row, 1, 7, "Інформація про замовника", fill=section_fill, font=section_font, alignment=left)

    info_rows = [
        ("Назва компанії/ФОП/ПІБ", company_data.get("companyName") or "—", company_fill, company_font),
        ("Номер компанії/ЄДРПОУ/ІПН", company_data.get("companyNumber") or "—", None, body_font),
        ("Номер телефону", company_data.get("contactPhone") or "—", None, body_font),
        ("Адреса доставки", company_data.get("deliveryAddress") or "—", None, body_font),
    ]
    store_link = company_data.get("storeLink") or "—"
    if store_link not in {"", "—"}:
        info_rows.append(("Посилання на магазин", store_link, link_fill, link_font))

    for label, value, value_fill, value_font in info_rows:
        row += 1
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = body_font
        ws[f"A{row}"].alignment = left
        ws[f"A{row}"].border = border
        ws[f"A{row}"].fill = section_fill

        ws.merge_cells(f"B{row}:G{row}")
        value_cell = ws[f"B{row}"]
        value_cell.value = value
        value_cell.font = value_font
        value_cell.alignment = left
        value_cell.border = border
        if value_fill is not None:
            value_cell.fill = value_fill
        _style_row(row, 2, 7, fill=value_fill, font=value_font, alignment=left)
        if label == "Посилання на магазин" and value not in {"", "—"}:
            value_cell.hyperlink = value
        estimated_lines = max(1, len(str(value)) // 58 + 1)
        ws.row_dimensions[row].height = max(22, estimated_lines * 16)

    row += 2
    _merge_row(row, 1, 7, f"Номер накладної: {invoice_number}", fill=company_fill, font=section_font, alignment=left)
    ws.row_dimensions[row].height = 24

    row += 1
    _merge_row(row, 1, 7, f"Дата створення: {created_at_label}", fill=section_fill, font=body_font, alignment=left)
    ws.row_dimensions[row].height = 22

    row += 2
    pricing_note = pricing.get("note") or "Автоматичні оптові tier-знижки застосовані за загальною кількістю."
    pricing_fill = manual_note_fill if pricing.get("has_manual_prices") else auto_note_fill
    _merge_row(row, 1, 7, pricing_note, fill=pricing_fill, font=emphasis_font, alignment=left)
    ws.row_dimensions[row].height = max(24, (len(pricing_note) // 72 + 1) * 17)

    row += 2
    _merge_row(row, 1, 7, "Позиції накладної", fill=section_fill, font=section_font, alignment=left)

    header_row = row + 1
    headers = ["№", "Назва товару", "Розмір", "Колір", "Кількість", "Ціна за од.", "Сума"]
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=column, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    current_row = header_row + 1
    for index, item in enumerate(normalized_items, start=1):
        size_cell_value = item["size"]
        if item.get("fit"):
            size_cell_value = f'{size_cell_value} / {item["fit"]}'.strip()
        ws.cell(row=current_row, column=1, value=index)
        ws.cell(row=current_row, column=2, value=item["display_title"])
        ws.cell(row=current_row, column=3, value=size_cell_value)
        ws.cell(row=current_row, column=4, value=item["color"])
        ws.cell(row=current_row, column=5, value=item["quantity"])
        ws.cell(row=current_row, column=6, value=float(item["unit_price"]))
        ws.cell(row=current_row, column=7, value=float(item["line_total"]))

        row_fill = manual_note_fill if item.get("pricing_mode") == "manual" else (zebra_fill if index % 2 == 0 else None)
        for column in range(1, 8):
            cell = ws.cell(row=current_row, column=column)
            cell.font = body_font
            cell.border = border
            if column in {1, 3, 4, 5}:
                cell.alignment = center
            elif column in {6, 7}:
                cell.alignment = right
                cell.number_format = '#,##0.00 "₴"'
            else:
                cell.alignment = left
            if row_fill is not None:
                cell.fill = row_fill

        estimated_lines = max(
            1,
            len(str(item["display_title"])) // 34 + 1,
            len(str(size_cell_value)) // 20 + 1,
        )
        ws.row_dimensions[current_row].height = max(24, estimated_lines * 16)
        current_row += 1

    summary_row = current_row
    ws.merge_cells(f"A{summary_row}:F{summary_row}")
    ws[f"A{summary_row}"] = "РАЗОМ:"
    ws[f"A{summary_row}"].font = section_font
    ws[f"A{summary_row}"].fill = summary_fill
    ws[f"A{summary_row}"].alignment = right
    ws[f"A{summary_row}"].border = border
    _style_row(summary_row, 1, 6, fill=summary_fill, font=section_font, alignment=right)
    ws[f"G{summary_row}"] = float(totals["total_amount"])
    ws[f"G{summary_row}"].font = section_font
    ws[f"G{summary_row}"].fill = summary_fill
    ws[f"G{summary_row}"].alignment = right
    ws[f"G{summary_row}"].border = border
    ws[f"G{summary_row}"].number_format = '#,##0.00 "₴"'
    ws.row_dimensions[summary_row].height = 26

    stats_row = summary_row + 2
    _merge_row(stats_row, 1, 7, "Статистика замовлення", fill=section_fill, font=section_font, alignment=left)

    stats = [
        ("Футболки", f'{totals["total_tshirts"]} шт.'),
        ("Худі", f'{totals["total_hoodies"]} шт.'),
        ("Режим розрахунку", "Ручні ціни без оптового перерахунку" if pricing.get("has_manual_prices") else "Автоматичні tier-знижки"),
        ("Загальна сума", float(totals["total_amount"])),
    ]
    for offset, (label, value) in enumerate(stats, start=1):
        row_index = stats_row + offset
        ws[f"A{row_index}"] = label
        ws[f"A{row_index}"].font = body_font
        ws[f"A{row_index}"].alignment = left
        ws[f"A{row_index}"].border = border
        ws[f"A{row_index}"].fill = section_fill

        ws.merge_cells(f"B{row_index}:G{row_index}")
        value_cell = ws[f"B{row_index}"]
        value_cell.value = value
        value_font = emphasis_font if label == "Загальна сума" else body_font
        value_alignment = right if label == "Загальна сума" else left
        value_cell.font = value_font
        value_cell.alignment = value_alignment
        value_cell.border = border
        value_fill = summary_fill if label == "Загальна сума" else (pricing_fill if label == "Режим розрахунку" else None)
        if value_fill is not None:
            value_cell.fill = value_fill
        _style_row(row_index, 2, 7, fill=value_fill, font=value_font, alignment=value_alignment)
        if label == "Загальна сума":
            value_cell.number_format = '#,##0.00 "₴"'
        ws.row_dimensions[row_index].height = 22

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:G{summary_row - 1}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16

    for column in range(1, 8):
        letter = get_column_letter(column)
        if ws.column_dimensions[letter].width < 12:
            ws.column_dimensions[letter].width = 12

    return wb
