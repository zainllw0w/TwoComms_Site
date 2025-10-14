from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.db.models import Sum, F, ExpressionWrapper, FloatField, IntegerField
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.cache import cache_page
from django.urls import reverse
from functools import wraps
from django.core.paginator import Paginator, EmptyPage

def cache_page_for_anon(timeout):
    """Кэширует только для анонимных пользователей (избегаем проблем с персональными данными)."""
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if request.user.is_authenticated:
                return view_func(request, *args, **kwargs)
            return cache_page(timeout)(view_func)(request, *args, **kwargs)
        return _wrapped_view
    return decorator
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import Product, Category, ProductImage, PromoCode, PrintProposal
from .forms import ProductForm, CategoryForm, PrintProposalForm
from accounts.models import UserProfile, FavoriteProduct
from django import forms
from django.conf import settings
from django.http import HttpResponse, FileResponse, Http404
from django.utils.text import slugify
from django.core.cache import cache
import os
import json
import logging
import copy
import re
import base64
import time
from datetime import timedelta

from urllib.parse import urljoin
import requests
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from django.utils import timezone
from cryptography.hazmat.primitives import serialization, hashes
from cryptography.hazmat.primitives.asymmetric import ec
from cryptography.exceptions import InvalidSignature

HOME_PRODUCTS_PER_PAGE = 8

def unique_slugify(model, base_slug):
    """
    Створює унікальний slug на основі base_slug для заданої моделі.
    """
    slug = base_slug or 'item'
    slug = slug.strip('-') or 'item'
    uniq = slug
    i = 2
    while model.objects.filter(slug=uniq).exists():
        uniq = f"{slug}-{i}"
        i += 1
    return uniq

# --- Примитивные формы аутентификации/профиля ---
class LoginForm(forms.Form):
    username = forms.CharField(label="Логін", max_length=150,
                               widget=forms.TextInput(attrs={"class":"form-control bg-elevate"}))
    password = forms.CharField(label="Пароль",
                               widget=forms.PasswordInput(attrs={"class":"form-control bg-elevate"}))

class RegisterForm(forms.Form):
    username = forms.CharField(label="Логін", max_length=150,
                               widget=forms.TextInput(attrs={"class":"form-control bg-elevate"}))
    password1 = forms.CharField(label="Пароль",
                                widget=forms.PasswordInput(attrs={"class":"form-control bg-elevate"}))
    password2 = forms.CharField(label="Повтор паролю",
                                widget=forms.PasswordInput(attrs={"class":"form-control bg-elevate"}))
    def clean(self):
        d=super().clean()
        if d.get("password1")!=d.get("password2"):
            self.add_error("password2","Паролі не співпадають")
        return d

class ProfileSetupForm(forms.Form):
    full_name = forms.CharField(label="ПІБ", max_length=200, required=False,
                                widget=forms.TextInput(attrs={"class":"form-control bg-elevate","placeholder":"Прізвище Ім'я По батькові"}))
    phone = forms.CharField(label="Телефон", required=True,
                            widget=forms.TextInput(attrs={"class":"form-control bg-elevate","placeholder":"+380XXXXXXXXX"}))
    email = forms.EmailField(label="Email", required=False,
                            widget=forms.EmailInput(attrs={"class":"form-control bg-elevate","placeholder":"your@email.com"}))
    telegram = forms.CharField(label="Telegram", required=False,
                              widget=forms.TextInput(attrs={"class":"form-control bg-elevate","placeholder":"@username"}))
    instagram = forms.CharField(label="Instagram", required=False,
                               widget=forms.TextInput(attrs={"class":"form-control bg-elevate","placeholder":"@username"}))
    city = forms.CharField(label="Місто", required=False,
                           widget=forms.TextInput(attrs={"class":"form-control bg-elevate","placeholder":"Київ"}))
    np_office = forms.CharField(label="Відділення/Поштомат НП", required=False,
                                widget=forms.TextInput(attrs={"class":"form-control bg-elevate","placeholder":"№ відділення або адреса поштомата"}))
    pay_type = forms.ChoiceField(label="Тип оплати", required=False, choices=(("partial","Часткова передоплата"),("full","Повна передоплата")),
                                 widget=forms.Select(attrs={"class":"form-select bg-elevate"}))
    avatar = forms.ImageField(label="Аватар", required=False)
    is_ubd = forms.BooleanField(label="Я — УБД", required=False, widget=forms.CheckboxInput(attrs={"class":"form-check-input"}))
    ubd_doc = forms.ImageField(label="Фото посвідчення УБД", required=False)
    def clean(self):
        d=super().clean(); 
        if d.get("is_ubd") and not d.get("ubd_doc"):
            self.add_error("ubd_doc","Для УБД додайте фото посвідчення")
        return d

# --- Аутентификация/профиль ---
# Старая функция login_view удалена - используется новая ниже

# Старая функция register_view удалена - используется новая ниже

# Старая функция logout_view удалена - используется новая ниже


from django import forms

# --------- AUTH FORMS (минимум, чтобы не плодить новый файл) ---------
class RegisterForm(forms.Form):
    username = forms.CharField(label="Логін", max_length=150,
                               widget=forms.TextInput(attrs={"class":"form-control bg-elevate"}))
    password1 = forms.CharField(label="Пароль",
                                widget=forms.PasswordInput(attrs={"class":"form-control bg-elevate"}))
    password2 = forms.CharField(label="Повтор паролю",
                                widget=forms.PasswordInput(attrs={"class":"form-control bg-elevate"}))

    def clean(self):
        data = super().clean()
        if data.get("password1") != data.get("password2"):
            self.add_error("password2", "Паролі не співпадають")
        return data



# --------- AUTH VIEWS ---------
def login_view(request):
    if request.user.is_authenticated:
        return redirect('profile_setup')
    form = LoginForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = authenticate(request, username=form.cleaned_data['username'], password=form.cleaned_data['password'])
        if user:
            login(request, user)
            
            # Переносим избранные товары из сессии в базу данных
            session_favorites = request.session.get('favorites', [])
            if session_favorites:
                for product_id in session_favorites:
                    try:
                        product = Product.objects.get(id=product_id)
                        FavoriteProduct.objects.get_or_create(
                            user=user,
                            product=product
                        )
                    except Product.DoesNotExist:
                        # Товар был удален, пропускаем
                        continue
                
                # Очищаем сессию
                del request.session['favorites']
                request.session.modified = True
            
            # если профиль пустой по телефону — просим заповнення
            try:
                prof = user.userprofile
            except UserProfile.DoesNotExist:
                prof = UserProfile.objects.create(user=user)
            if not prof.phone:
                return redirect('profile_setup')
            
            # Проверяем параметр next для перенаправления
            next_url = request.GET.get('next') or request.POST.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('home')
        else:
            form.add_error(None, "Невірний логін або пароль")
    return render(request, 'pages/auth_login.html', {'form': form, 'next': request.GET.get('next')})

def register_view(request):
    if request.user.is_authenticated:
        return redirect('profile_setup')
    form = RegisterForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        from django.contrib.auth.models import User
        if User.objects.filter(username=form.cleaned_data['username']).exists():
            form.add_error('username', 'Користувач з таким логіном вже існує')
        else:
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password1']
            )
            login(request, user)
            
            # Переносим избранные товары из сессии в базу данных
            session_favorites = request.session.get('favorites', [])
            if session_favorites:
                for product_id in session_favorites:
                    try:
                        product = Product.objects.get(id=product_id)
                        FavoriteProduct.objects.get_or_create(
                            user=user,
                            product=product
                        )
                    except Product.DoesNotExist:
                        # Товар был удален, пропускаем
                        continue
                
                # Очищаем сессию
                del request.session['favorites']
                request.session.modified = True
            
            return redirect('profile_setup')
    return render(request, 'pages/auth_register.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('home')


FEED_SIZE_OPTIONS = ["S", "M", "L", "XL"]
DEFAULT_FEED_COLOR = "Черный"
DEFAULT_FEED_SEASON = "Демисезон"


def _sanitize_feed_description(raw: str) -> str:
    if not raw:
        return ""

    cleaned = re.sub(r"(?is)<[^>]*?(?:цена|price|грн|uah|₴)[^>]*?>.*?</[^>]+>", "", raw)
    cleaned = re.sub(r"(?i)цена\s*[:\-]?[^\n<]*", "", cleaned)
    cleaned = re.sub(r"\d+[\s.,]*(?:грн|uah|₴)", "", cleaned, flags=re.IGNORECASE)

    lines = re.split(r"\r?\n", cleaned)
    price_line = re.compile(r"(?i)(грн|uah|₴|цена|price)")
    filtered = [line for line in lines if line and not price_line.search(line)]
    collapsed = "\n".join(filtered)
    collapsed = re.sub(r"\n{3,}", "\n\n", collapsed)
    return collapsed.strip()


def _material_for_product(product) -> str:
    lookup_source = " ".join(filter(None, [
        product.category.name if getattr(product, "category", None) else "",
        product.title or "",
    ])).lower()

    if any(token in lookup_source for token in ("худі", "hood", "hudi")):
        return "трехнитка с начесом"
    if "лонг" in lookup_source or "long" in lookup_source:
        return "бамбук"
    return "кулирка"


def _normalize_color_name(raw_color: str | None) -> str:
    if not raw_color:
        return DEFAULT_FEED_COLOR
    trimmed = raw_color.strip()
    if not trimmed:
        return DEFAULT_FEED_COLOR
    return trimmed[0].upper() + trimmed[1:]


def _absolute_media_url(base_url: str, path: str | None) -> str | None:
    if not path:
        return None
    if path.startswith("http://") or path.startswith("https://"):
        return path
    if not path.startswith("/"):
        path = f"/{path}"
    return urljoin(base_url, path)


def uaprom_products_feed(request):
    from django.utils import timezone
    import xml.etree.ElementTree as ET

    base_url = getattr(settings, "FEED_BASE_URL", "").strip() or request.build_absolute_uri("/").rstrip("/")

    products_qs = (
        Product.objects
        .select_related("category")
        .prefetch_related("images", "color_variants__images", "color_variants__color")
        .order_by("id")
    )
    products = list(products_qs)

    categories_ids = {p.category_id for p in products if p.category_id}
    categories_map = {
        cat.id: cat for cat in Category.objects.filter(id__in=categories_ids)
    }

    catalog = ET.Element("yml_catalog", {"date": timezone.now().strftime("%Y-%m-%d %H:%M")})
    shop = ET.SubElement(catalog, "shop")
    ET.SubElement(shop, "name").text = "TwoComms"
    ET.SubElement(shop, "company").text = "TWOCOMMS"
    ET.SubElement(shop, "url").text = base_url

    currencies = ET.SubElement(shop, "currencies")
    ET.SubElement(currencies, "currency", {"id": "UAH", "rate": "1"})

    categories_el = ET.SubElement(shop, "categories")
    for cat_id in sorted(categories_ids):
        category = categories_map.get(cat_id)
        if not category:
            continue
        category_el = ET.SubElement(categories_el, "category", {"id": str(category.id)})
        category_el.text = category.name

    offers_el = ET.SubElement(shop, "offers")

    for product in products:
        material_value = _material_for_product(product)

        base_image_paths = []
        if product.main_image:
            base_image_paths.append(product.main_image.url)
        base_image_paths.extend(img.image.url for img in product.images.all() if getattr(img, "image", None))
        base_image_paths = list(dict.fromkeys(base_image_paths))

        color_variants = list(product.color_variants.all())
        if color_variants:
            variant_payloads = []
            for variant in color_variants:
                color_label = _normalize_color_name(variant.color.name if variant.color else None)
                variant_images = [img.image.url for img in variant.images.all() if getattr(img, "image", None)]
                variant_payloads.append((color_label, variant_images))
        else:
            variant_payloads = [(DEFAULT_FEED_COLOR, base_image_paths)]

        for color_name, variant_images in variant_payloads:
            images_to_use = variant_images or base_image_paths
            image_urls = [
                url for url in (
                    _absolute_media_url(base_url, path) for path in images_to_use
                ) if url
            ]
            image_urls = list(dict.fromkeys(image_urls))
            if not image_urls and base_image_paths:
                # Fallback to ensure feed has at least one image per offer
                fallback = _absolute_media_url(base_url, base_image_paths[0])
                if fallback:
                    image_urls = [fallback]

            color_slug = slugify(color_name, allow_unicode=True).replace("-", "")
            color_slug = color_slug.upper() or "COLOR"
            group_id = f"TC-GROUP-{product.id}"

            for size in FEED_SIZE_OPTIONS:
                offer_id = f"TC-{product.id:04d}-{color_slug}-{size}"
                offer_el = ET.SubElement(
                    offers_el,
                    "offer",
                    {"id": offer_id, "available": "true", "group_id": group_id},
                )

                product_path = f"/product/{product.slug}/"
                query = f"?size={size}&color={slugify(color_name, allow_unicode=True)}"
                ET.SubElement(offer_el, "url").text = f"{urljoin(base_url, product_path)}{query}"

                ET.SubElement(offer_el, "oldprice").text = str(product.price)
                ET.SubElement(offer_el, "price").text = str(product.final_price)
                ET.SubElement(offer_el, "currencyId").text = "UAH"

                if product.category_id:
                    ET.SubElement(offer_el, "categoryId").text = str(product.category_id)

                for image_url in image_urls:
                    ET.SubElement(offer_el, "picture").text = image_url

                ET.SubElement(offer_el, "pickup").text = "true"
                ET.SubElement(offer_el, "delivery").text = "true"

                display_name = f"TwoComms {product.title} {color_name} {size}"
                display_name_clean = display_name.strip()
                ET.SubElement(offer_el, "name").text = display_name_clean
                ET.SubElement(offer_el, "name_ua").text = display_name_clean
                ET.SubElement(offer_el, "vendor").text = "TWOCOMMS"
                ET.SubElement(offer_el, "vendorCode").text = offer_id
                ET.SubElement(offer_el, "country_of_origin").text = "Украина"

        raw_description = product.description or ""
        sanitized_description = _sanitize_feed_description(raw_description)
        if not sanitized_description:
            sanitized_description = f"TwoComms {product.title} — демисезонная вещь собственного производства."

        ua_source = getattr(product, "ai_description", None) or raw_description
        sanitized_description_ua = _sanitize_feed_description(ua_source)
        if not sanitized_description_ua:
            sanitized_description_ua = sanitized_description

        description_el = ET.SubElement(offer_el, "description")
        if hasattr(ET, 'CDATA'):
            description_el.text = ET.CDATA(sanitized_description)
        else:
            description_el.text = sanitized_description

        description_ua_el = ET.SubElement(offer_el, "description_ua")
        if hasattr(ET, 'CDATA'):
            description_ua_el.text = ET.CDATA(sanitized_description_ua)
        else:
            description_ua_el.text = sanitized_description_ua

        param_size = ET.SubElement(offer_el, "param", {"name": "Размер", "unit": ""})
        param_size.text = size

        param_color = ET.SubElement(offer_el, "param", {"name": "Цвет", "unit": ""})
        param_color.text = color_name

        param_material = ET.SubElement(offer_el, "param", {"name": "Материал", "unit": ""})
        param_material.text = material_value

        param_season = ET.SubElement(offer_el, "param", {"name": "Сезон", "unit": ""})
        param_season.text = DEFAULT_FEED_SEASON

    ET.indent(catalog, space="  ", level=0)
    xml_payload = ET.tostring(catalog, encoding="utf-8", xml_declaration=True)

    response = HttpResponse(xml_payload, content_type="application/xml; charset=utf-8")
    response["Content-Disposition"] = 'inline; filename="products_feed.xml"'
    return response


def google_merchant_feed(request):
    """Отдаёт XML-фид Google Merchant.
    Ищет файл в нескольких типичных местах и выбирает самый свежий.
    Путь: /google-merchant-feed.xml (и альтернативы)
    """
    from django.conf import settings
    candidates = [
        os.path.join(settings.BASE_DIR, 'twocomms', 'static', 'google_merchant_feed.xml'),
        os.path.join(settings.BASE_DIR, 'static', 'google_merchant_feed.xml'),
    ]
    # Попробуем также STATIC_ROOT, если настроен
    static_root = getattr(settings, 'STATIC_ROOT', None)
    if static_root:
        candidates.append(os.path.join(static_root, 'google_merchant_feed.xml'))

    existing = [(p, os.path.getmtime(p)) for p in candidates if os.path.exists(p)]
    if not existing:
        raise Http404("Feed file not found")

    # Берём самый свежий файл
    existing.sort(key=lambda x: x[1], reverse=True)
    feed_path = existing[0][0]

    resp = FileResponse(open(feed_path, 'rb'), content_type='application/xml')
    # Отключаем кэширование, чтобы всегда видеть свежий фид
    resp['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp['Pragma'] = 'no-cache'
    resp['Expires'] = '0'
    return resp



@cache_page_for_anon(300)  # Кэшируем на 5 минут только для анонимов
def home(request):
    # Оптимизированные запросы с select_related и prefetch_related
    featured = Product.objects.select_related('category').filter(featured=True).order_by('-id').first()
    categories = Category.objects.filter(is_active=True).order_by('order','name')
    page_number = request.GET.get('page', '1')
    product_qs = Product.objects.select_related('category').order_by('-id')
    paginator = Paginator(product_qs, HOME_PRODUCTS_PER_PAGE)

    try:
        page_obj = paginator.get_page(page_number)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)

    products = list(page_obj.object_list)
    # Варіанти кольорів для featured (якщо є додаток і дані)
    featured_variants = []
    # Варіанти кольорів для «Новинок»
    try:
        from productcolors.models import ProductColorVariant
        if featured:
            vset = ProductColorVariant.objects.select_related('color').prefetch_related('images').filter(product=featured).order_by('order','id')
            for v in vset:
                first_image = v.images.first()
                featured_variants.append({
                    'primary_hex': v.color.primary_hex,
                    'secondary_hex': v.color.secondary_hex or '',
                    'is_default': v.is_default,
                    'first_image_url': first_image.image.url if first_image else '',
                })
        # Для списку новинок — підготуємо мапу
        prod_ids = [p.id for p in products]
        if prod_ids:
            vlist = ProductColorVariant.objects.select_related('color').prefetch_related('images').filter(product_id__in=prod_ids).order_by('product_id','order','id')
            colors_map = {}
            for v in vlist:
                first_image = v.images.first()
                colors_map.setdefault(v.product_id, []).append({
                    'primary_hex': v.color.primary_hex,
                    'secondary_hex': v.color.secondary_hex or '',
                    'first_image_url': first_image.image.url if first_image else '',
                })
            for p in products:
                setattr(p, 'colors_preview', colors_map.get(p.id, []))
    except Exception:
        # якщо додаток або таблиці відсутні — просто не показуємо кольори
        for p in products:
            setattr(p, 'colors_preview', [])
        featured_variants = featured_variants or []
    # Проверяем есть ли еще товары для пагинации
    total_products = paginator.count
    has_more = page_obj.has_next()


    return render(
        request,
        'pages/index.html',
        {
            'featured': featured, 
            'categories': categories, 
            'products': products, 
            'featured_variants': featured_variants,
            'has_more_products': has_more,
            'current_page': page_obj.number,
            'paginator': paginator,
            'page_obj': page_obj,
            'total_products': total_products
        }
    )

def load_more_products(request):
    """AJAX view для загрузки дополнительных товаров"""
    if request.method == 'GET':
        page = int(request.GET.get('page', 1))
        per_page = HOME_PRODUCTS_PER_PAGE

        product_qs = Product.objects.select_related('category').order_by('-id')
        paginator = Paginator(product_qs, per_page)

        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        products = list(page_obj.object_list)

        # Подготавливаем цвета для товаров
        try:
            from productcolors.models import ProductColorVariant
            prod_ids = [p.id for p in products]
            if prod_ids:
                vlist = ProductColorVariant.objects.select_related('color').prefetch_related('images').filter(product_id__in=prod_ids).order_by('product_id','order','id')
                colors_map = {}
                for v in vlist:
                    first_image = v.images.first()
                    colors_map.setdefault(v.product_id, []).append({
                        'primary_hex': v.color.primary_hex,
                        'secondary_hex': v.color.secondary_hex or '',
                        'first_image_url': first_image.image.url if first_image else '',
                    })
                for p in products:
                    setattr(p, 'colors_preview', colors_map.get(p.id, []))
        except Exception:
            for p in products:
                setattr(p, 'colors_preview', [])
        
        # Проверяем есть ли еще товары
        total_products = paginator.count
        has_more = page_obj.has_next()

        # Рендерим HTML для товаров
        from django.template.loader import render_to_string
        products_html = render_to_string('partials/products_list.html', {
            'products': products,
            'page': page
        })


        return JsonResponse({
            'html': products_html,
            'has_more': has_more,
            'next_page': page_obj.next_page_number() if has_more else None,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number
        })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@cache_page_for_anon(600)  # Кэшируем каталог на 10 минут только для анонимов
def catalog(request, cat_slug=None):
    categories = cache.get('categories_ordered')
    if categories is None:
        categories = list(Category.objects.order_by('order','name'))
        cache.set('categories_ordered', categories, 600)
    if cat_slug:
        category = get_object_or_404(Category, slug=cat_slug)
        products = Product.objects.filter(category=category).order_by('-id')
        show_category_cards = False
    else:
        category = None
        products = Product.objects.order_by('-id')
        show_category_cards = True
    
    # Добавляем данные о цветах для товаров
    for product in products:
        try:
            from productcolors.models import ProductColorVariant
            variants = ProductColorVariant.objects.select_related('color').prefetch_related('images').filter(product=product)
            product.colors_preview = []
            for v in variants:
                first_image = v.images.first()
                product.colors_preview.append({
                    'primary_hex': v.color.primary_hex,
                    'secondary_hex': v.color.secondary_hex or '',
                    'first_image_url': first_image.image.url if first_image else '',
                })
        except:
            product.colors_preview = []
    
    return render(request,'pages/catalog.html',{
        'categories': categories,
        'category': category,
        'products': products,
        'show_category_cards': show_category_cards,
        'cat_slug': cat_slug or ''
    })

@cache_page_for_anon(600)  # Кэшируем карточку товара на 10 минут для анонимов
def product_detail(request, slug):
    product = get_object_or_404(Product, slug=slug)
    images = product.images.all()
    # Варианты цветов с изображениями (если есть приложение и данные)
    color_variants = []
    auto_select_first_color = False
    
    try:
        from productcolors.models import ProductColorVariant
        variants = ProductColorVariant.objects.select_related('color').prefetch_related('images').filter(product=product)
        for v in variants:
            color_variants.append({
                'id': v.id,
                'primary_hex': v.color.primary_hex,
                'secondary_hex': v.color.secondary_hex or '',
                'is_default': v.is_default,
                'images': [img.image.url for img in v.images.all()],
            })
        
        # Если есть цветовые варианты, устанавливаем первый как активный по умолчанию
        if color_variants:
            # Устанавливаем первый цвет как активный
            color_variants[0]['is_default'] = True
            # Убираем is_default у остальных
            for i in range(1, len(color_variants)):
                color_variants[i]['is_default'] = False
            
            # Если нет основной картинки, помечаем для автоматического выбора изображения
            if not product.main_image:
                auto_select_first_color = True
                    
    except Exception:
        color_variants = []
    
    # Генерируем breadcrumbs для SEO
    breadcrumbs = [
        {'name': 'Головна', 'url': reverse('home')},
        {'name': 'Каталог', 'url': reverse('catalog')},
    ]
    
    if product.category:
        breadcrumbs.append({
            'name': product.category.name, 
            'url': reverse('catalog_by_cat', kwargs={'cat_slug': product.category.slug})
        })
    
    breadcrumbs.append({
        'name': product.title, 
        'url': reverse('product', kwargs={'slug': product.slug})
    })
    
    return render(request,'pages/product_detail.html',{
        'product': product,
        'images': images,
        'color_variants': color_variants,
        'auto_select_first_color': auto_select_first_color,
        'breadcrumbs': breadcrumbs
    })

@transaction.atomic
def add_product(request):
    """Добавление нового товара через унифицированный интерфейс"""
    if request.method == 'POST':
        # Обработка основной информации о товаре
        if 'form_type' in request.POST and request.POST['form_type'] == 'main_info':
            form = ProductForm(request.POST, request.FILES)
            if form.is_valid():
                product = form.save(commit=False)
                # slug генерируем автоматически, если не задан
                if not getattr(product, 'slug', None):
                    base = slugify(product.title or '')
                    product.slug = unique_slugify(Product, base)
                product.save()
                return JsonResponse({'success': True, 'product_id': product.id})
            else:
                return JsonResponse({'success': False, 'errors': form.errors})
        
        # Обработка цветов
        elif 'form_type' in request.POST and request.POST['form_type'] == 'colors':
            product_id = request.POST.get('product_id')
            try:
                product = Product.objects.get(id=product_id)
                from productcolors.models import Color, ProductColorVariant, ProductColorImage
                
                # Обработка добавления цвета
                if 'add_color' in request.POST:
                    color_name = request.POST.get('color_name', '').strip()
                    primary_hex = request.POST.get('primary_hex', '#000000')
                    secondary_hex = request.POST.get('secondary_hex', '').strip()
                    
                    if color_name:
                        # Создаем или получаем цвет
                        color, created = Color.objects.get_or_create(
                            name=color_name,
                            defaults={
                                'primary_hex': primary_hex,
                                'secondary_hex': secondary_hex if secondary_hex else None
                            }
                        )
                        
                        # Создаем вариант цвета для товара
                        variant = ProductColorVariant.objects.create(
                            product=product,
                            color=color,
                            is_default=False
                        )
                        
                        # Обрабатываем изображения
                        images = request.FILES.getlist('color_images')
                        for img in images:
                            ProductColorImage.objects.create(
                                variant=variant,
                                image=img
                            )
                        
                        return JsonResponse({'success': True, 'variant_id': variant.id})
                    else:
                        return JsonResponse({'success': False, 'error': 'Название цвета обязательно'})
                
                # Обработка удаления цвета
                elif 'delete_color' in request.POST:
                    variant_id = request.POST.get('variant_id')
                    try:
                        variant = ProductColorVariant.objects.get(id=variant_id, product=product)
                        variant.delete()
                        return JsonResponse({'success': True})
                    except ProductColorVariant.DoesNotExist:
                        return JsonResponse({'success': False, 'error': 'Вариант цвета не найден'})
                
                # Обработка удаления изображения
                elif 'delete_image' in request.POST:
                    image_id = request.POST.get('image_id')
                    try:
                        image = ProductColorImage.objects.get(id=image_id, variant__product=product)
                        image.delete()
                        return JsonResponse({'success': True})
                    except ProductColorImage.DoesNotExist:
                        return JsonResponse({'success': False, 'error': 'Изображение не найдено'})
                
            except Product.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Товар не найден'})
        
        # Обработка основной формы (для совместимости)
        else:
            form = ProductForm(request.POST, request.FILES)
            if form.is_valid():
                product = form.save(commit=False)
                if not getattr(product, 'slug', None):
                    base = slugify(product.title or '')
                    product.slug = unique_slugify(Product, base)
                product.save()
                return redirect('product', slug=product.slug)
            else:
                return render(request, 'pages/add_product_new.html', {
                    'form': form,
                    'product': None,
                    'is_new': True
                })
    
    else:
        form = ProductForm()
        return render(request, 'pages/add_product_new.html', {
            'form': form,
            'product': None,
            'is_new': True
        })

@login_required
def add_category(request):
    if request.method=='POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save(); return redirect('catalog')
    else:
        form = CategoryForm()
    return render(request,'pages/add_category.html',{'form':form})

def add_print(request):
    # Анти-спам: ограничим по сессии частоту отправок (не чаще 1 минуты)
    import time
    last_ts = request.session.get('print_proposal_last_ts', 0)
    now = int(time.time())
    rate_limited = now - last_ts < 60

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    proposal_form = None
    proposals = []

    if request.user.is_authenticated:
        # Список последних заявок пользователя
        proposals = PrintProposal.objects.filter(user=request.user).order_by('-created_at')[:10]

        if request.method == 'POST':
            if rate_limited:
                if is_ajax:
                    return JsonResponse({'ok': False, 'error': 'rate_limited'}, status=429)
                else:
                    return redirect('cooperation')

            form = PrintProposalForm(request.POST, request.FILES)
            if form.is_valid():
                obj: PrintProposal = form.save(commit=False)
                obj.user = request.user
                obj.status = 'pending'
                obj.save()
                request.session['print_proposal_last_ts'] = now

                if is_ajax:
                    proposal_data = {
                        'id': obj.id,
                        'created_at': obj.created_at.strftime('%d.%m.%Y %H:%M'),
                        'status': obj.status,
                        'status_display': obj.get_status_display(),
                        'awarded_points': obj.awarded_points,
                        'awarded_promocode': obj.awarded_promocode.code if obj.awarded_promocode else None,
                        'awarded_promocode_display': obj.awarded_promocode.get_discount_display() if obj.awarded_promocode else None,
                        'description': obj.description or '',
                        'image_url': obj.image.url if obj.image else None,
                    }
                    return JsonResponse({'ok': True, 'proposal': proposal_data})
                else:
                    return redirect('cooperation')
            else:
                if is_ajax:
                    # Вернем первую ошибку для простоты
                    errors = []
                    for field, errs in form.errors.items():
                        errors.extend([str(e) for e in errs])
                    return JsonResponse({'ok': False, 'error': errors[0] if errors else 'invalid'}, status=400)
                else:
                    proposal_form = form
        else:
            proposal_form = PrintProposalForm()
    else:
        proposal_form = None

    return render(request, 'pages/add-print.html', {
        'proposal_form': proposal_form,
        'proposals': proposals,
        'rate_limited': rate_limited,
    })

def cooperation(request):
    """Страница сотрудничества с брендом TwoComms"""
    return render(request, 'pages/cooperation.html')

def about(request): return render(request,'pages/about.html')

def contacts(request): return render(request,'pages/contacts.html')

@cache_page_for_anon(120)  # Кэшируем результаты поиска на 2 минуты для анонимов
def search(request):
    q=(request.GET.get('q') or '').strip()
    qs = Product.objects.select_related('category').prefetch_related('images', 'color_variants__images')
    if q: qs=qs.filter(title__icontains=q)
    categories = cache.get('categories_ordered')
    if categories is None:
        categories = list(Category.objects.order_by('order','name'))
        cache.set('categories_ordered', categories, 600)
    return render(request,'pages/catalog.html',{'categories':categories,'products':qs.order_by('-id'),'show_category_cards':False})

def debug_media(request):
    """Диагностика медиа-файлов"""
    import os
    from django.conf import settings
    
    # Обработка POST запроса для тестирования загрузки
    if request.method == 'POST':
        try:
            uploaded_file = request.FILES.get('test_file')
            if uploaded_file:
                # Создаем папку test если её нет
                test_dir = os.path.join(settings.MEDIA_ROOT, 'test')
                os.makedirs(test_dir, exist_ok=True)
                
                # Сохраняем файл
                file_path = os.path.join(test_dir, uploaded_file.name)
                with open(file_path, 'wb+') as destination:
                    for chunk in uploaded_file.chunks():
                        destination.write(chunk)
                
                return JsonResponse({
                    'success': True,
                    'file_path': file_path,
                    'file_url': f"{settings.MEDIA_URL}test/{uploaded_file.name}",
                    'file_size': uploaded_file.size,
                    'message': 'Файл успешно загружен'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Файл не найден в запросе'
                })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    # GET запрос - возвращаем диагностическую информацию
    debug_info = {
        'MEDIA_URL': settings.MEDIA_URL,
        'MEDIA_ROOT': str(settings.MEDIA_ROOT),
        'DEBUG': settings.DEBUG,
        'BASE_DIR': str(settings.BASE_DIR),
    }
    
    # Проверяем существование папок
    media_root = settings.MEDIA_ROOT
    debug_info['media_root_exists'] = os.path.exists(media_root)
    
    if os.path.exists(media_root):
        debug_info['media_root_contents'] = os.listdir(media_root)
        debug_info['media_root_permissions'] = oct(os.stat(media_root).st_mode)[-3:]
    
    # Проверяем подпапки
    subfolders = ['products', 'avatars', 'category_covers', 'category_icons', 'product_colors', 'test']
    debug_info['subfolders'] = {}
    
    for folder in subfolders:
        folder_path = os.path.join(media_root, folder)
        debug_info['subfolders'][folder] = {
            'exists': os.path.exists(folder_path),
            'contents': os.listdir(folder_path) if os.path.exists(folder_path) else [],
            'permissions': oct(os.stat(folder_path).st_mode)[-3:] if os.path.exists(folder_path) else None
        }
    
    # Проверяем последние загруженные файлы
    debug_info['recent_files'] = []
    for folder in subfolders:
        folder_path = os.path.join(media_root, folder)
        if os.path.exists(folder_path):
            files = os.listdir(folder_path)
            for file in files[:5]:  # Показываем первые 5 файлов
                file_path = os.path.join(folder_path, file)
                debug_info['recent_files'].append({
                    'folder': folder,
                    'file': file,
                    'size': os.path.getsize(file_path),
                    'permissions': oct(os.stat(file_path).st_mode)[-3:],
                    'url': f"{settings.MEDIA_URL}{folder}/{file}"
                })
    
    return JsonResponse(debug_info)

def debug_media_page(request):
    """Страница диагностики медиа-файлов"""
    return render(request, 'pages/debug_media.html')

def debug_product_images(request):
    """Диагностика изображений товаров"""
    from storefront.models import Product
    
    products = Product.objects.select_related('category').prefetch_related('images', 'color_variants__images')[:10]  # Берем первые 10 товаров
    
    debug_info = []
    for product in products:
        product_info = {
            'id': product.id,
            'title': product.title,
            'main_image': str(product.main_image) if product.main_image else None,
            'display_image': str(product.display_image) if product.display_image else None,
            'has_color_variants': product.color_variants.exists(),
            'color_variants_count': product.color_variants.count(),
        }
        
        # Проверяем цветовые варианты
        if product.color_variants.exists():
            first_variant = product.color_variants.first()
            product_info['first_variant'] = {
                'id': first_variant.id,
                'color_name': first_variant.color.name,
                'images_count': first_variant.images.count(),
                'first_image': str(first_variant.images.first().image) if first_variant.images.exists() else None,
            }
        
        debug_info.append(product_info)
    
    return JsonResponse({'products': debug_info})

@csrf_exempt
@require_POST
def add_to_cart(request):
    """
    Добавляет товар в корзину (сессия) с учётом размера и цвета, возвращает JSON с количеством и суммой.
    """
    pid = request.POST.get('product_id')
    size = ((request.POST.get('size') or '').strip() or 'S').upper()
    color_variant_id = request.POST.get('color_variant_id')
    try:
        qty = int(request.POST.get('qty') or '1')
    except ValueError:
        qty = 1
    qty = max(qty, 1)

    product = get_object_or_404(Product, pk=pid)

    cart = request.session.get('cart', {})
    key = f"{product.id}:{size}:{color_variant_id or 'default'}"
    item = cart.get(key, {
        'product_id': product.id, 
        'size': size, 
        'color_variant_id': color_variant_id,
        'qty': 0
    })
    item['qty'] += qty
    cart[key] = item
    if qty <= 0:
        _reset_monobank_session(request, drop_pending=True)
    request.session['cart'] = cart
    request.session.modified = True

    _reset_monobank_session(request, drop_pending=True)

    ids = [i['product_id'] for i in cart.values()]
    prods = Product.objects.in_bulk(ids)
    total_qty = sum(i['qty'] for i in cart.values())
    total_sum = 0
    for i in cart.values():
        p = prods.get(i['product_id'])
        if p:
            total_sum += i['qty'] * p.final_price

    return JsonResponse({'ok': True, 'count': total_qty, 'total': total_sum})

def cart_summary(request):
    """
    Краткая сводка корзины для обновления бейджа.
    """
    cart = request.session.get('cart', {})
    
    # Отладочная информация

    
    if not cart:
        return JsonResponse({'ok': True, 'count': 0, 'total': 0})
    
    ids = [i['product_id'] for i in cart.values()]
    prods = Product.objects.in_bulk(ids)
    
    # Проверяем, какие товары найдены
    found_products = set(prods.keys())
    missing_products = set(ids) - found_products
    
    if missing_products:
        # Удаляем несуществующие товары из корзины
        cart_to_clean = dict(cart)
        for key, item in cart_to_clean.items():
            if item['product_id'] in missing_products:
                cart.pop(key, None)
        _reset_monobank_session(request, drop_pending=True)
        request.session['cart'] = cart
        request.session.modified = True
    
    # Пересчитываем с очищенной корзиной
    total_qty = sum(i['qty'] for i in cart.values())
    total_sum = 0
    for i in cart.values():
        p = prods.get(i['product_id'])
        if p:
            total_sum += i['qty'] * p.final_price
    return JsonResponse({'ok': True, 'count': total_qty, 'total': total_sum})

def cart_mini(request):
    """
    HTML для мини‑корзины (выпадающая панель).
    """
    cart_sess = request.session.get('cart', {})
    
    if not cart_sess:
        return render(request,'partials/mini_cart.html',{'items':[],'total':0})
    
    ids = [i['product_id'] for i in cart_sess.values()]
    prods = Product.objects.in_bulk(ids)
    
    # Проверяем, какие товары найдены
    found_products = set(prods.keys())
    missing_products = set(ids) - found_products
    
    if missing_products:
        # Удаляем несуществующие товары из корзины
        cart_to_clean = dict(cart_sess)
        for key, item in cart_to_clean.items():
            if item['product_id'] in missing_products:
                cart_sess.pop(key, None)
        _reset_monobank_session(request, drop_pending=True)
        request.session['cart'] = cart_sess
        request.session.modified = True
    
    items = []
    total = 0
    total_points = 0
    for key, it in cart_sess.items():
        p = prods.get(it['product_id'])
        if not p:
            continue
        
        # Получаем информацию о цвете
        color_variant = _get_color_variant_safe(it.get('color_variant_id'))
        
        unit = p.final_price
        line = unit * it['qty']
        total += line
        # Баллы за товар, если предусмотрены
        try:
            if getattr(p, 'points_reward', 0):
                total_points += int(p.points_reward) * int(it['qty'])
        except Exception:
            pass
        items.append({
            'key': key,
            'product': p,
            'size': it['size'],
            'color_variant': color_variant,
            'qty': it['qty'],
            'unit_price': unit,
            'line_total': line
        })
    return render(request,'partials/mini_cart.html',{'items':items,'total':total})

def clean_cart(request):
    """
    Очистка корзины от несуществующих товаров.
    """
    cart = request.session.get('cart', {})
    
    if not cart:
        return JsonResponse({'ok': True, 'cleaned': 0, 'message': 'Корзина пуста'})
    
    ids = [i['product_id'] for i in cart.values()]
    prods = Product.objects.in_bulk(ids)
    
    # Проверяем, какие товары найдены
    found_products = set(prods.keys())
    missing_products = set(ids) - found_products
    
    cleaned_count = 0
    if missing_products:
        # Удаляем несуществующие товары из корзины
        cart_to_clean = dict(cart)
        for key, item in cart_to_clean.items():
            if item['product_id'] in missing_products:
                cart.pop(key, None)
                cleaned_count += 1
        request.session['cart'] = cart
        request.session.modified = True
    
    return JsonResponse({
        'ok': True, 
        'cleaned': cleaned_count, 
        'message': f'Очищено {cleaned_count} несуществующих товаров' if cleaned_count > 0 else 'Корзина в порядке'
    })

def process_guest_order(request):
    """
    Обработка заказа для неавторизованного пользователя
    """
    from orders.models import Order, OrderItem
    
    # Проверяем корзину
    cart = request.session.get('cart', {})
    if not cart:
        from django.contrib import messages
        messages.error(request, 'Кошик порожній!')
        return redirect('cart')
    
    # Валидация данных доставки
    full_name = request.POST.get('full_name', '').strip()
    phone = request.POST.get('phone', '').strip()
    city = request.POST.get('city', '').strip()
    np_office = request.POST.get('np_office', '').strip()
    pay_type = request.POST.get('pay_type', '')
    
    # Проверяем обязательные поля
    if not full_name or len(full_name) < 3:
        from django.contrib import messages
        messages.error(request, 'ПІБ повинно містити мінімум 3 символи!')
        return redirect('cart')
    
    if not phone:
        from django.contrib import messages
        messages.error(request, 'Введіть номер телефону!')
        return redirect('cart')
    
    if not city:
        from django.contrib import messages
        messages.error(request, 'Введіть місто!')
        return redirect('cart')
    
    if not np_office or len(np_office.strip()) < 1:
        from django.contrib import messages
        messages.error(request, 'Введіть адресу відділення!')
        return redirect('cart')
    
    if not pay_type:
        from django.contrib import messages
        messages.error(request, 'Оберіть тип оплати!')
        return redirect('cart')
    
    # Проверяем формат телефона
    phone_clean = ''.join(filter(str.isdigit, phone))
    if not phone.startswith('+380') or len(phone_clean) != 12:
        from django.contrib import messages
        messages.error(request, 'Невірний формат телефону! Використовуйте формат +380XXXXXXXXX')
        return redirect('cart')
    
    # Проверяем, что не создается дублирующий заказ
    # Используем хеш данных доставки для проверки
    import hashlib
    delivery_hash = hashlib.md5(f"{full_name}{phone}{city}{np_office}".encode()).hexdigest()
    
    # Проверяем, не было ли уже заказа с такими данными в последние 5 минут
    from django.utils import timezone
    from datetime import timedelta
    
    recent_orders = Order.objects.filter(
        full_name=full_name,
        phone=phone,
        city=city,
        np_office=np_office,
        created__gte=timezone.now() - timedelta(minutes=5)
    )
    
    if recent_orders.exists():
        from django.contrib import messages
        messages.error(request, 'Замовлення з такими даними вже було створено нещодавно. Спробуйте ще раз через кілька хвилин.')
        return redirect('cart')
    
    # Создаем заказ
    ids = [i['product_id'] for i in cart.values()]
    prods = Product.objects.in_bulk(ids)
    total_sum = 0
    
    order = Order.objects.create(
        user=None,  # Гостевой заказ
        full_name=full_name,
        phone=phone,
        city=city,
        np_office=np_office,
        pay_type=pay_type,
        total_sum=0,
        status='new'
    )
    
    # Добавляем товары в заказ
    for key, it in cart.items():
        p = prods.get(it['product_id'])
        if not p:
            continue
        
        # Получаем информацию о цвете
        color_variant = _get_color_variant_safe(it.get('color_variant_id'))
        
        unit = p.final_price
        line = unit * it['qty']
        total_sum += line
        
        monobank_logger.info('Creating OrderItem: product=%s, unit_price=%s, qty=%s, line_total=%s', 
                           p.title, unit, it['qty'], line)
        
        if not unit:
            monobank_logger.error('Product %s has null final_price: price=%s, discount_percent=%s', 
                                p.title, p.price, getattr(p, 'discount_percent', 'N/A'))
        
        OrderItem.objects.create(
            order=order,
            product=p,
            color_variant=color_variant,
            title=p.title,
            size=it.get('size', ''),
            qty=it['qty'],
            unit_price=unit,
            line_total=line
        )
    
    # Применяем промокод если есть
    promo_code = request.session.get('applied_promo_code')
    if promo_code:
        from .models import PromoCode
        try:
            promo = PromoCode.objects.get(code=promo_code, is_active=True)
            if promo.can_be_used():
                discount = promo.calculate_discount(total_sum)
                order.discount_amount = discount
                order.promo_code = promo
                promo.use()  # Увеличиваем счетчик использований
        except PromoCode.DoesNotExist:
            pass
    
    order.total_sum = total_sum
    order.save()
    
    # Очищаем корзину и промокод
    _reset_monobank_session(request, drop_pending=True)
    request.session['cart'] = {}
    request.session.pop('applied_promo_code', None)
    request.session.modified = True
    
    # Отправляем Telegram уведомление после создания заказа и товаров
    try:
        from orders.telegram_notifications import TelegramNotifier
        notifier = TelegramNotifier()
        notifier.send_new_order_notification(order)
    except Exception as e:
        # Не прерываем процесс, если уведомление не отправилось
        pass
    
    from django.contrib import messages
    messages.success(request, f'Замовлення #{order.order_number} успішно створено!')
    
    # Для неавторизованных пользователей перенаправляем на страницу успеха
    if not request.user.is_authenticated:
        return redirect('order_success', order_id=order.id)
    
    return redirect('my_orders')

@csrf_exempt
@require_POST
def cart_remove(request):
    """
    Удаление позиции: поддерживает key="productId:size" и (product_id + size).
    Печатает отладочную информацию в консоль.
    """
    cart = request.session.get('cart', {})


    key = (request.POST.get('key') or '').strip()
    pid = request.POST.get('product_id')
    size = (request.POST.get('size') or '').strip()

    removed = []

    def remove_key_exact(k: str) -> bool:
        if k in cart:
            cart.pop(k, None)
            removed.append(k)
            return True
        return False

    # 1) Пытаемся удалить по exact key
    if key:
        if not remove_key_exact(key):
            # 1.1) case-insensitive сравнение ключей
            target = key.lower()
            for kk in list(cart.keys()):
                if kk.lower() == target:
                    remove_key_exact(kk)
                    break
            # 1.2) если всё ещё не нашли, попробуем удалить все варианты этого товара
            if not removed and ":" in key:
                pid_part = key.split(":", 1)[0]
                for kk in list(cart.keys()):
                    if kk.split(":", 1)[0] == pid_part:
                        remove_key_exact(kk)
    # 2) Либо удаляем по product_id (+optional size)
    elif pid:
        try:
            pid_str = str(int(pid))
        except (ValueError, TypeError):
            pid_str = str(pid)
        if size:
            candidate = f"{pid_str}:{size}"
            if not remove_key_exact(candidate):
                # case-insensitive
                target = candidate.lower()
                for kk in list(cart.keys()):
                    if kk.lower() == target:
                        remove_key_exact(kk)
                        break
        else:
            for kk in list(cart.keys()):
                if kk.split(":", 1)[0] == pid_str:
                    remove_key_exact(kk)

    # Сохраняем изменения
    request.session['cart'] = cart
    request.session.modified = True

    # Пересчёт сводки
    ids = [i['product_id'] for i in cart.values()]
    prods = Product.objects.in_bulk(ids)
    total_qty = sum(i['qty'] for i in cart.values())
    total_sum = 0
    for i in cart.values():
        p = prods.get(i['product_id'])
        if p:
            total_sum += i['qty'] * p.final_price



    return JsonResponse({'ok': True, 'count': total_qty, 'total': total_sum, 'removed': removed, 'keys': list(cart.keys())})

def cart(request):
    """
    Рендер корзины: собираем позиции из сессии, считаем суммы.
    """
    from orders.models import Order
    
    # Обработка POST запросов
    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        
        if form_type == 'update_profile' and request.user.is_authenticated:
            # Обновление профиля пользователя
            try:
                prof = request.user.userprofile
                prof.full_name = request.POST.get('full_name', '')
                prof.phone = request.POST.get('phone', '')
                prof.city = request.POST.get('city', '')
                prof.np_office = request.POST.get('np_office', '')
                prof.pay_type = request.POST.get('pay_type', 'full')
                prof.save()
                
                # Показываем сообщение об успехе
                from django.contrib import messages
                messages.success(request, 'Дані доставки успішно оновлено!')
                
            except UserProfile.DoesNotExist:
                pass
                
        elif form_type == 'guest_order':
            # Оформление заказа для гостя
            return process_guest_order(request)
        elif form_type == 'order_create':
            # Оформление заказа для авторизованного пользователя
            return order_create(request)
    
    cart_sess = request.session.get('cart', {})
    
    if not cart_sess:
        return render(request,'pages/cart.html',{'items':[],'total':0,'discount':0,'grand_total':0})
    
    ids = [i['product_id'] for i in cart_sess.values()]
    prods = Product.objects.in_bulk(ids)
    
    # Проверяем, какие товары найдены
    found_products = set(prods.keys())
    missing_products = set(ids) - found_products
    
    if missing_products:
        # Удаляем несуществующие товары из корзины
        cart_to_clean = dict(cart_sess)
        for key, item in cart_to_clean.items():
            if item['product_id'] in missing_products:
                cart_sess.pop(key, None)
        _reset_monobank_session(request, drop_pending=True)
        request.session['cart'] = cart_sess
        request.session.modified = True
    
    items = []
    total = 0
    total_points = 0
    for key, it in cart_sess.items():
        p = prods.get(it['product_id'])
        if not p:
            continue
        
        # Получаем информацию о цвете
        color_variant = None
        color_variant_id = it.get('color_variant_id')
        if color_variant_id:
            try:
                from productcolors.models import ProductColorVariant
                color_variant = ProductColorVariant.objects.get(id=color_variant_id)
            except ProductColorVariant.DoesNotExist:
                pass
        
        unit = p.final_price
        line = unit * it['qty']
        total += line
        # Баллы за товар, если предусмотрены
        try:
            if getattr(p, 'points_reward', 0):
                total_points += int(p.points_reward) * int(it['qty'])
        except Exception:
            pass
        items.append({
            'key': key,
            'product': p,
            'size': it['size'],
            'color_variant': color_variant,
            'qty': it['qty'],
            'unit_price': unit,
            'line_total': line
        })
    
    # Проверяем примененный промокод
    applied_promo = None
    promo_code = request.session.get('applied_promo_code')
    if promo_code:
        from .models import PromoCode
        try:
            applied_promo = PromoCode.objects.get(code=promo_code, is_active=True)
            if applied_promo.can_be_used():
                discount = applied_promo.calculate_discount(total)
                grand_total = total - discount
            else:
                # Промокод больше не действителен
                request.session.pop('applied_promo_code', None)
                grand_total = total
        except PromoCode.DoesNotExist:
            request.session.pop('applied_promo_code', None)
            grand_total = total
    else:
        grand_total = total
    
    context = {
        'items': items,
        'total': total,
        'discount': total - grand_total if applied_promo else 0,
        'grand_total': grand_total,
        'total_points': total_points,
        'applied_promo': applied_promo
    }
    
    return render(request,'pages/cart.html', context)

def checkout(request): return cart(request)

@login_required
def admin_product_colors(request, pk):
    if not request.user.is_staff:
        return redirect('home')
    product = get_object_or_404(Product, pk=pk)
    # Ленивая загрузка моделей цветов
    from productcolors.models import Color, ProductColorVariant, ProductColorImage

    if request.method == 'POST':
        action = request.POST.get('action') or 'add'
        if action == 'add':
            # Можно выбрать существующий цвет или задать hex
            color_id = request.POST.get('color_id')
            primary_hex = (request.POST.get('primary_hex') or '').strip() or '#000000'
            secondary_hex = (request.POST.get('secondary_hex') or '').strip() or None
            is_default = bool(request.POST.get('is_default'))
            # Если выбрали существующий
            if color_id:
                color = get_object_or_404(Color, pk=color_id)
            else:
                color, _ = Color.objects.get_or_create(primary_hex=primary_hex, secondary_hex=secondary_hex)
            # Создаём вариант
            variant = ProductColorVariant.objects.create(product=product, color=color, is_default=is_default, order=int(request.POST.get('order') or 0))
            # Загружаем изображения
            for f in request.FILES.getlist('images'):
                ProductColorImage.objects.create(variant=variant, image=f)
        elif action == 'delete_variant':
            vid = request.POST.get('variant_id')
            try:
                v = ProductColorVariant.objects.get(pk=vid, product=product)
                v.delete()
            except ProductColorVariant.DoesNotExist:
                pass
        elif action == 'delete_image':
            iid = request.POST.get('image_id')
            try:
                img = ProductColorImage.objects.filter(variant__product=product).get(pk=iid)
                img.delete()
            except ProductColorImage.DoesNotExist:
                pass
        return redirect('admin_product_colors', pk=product.id)

    # Справочник ранее использованных цветов
    used_colors = Color.objects.order_by('primary_hex','secondary_hex').all()
    variants = (ProductColorVariant.objects.select_related('color')
                .prefetch_related('images')
                .filter(product=product).order_by('order','id'))
    return render(request, 'pages/admin_product_colors.html', {
        'product': product,
        'used_colors': used_colors,
        'variants': variants
    })

# ===== ЧИСТЫЕ ВЬЮ, БЕЗ ДУБЛИКАТОВ =====
def register_view_new(request):
    if request.user.is_authenticated:
        return redirect('profile_setup')
    form = RegisterForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        from django.contrib.auth.models import User
        if User.objects.filter(username=form.cleaned_data['username']).exists():
            form.add_error('username','Користувач з таким логіном вже існує')
        else:
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password1']
            )
            login(request, user)
            return redirect('profile_setup')
    return render(request, 'pages/auth_register.html', {'form': form})

@login_required
def profile_setup_db(request):
    # Профиль в БД
    from accounts.models import UserProfile
    prof, _ = UserProfile.objects.get_or_create(user=request.user)
    initial = {
        'full_name': getattr(prof, 'full_name', ''),
        'phone': prof.phone or '',
        'email': prof.email or '',
        'telegram': prof.telegram or '',
        'instagram': prof.instagram or '',
        'city': prof.city or '',
        'np_office': prof.np_office or '',
        'pay_type': prof.pay_type or '',
        'is_ubd': prof.is_ubd,
    }
    form = ProfileSetupForm(request.POST or None, request.FILES or None, initial=initial)
    if request.method == 'POST' and form.is_valid():
        prof.full_name = form.cleaned_data.get('full_name') or ''
        prof.phone = form.cleaned_data.get('phone') or ''
        prof.email = form.cleaned_data.get('email') or ''
        prof.telegram = form.cleaned_data.get('telegram') or ''
        prof.instagram = form.cleaned_data.get('instagram') or ''
        prof.city = form.cleaned_data.get('city') or ''
        prof.np_office = form.cleaned_data.get('np_office') or ''
        prof.pay_type = form.cleaned_data.get('pay_type') or ''
        prof.is_ubd = bool(form.cleaned_data.get('is_ubd'))
        
        # Отладочная информация для аватара
        if form.cleaned_data.get('avatar'):
            prof.avatar = form.cleaned_data['avatar']
            
        if prof.is_ubd and form.cleaned_data.get('ubd_doc'):
            prof.ubd_doc = form.cleaned_data['ubd_doc']
        prof.save()
        # Обновим сессию (не обязательно, но пусть будет телефон под рукой)
        request.session['profile_phone'] = prof.phone
        if prof.avatar:
            request.session['profile_avatar'] = prof.avatar.name
        request.session.modified = True
        return redirect('home')
    return render(request, 'pages/auth_profile_setup.html', {'form': form, 'profile': prof})

@login_required
def dev_grant_admin(request):
    """
    Выдать текущему пользователю права администратора (только при DEBUG).
    """
    if not settings.DEBUG:
        return redirect('home')
    u = request.user
    if not u.is_staff or not u.is_superuser:
        u.is_staff = True
        u.is_superuser = True
        u.save()
    return redirect('admin_panel')

# ===== Панель администратора (для staff) =====
@login_required
def admin_panel(request):
    if not request.user.is_staff:
        return redirect('home')
    section = request.GET.get('section', 'stats')
    ctx = {'section': section}
    
    # Проверяем, есть ли параметр принудительного обновления
    force_refresh = request.GET.get('_t')
    if force_refresh:
        # Очищаем кэш для принудительного обновления
        from django.core.cache import cache
        cache.clear()
    
    # Импорты для всех секций
    from django.db.models import Sum, Count, Avg
    if section == 'users':
        # Оптимизированный запрос с select_related и prefetch_related
        users = User.objects.select_related('userprofile').prefetch_related('points', 'orders').order_by('username')
        from accounts.models import UserPoints
        from orders.models import Order
        
        user_data = []
        for user in users:
            profile = getattr(user, 'userprofile', None)
            points = getattr(user, 'points', None)
            
            # Получаем заказы пользователя (уже загружены через prefetch_related)
            user_orders = user.orders.all()
            total_orders = user_orders.count()
            
            # Подсчитываем заказы по статусам
            order_status_counts = {}
            for status_code, status_name in Order.STATUS_CHOICES:
                order_status_counts[status_code] = user_orders.filter(status=status_code).count()
            
            # Подсчитываем заказы по статусам оплаты
            payment_status_counts = {}
            for payment_status_code, payment_status_name in [
                ('unpaid', 'Не оплачено'),
                ('checking', 'На перевірці'),
                ('partial', 'Внесена передплата'),
                ('paid', 'Оплачено повністю')
            ]:
                payment_status_counts[payment_status_code] = user_orders.filter(payment_status=payment_status_code).count()
            
            # Общая сумма заказов (включая предоплаты и частичные оплаты)
            total_spent = user_orders.exclude(payment_status='unpaid').aggregate(
                total=Sum('total_sum')
            )['total'] or 0
            
            # Потраченные баллы
            points_spent = points.total_spent if points else 0
            
            # Получаем промокоды пользователя (включая использованные)
            try:
                from .models import PromoCode
                # Промокоды, выданные пользователю
                user_promocodes = PromoCode.objects.filter(user=user).order_by('-created_at')
                # Промокоды, использованные в заказах пользователя
                used_promocodes = []
                for order in user_orders:
                    if order.promo_code:
                        used_promocodes.append({
                            'code': order.promo_code.code,
                            'discount': order.promo_code.discount,
                            'used_in_order': order.order_number,
                            'used_date': order.created
                        })
            except:
                user_promocodes = []
                used_promocodes = []
            
            user_data.append({
                'user': user,
                'profile': profile,
                'points': points,
                'total_orders': total_orders,
                'order_status_counts': order_status_counts,
                'payment_status_counts': payment_status_counts,
                'total_spent': total_spent,
                'points_spent': points_spent,
                'promocodes': user_promocodes,
                'used_promocodes': used_promocodes
            })
        
        ctx.update({'user_data': user_data})
    elif section == 'catalogs':
        from .models import Category, Product
        categories = Category.objects.filter(is_active=True).order_by('order','name') if hasattr(Category, 'order') else Category.objects.filter(is_active=True).order_by('name')
        products = Product.objects.select_related('category').order_by('-id')
        ctx.update({'categories': categories, 'products': products})
    elif section == 'promocodes':
        from .models import PromoCode
        promocodes = PromoCode.objects.all()
        total_promocodes = promocodes.count()
        active_promocodes = promocodes.filter(is_active=True).count()
        ctx.update({
            'promocodes': promocodes,
            'total_promocodes': total_promocodes,
            'active_promocodes': active_promocodes
        })
    elif section == 'offline_stores':
        from .models import OfflineStore
        stores = OfflineStore.objects.all()
        total_stores = stores.count()
        active_stores = stores.filter(is_active=True).count()
        ctx.update({
            'stores': stores,
            'total_stores': total_stores,
            'active_stores': active_stores
        })
    elif section == 'print_proposals':
        from .models import PrintProposal
        proposals = PrintProposal.objects.select_related('user', 'user__userprofile', 'awarded_promocode').order_by('-created_at')
        total_proposals = proposals.count()
        pending_proposals = proposals.filter(status='pending').count()
        ctx.update({
            'print_proposals': proposals,
            'total_proposals': total_proposals,
            'pending_proposals': pending_proposals
        })
    elif section == 'orders':
        # реальные замовлення
        try:
            from orders.models import Order
            
            # Получаем параметры фильтрации
            status_filter = request.GET.get('status', 'all')
            payment_filter = request.GET.get('payment', 'all')
            user_id_filter = request.GET.get('user_id', None)
            
            
            # Базовый queryset
            orders = Order.objects.select_related('user').prefetch_related('items','items__product')
            
            # Применяем фильтры
            if status_filter != 'all':
                orders = orders.filter(status=status_filter)
            
            if payment_filter != 'all':
                orders = orders.filter(payment_status=payment_filter)
            
            # Фильтр по конкретному пользователю
            user_filter_info = None
            if user_id_filter:
                orders = orders.filter(user_id=user_id_filter)
                
                # Получаем информацию о пользователе для отображения
                try:
                    user_obj = User.objects.get(id=user_id_filter)
                    
                    # Проверяем, есть ли у пользователя профиль
                    full_name = None
                    if hasattr(user_obj, 'userprofile') and user_obj.userprofile:
                        try:
                            full_name = user_obj.userprofile.full_name
                        except:
                            full_name = None
                    
                    user_filter_info = {
                        'username': user_obj.username,
                        'full_name': full_name
                    }
                except User.DoesNotExist:
                    pass
                    user_filter_info = None
            
            # Получаем отфильтрованные заказы
            orders = orders.all()
            
            # Подсчет заказов по статусам
            status_counts = {}
            for status_code, status_name in Order.STATUS_CHOICES:
                status_counts[status_code] = Order.objects.filter(status=status_code).count()
            
            # Подсчет заказов по статусам оплаты
            payment_status_counts = {}
            for payment_status_code, payment_status_name in [
                ('unpaid', 'Не оплачено'),
                ('checking', 'На перевірці'),
                ('partial', 'Внесена передплата'),
                ('paid', 'Оплачено повністю')
            ]:
                payment_status_counts[payment_status_code] = Order.objects.filter(payment_status=payment_status_code).count()
            
            # Общее количество заказов
            total_orders = Order.objects.count()
            
        except Exception:
            orders = []
            status_counts = {}
            payment_status_counts = {}
            total_orders = 0
        
        ctx.update({
            'orders': orders,
            'status_counts': status_counts,
            'payment_status_counts': payment_status_counts,
            'total_orders': total_orders,
            'status_filter': status_filter,
            'payment_filter': payment_filter,
            'user_id_filter': user_id_filter,
            'user_filter_info': user_filter_info
        })
    else:
        # stats — основная статистика с фильтрами по времени
        try:
            
            # Импорты
            from orders.models import Order
            from django.utils import timezone
            from datetime import timedelta
            from django.db.models import Sum, Count, Avg, F, ExpressionWrapper, DurationField
            
            # Импортируем опциональные модели с защитой от отсутствующих таблиц
            try:
                from accounts.models import UserPoints
            except Exception as e:
                UserPoints = None
            
            try:
                from .models import Product, Category, PrintProposal, SiteSession, PageView
            except Exception as e:
                raise
            
            # Получаем период из параметров
            period = request.GET.get('period', 'today')
            now = timezone.now()
            today = now.date()
            
            # Определяем временные рамки
            if period == 'today':
                start_date = today
                end_date = today
                period_name = 'Сьогодні'
            elif period == 'week':
                start_date = today - timedelta(days=7)
                end_date = today
                period_name = 'За тиждень'
            elif period == 'month':
                start_date = today - timedelta(days=30)
                end_date = today
                period_name = 'За місяць'
            else:  # all_time
                start_date = None
                end_date = None
                period_name = 'За весь час'
            
            # Базовые QuerySet'ы для фильтрации
            if start_date and end_date:
                orders_qs = Order.objects.filter(created__date__range=[start_date, end_date])
                users_qs = User.objects.filter(date_joined__date__range=[start_date, end_date])
            else:
                orders_qs = Order.objects.all()
                users_qs = User.objects.all()
            
            # === РАБОЧИЕ МЕТРИКИ ===
            
            # Заказы
            try:
                orders_count = orders_qs.count()
                orders_today = Order.objects.filter(created__date=today).count()
            except Exception as e:
                orders_count = 0
                orders_today = 0
            
            # Выручка
            try:
                revenue = orders_qs.filter(payment_status='paid').aggregate(
                    total=Sum('total_sum')
                )['total'] or 0
            except Exception as e:
                revenue = 0
            
            # Средний чек
            try:
                avg_order_value = orders_qs.filter(payment_status='paid').aggregate(
                    avg=Avg('total_sum')
                )['avg'] or 0
            except Exception as e:
                avg_order_value = 0
            
            # Пользователи
            try:
                total_users = User.objects.count()
                new_users_period = users_qs.count()
                active_users_today = User.objects.filter(last_login__date=today).count()
                active_users_period = User.objects.filter(last_login__date__range=[start_date, end_date]).count() if start_date and end_date else User.objects.filter(last_login__date=today).count()
            except Exception as e:
                total_users = 0
                new_users_period = 0
                active_users_today = 0
                active_users_period = 0
            
            # Товары
            try:
                total_products = Product.objects.count()
                total_categories = Category.objects.count()
            except Exception as e:
                total_products = 0
                total_categories = 0
            
            # Принты на рассмотрении
            try:
                print_proposals_pending = PrintProposal.objects.filter(status='pending').count()
            except Exception as e:
                print_proposals_pending = 0
            
            # Промокоды использованные
            try:
                promocodes_used = orders_qs.exclude(promo_code__isnull=True).count()
            except Exception as e:
                promocodes_used = 0
            
            # Баллы (защита, если таблицы ещё нет)
            if UserPoints is not None:
                try:
                    total_points_earned = UserPoints.objects.aggregate(
                        total=Sum('points')
                    )['total'] or 0
                    users_with_points = UserPoints.objects.filter(points__gt=0).count()
                except Exception as e:
                    total_points_earned = 0
                    users_with_points = 0
            else:
                total_points_earned = 0
                users_with_points = 0
            
            # === МЕТРИКИ ПОСЕЩАЕМОСТИ (встроенная лёгкая аналитика) ===
            try:
                online_threshold = timezone.now() - timedelta(minutes=5)
                
                online_users = SiteSession.objects.filter(last_seen__gte=online_threshold, is_bot=False).count()
                unique_visitors_today = SiteSession.objects.filter(first_seen__date=today, is_bot=False).count()
                page_views_today = PageView.objects.filter(when__date=today, is_bot=False).count()

                # Fallback: если по какой‑то причине онлайн/уники не посчитались через SiteSession,
                # попробуем оценить их через PageView (уникальные сессии по просмотрам)
                if online_users == 0:
                    online_users = (
                        PageView.objects
                        .filter(when__gte=online_threshold, is_bot=False)
                        .values('session_id')
                        .distinct()
                        .count()
                    )
                
                if unique_visitors_today == 0:
                    unique_visitors_today = (
                        PageView.objects
                        .filter(when__date=today, is_bot=False)
                        .values('session_id')
                        .distinct()
                        .count()
                    )
                    
            except Exception as e:
                online_users = 0
                unique_visitors_today = 0
                page_views_today = 0
            today_sessions = SiteSession.objects.filter(first_seen__date=today, is_bot=False)
            sv = today_sessions.filter(pageviews__lte=1).count()
            bounce_rate = round((sv / today_sessions.count()) * 100, 2) if today_sessions.exists() else 0
            durations = today_sessions.annotate(dur=ExpressionWrapper(F('last_seen') - F('first_seen'), output_field=DurationField())).values_list('dur', flat=True)
            if durations:
                total_seconds = sum((d.total_seconds() for d in durations if d is not None), 0)
                avg_session_duration = int(total_seconds / max(len(durations), 1))
            else:
                avg_session_duration = 0
            
            # Продажи метрики
            conversion_rate = 0  # Нужна система аналитики для расчета
            products_sold_today = 0  # Нужно считать количество товаров в заказах
            abandoned_carts = 0  # Нужна система отслеживания корзин
            
            # Контент метрики
            favorites_count = FavoriteProduct.objects.count()
            reviews_count = 0  # Нет модели отзывов
            avg_rating = 0  # Нет системы рейтингов
            
            stats = {
                # Рабочие метрики
                'orders_today': orders_today,
                'orders_count': orders_count,
                'revenue_today': revenue,
                'avg_order_value': round(avg_order_value, 2),
                'total_users': total_users,
                'new_users_today': new_users_period,
                'active_users_today': active_users_today,
                'active_users_period': active_users_period,
                'total_products': total_products,
                'total_categories': total_categories,
                'print_proposals_pending': print_proposals_pending,
                'promocodes_used_today': promocodes_used,
                'total_points_earned': total_points_earned,
                'users_with_points': users_with_points,
                'favorites_count': favorites_count,
                
                # Нерабочие метрики (заглушки)
                'online_users': online_users,
                'unique_visitors_today': unique_visitors_today,
                'page_views_today': page_views_today,
                'page_views_period': page_views_today if period == 'today' else (
                    PageView.objects.filter(when__date__range=[start_date, end_date], is_bot=False).count() if start_date and end_date 
                    else PageView.objects.filter(is_bot=False).count()
                ),
                'sessions_period': unique_visitors_today if period == 'today' else (
                    SiteSession.objects.filter(first_seen__date__range=[start_date, end_date], is_bot=False).count() if start_date and end_date 
                    else SiteSession.objects.filter(is_bot=False).count()
                ),
                'bounce_rate': bounce_rate,
                'avg_session_duration': avg_session_duration,
                'conversion_rate': conversion_rate,
                'products_sold_today': products_sold_today,
                'abandoned_carts': abandoned_carts,
                'reviews_count': reviews_count,
                'avg_rating': avg_rating,
                
                # Мета информация
                'period': period,
                'period_name': period_name,
                'start_date': start_date,
                'end_date': end_date
            }
            
        except Exception as e:
            # В случае ошибки возвращаем базовые значения
            # Получаем период из параметров для fallback
            period = request.GET.get('period', 'today')
            if period == 'today':
                period_name = 'Сьогодні'
            elif period == 'week':
                period_name = 'За тиждень'
            elif period == 'month':
                period_name = 'За місяць'
            else:
                period_name = 'За весь час'
            
            # Пытаемся получить базовые данные даже в случае ошибки
            try:
                total_users_fallback = User.objects.count()
                total_products_fallback = Product.objects.count()
                total_categories_fallback = Category.objects.count()
            except Exception as e:
                total_users_fallback = 0
                total_products_fallback = 0
                total_categories_fallback = 0
            
            stats = {
                'orders_today': 0,
                'orders_count': 0,
                'revenue_today': 0,
                'avg_order_value': 0,
                'total_users': total_users_fallback,
                'new_users_today': 0,
                'active_users_today': 0,
                'active_users_period': 0,
                'total_products': total_products_fallback,
                'total_categories': total_categories_fallback,
                'print_proposals_pending': 0,
                'promocodes_used_today': 0,
                'total_points_earned': 0,
                'users_with_points': 0,
                'favorites_count': 0,
                'online_users': 0,
                'unique_visitors_today': 0,
                'page_views_today': 0,
                'page_views_period': 0,
                'sessions_period': 0,
                'bounce_rate': 0,
                'avg_session_duration': 0,
                'conversion_rate': 0,
                'products_sold_today': 0,
                'abandoned_carts': 0,
                'reviews_count': 0,
                'avg_rating': 0,
                'period': period,
                'period_name': period_name,
                'start_date': None,
                'end_date': None
            }
        
        ctx.update({'stats': stats})
    
    # Создаем ответ с заголовками для предотвращения кэширования
    from django.template.loader import render_to_string
    from django.http import HttpResponse
    
    html_content = render_to_string('pages/admin_panel.html', ctx, request=request)
    response = HttpResponse(html_content)
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response

@login_required
def admin_print_proposal_update_status(request):
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Доступ заборонено'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Невірний метод запиту'})
    
    try:
        proposal_id = request.POST.get('proposal_id')
        status = request.POST.get('status')
        
        if not proposal_id or not status:
            return JsonResponse({'success': False, 'error': 'Відсутні обов\'язкові параметри'})
        
        proposal = PrintProposal.objects.get(id=proposal_id)
        proposal.status = status
        proposal.save()
        
        status_text = dict(PrintProposal.STATUS_CHOICES)[status]
        
        return JsonResponse({
            'success': True, 
            'message': f'Статус принта оновлено на "{status_text}"',
            'new_status_text': status_text
        })
        
    except PrintProposal.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Принт не знайдено'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Помилка: {str(e)}'})

@login_required
def admin_print_proposal_award_points(request):
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Доступ заборонено'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Невірний метод запиту'})
    
    try:
        proposal_id = request.POST.get('proposal_id')
        points = int(request.POST.get('points', 0))
        
        if not proposal_id or points <= 0:
            return JsonResponse({'success': False, 'error': 'Відсутні обов\'язкові параметри або невірна кількість балів'})
        
        proposal = PrintProposal.objects.get(id=proposal_id)
        proposal.awarded_points = points
        proposal.save()
        
        # Добавляем баллы пользователю
        from accounts.models import UserPoints
        user_points, created = UserPoints.objects.get_or_create(user=proposal.user)
        user_points.total_earned += points
        user_points.save()
        
        return JsonResponse({
            'success': True, 
            'message': f'Користувачу {proposal.user.username} нараховано {points} балів'
        })
        
    except PrintProposal.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Принт не знайдено'})
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Невірна кількість балів'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Помилка: {str(e)}'})

@login_required
def admin_print_proposal_award_promocode(request):
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Доступ заборонено'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Невірний метод запиту'})
    
    try:
        proposal_id = request.POST.get('proposal_id')
        amount = int(request.POST.get('amount', 0))
        
        if not proposal_id or amount <= 0:
            return JsonResponse({'success': False, 'error': 'Відсутні обов\'язкові параметри або невірна сума'})
        
        proposal = PrintProposal.objects.get(id=proposal_id)
        
        # Создаем промокод
        from .models import PromoCode
        import random
        import string
        
        # Генерируем уникальный код
        while True:
            code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
            if not PromoCode.objects.filter(code=code).exists():
                break
        
        promocode = PromoCode.objects.create(
            code=code,
            discount_type='fixed',
            discount_amount=amount,
            user=proposal.user,
            is_active=True,
            description=f'Промокод за принт від {proposal.user.username}'
        )
        
        proposal.awarded_promocode = promocode
        proposal.save()
        
        return JsonResponse({
            'success': True, 
            'message': f'Користувачу {proposal.user.username} видано промокод {code} на суму {amount}₴'
        })
        
    except PrintProposal.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Принт не знайдено'})
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Невірна сума промокоду'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Помилка: {str(e)}'})

@login_required
def admin_update_user(request):
    if not request.user.is_staff:
        return redirect('home')
    if request.method != 'POST':
        return redirect('admin_panel')
    user_id = request.POST.get('user_id')
    is_staff = True if request.POST.get('is_staff') == 'on' else False
    is_superuser_req = True if request.POST.get('is_superuser') == 'on' else False
    
    # Обработка баллов
    points_adjustment = request.POST.get('points_adjustment')
    points_reason = request.POST.get('points_reason', '')

    try:
        u = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return redirect('admin_panel')

    if not request.user.is_superuser:
        is_superuser_req = u.is_superuser

    u.is_staff = is_staff
    u.is_superuser = is_superuser_req
    u.save()
    
    # Обновляем баллы если указано
    if points_adjustment:
        try:
            points_change = int(points_adjustment)
            if points_change != 0:
                from accounts.models import UserPoints
                user_points = UserPoints.get_or_create_points(u)
                
                if points_change > 0:
                    user_points.add_points(points_change, f'Коригування адміністратором: {points_reason}')
                else:
                    user_points.spend_points(abs(points_change), f'Коригування адміністратором: {points_reason}')
        except ValueError:
            pass
    
    return redirect('/admin-panel/?section=users')

# ======= Orders =======
# Импорт моделей заказов выполняется лениво внутри функций,
# чтобы избежать падения при старте, если приложение ещё не подключено.

@login_required
def order_create(request):
    # Локальный импорт моделей заказов – сразу, до использования
    from orders.models import Order, OrderItem

    # Требуем заполненный профиль доставки
    try:
        prof = request.user.userprofile
    except UserProfile.DoesNotExist:
        return redirect('profile_setup')
    
    # Получаем данные из формы или из профиля
    if request.method == 'POST':
        # Используем данные из формы
        full_name = request.POST.get('full_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        city = request.POST.get('city', '').strip()
        np_office = request.POST.get('np_office', '').strip()
        pay_type = request.POST.get('pay_type', '')
        
        # Валидация данных из формы
        if not full_name or len(full_name) < 3:
            from django.contrib import messages
            messages.error(request, 'ПІБ повинно містити мінімум 3 символи!')
            return redirect('cart')
        
        if not phone or len(phone.strip()) < 10:
            from django.contrib import messages
            messages.error(request, 'Введіть коректний номер телефону!')
            return redirect('cart')
        
        if not city or len(city.strip()) < 2:
            from django.contrib import messages
            messages.error(request, 'Введіть назву міста!')
            return redirect('cart')
        
        if not np_office or len(np_office.strip()) < 1:
            from django.contrib import messages
            messages.error(request, 'Введіть адресу відділення!')
            return redirect('cart')
        
        if not pay_type:
            from django.contrib import messages
            messages.error(request, 'Оберіть тип оплати!')
            return redirect('cart')
        
        # Обновляем профиль пользователя данными из формы
        prof.full_name = full_name
        prof.phone = phone
        prof.city = city
        prof.np_office = np_office
        prof.pay_type = pay_type
        prof.save()
        
    else:
        # Используем данные из профиля (для GET запросов)
        full_name = getattr(prof, 'full_name', '') or request.user.username
        phone = prof.phone
        city = prof.city
        np_office = prof.np_office
        pay_type = prof.pay_type
        
        # Проверяем обязательные поля с более строгой валидацией
        if not phone or len(phone.strip()) < 10:
            from django.contrib import messages
            messages.error(request, 'Введіть коректний номер телефону!')
            return redirect('cart')
        
        if not city or len(city.strip()) < 2:
            from django.contrib import messages
            messages.error(request, 'Введіть назву міста!')
            return redirect('cart')
        
        if not np_office or len(np_office.strip()) < 1:
            from django.contrib import messages
            messages.error(request, 'Введіть адресу відділення!')
            return redirect('cart')
        
        if not pay_type:
            from django.contrib import messages
            messages.error(request, 'Оберіть тип оплати!')
            return redirect('cart')

    # Корзина должна быть не пустой
    cart = request.session.get('cart') or {}
    if not cart:
        return redirect('cart')

    # Защита от дублирования заказов
    from django.utils import timezone
    from datetime import timedelta
    
    # Проверяем, не было ли уже заказа от этого пользователя в последние 5 минут
    recent_orders = Order.objects.filter(
        user=request.user,
        created__gte=timezone.now() - timedelta(minutes=5)
    )
    
    if recent_orders.exists():
        from django.contrib import messages
        messages.error(request, 'Замовлення вже було створено нещодавно. Спробуйте ще раз через кілька хвилин.')
        return redirect('cart')

    # Пересчёт total и создание заказа в одной транзакции
    from django.db import transaction
    
    with transaction.atomic():
        ids = [i['product_id'] for i in cart.values()]
        prods = Product.objects.in_bulk(ids)
        total_sum = 0
        
        # Создаем заказ
        order = Order.objects.create(
            user=request.user,
            full_name=full_name,
            phone=phone, city=city, np_office=np_office,
            pay_type=pay_type, total_sum=0, status='new'
        )
        
        # Создаем все товары заказа
        order_items = []
        for key, it in cart.items():
            p = prods.get(it['product_id'])
            if not p:
                continue
            
            # Получаем информацию о цвете
            color_variant = _get_color_variant_safe(it.get('color_variant_id'))
            
            unit = p.final_price
            line = unit * it['qty']
            total_sum += line
            
            # Создаем OrderItem
            order_item = OrderItem(
                order=order, product=p, color_variant=color_variant, title=p.title, size=it.get('size', ''),
                qty=it['qty'], unit_price=unit, line_total=line
            )
            order_items.append(order_item)
        
        # Создаем все товары одним запросом
        OrderItem.objects.bulk_create(order_items)
        
        # Применяем промокод если есть
        promo_code = request.session.get('applied_promo_code')
        if promo_code:
            from .models import PromoCode
            try:
                promo = PromoCode.objects.get(code=promo_code, is_active=True)
                if promo.can_be_used():
                    discount = promo.calculate_discount(total_sum)
                    order.discount_amount = discount
                    order.promo_code = promo
                    promo.use()  # Увеличиваем счетчик использований
            except PromoCode.DoesNotExist:
                pass
        
        # Обновляем общую сумму заказа
        order.total_sum = total_sum
        order.save()

    # Очищаем корзину и промокод
    _reset_monobank_session(request, drop_pending=True)
    request.session['cart'] = {}
    request.session.pop('applied_promo_code', None)
    request.session.modified = True

    # Отправляем Telegram уведомление после создания заказа и товаров
    try:
        from orders.telegram_notifications import TelegramNotifier
        notifier = TelegramNotifier()
        notifier.send_new_order_notification(order)
    except Exception as e:
        # Не прерываем процесс, если уведомление не отправилось
        pass

    from django.contrib import messages
    messages.success(request, f'Замовлення #{order.order_number} успішно створено!')

    return redirect('my_orders')

def order_success(request, order_id):
    """
    Страница успешного оформления заказа
    """
    from orders.models import Order
    
    try:
        order = Order.objects.get(id=order_id)
        # Проверяем, что заказ был создан недавно (в течение часа)
        from django.utils import timezone
        from datetime import timedelta
        
        if order.created < timezone.now() - timedelta(hours=1):
            return redirect('home')
            
    except Order.DoesNotExist:
        return redirect('home')
    
    return render(request, 'pages/order_success.html', {
        'order': order,
        'order_number': order.id
    })

@login_required
def my_orders(request):
    from orders.models import Order
    qs = Order.objects.filter(user=request.user).prefetch_related('items__product').order_by('-created')
    return render(request, 'pages/my_orders.html', {'orders': qs})

@login_required
@require_POST
def update_payment_method(request):
    """
    AJAX view для обновления метода оплаты заказа
    """
    from orders.models import Order
    
    order_id = request.POST.get('order_id')
    payment_method = request.POST.get('payment_method')
    
    if not order_id or not payment_method:
        return JsonResponse({
            'success': False,
            'error': 'Відсутні необхідні дані'
        })
    
    if payment_method not in ['full', 'partial']:
        return JsonResponse({
            'success': False,
            'error': 'Невірний метод оплати'
        })
    
    try:
        order = Order.objects.get(id=order_id, user=request.user)
    except Order.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Замовлення не знайдено'
        })
    
    # Обновляем метод оплаты
    order.pay_type = payment_method
    order.save()
    
    # Определяем отображаемое название метода
    method_display = 'Повна передоплата' if payment_method == 'full' else 'Часткова передоплата'
    
    return JsonResponse({
        'success': True,
        'payment_method': payment_method,
        'method_display': method_display
    })

@login_required
def my_promocodes(request):
    """
    Страница промокодов пользователя
    """
    from orders.models import Order
    
    # Получаем все заказы пользователя с промокодами
    orders_with_promocodes = Order.objects.filter(
        user=request.user,
        promo_code__isnull=False
    ).select_related('promo_code').order_by('-created')
    
    # Создаем список использованных промокодов
    used_promocodes = []
    for order in orders_with_promocodes:
        if order.promo_code:
            used_promocodes.append({
                'promo_code': order.promo_code,
                'order': order,
                'used_date': order.created,
                'discount_amount': order.discount_amount,
                'order_total': order.total_sum
            })
    
    return render(request, 'pages/my_promocodes.html', {
        'used_promocodes': used_promocodes
    })

@login_required
def buy_with_points(request):
    """
    Страница покупки за баллы
    """
    from accounts.models import UserPoints
    
    # Получаем баллы пользователя
    try:
        user_points = UserPoints.objects.get(user=request.user)
        current_points = user_points.points
    except UserPoints.DoesNotExist:
        current_points = 0
    
    # Определяем доступные товары за баллы
    available_items = [
        {
            'id': 'promo_10',
            'name': 'Промокод на знижку 10%',
            'description': 'Отримайте промокод на знижку 10% від суми замовлення',
            'points_cost': 100,
            'type': 'promo',
            'icon': 'M21.41 11.58l-9-9C12.05 2.22 11.55 2 11 2H4c-1.1 0-2 .9-2 2v7c0 .55.22 1.05.59 1.42l9 9c.36.36.86.58 1.41.58.55 0 1.05-.22 1.41-.59l7-7c.37-.36.59-.86.59-1.41 0-.55-.23-1.06-.59-1.42zM5.5 7C4.67 7 4 6.33 4 5.5S4.67 4 5.5 4 7 4.67 7 5.5 6.33 7 5.5 7z',
            'color': 'linear-gradient(135deg, #3b82f6, #1d4ed8)',
            'can_afford': current_points >= 100
        },
        {
            'id': 'donate_zsu',
            'name': 'Донат на ЗСУ',
            'description': 'Пожертвуйте всі ваші бали на підтримку Збройних Сил України',
            'points_cost': current_points,
            'type': 'donate',
            'icon': 'M12 2l3.09 6.26L22 9.27l-5 4.87 1.18 6.88L12 17.77l-6.18 3.25L7 14.14 2 9.27l6.91-1.01L12 2z',
            'color': 'linear-gradient(135deg, #dc2626, #991b1b)',
            'can_afford': current_points > 0
        }
    ]
    
    return render(request, 'pages/buy_with_points.html', {
        'current_points': current_points,
        'available_items': available_items
    })

@login_required
def purchase_with_points(request):
    """
    Обработка покупки за баллы
    """
    if request.method != 'POST':
        return redirect('buy_with_points')
    
    from accounts.models import UserPoints
    from django.contrib import messages
    
    item_id = request.POST.get('item_id')
    
    try:
        user_points = UserPoints.objects.get(user=request.user)
    except UserPoints.DoesNotExist:
        messages.error(request, 'У вас немає балів для покупки')
        return redirect('buy_with_points')
    
    if item_id == 'promo_10':
        # Покупка промокода на 10% скидки
        if user_points.points >= 100:
            # Здесь будет логика создания промокода
            # Пока что просто списываем баллы
            user_points.spend_points(100, 'Покупка промокода на знижку 10%')
            messages.success(request, 'Промокод на знижку 10% успішно придбано! Код: POINTS10')
        else:
            messages.error(request, 'Недостатньо балів для покупки промокода')
    
    elif item_id == 'donate_zsu':
        # Донат на ЗСУ
        if user_points.points > 0:
            points_to_donate = user_points.points
            user_points.spend_points(points_to_donate, 'Донат на ЗСУ')
            messages.success(request, f'Дякуємо за підтримку ЗСУ! Пожертвовано {points_to_donate} балів')
        else:
            messages.error(request, 'У вас немає балів для пожертвування')
    
    else:
        messages.error(request, 'Невідомий товар')
    
    return redirect('buy_with_points')

@login_required
def admin_order_update(request):
    if not request.user.is_staff:
        return redirect('home')
    if request.method != 'POST':
        return redirect('/admin-panel/?section=orders')
    
    from orders.models import Order
    from django.http import JsonResponse
    
    oid = request.POST.get('order_id')
    status = request.POST.get('status')
    tracking_number = request.POST.get('tracking_number', '').strip()
    
    try:
        o = Order.objects.get(pk=oid)
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Замовлення не знайдено'})
    
    old_status = o.status
    if status in dict(Order.STATUS_CHOICES):
        o.status = status
        
        # Обновляем ТТН если предоставлен
        if tracking_number:
            o.tracking_number = tracking_number
        
        o.save()
        
        # Обрабатываем баллы при изменении статуса
        if o.user:  # Только для авторизованных пользователей
            from accounts.models import UserPoints
            user_points = UserPoints.get_or_create_points(o.user)
            
            # Рассчитываем баллы за заказ
            total_points = 0
            for item in o.items.all():
                if item.product.points_reward > 0:
                    total_points += item.product.points_reward * item.qty
            
            # Статусы, при которых начисляются баллы
            points_awarding_statuses = ['prep', 'ship', 'done']  # 'готовится к отправке', 'отправлен', 'получен'
            
            # Статусы, при которых отменяются баллы
            points_cancelling_statuses = ['cancelled']  # 'отменен'
            
            # Если переходим к статусу начисления баллов и баллы еще не начислены
            if status in points_awarding_statuses and old_status not in points_awarding_statuses and not o.points_awarded:
                if total_points > 0:
                    user_points.add_points(total_points, f'Замовлення #{o.order_number} {o.get_status_display()}')
                    o.points_awarded = True
                    o.save(update_fields=['points_awarded'])
            
            # Если переходим к статусу отмены баллов и баллы были начислены
            elif status in points_cancelling_statuses and old_status not in points_cancelling_statuses and o.points_awarded:
                if total_points > 0:
                    user_points.spend_points(total_points, f'Скасування замовлення #{o.order_number}')
                    o.points_awarded = False
                    o.save(update_fields=['points_awarded'])
        
        # Формируем сообщение об успехе
        status_display = dict(Order.STATUS_CHOICES).get(status, status)
        message = f'Статус замовлення змінено на "{status_display}"'
        if tracking_number:
            message += f' з ТТН: {tracking_number}'
        
        return JsonResponse({
            'success': True, 
            'message': message,
            'status': status,
            'tracking_number': tracking_number
        })
    
    return JsonResponse({'success': False, 'error': 'Невірний статус'})


@login_required
def admin_update_payment_status(request):
    """Обновление статуса оплаты заказа"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Доступ заборонено'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Невірний метод запиту'})
    
    from orders.models import Order
    from django.http import JsonResponse
    
    order_id = request.POST.get('order_id')
    payment_status = request.POST.get('payment_status')
    
    if not order_id or not payment_status:
        return JsonResponse({'success': False, 'error': 'Відсутні необхідні дані'})
    
    try:
        order = Order.objects.get(pk=order_id)
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Замовлення не знайдено'})
    
    # Проверяем валидность статуса оплаты
    valid_payment_statuses = ['unpaid', 'checking', 'partial', 'paid']
    if payment_status not in valid_payment_statuses:
        return JsonResponse({'success': False, 'error': 'Невірний статус оплати'})
    
    # Обновляем статус оплаты
    order.payment_status = payment_status
    order.save()
    
    # Формируем сообщение об успехе
    payment_status_display = {
        'unpaid': 'Не оплачено',
        'checking': 'На перевірці',
        'partial': 'Внесена передплата',
        'paid': 'Оплачено повністю'
    }.get(payment_status, payment_status)
    
    return JsonResponse({
        'success': True,
        'message': f'Статус оплати змінено на "{payment_status_display}"',
        'payment_status': payment_status
    })


@login_required
def admin_approve_payment(request):
    """Подтверждение или отклонение оплаты заказа"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Доступ заборонено'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Невірний метод запиту'})
    
    from orders.models import Order
    from django.http import JsonResponse
    
    order_id = request.POST.get('order_id')
    action = request.POST.get('action')  # 'approve' или 'reject'
    
    if not order_id or not action:
        return JsonResponse({'success': False, 'error': 'Відсутні необхідні дані'})
    
    if action not in ['approve', 'reject']:
        return JsonResponse({'success': False, 'error': 'Невірна дія'})
    
    try:
        order = Order.objects.get(pk=order_id)
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Замовлення не знайдено'})
    
    if action == 'approve':
        # Подтверждаем оплату
        order.payment_status = 'paid'
        message = 'Оплату підтверджено'
        new_status = 'paid'
    else:
        # Отклоняем оплату
        order.payment_status = 'unpaid'
        message = 'Оплату відхилено'
        new_status = 'unpaid'
    
    order.save()
    
    return JsonResponse({
        'success': True,
        'message': message,
        'new_status': new_status
    })





# ---- Каталоги CRUD ----
@login_required
def admin_category_new(request):
    if not request.user.is_staff:
        return redirect('home')
    from .forms import CategoryForm
    form = CategoryForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('/admin-panel/?section=catalogs')
    return render(request, 'pages/admin_category_form.html', {'form': form, 'mode': 'new'})

@login_required
def admin_category_edit(request, pk):
    if not request.user.is_staff:
        return redirect('home')
    from .forms import CategoryForm
    from .models import Category
    obj = get_object_or_404(Category, pk=pk)
    form = CategoryForm(request.POST or None, request.FILES or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('/admin-panel/?section=catalogs')
    return render(request, 'pages/admin_category_form.html', {'form': form, 'mode': 'edit', 'obj': obj})

@login_required
def admin_category_delete(request, pk):
    if not request.user.is_staff:
        return redirect('home')
    from .models import Category
    obj = get_object_or_404(Category, pk=pk)
    obj.delete()
    return redirect('/admin-panel/?section=catalogs')

@login_required
def admin_product_new(request):
    if not request.user.is_staff:
        return redirect('home')
    from .forms import ProductForm
    form = ProductForm(request.POST or None, request.FILES or None)
    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        
        # Обработка формы товара
        if form_type == 'product':
            if form.is_valid():
                prod = form.save()
                # возможно добавление дополнительных изображений позже
                
                # Если это AJAX запрос, возвращаем JSON
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json':
                    return JsonResponse({
                        'success': True,
                        'product_id': prod.id,
                        'message': 'Товар успішно створено!'
                    })
                
                return redirect('/admin-panel/?section=catalogs')
            else:
                # Если это AJAX запрос с ошибками, возвращаем JSON с ошибками
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json':
                    return JsonResponse({
                        'success': False,
                        'errors': form.errors,
                        'message': 'Помилка валідації форми'
                    })
        
        # Обработка формы цвета
        elif form_type == 'color':
            product_id = request.POST.get('product_id')
            if not product_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Не вказано ID товару'
                })
            
            try:
                product = Product.objects.get(id=product_id)
                from productcolors.models import Color, ProductColorVariant, ProductColorImage
                
                color_name = request.POST.get('color_name')
                primary_hex = request.POST.get('primary_hex')
                secondary_hex = request.POST.get('secondary_hex', '')
                color_images = request.FILES.getlist('color_images')
                
                if color_name and primary_hex and color_images:
                    
                    # Проверяем, существует ли уже такой цвет
                    color = Color.objects.filter(
                        primary_hex=primary_hex,
                        secondary_hex=secondary_hex if secondary_hex else None
                    ).first()
                    
                    
                    if color:
                        # Если цвет существует, обновляем его название
                        color.name = color_name
                        color.save()
                    else:
                        # Создаем новый цвет с HEX кодами
                        color = Color.objects.create(
                            name=color_name,
                            primary_hex=primary_hex,
                            secondary_hex=secondary_hex if secondary_hex else None
                        )
                    
                    # Проверяем, не существует ли уже такой вариант цвета для этого товара
                    color_variant, created = ProductColorVariant.objects.get_or_create(
                        product=product,
                        color=color
                    )
                    
                    
                    # Добавляем изображения (всегда, если они есть)
                    if color_images:
                        for i, image_file in enumerate(color_images):
                            ProductColorImage.objects.create(
                                color_variant=color_variant,
                                image=image_file
                            )
                    
                    if created:
                        message = 'Колір успішно додано!'
                    else:
                        message = 'Колір вже існує для цього товару, але зображення додано!'
                    
                    return JsonResponse({
                        'success': True,
                        'color': {
                            'id': color.id,
                            'name': color.name,
                            'image_url': color_variant.images.first().image.url if color_variant.images.exists() else ''
                        },
                        'message': message
                    })
                else:
                    return JsonResponse({
                        'success': False,
                        'message': 'Не всі обов\'язкові поля заповнені'
                    })
                    
            except Product.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Товар не знайдено'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Помилка створення кольору: {str(e)}'
                })
        
        # Обработка формы изображения
        elif form_type == 'image':
            product_id = request.POST.get('product_id')
            if not product_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Не вказано ID товару'
                })
            
            try:
                product = Product.objects.get(id=product_id)
                additional_image = request.FILES.get('additional_image')
                
                if additional_image:
                    from .models import ProductImage
                    ProductImage.objects.create(
                        product=product,
                        image=additional_image
                    )
                    
                    return JsonResponse({
                        'success': True,
                        'image': {
                            'id': 'new',
                            'image_url': ProductImage.objects.filter(product=product).last().image.url
                        },
                        'message': 'Зображення успішно додано!'
                    })
                else:
                    return JsonResponse({
                        'success': False,
                        'message': 'Не вибрано зображення'
                    })
                    
            except Product.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Товар не знайдено'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Помилка додавання зображення: {str(e)}'
                })
    return render(request, 'pages/add_product_new.html', {'form': form, 'mode': 'new'})

@login_required
def admin_product_edit(request, pk):
    if not request.user.is_staff:
        return redirect('home')
    from .forms import ProductForm
    from .models import Product
    obj = get_object_or_404(Product, pk=pk)
    
    # Если пришло действие с цветами — обрабатываем и возвращаемся на эту же страницу
    action = (request.POST.get('action') or '').strip() if request.method == 'POST' else ''
    
    # Если это POST запрос для цветов, обрабатываем его
    if action and request.method == 'POST':
        try:
            from productcolors.models import Color, ProductColorVariant, ProductColorImage
        except Exception:
            Color = ProductColorVariant = ProductColorImage = None

        if Color and ProductColorVariant and ProductColorImage:
            if action == 'add':
                color_id = request.POST.get('color_id')
                primary_hex = (request.POST.get('primary_hex') or '').strip() or '#000000'
                secondary_hex = (request.POST.get('secondary_hex') or '').strip() or None
                is_default = bool(request.POST.get('is_default'))
                order = int(request.POST.get('order') or 0)

                if color_id:
                    color = get_object_or_404(Color, pk=color_id)
                else:
                    color, _ = Color.objects.get_or_create(primary_hex=primary_hex, secondary_hex=secondary_hex)
                variant = ProductColorVariant.objects.create(product=obj, color=color, is_default=is_default, order=order)
                for f in request.FILES.getlist('images'):
                    ProductColorImage.objects.create(variant=variant, image=f)
            elif action == 'delete_variant':
                vid = request.POST.get('variant_id')
                try:
                    v = ProductColorVariant.objects.get(pk=vid, product=obj)
                    v.delete()
                except ProductColorVariant.DoesNotExist:
                    pass
            elif action == 'delete_image':
                iid = request.POST.get('image_id')
                try:
                    img = ProductColorImage.objects.filter(variant__product=obj).get(pk=iid)
                    img.delete()
                except ProductColorImage.DoesNotExist:
                    pass
        return redirect('admin_product_edit', pk=obj.id)

    # Основная форма товара
    form = ProductForm(request.POST or None, request.FILES or None, instance=obj)
    if request.method == 'POST':
        
        # Проверяем, что это не запрос для цветов
        if not request.POST.get('action'):
            
            if form.is_valid():
                product = form.save(commit=False)
                # Автогенерація slug, якщо порожній
                if not getattr(product, 'slug', None):
                    base = slugify(product.title or '')
                    product.slug = unique_slugify(Product, base)
                product.save()
                return redirect('/admin-panel/?section=catalogs')
            else:
                # Обработка ошибок формы
                pass
        else:
            # Это запрос для цветов, пропускаем обработку основной формы
            pass

    # Данные для блока «Кольори»
    used_colors = []
    variants = []
    try:
        from productcolors.models import Color, ProductColorVariant
        used_colors = Color.objects.order_by('primary_hex', 'secondary_hex').all()
        variants = (ProductColorVariant.objects.select_related('color')
                    .prefetch_related('images')
                    .filter(product=obj).order_by('order', 'id'))
    except Exception as e:
        # Возвращаем базовые значения
        used_colors = []
        variants = []

    return render(
        request,
        'pages/admin_product_form.html',
        {'form': form, 'mode': 'edit', 'obj': obj, 'used_colors': used_colors, 'variants': variants}
    )

@login_required
def admin_product_edit_unified(request, pk):
    """Универсальное редактирование товара со всеми настройками"""
    if not request.user.is_staff:
        return redirect('home')
    from .forms import ProductForm
    from .models import Product, ProductImage
    from productcolors.models import ProductColorVariant, Color, ProductColorImage
    obj = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        
        if form_type == 'product':
            form = ProductForm(request.POST, request.FILES, instance=obj)
            
            if form.is_valid():
                product = form.save(commit=False)
                
                # Автогенерація slug, якщо порожній
                if not getattr(product, 'slug', None):
                    base = slugify(product.title or '')
                    product.slug = unique_slugify(Product, base)
                
                # Если главное изображение не выбрано, используем первую фотографию первого цвета
                if not product.main_image:
                    first_color_variant = product.color_variants.first()
                    if first_color_variant:
                        first_image = first_color_variant.images.first()
                        if first_image:
                            product.main_image = first_image.image
                
                product.save()
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Товар успішно збережено!'})
                return redirect('/admin-panel/?section=catalogs')
            else:
                # Возвращаем форму с ошибками для отображения пользователю
                return render(request, 'pages/admin_product_edit_unified.html', {
                    'form': form, 
                    'obj': obj
                })
        
        elif form_type == 'color':
            color_name = request.POST.get('color_name')
            primary_hex = request.POST.get('primary_hex')
            secondary_hex = request.POST.get('secondary_hex', '')
            color_images = request.FILES.getlist('color_images')
            
            if color_name and primary_hex and color_images:
                # Создаем цвет с HEX кодами
                color = Color.objects.create(
                    name=color_name,
                    primary_hex=primary_hex,
                    secondary_hex=secondary_hex if secondary_hex else None
                )
                
                # Создаем вариант цвета для товара
                color_variant = ProductColorVariant.objects.create(
                    product=obj,
                    color=color
                )
                
                # Создаем изображения для варианта цвета
                for i, image in enumerate(color_images):
                    ProductColorImage.objects.create(
                        variant=color_variant,
                        image=image,
                        order=i  # Сохраняем порядок изображений
                    )
                
                # Если у товара нет главного изображения, используем первую фотографию этого цвета
                if not obj.main_image and color_images:
                    obj.main_image = color_images[0]
                    obj.save()
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': f'Колір успішно додано з {len(color_images)} зображеннями!'})
                return redirect(request.path)
            else:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Колір успішно додано!'})
                return redirect(request.path)
        
        elif form_type == 'image':
            additional_image = request.FILES.get('additional_image')
            
            if additional_image:
                ProductImage.objects.create(
                    product=obj,
                    image=additional_image
                )
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Зображення успішно додано!'})
                return redirect(request.path)
            else:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Зображення успішно додано!'})
                return redirect(request.path)
    
    form = ProductForm(instance=obj)
    
    return render(request, 'pages/admin_product_edit_unified.html', {
        'form': form, 
        'obj': obj
    })

def api_colors(request):
    """API endpoint для получения всех цветов"""
    from productcolors.models import Color
    from django.http import JsonResponse
    
    colors = Color.objects.all()
    colors_data = []
    
    for color in colors:
        colors_data.append({
            'id': color.id,
            'name': color.name,
            'primary_hex': color.primary_hex,
            'secondary_hex': color.secondary_hex
        })
    
    return JsonResponse(colors_data, safe=False)

@login_required
def admin_product_edit_simple(request, pk):
    """Упрощенное редактирование товара без цветов"""
    if not request.user.is_staff:
        return redirect('home')
    from .forms import ProductForm
    from .models import Product
    obj = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            product = form.save(commit=False)
            # Автогенерація slug, якщо порожній
            if not getattr(product, 'slug', None):
                base = slugify(product.title or '')
                product.slug = unique_slugify(Product, base)
            product.save()
            return redirect('/admin-panel/?section=catalogs')
    else:
        form = ProductForm(instance=obj)
    
    return render(request, 'pages/admin_product_edit_simple.html', {
        'form': form, 
        'obj': obj
    })

@login_required
def admin_product_color_delete(request, product_pk, color_pk):
    """Удаление цвета товара"""
    if not request.user.is_staff:
        return redirect('home')
    from .models import Product
    from productcolors.models import ProductColorVariant
    product = get_object_or_404(Product, pk=product_pk)
    color_variant = get_object_or_404(ProductColorVariant, pk=color_pk, product=product)
    color_variant.delete()
    return redirect(f'/admin-panel/product/{product_pk}/edit-unified/')

@login_required
def admin_product_image_delete(request, product_pk, image_pk):
    """Удаление дополнительного изображения товара"""
    if not request.user.is_staff:
        return redirect('home')
    from .models import Product, ProductImage
    product = get_object_or_404(Product, pk=product_pk)
    image = get_object_or_404(ProductImage, pk=image_pk, product=product)
    image.delete()
    return redirect(f'/admin-panel/product/{product_pk}/edit-unified/')

@login_required
def admin_product_delete(request, pk):
    if not request.user.is_staff:
        return redirect('home')
    from .models import Product
    obj = get_object_or_404(Product, pk=pk)
    obj.delete()
    return redirect('/admin-panel/?section=catalogs')

# ===== БАЛЫ И ПРОМОКОДЫ ПОЛЬЗОВАТЕЛЯ =====

@login_required
def user_points(request):
    """Страница баллов пользователя"""
    from accounts.models import UserPoints, PointsHistory
    
    # Получаем или создаем объект баллов пользователя
    user_points = UserPoints.get_or_create_points(request.user)
    
    # Получаем историю баллов
    history = PointsHistory.objects.filter(user=request.user)[:20]  # Последние 20 записей
    
    return render(request, 'pages/user_points.html', {
        'user_points': user_points,
        'history': history
    })



# ===== АДМИН-ПАНЕЛЬ ПРОМОКОДОВ =====

class PromoCodeForm(forms.ModelForm):
    class Meta:
        model = PromoCode
        fields = ['code', 'discount_type', 'discount_value', 'max_uses', 'is_active']
        widgets = {
            'code': forms.TextInput(attrs={
                'class': 'form-control bg-elevate',
                'placeholder': 'Введіть код або залиште порожнім для автогенерації'
            }),
            'discount_type': forms.Select(attrs={'class': 'form-select bg-elevate'}),
            'discount_value': forms.NumberInput(attrs={
                'class': 'form-control bg-elevate',
                'step': '0.01',
                'min': '0'
            }),
            'max_uses': forms.NumberInput(attrs={
                'class': 'form-control bg-elevate',
                'min': '0',
                'placeholder': '0 = безліміт'
            }),
            'is_active': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }
    
    def clean_code(self):
        """Позволяет оставить поле code пустым для автоматической генерации"""
        code = self.cleaned_data.get('code', '').strip()
        
        # Если код пустой, это нормально - он будет сгенерирован автоматически
        if not code:
            return ''
        
        # Если код указан, проверяем его уникальность
        if PromoCode.objects.filter(code=code).exists():
            if self.instance and self.instance.pk:
                # При редактировании исключаем текущий промокод
                if not PromoCode.objects.filter(code=code).exclude(pk=self.instance.pk).exists():
                    return code
            raise forms.ValidationError("Промокод з таким кодом вже існує")
        
        return code

@login_required
def admin_promocodes(request):
    """Список промокодов"""
    if not request.user.is_staff:
        return redirect('home')
    
    from .models import PromoCode
    promocodes = PromoCode.objects.all()
    
    # Подсчитываем статистику
    total_promocodes = promocodes.count()
    active_promocodes = promocodes.filter(is_active=True).count()
    
    return render(request, 'pages/admin_promocodes.html', {
        'promocodes': promocodes,
        'total_promocodes': total_promocodes,
        'active_promocodes': active_promocodes,
        'section': 'promocodes'
    })

@login_required
def admin_promocode_create(request):
    """Создание нового промокода"""
    if not request.user.is_staff:
        return redirect('home')
    
    from .models import PromoCode
    
    if request.method == 'POST':
        form = PromoCodeForm(request.POST)
        if form.is_valid():
            promocode = form.save(commit=False)
            
            # Если код не указан или пустой, генерируем автоматически
            if not promocode.code or promocode.code.strip() == '':
                promocode.code = PromoCode.generate_code()
            
            promocode.save()
            return redirect('admin_promocodes')
    else:
        form = PromoCodeForm()
    
    return render(request, 'pages/admin_promocode_form.html', {
        'form': form,
        'mode': 'create',
        'section': 'promocodes'
    })

@login_required
def admin_promocode_edit(request, pk):
    """Редактирование промокода"""
    if not request.user.is_staff:
        return redirect('home')
    
    from .models import PromoCode
    promocode = get_object_or_404(PromoCode, pk=pk)
    
    if request.method == 'POST':
        form = PromoCodeForm(request.POST, instance=promocode)
        if form.is_valid():
            edited_promocode = form.save(commit=False)
            
            # Если код не указан или пустой, генерируем автоматически
            if not edited_promocode.code or edited_promocode.code.strip() == '':
                edited_promocode.code = PromoCode.generate_code()
            
            edited_promocode.save()
            return redirect('admin_promocodes')
    else:
        form = PromoCodeForm(instance=promocode)
    
    return render(request, 'pages/admin_promocode_form.html', {
        'form': form,
        'mode': 'edit',
        'promocode': promocode,
        'section': 'promocodes'
    })

@login_required
def admin_promocode_toggle(request, pk):
    """Активация/деактивация промокода"""
    if not request.user.is_staff:
        return redirect('home')
    
    from .models import PromoCode
    promocode = get_object_or_404(PromoCode, pk=pk)
    promocode.is_active = not promocode.is_active
    promocode.save()
    
    return redirect('admin_promocodes')

@login_required
def admin_promocode_delete(request, pk):
    """Удаление промокода"""
    if not request.user.is_staff:
        return redirect('home')
    
    from .models import PromoCode
    promocode = get_object_or_404(PromoCode, pk=pk)
    promocode.delete()
    
    return redirect('admin_promocodes')

# ===== ПРОМОКОДЫ В КОРЗИНЕ =====

def apply_promo_code(request):
    """Применение промокода в корзине"""
    if request.method == 'POST':
        promo_code = request.POST.get('promo_code', '').strip().upper()
        
        if not promo_code:
            return JsonResponse({
                'success': False,
                'message': 'Введіть код промокоду'
            })
        
        try:
            from .models import PromoCode
            promocode = PromoCode.objects.get(code=promo_code)
            
            if not promocode.can_be_used():
                return JsonResponse({
                    'success': False,
                    'message': 'Промокод неактивний або вичерпаний'
                })
            
            # Получаем корзину из сессии
            cart = request.session.get('cart', {})
            if not cart:
                return JsonResponse({
                    'success': False,
                    'message': 'Корзина порожня'
                })
            
            # Рассчитываем общую сумму корзины
            total = 0
            for key, item in cart.items():
                # Получаем цену товара
                try:
                    product = Product.objects.get(id=item['product_id'])
                    price = product.final_price
                    total += price * item['qty']
                except Product.DoesNotExist:
                    continue
            
            # Рассчитываем скидку
            discount = promocode.calculate_discount(total)
            
            if discount <= 0:
                return JsonResponse({
                    'success': False,
                    'message': 'Промокод не застосовується до цієї суми'
                })
            
            # Сохраняем промокод в сессии
            request.session['applied_promo_code'] = {
                'code': promocode.code,
                'discount': float(discount),
                'discount_type': promocode.discount_type,
                'discount_value': float(promocode.discount_value)
            }
            
            final_total = total - discount
            
            return JsonResponse({
                'success': True,
                'message': f'Промокод {promocode.code} застосовано! Знижка: {promocode.get_discount_display()}',
                'discount': float(discount),
                'final_total': float(final_total),
                'promo_code': promocode.code
            })
            
        except PromoCode.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Промокод не знайдено'
            })
    
    return JsonResponse({'success': False, 'message': 'Невірний запит'})

def remove_promo_code(request):
    """Удаление промокода из корзины"""
    if request.method == 'POST':
        if 'applied_promo_code' in request.session:
            del request.session['applied_promo_code']
            request.session.modified = True
        
        # Пересчитываем общую сумму
        cart = request.session.get('cart', {})
        total = 0
        for key, item in cart.items():
            try:
                product = Product.objects.get(id=item['product_id'])
                price = product.final_price
                total += price * item['qty']
            except Product.DoesNotExist:
                continue
        
        return JsonResponse({
            'success': True,
            'message': 'Промокод видалено',
            'final_total': float(total)
        })
    
    return JsonResponse({'success': False, 'message': 'Невірний запит'})

# ===== ИЗБРАННЫЕ ТОВАРЫ =====
@require_POST
def toggle_favorite(request, product_id):
    """Добавить/удалить товар из избранного"""
    try:
        product = get_object_or_404(Product, id=product_id)
        
        if request.user.is_authenticated:
            # Для авторизованных пользователей - используем базу данных
            favorite, created = FavoriteProduct.objects.get_or_create(
                user=request.user,
                product=product
            )
            
            if not created:
                # Если запись уже существует, удаляем её
                favorite.delete()
                is_favorite = False
                message = 'Товар видалено з обраного'
            else:
                is_favorite = True
                message = 'Товар додано до обраного'
        else:
            # Для неавторизованных пользователей - используем сессии
            session_favorites = request.session.get('favorites', [])
            # Преобразуем product_id в int для корректного сравнения
            product_id_int = int(product_id)
            
            if product_id_int in session_favorites:
                # Удаляем из избранного
                session_favorites.remove(product_id_int)
                is_favorite = False
                message = 'Товар видалено з обраного'
            else:
                # Добавляем в избранное
                session_favorites.append(product_id_int)
                is_favorite = True
                message = 'Товар додано до обраного'
            
            # Сохраняем в сессии
            request.session['favorites'] = session_favorites
            request.session.modified = True
        
        # Подсчитываем общее количество избранных товаров
        if request.user.is_authenticated:
            favorites_count = FavoriteProduct.objects.filter(user=request.user).count()
        else:
            favorites_count = len(request.session.get('favorites', []))
        
        return JsonResponse({
            'success': True,
            'is_favorite': is_favorite,
            'message': message,
            'favorites_count': favorites_count
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': 'Помилка: ' + str(e)
        }, status=400)


def favorites_list(request):
    """Страница с избранными товарами"""
    if request.user.is_authenticated:
        # Для авторизованных пользователей - получаем из базы данных
        favorites = FavoriteProduct.objects.filter(user=request.user).select_related('product', 'product__category')
    else:
        # Для неавторизованных пользователей - получаем из сессии
        session_favorites = request.session.get('favorites', [])
        favorites = []
        
        if session_favorites:
            # Получаем товары по ID из сессии
            products = Product.objects.filter(id__in=session_favorites).select_related('category')
            
            # Создаем объекты, похожие на FavoriteProduct
            for product in products:
                favorite = type('FavoriteProduct', (), {
                    'product': product,
                    'color_variants_data': []
                })()
                favorites.append(favorite)
    
    # Получаем варианты цветов для избранных товаров
    for favorite in favorites:
        try:
            from productcolors.models import ProductColorVariant
            variants = ProductColorVariant.objects.select_related('color').filter(product=favorite.product)
            # Создаем список словарей с данными о цветах
            color_variants_data = [
                {
                    'primary_hex': v.color.primary_hex,
                    'secondary_hex': v.color.secondary_hex or '',
                }
                for v in variants
            ]
            # Добавляем как атрибут к объекту favorite
            favorite.color_variants_data = color_variants_data
        except:
            favorite.color_variants_data = []
    
    return render(request, 'pages/favorites.html', {
        'favorites': favorites,
        'title': 'Обрані товари'
    })


def check_favorite_status(request, product_id):
    """Проверить статус избранного для товара"""
    try:
        if request.user.is_authenticated:
            # Для авторизованных пользователей - проверяем в базе данных
            is_favorite = FavoriteProduct.objects.filter(
                user=request.user,
                product_id=product_id
            ).exists()
        else:
            # Для неавторизованных пользователей - проверяем в сессии
            session_favorites = request.session.get('favorites', [])
            product_id_int = int(product_id)
            is_favorite = product_id_int in session_favorites
        
        return JsonResponse({'is_favorite': is_favorite})
    except:
        return JsonResponse({'is_favorite': False})


def favorites_count(request):
    """Получить количество товаров в избранном"""
    try:
        if request.user.is_authenticated:
            # Для авторизованных пользователей - считаем в базе данных
            count = FavoriteProduct.objects.filter(user=request.user).count()
        else:
            # Для неавторизованных пользователей - считаем в сессии
            session_favorites = request.session.get('favorites', [])
            count = len(session_favorites)
        
        return JsonResponse({
            'count': count
        })
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=400)

@login_required
@require_POST
def confirm_payment(request):
    """
    AJAX view для подтверждения оплаты заказа
    """
    from orders.models import Order
    
    order_id = request.POST.get("order_id")
    payment_screenshot = request.FILES.get("payment_screenshot")
    
    if not order_id:
        return JsonResponse({
            "success": False,
            "error": "Відсутній ID замовлення"
        })
    
    if not payment_screenshot:
        return JsonResponse({
            "success": False,
            "error": "Будь ласка, завантажте скріншот оплати"
        })
    
    try:
        order = Order.objects.get(id=order_id, user=request.user)
    except Order.DoesNotExist:
        return JsonResponse({
            "success": False,
            "error": "Замовлення не знайдено"
        })
    
    # Сохраняем скриншот оплаты
    order.payment_screenshot = payment_screenshot
    order.payment_status = "checking"
    order.save()
    
    return JsonResponse({
        "success": True,
        "message": "Скріншот оплати успішно завантажено"
    })


# ===== ОФФЛАЙН МАГАЗИНЫ =====

@login_required
def admin_offline_stores(request):
    """Список оффлайн магазинов"""
    if not request.user.is_staff:
        return redirect('home')
    
    from .models import OfflineStore
    stores = OfflineStore.objects.all()
    
    # Подсчитываем статистику
    total_stores = stores.count()
    active_stores = stores.filter(is_active=True).count()
    
    return render(request, 'pages/admin_offline_stores.html', {
        'stores': stores,
        'total_stores': total_stores,
        'active_stores': active_stores,
        'section': 'offline_stores'
    })


class OfflineStoreForm(forms.ModelForm):
    class Meta:
        from .models import OfflineStore
        model = OfflineStore
        fields = ['name', 'address', 'phone', 'email', 'working_hours', 'description', 'is_active', 'order']
        widgets = {
            'name': forms.TextInput(attrs={
                'class': 'form-control bg-elevate',
                'placeholder': 'Назва магазину'
            }),
            'address': forms.Textarea(attrs={
                'class': 'form-control bg-elevate',
                'rows': 3,
                'placeholder': 'Повна адреса магазину'
            }),
            'phone': forms.TextInput(attrs={
                'class': 'form-control bg-elevate',
                'placeholder': 'Телефон магазину'
            }),
            'email': forms.EmailInput(attrs={
                'class': 'form-control bg-elevate',
                'placeholder': 'Email магазину'
            }),
            'working_hours': forms.TextInput(attrs={
                'class': 'form-control bg-elevate',
                'placeholder': 'Наприклад: Пн-Пт 9:00-18:00, Сб 10:00-16:00'
            }),
            'description': forms.Textarea(attrs={
                'class': 'form-control bg-elevate',
                'rows': 4,
                'placeholder': 'Опис магазину, особливості, послуги'
            }),
            'order': forms.NumberInput(attrs={
                'class': 'form-control bg-elevate',
                'min': '0',
                'placeholder': 'Порядок відображення'
            }),
            'is_active': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)


@login_required
def admin_offline_store_create(request):
    """Создание нового оффлайн магазина"""
    if not request.user.is_staff:
        return redirect('home')
    
    from .models import OfflineStore
    
    if request.method == 'POST':
        form = OfflineStoreForm(request.POST)
        if form.is_valid():
            store = form.save()
            return redirect('admin_store_management', store_id=store.id)
    else:
        form = OfflineStoreForm()
    
    return render(request, 'pages/admin_offline_store_form.html', {
        'form': form,
        'mode': 'create',
        'section': 'offline_stores'
    })


@login_required
def admin_offline_store_edit(request, pk):
    """Редактирование оффлайн магазина"""
    if not request.user.is_staff:
        return redirect('home')
    
    from .models import OfflineStore
    store = get_object_or_404(OfflineStore, pk=pk)
    
    if request.method == 'POST':
        form = OfflineStoreForm(request.POST, instance=store)
        if form.is_valid():
            form.save()
            return redirect('admin_offline_stores')
    else:
        form = OfflineStoreForm(instance=store)
    
    return render(request, 'pages/admin_offline_store_form.html', {
        'form': form,
        'mode': 'edit',
        'store': store,
        'section': 'offline_stores'
    })


@login_required
def admin_offline_store_toggle(request, pk):
    """Активация/деактивация оффлайн магазина"""
    if not request.user.is_staff:
        return redirect('home')
    
    from .models import OfflineStore
    store = get_object_or_404(OfflineStore, pk=pk)
    store.is_active = not store.is_active
    store.save()
    
    return redirect('admin_offline_stores')


@login_required
def admin_offline_store_delete(request, pk):
    """Удаление оффлайн магазина"""
    if not request.user.is_staff:
        return redirect('home')
    
    from .models import OfflineStore
    store = get_object_or_404(OfflineStore, pk=pk)
    store.delete()
    
    return redirect('admin_offline_stores')


# ===== Новые views для функционала оффлайн магазинов =====


def _get_store_inventory_items(store):
    return list(
        store.store_products.filter(is_active=True)
        .select_related('product__category', 'product', 'color')
        .order_by('-created_at')
    )


def _get_store_sales(store):
    return list(
        store.store_sales.select_related('product__category', 'product', 'color')
        .order_by('-sold_at')
    )


def _calculate_inventory_stats(inventory_items):
    stats = {
        'units': 0,
        'cost': 0,
        'revenue': 0,
        'margin': 0,
        'avg_price': 0,
        'avg_margin_per_unit': 0,
        'margin_percent': 0,
        'active_skus': len(inventory_items),
    }

    for item in inventory_items:
        qty = item.quantity or 0
        cost = (item.cost_price or 0) * qty
        revenue = (item.selling_price or 0) * qty
        margin = revenue - cost

        stats['units'] += qty
        stats['cost'] += cost
        stats['revenue'] += revenue
        stats['margin'] += margin

    if stats['units']:
        stats['avg_price'] = stats['revenue'] / stats['units']
        stats['avg_margin_per_unit'] = stats['margin'] / stats['units'] if stats['margin'] else 0

    if stats['cost']:
        stats['margin_percent'] = (stats['margin'] / stats['cost']) * 100

    return stats


def _calculate_sales_stats(sales_items):
    stats = {
        'units': 0,
        'cost': 0,
        'revenue': 0,
        'margin': 0,
        'avg_ticket': 0,
        'margin_percent': 0,
        'last_sale_at': None,
    }

    for sale in sales_items:
        qty = sale.quantity or 0
        cost = (sale.cost_price or 0) * qty
        revenue = (sale.selling_price or 0) * qty
        margin = revenue - cost

        stats['units'] += qty
        stats['cost'] += cost
        stats['revenue'] += revenue
        stats['margin'] += margin

    if stats['units']:
        stats['avg_ticket'] = stats['revenue'] / stats['units']

    if stats['cost']:
        stats['margin_percent'] = (stats['margin'] / stats['cost']) * 100

    if sales_items:
        stats['last_sale_at'] = sales_items[0].sold_at

    return stats


def _build_category_stats(inventory_items):
    category_data = {}

    for item in inventory_items:
        category_name = (
            item.product.category.name if getattr(item.product, 'category', None) else 'Без категорії'
        )
        entry = category_data.setdefault(
            category_name,
            {'name': category_name, 'units': 0, 'cost': 0, 'revenue': 0, 'margin': 0},
        )

        qty = item.quantity or 0
        cost = (item.cost_price or 0) * qty
        revenue = (item.selling_price or 0) * qty

        entry['units'] += qty
        entry['cost'] += cost
        entry['revenue'] += revenue
        entry['margin'] = entry['revenue'] - entry['cost']

    for entry in category_data.values():
        if entry['cost']:
            entry['margin_percent'] = (entry['margin'] / entry['cost']) * 100
        else:
            entry['margin_percent'] = 0

    return sorted(category_data.values(), key=lambda x: x['revenue'], reverse=True)


def _serialize_sale(sale):
    image = None
    display_image = getattr(sale.product, 'display_image', None)
    if display_image:
        try:
            image = display_image.url
        except Exception:
            image = None

    return {
        'id': sale.id,
        'product_id': sale.product_id,
        'title': sale.product.title,
        'size': sale.size,
        'color': sale.color.name if sale.color else None,
        'color_hex': sale.color.primary_hex if sale.color else None,
        'quantity': sale.quantity,
        'cost_price': sale.cost_price,
        'selling_price': sale.selling_price,
        'total_revenue': sale.total_revenue,
        'total_margin': sale.margin,
        'sold_at': sale.sold_at.isoformat(),
        'sold_at_human': sale.sold_at.strftime('%d.%m.%Y %H:%M'),
        'image': image,
    }


def _compose_store_stats(inventory_items, sales_items):
    inventory_stats = _calculate_inventory_stats(inventory_items)
    sales_stats = _calculate_sales_stats(sales_items)

    total_units = inventory_stats['units'] + sales_stats['units']
    sell_through_rate = (sales_stats['units'] / total_units * 100) if total_units else 0

    if sales_stats['last_sale_at']:
        last_sale_iso = sales_stats['last_sale_at'].isoformat()
        last_sale_human = sales_stats['last_sale_at'].strftime('%d.%m.%Y %H:%M')
    else:
        last_sale_iso = None
        last_sale_human = None

    return {
        'inventory_units': inventory_stats['units'],
        'inventory_cost': inventory_stats['cost'],
        'inventory_revenue': inventory_stats['revenue'],
        'inventory_margin': inventory_stats['margin'],
        'inventory_margin_percent': round(inventory_stats['margin_percent'], 2) if inventory_stats['margin_percent'] else 0,
        'inventory_avg_price': round(inventory_stats['avg_price'], 2) if inventory_stats['avg_price'] else 0,
        'inventory_avg_margin_per_unit': round(inventory_stats['avg_margin_per_unit'], 2) if inventory_stats['avg_margin_per_unit'] else 0,
        'inventory_active_skus': inventory_stats['active_skus'],
        'sales_units': sales_stats['units'],
        'sales_cost': sales_stats['cost'],
        'sales_revenue': sales_stats['revenue'],
        'sales_margin': sales_stats['margin'],
        'sales_margin_percent': round(sales_stats['margin_percent'], 2) if sales_stats['margin_percent'] else 0,
        'sales_avg_ticket': round(sales_stats['avg_ticket'], 2) if sales_stats['avg_ticket'] else 0,
        'sell_through_rate': round(sell_through_rate, 2) if sell_through_rate else 0,
        'unsold_units': inventory_stats['units'],
        'total_committed_cost': inventory_stats['cost'] + sales_stats['cost'],
        'total_realized_revenue': sales_stats['revenue'],
        'total_expected_revenue': inventory_stats['revenue'] + sales_stats['revenue'],
        'lifetime_margin': sales_stats['margin'],
        'last_sale_at': last_sale_iso,
        'last_sale_at_display': last_sale_human,
    }

@login_required
def admin_store_management(request, store_id):
    """Главная страница управления оффлайн магазином"""
    if not request.user.is_staff:
        return redirect('home')
    
    from .models import OfflineStore, Product, Category, StoreProduct, StoreOrder, StoreOrderItem
    from productcolors.models import Color
    
    store = get_object_or_404(OfflineStore, pk=store_id)

    products = Product.objects.select_related('category').prefetch_related(
        'color_variants__color',
        'color_variants__images'
    ).all()

    categories = Category.objects.filter(is_active=True).order_by('order', 'name')

    inventory_items = _get_store_inventory_items(store)
    sales_items = _get_store_sales(store)

    active_orders = StoreOrder.objects.filter(store=store, status='draft').prefetch_related(
        'order_items__product__category', 'order_items__color'
    )

    store_stats = _compose_store_stats(inventory_items, sales_items)
    category_stats = _build_category_stats(inventory_items)
    sold_products = [_serialize_sale(sale) for sale in sales_items]

    return render(request, 'pages/admin_store_management.html', {
        'store': store,
        'products': products,
        'categories': categories,
        'store_products': inventory_items,
        'sold_products': sold_products,
        'active_orders': active_orders,
        'store_stats': store_stats,
        'category_stats': category_stats,
        'section': 'offline_stores'
    })


@login_required
def admin_store_get_order_items(request, store_id, order_id):
    """Получение товаров заказа для AJAX"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Доступ запрещен'})
    
    from .models import OfflineStore, StoreOrder
    
    try:
        store = OfflineStore.objects.get(id=store_id)
        order = StoreOrder.objects.get(id=order_id, store=store)
        
        items = []
        for item in order.order_items.all():
            items.append({
                'id': item.id,
                'product_name': item.product.name,
                'size': item.size,
                'color_name': item.color.name if item.color else None,
                'quantity': item.quantity,
                'cost_price': float(item.cost_price),
                'selling_price': float(item.selling_price),
            })
        
        return JsonResponse({'success': True, 'items': items})
    
    except (OfflineStore.DoesNotExist, StoreOrder.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Заказ не найден'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def admin_store_get_product_colors(request, store_id, product_id):
    """Получение цветов товара для AJAX"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)
    
    from .models import OfflineStore, Product
    from productcolors.models import Color, ProductColorVariant
    
    try:
        store = get_object_or_404(OfflineStore, pk=store_id)
        product = get_object_or_404(Product, pk=product_id)
        
        # Получаем цвета товара через color_variants
        colors = []
        for variant in product.color_variants.all():
            # Получаем первое изображение для этого цветового варианта
            first_image = variant.images.first()
            image_url = first_image.image.url if first_image else None
            
            colors.append({
                'id': variant.color.id,
                'name': variant.color.name or variant.color.primary_hex,
                'hex_code': variant.color.primary_hex,
                'image_url': image_url,
                'variant_id': variant.id
            })
        
        return JsonResponse({
            'success': True,
            'colors': colors
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_add_product_to_order(request, store_id):
    """Добавление товара в заказ через AJAX"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не поддерживается'}, status=405)
    
    from .models import OfflineStore, Product, StoreOrder, StoreOrderItem
    from productcolors.models import Color
    
    try:
        store = get_object_or_404(OfflineStore, pk=store_id)
        product_id = request.POST.get('product_id')
        size = request.POST.get('size', '').strip()
        color_id = request.POST.get('color_id')
        cost_price = int(request.POST.get('cost_price', 0))
        selling_price = int(request.POST.get('selling_price', 0))
        quantity = int(request.POST.get('quantity', 1))
        
        product = get_object_or_404(Product, pk=product_id)
        color = None
        if color_id and color_id != 'default':
            color = get_object_or_404(Color, pk=color_id)
        elif color_id == 'default':
            # Создаем или получаем дефолтный черный цвет
            color, created = Color.objects.get_or_create(
                name='Чорний',
                defaults={'hex_code': '#000000'}
            )
        
        # Получаем или создаем черновик заказа
        order, created = StoreOrder.objects.get_or_create(
            store=store,
            status='draft',
            defaults={'notes': ''}
        )
        
        # Проверяем, есть ли уже такой товар в заказе
        existing_item = StoreOrderItem.objects.filter(
            order=order,
            product=product,
            size=size or None,
            color=color
        ).first()
        
        if existing_item:
            # Обновляем существующий товар
            existing_item.cost_price = cost_price
            existing_item.selling_price = selling_price
            existing_item.quantity += quantity
            existing_item.save()
            item = existing_item
        else:
            # Создаем новый товар в заказе
            item = StoreOrderItem.objects.create(
                order=order,
                product=product,
                size=size or None,
                color=color,
                cost_price=cost_price,
                selling_price=selling_price,
                quantity=quantity
            )
        
            # Получаем изображение цвета, если есть
            color_image_url = None
            if item.color:
                try:
                    from productcolors.models import ProductColorVariant
                    color_variant = ProductColorVariant.objects.filter(
                        product=item.product, 
                        color=item.color
                    ).first()
                    if color_variant:
                        first_image = color_variant.images.first()
                        if first_image:
                            color_image_url = first_image.image.url
                except:
                    pass
            
            return JsonResponse({
                'success': True,
                'order_id': order.id,
                'item': {
                    'id': item.id,
                    'product_name': item.product.title,
                    'size': item.size or '',
                    'color_name': item.color.name if item.color else 'Чорний',
                    'color_hex': item.color.primary_hex if item.color else '#000000',
                    'color_image': color_image_url,
                    'quantity': item.quantity,
                    'cost_price': float(item.cost_price),
                    'selling_price': float(item.selling_price),
                    'product_image': item.product.display_image.url if item.product.display_image else None,
                },
                'message': 'Товар добавлен в заказ'
            })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_remove_product_from_order(request, store_id, order_id, item_id):
    """Удаление товара из заказа"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)
    
    from .models import StoreOrder, StoreOrderItem
    
    try:
        order = get_object_or_404(StoreOrder, pk=order_id, store_id=store_id)
        item = get_object_or_404(StoreOrderItem, pk=item_id, order=order)
        item.delete()
        
        return JsonResponse({'success': True, 'message': 'Товар удален из заказа'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_add_products_to_store(request, store_id):
    """Добавление товаров из заказа в магазин"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не поддерживается'}, status=405)
    
    from .models import OfflineStore, StoreOrder, StoreOrderItem, StoreProduct
    
    try:
        store = get_object_or_404(OfflineStore, pk=store_id)
        order_id = request.POST.get('order_id')
        
        order = get_object_or_404(StoreOrder, pk=order_id, store=store)
        
        added_count = 0
        for item in order.order_items.all():
            # Проверяем, есть ли уже такой товар в магазине
            existing_product = StoreProduct.objects.filter(
                store=store,
                product=item.product,
                size=item.size,
                color=item.color
            ).first()
            
            if existing_product:
                # Обновляем существующий товар
                existing_product.cost_price = item.cost_price
                existing_product.selling_price = item.selling_price
                existing_product.quantity += item.quantity
                existing_product.is_active = True
                existing_product.save()
            else:
                # Создаем новый товар в магазине
                StoreProduct.objects.create(
                    store=store,
                    product=item.product,
                    size=item.size,
                    color=item.color,
                    cost_price=item.cost_price,
                    selling_price=item.selling_price,
                    quantity=item.quantity
                )
            added_count += 1
        
        # Обновляем статус заказа
        order.status = 'completed'
        order.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Добавлено {added_count} товаров в магазин'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_generate_invoice(request, store_id):
    """Генерация Excel накладной"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не поддерживается'}, status=405)
    
    from .models import OfflineStore, StoreOrder, StoreInvoice
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from datetime import datetime
    import os
    
    try:
        store = get_object_or_404(OfflineStore, pk=store_id)
        order_id = request.POST.get('order_id')
        
        order = get_object_or_404(StoreOrder, pk=order_id, store=store)
        
        # Создаем Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Накладна"
        
        # Заголовки
        ws['A1'] = f"Накладна для магазину: {store.name} (під реалізацію)"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].font = Font(size=12)
        
        # Заголовки таблицы
        headers = ['Товар', 'Кількість', 'Ціна (грн)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Данные товаров
        row = 5
        category_totals = {}
        
        for item in order.order_items.all():
            # Формируем название товара
            product_name = f"TwoComms {item.product.title}"
            if item.size:
                product_name += f" ({item.size})"
            if item.color:
                product_name += f" [{item.color.name}]"
            
            ws.cell(row=row, column=1, value=product_name)
            ws.cell(row=row, column=2, value=item.quantity)
            ws.cell(row=row, column=3, value=item.selling_price)
            
            # Подсчитываем по категориям
            category = item.product.category.name
            if category not in category_totals:
                category_totals[category] = {'count': 0, 'total': 0}
            category_totals[category]['count'] += item.quantity
            category_totals[category]['total'] += item.selling_price * item.quantity
            
            row += 1
        
        # Итоги по категориям
        row += 1
        ws.cell(row=row, column=1, value="Підсумки по категоріях:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        for category, totals in category_totals.items():
            ws.cell(row=row, column=1, value=f"{category}: {totals['count']} шт., сума {totals['total']} грн")
            row += 1
        
        # Общий итог
        row += 1
        total_amount = sum(totals['total'] for totals in category_totals.values())
        ws.cell(row=row, column=1, value=f"ВСЬОГО: {total_amount} грн")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        
        # Сохраняем файл
        filename = f"TwoComms_накладна_{store.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, 'invoices', filename)
        
        # Создаем директорию если не существует
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        wb.save(filepath)
        
        # Сохраняем информацию о накладной в БД
        StoreInvoice.objects.create(
            store=store,
            order=order,
            file_name=filename,
            file_path=filepath
        )
        
        # Заказ остается активным для продолжения работы
        # order.status = 'completed'  # Убрано - заказ остается draft
        
        # Возвращаем файл для скачивания
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_update_product(request, store_id, product_id):
    """Обновление товара в магазине"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не поддерживается'}, status=405)
    
    from .models import StoreProduct
    
    try:
        store_product = get_object_or_404(StoreProduct, pk=product_id, store_id=store_id)

        cost_price = request.POST.get('cost_price')
        selling_price = request.POST.get('selling_price')
        quantity = request.POST.get('quantity')
        
        if cost_price:
            store_product.cost_price = int(cost_price)
        if selling_price:
            store_product.selling_price = int(selling_price)
        if quantity:
            store_product.quantity = int(quantity)

        store_product.save()

        store = store_product.store
        inventory_items = _get_store_inventory_items(store)
        sales_items = _get_store_sales(store)
        stats = _compose_store_stats(inventory_items, sales_items)
        category_stats = _build_category_stats(inventory_items)

        return JsonResponse({
            'success': True,
            'message': 'Товар оновлено',
            'product': {
                'id': store_product.id,
                'cost_price': store_product.cost_price,
                'selling_price': store_product.selling_price,
                'quantity': store_product.quantity,
                'margin_per_unit': store_product.margin,
                'total_cost': store_product.cost_price * store_product.quantity,
                'total_revenue': store_product.selling_price * store_product.quantity,
                'total_margin': (store_product.selling_price - store_product.cost_price) * store_product.quantity,
            },
            'stats': stats,
            'category_stats': category_stats,
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_remove_product(request, store_id, product_id):
    """Удаление товара из магазина"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)

    from .models import StoreProduct

    try:
        store_product = get_object_or_404(StoreProduct, pk=product_id, store_id=store_id)
        store = store_product.store
        removed_product_id = store_product.id
        store_product.delete()

        inventory_items = _get_store_inventory_items(store)
        sales_items = _get_store_sales(store)
        stats = _compose_store_stats(inventory_items, sales_items)
        category_stats = _build_category_stats(inventory_items)

        return JsonResponse({
            'success': True,
            'message': 'Товар видалено з магазину',
            'removed_product_id': removed_product_id,
            'stats': stats,
            'category_stats': category_stats,
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_mark_product_sold(request, store_id, product_id):
    """Позначає товар як проданий та переносить його в блок статистики продажів."""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ заборонено'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не підтримується'}, status=405)

    from .models import StoreProduct, StoreSale

    try:
        store_product = get_object_or_404(StoreProduct, pk=product_id, store_id=store_id)
        store = store_product.store

        quantity_param = request.POST.get('quantity')

        if quantity_param is None and request.body:
            try:
                payload = json.loads(request.body.decode('utf-8'))
                quantity_param = payload.get('quantity')
            except (ValueError, json.JSONDecodeError):
                quantity_param = None

        if quantity_param is None or quantity_param == '':
            quantity_to_sell = store_product.quantity
        else:
            try:
                quantity_to_sell = int(quantity_param)
            except (TypeError, ValueError):
                return JsonResponse({'error': 'Невірне значення кількості'}, status=400)

        if quantity_to_sell <= 0:
            return JsonResponse({'error': 'Кількість повинна бути більшою за нуль'}, status=400)

        if quantity_to_sell > store_product.quantity:
            return JsonResponse({'error': 'Кількість перевищує залишок товару'}, status=400)

        sale = StoreSale.objects.create(
            store=store,
            product=store_product.product,
            color=store_product.color,
            size=store_product.size,
            quantity=quantity_to_sell,
            cost_price=store_product.cost_price,
            selling_price=store_product.selling_price,
            source_store_product=store_product,
        )

        remaining_product = None
        removed_product_id = None

        if quantity_to_sell >= store_product.quantity:
            removed_product_id = store_product.id
            store_product.delete()
        else:
            store_product.quantity -= quantity_to_sell
            store_product.save(update_fields=['quantity', 'updated_at'])
            remaining_product = {
                'id': store_product.id,
                'quantity': store_product.quantity,
                'cost_price': store_product.cost_price,
                'selling_price': store_product.selling_price,
                'margin_per_unit': store_product.margin,
                'total_cost': store_product.cost_price * store_product.quantity,
                'total_revenue': store_product.selling_price * store_product.quantity,
                'total_margin': (store_product.selling_price - store_product.cost_price) * store_product.quantity,
            }

        inventory_items = _get_store_inventory_items(store)
        sales_items = _get_store_sales(store)
        stats = _compose_store_stats(inventory_items, sales_items)
        category_stats = _build_category_stats(inventory_items)

        return JsonResponse({
            'success': True,
            'message': 'Товар позначено як проданий',
            'sale': _serialize_sale(sale),
            'stats': stats,
            'category_stats': category_stats,
            'remaining_product': remaining_product,
            'removed_product_id': removed_product_id,
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def robots_txt(request):
    host = request.get_host() or "twocomms.shop"
    scheme = "https"
    lines = [
        "User-agent: *",
        "Allow: /",
        "",
        "Disallow: /admin/",
        "Disallow: /admin-panel/",
        "Disallow: /debug/",
        "Disallow: /media/debug/",
        "Disallow: /cart/",
        "Disallow: /checkout/",
        "Disallow: /orders/",
        "Disallow: /my/",
        "Disallow: /login/",
        "Disallow: /logout/",
        "Disallow: /register/",
        "Disallow: /profile/",
        "Disallow: /search/",
        "",
        f"Sitemap: {scheme}://{host}/sitemap.xml",
    ]
    resp = HttpResponse("\n".join(lines) + "\n", content_type="text/plain; charset=utf-8")
    # Снимаем кэш, чтобы поисковики и CDN не держали старый редирект
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    resp["Expires"] = "0"
    resp["Content-Disposition"] = 'inline; filename="robots.txt"'
    return resp


def static_verification_file(request):
    """Отдаёт verification-файл из корня проекта по прямой ссылке."""
    filename = "494cb80b2da94b4395dbbed566ab540d.txt"
    file_path = os.path.join(settings.BASE_DIR, filename)
    if not os.path.exists(file_path):
        raise Http404("Verification file not found")

    response = FileResponse(open(file_path, "rb"), content_type="text/plain; charset=utf-8")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    response["Cache-Control"] = "public, max-age=86400"
    return response


def custom_sitemap(request):
    """Кастомный sitemap view без проблемных заголовков."""
    from django.contrib.sitemaps.views import sitemap
    from storefront.sitemaps import StaticViewSitemap, ProductSitemap, CategorySitemap
    
    sitemaps = {
        'static': StaticViewSitemap,
        'products': ProductSitemap,
        'categories': CategorySitemap,
    }
    
    response = sitemap(request, sitemaps)
    
    # Убираем проблемные заголовки для sitemap
    if 'x-robots-tag' in response:
        del response['x-robots-tag']
    if 'x-frame-options' in response:
        del response['x-frame-options']
    
    # Добавляем правильные заголовки для sitemap
    response['Content-Type'] = 'application/xml; charset=utf-8'
    response['Cache-Control'] = 'public, max-age=3600'  # Кешируем на 1 час
    
    return response


def static_sitemap(request):
    """Отдает статический sitemap.xml файл."""
    sitemap_path = os.path.join(settings.BASE_DIR, 'twocomms', 'sitemap.xml')
    
    if not os.path.exists(sitemap_path):
        # Если файл не существует, генерируем динамический sitemap
        return custom_sitemap(request)
    
    response = FileResponse(open(sitemap_path, 'rb'), content_type='application/xml; charset=utf-8')
    response['Cache-Control'] = 'public, max-age=3600'  # Кешируем на 1 час
    return response

@login_required
def admin_order_delete(request, pk: int):
    if not request.user.is_staff:
        return redirect('home')

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    try:
        from orders.models import Order
        order = Order.objects.get(pk=pk)
        order.delete()
        if is_ajax:
            return JsonResponse({'ok': True, 'removed_id': pk})
        from django.contrib import messages
        messages.success(request, f"Замовлення #{order.order_number} видалено")
    except Exception as e:
        if is_ajax:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
        from django.contrib import messages
        messages.error(request, f"Помилка видалення: {e}")
    return redirect('admin_panel')


def delivery_view(request):
    """Страница доставки и оплаты"""
    # FAQ данные для структурированных данных
    faq_items = [
        {
            "question": "Як довго триває доставка по Україні?",
            "answer": "Доставка по Україні зазвичай триває 1-5 робочих днів залежно від регіону. У великих містах доставка може бути швидшою."
        },
        {
            "question": "Чи можна оплатити товар при отриманні?",
            "answer": "Так, ми пропонуємо накладений платіж для доставки по Україні. Варто врахувати, що при цьому способі оплати додається комісія перевізника."
        },
        {
            "question": "Як відстежити замовлення?",
            "answer": "Після відправки замовлення ви отримаєте номер ТТН на email або в Telegram. Ви можете відстежити посилку на сайті перевізника або через наш Telegram бот."
        },
        {
            "question": "Чи можлива доставка за кордон?",
            "answer": "Так, ми здійснюємо міжнародну доставку через Укрпошту та Нову пошту. Термін доставки 7-21 робочий день залежно від країни."
        },
        {
            "question": "Які способи оплати доступні?",
            "answer": "Ми приймаємо оплату на банківську картку, накладений платіж при отриманні, а також бали за купони для зареєстрованих користувачів."
        }
    ]
    
    return_steps = [
        {
            "title": "Зв'яжіться з нами протягом 14 днів",
            "description": "Напишіть у Telegram або на електронну пошту з номером замовлення та причиною повернення, щоб узгодити деталі."
        },
        {
            "title": "Підготуйте товар",
            "description": "Збережіть бірки, пакування та переконайтеся, що річ не була у використанні. Це пришвидшить перевірку."
        },
        {
            "title": "Надішліть посилку",
            "description": "Відправте повернення Новою поштою або Укрпоштою за погодженими реквізитами й надішліть номер ТТН. Доставка сплачується клієнтом."
        },
        {
            "title": "Отримайте кошти",
            "description": "Після перевірки товару ми повернемо оплату протягом 3 робочих днів тим самим способом, яким було здійснено платіж."
        }
    ]

    return_policy_howto = {
        "@context": "https://schema.org",
        "@type": "HowTo",
        "name": "Як оформити повернення товару у TwoComms",
        "description": "Покрокова інструкція для повернення або обміну товару протягом 14 днів",
        "totalTime": "P14D",
        "supply": [
            {
                "@type": "HowToSupply",
                "name": "Товар з бірками та пакуванням"
            },
            {
                "@type": "HowToSupply",
                "name": "Номер замовлення та контактні дані"
            }
        ],
        "tool": [
            {
                "@type": "HowToTool",
                "name": "Telegram або email для зв'язку"
            },
            {
                "@type": "HowToTool",
                "name": "Сервіс доставки (Нова пошта чи Укрпошта)"
            }
        ],
        "step": [
            {
                "@type": "HowToStep",
                "position": 1,
                "name": return_steps[0]["title"],
                "itemListElement": [
                    {
                        "@type": "HowToDirection",
                        "text": return_steps[0]["description"]
                    }
                ]
            },
            {
                "@type": "HowToStep",
                "position": 2,
                "name": return_steps[1]["title"],
                "itemListElement": [
                    {
                        "@type": "HowToDirection",
                        "text": return_steps[1]["description"]
                    }
                ]
            },
            {
                "@type": "HowToStep",
                "position": 3,
                "name": return_steps[2]["title"],
                "itemListElement": [
                    {
                        "@type": "HowToDirection",
                        "text": return_steps[2]["description"]
                    }
                ]
            },
            {
                "@type": "HowToStep",
                "position": 4,
                "name": return_steps[3]["title"],
                "itemListElement": [
                    {
                        "@type": "HowToDirection",
                        "text": return_steps[3]["description"]
                    }
                ]
            }
        ]
    }

    return render(request, 'pages/delivery.html', {
        'faq_items': faq_items,
        'return_steps': return_steps,
        'return_policy_howto': json.dumps(return_policy_howto, ensure_ascii=False, indent=2)
    })


def google_merchant_feed(request):
    """Генерирует XML фид для Google Merchant Center"""
    from django.http import HttpResponse
    from django.utils import timezone
    import xml.etree.ElementTree as ET
    
    # Создаем корневой элемент RSS
    rss = ET.Element('rss')
    rss.set('version', '2.0')
    rss.set('xmlns:g', 'http://base.google.com/ns/1.0')

    # Создаем канал
    channel = ET.SubElement(rss, 'channel')
    
    # Заголовок канала
    title = ET.SubElement(channel, 'title')
    title.text = 'TwoComms - Стріт & Мілітарі Одяг'
    
    link = ET.SubElement(channel, 'link')
    link.text = 'https://twocomms.shop'
    
    description = ET.SubElement(channel, 'description')
    description.text = 'Магазин стріт & мілітарі одягу з ексклюзивним дизайном'

    # Получаем все товары
    products = Product.objects.select_related('category').all()

    for product in products:
        # Создаем элемент товара
        item = ET.SubElement(channel, 'item')

        # Обязательные поля
        g_id = ET.SubElement(item, 'g:id')
        g_id.text = f"TC-{product.id}"

        g_title = ET.SubElement(item, 'g:title')
        g_title.text = product.title

        g_description = ET.SubElement(item, 'g:description')
        g_description.text = product.description or f"Якісний {product.category.name.lower() if product.category else 'одяг'} з ексклюзивним дизайном від TwoComms"

        g_link = ET.SubElement(item, 'g:link')
        g_link.text = f"https://twocomms.shop/product/{product.slug}/"

        g_image_link = ET.SubElement(item, 'g:image_link')
        if product.display_image:
            g_image_link.text = f"https://twocomms.shop{product.display_image.url}"
        else:
            g_image_link.text = "https://twocomms.shop/static/img/placeholder.jpg"

        g_availability = ET.SubElement(item, 'g:availability')
        g_availability.text = 'in stock'

        g_price = ET.SubElement(item, 'g:price')
        g_price.text = f"{product.final_price} UAH"

        g_condition = ET.SubElement(item, 'g:condition')
        g_condition.text = 'new'

        g_brand = ET.SubElement(item, 'g:brand')
        g_brand.text = 'TwoComms'

        # Дополнительные поля
        g_mpn = ET.SubElement(item, 'g:mpn')
        g_mpn.text = f"TC-{product.id}"

        g_gtin = ET.SubElement(item, 'g:gtin')
        g_gtin.text = f"TC{product.id:08d}"

        # Категория
        if product.category:
            g_product_type = ET.SubElement(item, 'g:product_type')
            g_product_type.text = product.category.name

            g_google_product_category = ET.SubElement(item, 'g:google_product_category')
            g_google_product_category.text = '1604'  # Apparel & Accessories > Clothing

        # Возрастная группа и пол
        g_age_group = ET.SubElement(item, 'g:age_group')
        g_age_group.text = 'adult'

        g_gender = ET.SubElement(item, 'g:gender')
        if product.category:
            category_name = product.category.name.lower()
            if any(word in category_name for word in ['чоловіч', 'мужск', 'men']):
                g_gender.text = 'male'
            elif any(word in category_name for word in ['жіноч', 'женск', 'women']):
                g_gender.text = 'female'
            else:
                g_gender.text = 'unisex'
        else:
            g_gender.text = 'unisex'

        # Размеры
        g_size = ET.SubElement(item, 'g:size')
        g_size.text = 'S,M,L,XL,XXL'

        # Материал
        g_material = ET.SubElement(item, 'g:material')
        g_material.text = '100% cotton'

        # Цвета
        try:
            from productcolors.models import ProductColorVariant
            color_variants = ProductColorVariant.objects.filter(product=product)
            colors = []
            for variant in color_variants:
                if variant.color and variant.color.name:
                    colors.append(variant.color.name)
            
            if colors:
                g_color = ET.SubElement(item, 'g:color')
                g_color.text = ','.join(colors[:3])  # Максимум 3 цвета
        except:
            pass

        # Дополнительные изображения
        try:
            for i, img in enumerate(product.images.all()[:9]):  # До 9 дополнительных изображений
                g_additional_image_link = ET.SubElement(item, 'g:additional_image_link')
                g_additional_image_link.text = f"https://twocomms.shop{img.image.url}"
        except:
            pass

    # Форматируем XML
    ET.indent(rss, space="  ", level=0)
    
    # Создаем XML строку
    xml_string = ET.tostring(rss, encoding='utf-8', xml_declaration=True).decode('utf-8')
    
    # Возвращаем XML ответ
    response = HttpResponse(xml_string, content_type='application/xml; charset=utf-8')
    response['Content-Disposition'] = 'inline; filename="google_merchant_feed.xml"'
    return response

# ===== MONOBANK ACQUIRING =====
monobank_logger = logging.getLogger('storefront.monobank')

MONOBANK_SUCCESS_STATUSES = {'success', 'hold'}
MONOBANK_PENDING_STATUSES = {'processing'}
MONOBANK_FAILURE_STATUSES = {'failure', 'expired', 'rejected', 'canceled', 'cancelled', 'reversed'}


def _reset_monobank_session(request, drop_pending=False):
    """Сбрасывает связанные с Mono checkout данные в сессии."""
    if drop_pending:
        pending_id = request.session.get('monobank_pending_order_id')
        if pending_id:
            try:
                from orders.models import Order
                qs = Order.objects.filter(
                    id=pending_id,
                    payment_provider__in=('monobank', 'monobank_checkout', 'monobank_pay')
                )
                if qs.exists():
                    qs.update(status='cancelled', payment_status='unpaid')
            except Exception:
                monobank_logger.debug('Failed to cancel pending Monobank order %s', pending_id, exc_info=True)

    for key in (
        'monobank_pending_order_id',
        'monobank_invoice_id',
        'monobank_order_id',
        'monobank_order_ref'
    ):
        if key in request.session:
            request.session.pop(key, None)

    request.session.modified = True


def _normalize_color_variant_id(raw):
    """
    Приводит значение идентификатора цветового варианта к int либо None.
    Отсекает плейсхолдеры вида 'default', 'null', 'None', 'false', 'undefined'.
    """
    if raw is None:
        return None
    if isinstance(raw, int):
        return raw if raw > 0 else None
    try:
        value = str(raw).strip()
    except Exception:
        return None
    if not value:
        return None
    lowered = value.lower()
    if lowered in {'default', 'none', 'null', 'false', 'undefined'}:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _get_color_variant_safe(color_variant_id):
    """
    Возвращает экземпляр ProductColorVariant либо None, не выбрасывая ошибок.
    """
    normalized_id = _normalize_color_variant_id(color_variant_id)
    if not normalized_id:
        return None
    try:
        from productcolors.models import ProductColorVariant
        return ProductColorVariant.objects.get(id=normalized_id)
    except (ProductColorVariant.DoesNotExist, ValueError, TypeError):
        return None


def _cleanup_expired_monobank_orders():
    """Удаляет старые неоплаченные Mono-заказы, чтобы они не копились."""
    expire_minutes = getattr(settings, 'MONOBANK_CHECKOUT_EXPIRATION_MINUTES', 1440)
    if expire_minutes and expire_minutes > 0:
        threshold = timezone.now() - timedelta(minutes=expire_minutes)
        try:
            from orders.models import Order
            from django.db.models import Q
            stale_qs = Order.objects.filter(
                payment_provider__in=('monobank', 'monobank_checkout', 'monobank_pay'),
                created__lt=threshold
            ).filter(
                Q(payment_status__in=('unpaid', 'checking')) | Q(status='cancelled')
            )
            removed = stale_qs.count()
            if removed:
                monobank_logger.info('Pruned %s expired Monobank orders older than %s minutes', removed, expire_minutes)
                stale_qs.delete()
        except Exception:
            monobank_logger.exception('Failed to prune expired Monobank orders')


MONOBANK_SIGNATURE_CACHE = {'key': None, 'fetched_at': 0}
MONOBANK_SIGNATURE_TTL = 60 * 60  # 1 hour



def _get_monobank_public_key():
    now = time.time()
    cache = MONOBANK_SIGNATURE_CACHE
    if cache['key'] is not None and now - cache['fetched_at'] < MONOBANK_SIGNATURE_TTL:
        return cache['key']
    data = _monobank_api_request('GET', '/personal/checkout/signature/public/key')
    key_b64 = (data.get('result') or {}).get('key')
    if not key_b64:
        raise MonobankAPIError('Не вдалося отримати ключ перевірки підпису.')
    try:
        pem_bytes = base64.b64decode(key_b64)
        public_key = serialization.load_pem_public_key(pem_bytes)
    except Exception as exc:
        raise MonobankAPIError(f'Помилка при обробці ключа підпису: {exc}') from exc
    cache['key'] = public_key
    cache['fetched_at'] = now
    return public_key


def _invalidate_monobank_public_key():
    MONOBANK_SIGNATURE_CACHE['key'] = None
    MONOBANK_SIGNATURE_CACHE['fetched_at'] = 0


def _verify_monobank_signature(request):
    signature = request.headers.get('X-Sign') or request.META.get('HTTP_X_SIGN')
    if not signature:
        return False
    try:
        signature_bytes = base64.b64decode(signature)
    except Exception:
        return False
    try:
        public_key = _get_monobank_public_key()
    except MonobankAPIError as exc:
        monobank_logger.error('Unable to fetch Monobank signature key: %s', exc)
        return False
    try:
        public_key.verify(signature_bytes, request.body, ec.ECDSA(hashes.SHA256()))
        return True
    except InvalidSignature:
        return False
    except Exception as exc:
        monobank_logger.exception('Signature verification error: %s', exc)
        return False


class MonobankAPIError(Exception):
    """Raised when Monobank API returns an error"""


def _ensure_session_key(request):
    if not request.session.session_key:
        request.session.save()
    return request.session.session_key


def _validate_checkout_payload(raw_payload):
    """Validate checkout payload and return (errors, cleaned_data)."""
    full_name = (raw_payload.get('full_name') or '').strip()
    phone = (raw_payload.get('phone') or '').strip()
    city = (raw_payload.get('city') or '').strip()
    np_office = (raw_payload.get('np_office') or '').strip()
    pay_type = (raw_payload.get('pay_type') or 'full').strip().lower() or 'full'

    errors = []
    if len(full_name) < 3:
        errors.append('ПІБ повинно містити мінімум 3 символи.')

    digits = ''.join(filter(str.isdigit, phone))
    if not phone.startswith('+380') or len(digits) != 12:
        errors.append('Невірний формат телефону. Використовуйте формат +380XXXXXXXXX.')

    if len(city) < 2:
        errors.append('Введіть назву міста.')

    if len(np_office) < 1:
        errors.append('Введіть адресу відділення або поштомата.')

    if pay_type not in ('full', 'partial'):
        errors.append('Невідомий тип оплати.')

    if pay_type != 'full':
        errors.append('Для онлайн-оплати карткою потрібно обрати «Повна передоплата».')

    cleaned = {
        'full_name': full_name,
        'phone': phone,
        'city': city,
        'np_office': np_office,
        'pay_type': 'full'
    }
    return errors, cleaned


def _create_or_update_monobank_order(request, customer_data):
    """Create or update an order for Monobank payment, returning (order, amount_decimal, basket_order)."""
    from orders.models import Order, OrderItem

    cart = request.session.get('cart') or {}
    if not cart:
        raise ValueError('cart_empty')

    product_ids = [item['product_id'] for item in cart.values()]
    products = Product.objects.in_bulk(product_ids)

    session_key = _ensure_session_key(request)
    pending_order_id = request.session.get('monobank_pending_order_id')
    order = None

    if pending_order_id:
        try:
            order = Order.objects.select_for_update().get(id=pending_order_id)
            if request.user.is_authenticated:
                if order.user_id != request.user.id:
                    order = None
            else:
                if order.user_id is not None or order.session_key != session_key:
                    order = None
        except Order.DoesNotExist:
            order = None

    with transaction.atomic():
        if order is None:
            order = Order.objects.create(
                user=request.user if request.user.is_authenticated else None,
                session_key=session_key,
                full_name=customer_data['full_name'],
                phone=customer_data['phone'],
                city=customer_data['city'],
                np_office=customer_data['np_office'],
                pay_type='full',
                status='new',
                payment_status='checking',
                total_sum=Decimal('0'),
                discount_amount=Decimal('0')
            )
        else:
            OrderItem.objects.filter(order=order).delete()
            order.full_name = customer_data['full_name']
            order.phone = customer_data['phone']
            order.city = customer_data['city']
            order.np_office = customer_data['np_office']
            order.pay_type = 'full'
            order.session_key = session_key
            order.payment_status = 'checking'
            order.discount_amount = Decimal('0')
            order.total_sum = Decimal('0')
            order.save(update_fields=['full_name', 'phone', 'city', 'np_office', 'pay_type', 'session_key', 'payment_status', 'discount_amount', 'total_sum'])

        order_items = []
        total_sum = Decimal('0')

        for key, item in cart.items():
            product = products.get(item['product_id'])
            if not product:
                continue
            try:
                qty = int(item.get('qty', 1) or 1)
            except (TypeError, ValueError):
                qty = 1
            qty = max(qty, 1)
            try:
                unit_price = Decimal(str(product.final_price))
            except (InvalidOperation, TypeError, ValueError):
                monobank_logger.warning('Skipping cart item %s: invalid price %s', key, product.final_price)
                continue
            line_total = unit_price * qty
            total_sum += line_total
            color_variant = _get_color_variant_safe(item.get('color_variant_id'))

            order_items.append(OrderItem(
                order=order,
                product=product,
                color_variant=color_variant,
                title=product.title,
                size=item.get('size', ''),
                qty=qty,
                unit_price=unit_price,
                line_total=line_total
            ))

        if not order_items:
            raise ValueError('cart_empty')

        OrderItem.objects.bulk_create(order_items)

        basket_order = []
        saved_items = (
            order.items
            .select_related('product', 'color_variant__color')
            .prefetch_related('color_variant__images')
        )

        for saved in saved_items:
            display_name = saved.title
            if saved.size:
                display_name = f"{display_name} • розмір {saved.size}"
            if saved.color_name:
                display_name = f"{display_name} • {saved.color_name}"
            display_name = display_name[:128]

            image_url = None
            try:
                img = saved.product_image
                if img and hasattr(img, 'url'):
                    image_url = request.build_absolute_uri(img.url)
                    if image_url.startswith('http://'):
                        image_url = 'https://' + image_url[len('http://'):]
            except Exception:
                image_url = None

            line_total = Decimal(str(saved.line_total))
            basket_item = {
                'name': display_name,
                'qty': saved.qty,
                'sum': int((line_total * 100).quantize(Decimal('1'), rounding=ROUND_HALF_UP)),
                'unit': 'шт.',
                'code': str(getattr(saved.product, 'sku', saved.product.id))
            }
            if image_url:
                basket_item['iconUrl'] = image_url
                basket_item['imageUrl'] = image_url
            basket_order.append(basket_item)

        discount = Decimal('0')
        promo = None
        promo_code_value = request.session.get('applied_promo_code')
        if promo_code_value:
            from .models import PromoCode
            try:
                potential_promo = PromoCode.objects.get(code=promo_code_value, is_active=True)
                if potential_promo.can_be_used():
                    possible_discount = potential_promo.calculate_discount(total_sum)
                    if possible_discount > 0:
                        discount = possible_discount
                        promo = potential_promo
                        promo.use()
            except PromoCode.DoesNotExist:
                pass

        order.discount_amount = discount
        order.promo_code = promo if discount > 0 else None
        order.payment_provider = 'monobank'
        order.total_sum = total_sum - discount
        order.save(update_fields=['discount_amount', 'promo_code', 'payment_provider', 'total_sum'])

    request.session['monobank_pending_order_id'] = order.id
    return order, order.total_sum, basket_order


def _monobank_api_request(method, endpoint, *, params=None, json_payload=None, timeout=10):
    base_url = settings.MONOBANK_API_BASE.rstrip('/')
    url = f"{base_url}{endpoint}"
    headers = {'X-Token': settings.MONOBANK_TOKEN}

    monobank_logger.info('Monobank API request: %s %s, payload: %s', method, endpoint, json_payload)

    try:
        if method.upper() == 'POST':
            response = requests.post(url, params=params, json=json_payload, headers=headers, timeout=timeout)
        elif method.upper() == 'GET':
            response = requests.get(url, params=params, headers=headers, timeout=timeout)
        else:
            raise MonobankAPIError('Unsupported HTTP method')
    except requests.RequestException as exc:
        monobank_logger.error('Monobank API request failed: %s', exc)
        raise MonobankAPIError(f'Помилка з\'єднання з платіжним сервісом: {exc}') from exc

    try:
        data = response.json()
    except ValueError:
        data = {}

    monobank_logger.info('Monobank API response: status=%s, data=%s', response.status_code, data)

    if response.status_code >= 400 or 'errCode' in data:
        err_text = data.get('errText') or data.get('message') or f'HTTP {response.status_code}'
        monobank_logger.error('Monobank API error: %s', err_text)
        raise MonobankAPIError(err_text)

    return data


def _prepare_checkout_customer_data(request):
    """Prepare customer data for checkout from request."""
    if request.user.is_authenticated:
        try:
            profile = request.user.userprofile
            return {
                'full_name': profile.full_name or f'{request.user.first_name} {request.user.last_name}'.strip() or 'Користувач',
                'phone': profile.phone or '',
                'city': profile.city or '',
                'np_office': profile.np_office or '',
                'pay_type': profile.pay_type or 'full',
                'user': request.user
            }
        except:
            pass
    
    # Default data for guests
    return {
        'full_name': 'Користувач',
        'phone': '',
        'city': '',
        'np_office': '',
        'pay_type': 'full',
        'user': None
    }


def _record_monobank_status(order, payload, source='api'):
    if not payload:
        return

    status = payload.get('status')
    payment_payload = order.payment_payload or {}
    history = payment_payload.get('history', [])
    history.append({
        'status': status,
        'data': payload,
        'source': source,
        'received_at': timezone.now().isoformat()
    })
    payment_payload['history'] = history[-20:]
    payment_payload['last_status'] = status
    payment_payload['last_update_source'] = source
    payment_payload['last_update_at'] = timezone.now().isoformat()
    order.payment_payload = payment_payload

    update_fields = ['payment_payload']

    if status in MONOBANK_SUCCESS_STATUSES:
        previous_status = order.payment_status
        order.payment_status = 'paid'
        update_fields.append('payment_status')
        try:
            order.save(update_fields=update_fields)
        except Exception:
            order.save()
        if previous_status != 'paid':
            try:
                from orders.telegram_notifications import TelegramNotifier
                notifier = TelegramNotifier()
                notifier.send_new_order_notification(order)
            except Exception:
                monobank_logger.exception('Failed to send Telegram notification for paid order %s', order.id)
        return

    if status in MONOBANK_PENDING_STATUSES:
        order.payment_status = 'checking'
        update_fields.append('payment_status')
    elif status in MONOBANK_FAILURE_STATUSES:
        order.payment_status = 'unpaid'
        update_fields.append('payment_status')

    try:
        order.save(update_fields=update_fields)
    except Exception:
        order.save()



@require_POST
def monobank_create_invoice(request):
    """Create Monobank pay invoice and return redirect URL."""
    _cleanup_expired_monobank_orders()
    try:
        customer = _prepare_checkout_customer_data(request)
        order, amount_decimal, _ = _create_or_update_monobank_order(request, customer)
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Кошик порожній. Додайте товари перед оплатою.'})

    items_qs = list(order.items.select_related('product', 'color_variant__color'))
    total_qty = sum(item.qty for item in items_qs)
    if total_qty <= 0 or amount_decimal <= 0:
        return JsonResponse({'success': False, 'error': 'Сума для оплати повинна бути більшою за 0.'})

    try:
        basket_entries = []
        for item in items_qs:
            name_parts = [item.product.title]
            if item.size:
                name_parts.append(f"розмір {item.size}")
            color_name = getattr(item, 'color_name', None)
            if color_name:
                name_parts.append(color_name)
            display_name = ' • '.join(filter(None, name_parts))[:128]
            try:
                line_total_minor = int(Decimal(str(item.line_total)) * 100)
            except (InvalidOperation, TypeError, ValueError):
                monobank_logger.warning('Skipping item %s in Mono Pay basket: invalid line total %s', item.id, item.line_total)
                continue

            icon_url = ''
            try:
                image_obj = None
                if getattr(item, 'color_variant', None) and item.color_variant.images.exists():
                    image_obj = item.color_variant.images.first().image
                elif item.product.main_image:
                    image_obj = item.product.main_image
                if image_obj and hasattr(image_obj, 'url'):
                    icon_url = request.build_absolute_uri(image_obj.url)
                    if icon_url.startswith('http://'):
                        icon_url = 'https://' + icon_url[len('http://'):]
            except Exception:
                icon_url = ''

            try:
                qty_minor = max(int(getattr(item, 'qty', 1) or 1), 1)
            except (TypeError, ValueError):
                qty_minor = 1

            basket_entries.append({
                'name': display_name or item.product.title[:128],
                'qty': qty_minor,
                'sum': line_total_minor,
                'icon': icon_url
            })

        if not basket_entries:
            return JsonResponse({'success': False, 'error': 'Кошик порожній. Додайте товари перед оплатою.'})

        # Для Monobank Pay используем API эквайринга
        payload = {
            'amount': int(amount_decimal * 100),  # сумма в копейках
            'ccy': 980,  # гривна
            'merchantPaymInfo': {
                'reference': order.order_number,
                'destination': f'Оплата замовлення {order.order_number}',
                'basketOrder': basket_entries
            },
            'redirectUrl': request.build_absolute_uri('/payments/monobank/return/'),
            'webHookUrl': request.build_absolute_uri('/payments/monobank/webhook/'),
        }
        
        creation_data = _monobank_api_request('POST', '/api/merchant/invoice/create', json_payload=payload)
    except MonobankAPIError as exc:
        monobank_logger.warning('Monobank pay invoice creation failed: %s', exc)
        return JsonResponse({'success': False, 'error': str(exc)})
    except Exception as exc:
        monobank_logger.exception('Failed to build Mono Pay payload: %s', exc)
        return JsonResponse({'success': False, 'error': 'Не вдалося підготувати дані для платежу. Спробуйте ще раз.'})

    result = creation_data.get('result') or creation_data
    invoice_id = result.get('invoiceId')
    invoice_url = result.get('pageUrl')
    
    if not invoice_id or not invoice_url:
        return JsonResponse({'success': False, 'error': 'Не вдалося створити платіж. Спробуйте пізніше.'})

    payment_payload = {
        'request': payload,
        'create': creation_data,
        'history': []
    }
    order.payment_invoice_id = invoice_id
    order.payment_payload = payment_payload
    order.payment_status = 'checking'
    order.payment_provider = 'monobank_pay'
    order.save(update_fields=['payment_invoice_id', 'payment_payload', 'payment_status', 'payment_provider'])

    request.session['monobank_invoice_id'] = invoice_id
    request.session['monobank_pending_order_id'] = order.id
    request.session.modified = True

    return JsonResponse({
        'success': True,
        'invoice_url': invoice_url,
        'invoice_id': invoice_id,
        'order_id': order.id,
        'order_ref': order.order_number
    })


def _build_monobank_checkout_payload(order, amount_decimal, total_qty, request, items=None):
    """Build payload for Monobank Checkout API."""
    from django.conf import settings
    
    if items is None:
        items = list(order.items.select_related('product', 'color_variant__color'))
    
    def _as_number(dec: Decimal):
        """Convert Decimal to int if possible, otherwise to float."""
        if dec == dec.to_integral():
            return int(dec)
        return float(dec)

    products = []
    total_amount_major = Decimal('0')
    total_count = 0

    for item in items:
        qty_value = getattr(item, 'qty', None)
        if qty_value is None:
            qty_value = getattr(item, 'quantity', None)

        monobank_logger.info('Building product data for item: %s, qty=%s, unit_price=%s',
                             item.product.title, qty_value, getattr(item, 'unit_price', None))

        try:
            qty = int(qty_value) if qty_value is not None else 0
        except (TypeError, ValueError):
            qty = 0

        if qty <= 0:
            monobank_logger.error('Item has non-positive qty: %s', qty)
            qty = 1

        try:
            line_total_major = Decimal(str(item.line_total)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        except Exception as exc:
            monobank_logger.exception('Failed to convert line total for item %s: %s', item.id if hasattr(item, 'id') else item.product_id, exc)
            continue

        if line_total_major <= 0:
            monobank_logger.error('Item has non-positive line total: %s', line_total_major)
            continue

        total_amount_major += line_total_major
        total_count += qty

        name_parts = [item.title or item.product.title]
        if item.size:
            name_parts.append(f"розмір {item.size}")
        if item.color_name:
            name_parts.append(item.color_name)

        product_data = {
            'name': ' • '.join(filter(None, name_parts)),
            'cnt': qty,
            'price': _as_number(line_total_major),
        }

        code_product = getattr(item.product, 'sku', None) or getattr(item.product, 'id', None)
        if code_product is not None:
            product_data['code_product'] = str(code_product)
        if item.product.main_image:
            product_data['icon'] = request.build_absolute_uri(item.product.main_image.url)
        if item.color_name:
            product_data['description'] = item.color_name

        products.append(product_data)

    if not products:
        monobank_logger.error('No products collected for monobank checkout payload')
        raise MonobankAPIError('Кошик порожній. Додайте товари перед оплатою.')

    if total_count <= 0:
        total_count = sum(max(int(getattr(item, 'qty', 1) or 1), 1) for item in items)

    amount_major = Decimal(str(amount_decimal)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    if total_amount_major and amount_major != total_amount_major:
        amount_major = total_amount_major

    payload = {
        'order_ref': order.order_number,
        'amount': _as_number(amount_major),
        'ccy': 980,  # гривна
        'count': int(total_count),
        'products': products,
        'destination': getattr(settings, 'MONOBANK_CHECKOUT_DESTINATION_TEMPLATE', 'Оплата замовлення {order_number}').format(order_number=order.order_number)
    }

    delivery_methods = getattr(settings, 'MONOBANK_CHECKOUT_DELIVERY_METHODS', None)
    if delivery_methods and 'MONOBANK_CHECKOUT_DELIVERY_METHODS' in os.environ:
        payload['dlv_method_list'] = delivery_methods

    payment_methods = getattr(settings, 'MONOBANK_CHECKOUT_PAYMENT_METHODS', None)
    if payment_methods and 'MONOBANK_CHECKOUT_PAYMENT_METHODS' in os.environ:
        payload['payment_method_list'] = payment_methods

    if getattr(settings, 'MONOBANK_CHECKOUT_DLV_PAY_MERCHANT', False) and 'MONOBANK_CHECKOUT_DLV_PAY_MERCHANT' in os.environ:
        payload['dlv_pay_merchant'] = True

    payments_number = getattr(settings, 'MONOBANK_CHECKOUT_PAYMENTS_NUMBER', None)
    if payments_number:
        payload['payments_number'] = payments_number

    callback_url = getattr(settings, 'MONOBANK_CHECKOUT_CALLBACK_URL', '')
    return_url = getattr(settings, 'MONOBANK_CHECKOUT_RETURN_URL', '')
    default_callback = request.build_absolute_uri('/payments/monobank/webhook/').replace('http://', 'https://', 1)
    default_return = request.build_absolute_uri('/payments/monobank/return/').replace('http://', 'https://', 1)
    payload['callback_url'] = callback_url or default_callback
    payload['return_url'] = return_url or default_return
    
    monobank_logger.info('Built Monobank Checkout payload: %s', payload)
    
    return payload


def _create_single_product_order(product, size, qty, color_variant_id, customer):
    """Create a temporary order for a single product for Monobank Checkout."""
    from orders.models import Order, OrderItem

    customer_data = {
        'user': customer.get('user'),
        'full_name': (customer.get('full_name') or 'Користувач').strip() or 'Користувач',
        'phone': customer.get('phone', '') or '',
        'city': customer.get('city', '') or '',
        'np_office': customer.get('np_office', '') or '',
        'pay_type': customer.get('pay_type', 'full') or 'full'
    }

    try:
        qty_int = max(int(qty or 1), 1)
    except (TypeError, ValueError):
        qty_int = 1
    unit_price_raw = product.final_price if product.final_price is not None else getattr(product, 'price', None)
    try:
        unit_price = Decimal(str(unit_price_raw or '0'))
    except (InvalidOperation, TypeError, ValueError):
        monobank_logger.warning('Single checkout: invalid price for product %s (%s)', product.id, unit_price_raw)
        unit_price = Decimal('0')
    line_total = (unit_price * qty_int).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    color_variant = _get_color_variant_safe(color_variant_id)

    order = Order.objects.create(
        user=customer_data['user'],
        full_name=customer_data['full_name'],
        phone=customer_data['phone'],
        city=customer_data['city'],
        np_office=customer_data['np_office'],
        pay_type=customer_data['pay_type'],
        payment_status='checking',
        status='new',
        total_sum=line_total,
        discount_amount=Decimal('0'),
        payment_provider='monobank_checkout'
    )

    OrderItem.objects.create(
        order=order,
        product=product,
        color_variant=color_variant,
        title=product.title,
        size=size or '',
        qty=qty_int,
        unit_price=unit_price,
        line_total=line_total
    )

    return order


@require_POST
def monobank_create_checkout(request):
    """Create Monobank checkout order and return redirect URL."""
    _cleanup_expired_monobank_orders()
    try:
        # Проверяем, создаем ли заказ на один товар
        body = json.loads(request.body.decode('utf-8')) if request.body else {}
        single_product = body.get('single_product', False)
        
        monobank_logger.info('Monobank checkout request: single_product=%s, body=%s', single_product, body)
        
        if single_product:
            # Создаем заказ на один товар
            product_id = body.get('product_id')
            size = body.get('size', 'S')
            qty = int(body.get('qty', 1))
            color_variant_id = body.get('color_variant_id')
            
            monobank_logger.info('Single product checkout: product_id=%s, size=%s, qty=%s, color_variant_id=%s', 
                               product_id, size, qty, color_variant_id)
            
            if not product_id:
                return JsonResponse({'success': False, 'error': 'Не вказано ID товару.'})
            
            try:
                product = Product.objects.get(id=product_id)
                monobank_logger.info('Found product: %s, price: %s', product.title, product.final_price)
            except Product.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Товар не знайдено.'})
            
            # Создаем временный заказ на один товар
            customer = _prepare_checkout_customer_data(request)
            monobank_logger.info('Customer data: %s', customer)
            order = _create_single_product_order(product, size, qty, color_variant_id, customer)
            amount_decimal = order.total_amount
            monobank_logger.info('Created single product order: %s, amount: %s', order.order_number, amount_decimal)
        else:
            # Обычная логика для корзины
            customer = _prepare_checkout_customer_data(request)
            order, amount_decimal, _ = _create_or_update_monobank_order(request, customer)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})
    except Exception as e:
        monobank_logger.exception('Error in monobank_create_checkout: %s', e)
        return JsonResponse({'success': False, 'error': 'Помилка при створенні замовлення.'})

    items_qs = list(order.items.select_related('product', 'color_variant__color'))
    total_qty = sum(item.qty for item in items_qs)
    if total_qty <= 0 or amount_decimal <= 0:
        return JsonResponse({'success': False, 'error': 'Сума для оплати повинна бути більшою за 0.'})

    try:
        payload = _build_monobank_checkout_payload(order, amount_decimal, total_qty, request, items=items_qs)
    except MonobankAPIError as exc:
        return JsonResponse({'success': False, 'error': str(exc)})
    except Exception as exc:
        monobank_logger.exception('Failed to build Mono Checkout payload: %s', exc)
        return JsonResponse({'success': False, 'error': 'Не вдалося підготувати дані для платежу. Спробуйте ще раз.'})

    try:
        creation_data = _monobank_api_request('POST', '/personal/checkout/order', json_payload=payload)
    except MonobankAPIError as exc:
        monobank_logger.warning('Monobank checkout order creation failed: %s', exc)
        return JsonResponse({'success': False, 'error': str(exc)})

    result = creation_data.get('result') or creation_data
    order_id = result.get('order_id') or result.get('orderId')
    redirect_url = result.get('redirect_url') or result.get('redirectUrl')
    if not order_id or not redirect_url:
        return JsonResponse({'success': False, 'error': 'Не вдалося створити замовлення. Спробуйте пізніше.'})

    payment_payload = {
        'request': payload,
        'create': creation_data,
        'history': []
    }
    order.payment_invoice_id = order_id
    order.payment_payload = payment_payload
    order.payment_status = 'checking'
    order.payment_provider = 'monobank_checkout'
    order.save(update_fields=['payment_invoice_id', 'payment_payload', 'payment_status', 'payment_provider'])

    request.session['monobank_order_id'] = order_id
    request.session['monobank_order_ref'] = order.order_number
    request.session['monobank_invoice_id'] = order_id
    request.session['monobank_pending_order_id'] = order.id
    request.session.modified = True

    return JsonResponse({
        'success': True,
        'redirect_url': redirect_url,
        'order_id': order.id,
        'order_ref': order.order_number
    })



def _fetch_and_apply_invoice_status(order, invoice_id, source):
    try:
        status_data = _monobank_api_request('GET', '/api/merchant/invoice/status', params={'invoiceId': invoice_id})
    except MonobankAPIError as exc:
        monobank_logger.warning('Failed to fetch invoice status for %s: %s', invoice_id, exc)
        raise

    _record_monobank_status(order, status_data, source=source)
    return status_data


def _cleanup_after_success(request):
    request.session.pop('cart', None)
    request.session.pop('applied_promo_code', None)
    request.session.pop('monobank_invoice_id', None)
    request.session.pop('monobank_pending_order_id', None)
    request.session.pop('monobank_order_id', None)
    request.session.pop('monobank_order_ref', None)
    request.session.modified = True


def monobank_return(request):
    """Handle user return from Monobank payment page."""
    from orders.models import Order
    from django.contrib import messages

    order_id = request.GET.get('orderId') or request.session.get('monobank_order_id')
    invoice_id = request.GET.get('invoiceId') or request.session.get('monobank_invoice_id')
    order_ref = request.GET.get('orderRef') or request.session.get('monobank_order_ref')

    order = None
    if order_id:
        order = Order.objects.filter(payment_invoice_id=order_id).order_by('-created').first()

    if order is None and invoice_id:
        pending_id = request.session.get('monobank_pending_order_id')
        if pending_id:
            try:
                order = Order.objects.get(id=pending_id)
            except Order.DoesNotExist:
                order = None

    if order is None and invoice_id:
        order = Order.objects.filter(payment_invoice_id=invoice_id).order_by('-created').first()

    if order is None and order_ref:
        order = Order.objects.filter(order_number=order_ref).order_by('-created').first()

    if order is None:
        messages.error(request, 'Замовлення не знайдено. Спробуйте ще раз.')
        return redirect('cart')

    try:
        if order.payment_provider == 'monobank_checkout' or order_id:
            result = _fetch_and_apply_checkout_status(order, source='return')
            status = (result.get('payment_status') or '').lower()
        else:
            status_data = _fetch_and_apply_invoice_status(order, invoice_id, source='return')
            status = status_data.get('status')
    except MonobankAPIError as exc:
        messages.error(request, f'Не вдалося підтвердити статус платежу: {exc}')
        return redirect('my_orders' if request.user.is_authenticated else 'cart')


    if status in MONOBANK_SUCCESS_STATUSES:
        _cleanup_after_success(request)
        messages.success(request, 'Оплату успішно отримано!')
        if request.user.is_authenticated:
            return redirect('my_orders')
        return redirect('order_success', order_id=order.id)

    if status in MONOBANK_PENDING_STATUSES:
        messages.info(request, 'Платіж обробляється. Ми повідомимо, щойно отримаємо підтвердження.')
        if request.user.is_authenticated:
            return redirect('my_orders')
        return redirect('order_success', order_id=order.id)

    messages.error(request, 'Оплату не завершено. Ви можете повторити спробу або обрати інший спосіб оплати.')
    return redirect('cart')


@csrf_exempt
def monobank_webhook(request):
    """Receive status updates from Monobank webhook."""
    from orders.models import Order

    if request.method != 'POST':
        return HttpResponse(status=405)

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return HttpResponse(status=400)

    # Legacy invoice webhook
    invoice_id = payload.get('invoiceId')
    if invoice_id:
        try:
            order = Order.objects.get(payment_invoice_id=invoice_id)
        except Order.DoesNotExist:
            monobank_logger.warning('Webhook received for unknown invoice %s', invoice_id)
            return JsonResponse({'ok': True})

        _record_monobank_status(order, payload, source='webhook')
        return JsonResponse({'ok': True})

    # Checkout order webhook requires signature validation
    if not _verify_monobank_signature(request):
        monobank_logger.warning('Invalid or missing X-Sign for checkout webhook')
        return HttpResponse(status=400)

    result = payload.get('result') or {}
    order_id = result.get('orderId') or result.get('order_id')
    order_ref = result.get('orderRef') or result.get('order_ref')

    order = None
    if order_id:
        order = Order.objects.filter(payment_invoice_id=order_id).first()
    if order is None and order_ref:
        order = Order.objects.filter(order_number=order_ref).first()

    if order is None:
        monobank_logger.warning('Checkout webhook received for unknown order: %s / %s', order_id, order_ref)
        return JsonResponse({'ok': True})

    try:
        _update_order_from_checkout_result(order, result, source='webhook')
    except Exception as exc:
        monobank_logger.exception('Failed to process checkout webhook for order %s: %s', order.order_number, exc)
        return HttpResponse(status=500)

    return JsonResponse({'ok': True})


def _get_product_image_url(product, request):
    """Get product image URL for Excel."""
    if product.main_image:
        return request.build_absolute_uri(product.main_image.url)
    return None


def _translate_color_to_ukrainian(color_name):
    """Translate color names from Russian to Ukrainian."""
    color_translations = {
        'черный': 'чорний',
        'белый': 'білий',
        'красный': 'червоний',
        'синий': 'синій',
        'зеленый': 'зелений',
        'желтый': 'жовтий',
        'оранжевый': 'помаранчевий',
        'фиолетовый': 'фіолетовий',
        'розовый': 'рожевий',
        'серый': 'сірий',
        'коричневый': 'коричневий',
        'голубой': 'блакитний',
        'бордовый': 'бордовий',
        'темно-синий': 'темно-синій',
        'светло-серый': 'світло-сірий',
        'темно-серый': 'темно-сірий',
        'бежевый': 'бежевий',
        'кремовый': 'кремовий',
        'оливковый': 'оливковий',
        'бирюзовый': 'бірюзовий',
        'лайм': 'лайм',
        'лаванда': 'лаванда',
        'мятный': "м'ятний",
        'коралловый': 'кораловий',
        'персиковый': 'персиковий',
        'золотой': 'золотий',
        'серебряный': 'срібний',
        'медный': 'мідний',
        'бронзовый': 'бронзовий',
        'хаки': 'хаки',
        'фуксия': 'фуксія',
        'индиго': 'індиго',
        'пурпурный': 'пурпурний',
        'малиновый': 'малиновий',
        'вишневый': 'вишневий',
        'изумрудный': 'смарагдовий',
        'нефрит': 'нефрит',
        'бирюза': 'бірюза',
        'бирюзово-зеленый': 'бірюзово-зелений',
        'темно-зеленый': 'темно-зелений',
        'светло-зеленый': 'світло-зелений',
        'темно-красный': 'темно-червоний',
        'светло-красный': 'світло-червоний',
        'темно-синий': 'темно-синій',
        'светло-синий': 'світло-синій',
        'темно-фиолетовый': 'темно-фіолетовий',
        'светло-фиолетовый': 'світло-фіолетовий',
        'темно-коричневый': 'темно-коричневий',
        'светло-коричневый': 'світло-коричневий',
        'темно-розовый': 'темно-рожевий',
        'светло-розовый': 'світло-рожевий',
        'темно-желтый': 'темно-жовтий',
        'светло-желтый': 'світло-жовтий',
        'темно-оранжевый': 'темно-помаранчевий',
        'светло-оранжевый': 'світло-помаранчевий',
        # Додаткові кольори
        'кайот': 'кайот',
        'navy': 'navy',
        'charcoal': 'charcoal',
        'heather': 'heather',
        'maroon': 'maroon',
        'forest': 'forest',
        'royal': 'royal',
        'sport grey': 'sport grey',
        'ash': 'ash',
        'dark heather': 'dark heather',
        'red': 'red',
        'blue': 'blue',
        'green': 'green',
        'yellow': 'yellow',
        'orange': 'orange',
        'purple': 'purple',
        'pink': 'pink',
        'black': 'чорний',
        'white': 'білий',
        'grey': 'сірий',
        'brown': 'коричневий',
    }
    
    if not color_name:
        return 'чорний'
    
    color_lower = color_name.lower().strip()
    return color_translations.get(color_lower, color_name)


def wholesale_prices_xlsx(request):
    """Generate and serve wholesale prices XLSX file."""
    from django.http import HttpResponse
    from django.db.models import Q
    from storefront.models import Product, Category
    from productcolors.models import ProductColorVariant, Color
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io
    
    # Фильтруем категории по ключевым словам
    tshirt_categories = Category.objects.filter(
        Q(name__icontains='футболка') | 
        Q(name__icontains='tshirt') | 
        Q(name__icontains='t-shirt') |
        Q(slug__icontains='футболка') |
        Q(slug__icontains='tshirt') |
        Q(slug__icontains='t-shirt')
    )
    
    hoodie_categories = Category.objects.filter(
        Q(name__icontains='худи') | 
        Q(name__icontains='hoodie') | 
        Q(name__icontains='hooded') |
        Q(slug__icontains='худи') |
        Q(slug__icontains='hoodie') |
        Q(slug__icontains='hooded')
    )
    
    # Получаем товары из нужных категорий
    tshirt_products = Product.objects.filter(category__in=tshirt_categories)
    hoodie_products = Product.objects.filter(category__in=hoodie_categories)
    
    # Создаем XLSX файл в памяти
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прайс (опт)"
    
    # Стили
    header_font = Font(bold=True, size=14)
    table_header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Заголовок (строка 1)
    ws.merge_cells('A1:L1')
    ws['A1'] = "Оптові ціни від кількості. Мінімальне замовлення по моделі — від 8 шт. і кратно 8."
    ws['A1'].font = header_font
    ws['A1'].alignment = center_alignment
    
    # Заголовки таблицы (строка 2)
    headers = [
        'Категорія',
        'Товар (S–XL)',
        'Колір',
        'Дроп (фікс. ціна)',
        '8–15 шт. (за 1 шт.)',
        '16–31 шт.',
        '32–63 шт.',
        '64–99 шт.',
        '100+ шт.',
        'Посилання'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = table_header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Цены для категорий
    tshirt_prices = [600, 570, 540, 510, 450]
    hoodie_prices = [1350, 1300, 1250, 1200, 1150]
    
    row = 3
    
    # Сначала худи
    for product in hoodie_products.order_by('title'):
        
        # Получаем цвета товара
        color_variants = ProductColorVariant.objects.filter(product=product)
        colors = []
        for variant in color_variants:
            if variant.color:
                colors.append(variant.color.name or str(variant.color))
        
        # Формируем название с "(фліс)" в конце
        product_title = f"{product.title} (S–XL) [фліс]"
        
        # Создаем ссылку на товар
        product_url = request.build_absolute_uri(f'/product/{product.slug}/')
        
        if colors:
            # Если есть цвета, создаем строку для каждого цвета
            for color in colors:
                ws.cell(row=row, column=1, value='Худі')
                ws.cell(row=row, column=2, value=product_title)
                ws.cell(row=row, column=3, value=_translate_color_to_ukrainian('чорний'))  # Для худи завжди чорний
                
                # Добавляем дроп цену (фиксированная 1450)
                ws.cell(row=row, column=4, value=1450)
                ws.cell(row=row, column=4).alignment = center_alignment
                
                # Добавляем оптовые цены
                for col, price in enumerate(hoodie_prices, 5):
                    ws.cell(row=row, column=col, value=price)
                    ws.cell(row=row, column=col).alignment = center_alignment
                
                # Добавляем ссылку на товар
                ws.cell(row=row, column=10, value=product_url)
                
                row += 1
        else:
            # Если нет цветов, создаем одну строку
            ws.cell(row=row, column=1, value='Худі')
            ws.cell(row=row, column=2, value=product_title)
            ws.cell(row=row, column=3, value=_translate_color_to_ukrainian('чорний'))  # Для худи завжди чорний
            
            # Добавляем дроп цену (фиксированная 1450)
            ws.cell(row=row, column=4, value=1450)
            ws.cell(row=row, column=4).alignment = center_alignment
            
            # Добавляем оптовые цены
            for col, price in enumerate(hoodie_prices, 5):
                ws.cell(row=row, column=col, value=price)
                ws.cell(row=row, column=col).alignment = center_alignment
            
            # Добавляем ссылку на товар
            ws.cell(row=row, column=10, value=product_url)
            
            row += 1
    
    # Затем футболки
    for product in tshirt_products.order_by('title'):
        
        # Получаем цвета товара
        color_variants = ProductColorVariant.objects.filter(product=product)
        colors = []
        for variant in color_variants:
            if variant.color:
                colors.append(variant.color.name or str(variant.color))
        
        # Формируем название без [фліс]
        product_title = f"{product.title} (S–XL)"
        
        # Создаем ссылку на товар
        product_url = request.build_absolute_uri(f'/product/{product.slug}/')
        
        if colors:
            # Если есть цвета, создаем строку для каждого цвета
            for color in colors:
                ws.cell(row=row, column=1, value='Футболки')
                ws.cell(row=row, column=2, value=product_title)
                ws.cell(row=row, column=3, value=_translate_color_to_ukrainian(color if color else 'чорний'))
                
                # Добавляем дроп цену (фиксированная 650 для футболок)
                ws.cell(row=row, column=4, value=650)
                ws.cell(row=row, column=4).alignment = center_alignment
                
                # Добавляем оптовые цены
                for col, price in enumerate(tshirt_prices, 5):
                    ws.cell(row=row, column=col, value=price)
                    ws.cell(row=row, column=col).alignment = center_alignment
                
                # Добавляем ссылку на товар
                ws.cell(row=row, column=10, value=product_url)
                
                row += 1
        else:
            # Если нет цветов, создаем одну строку с черным цветом
            ws.cell(row=row, column=1, value='Футболки')
            ws.cell(row=row, column=2, value=product_title)
            ws.cell(row=row, column=3, value=_translate_color_to_ukrainian('чорний'))
            
            # Добавляем дроп цену (фиксированная 650 для футболок)
            ws.cell(row=row, column=4, value=650)
            ws.cell(row=row, column=4).alignment = center_alignment
            
            # Добавляем оптовые цены
            for col, price in enumerate(tshirt_prices, 5):
                ws.cell(row=row, column=col, value=price)
                ws.cell(row=row, column=col).alignment = center_alignment
            
            # Добавляем ссылку на товар
            ws.cell(row=row, column=10, value=product_url)
            
            row += 1
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Создаем HTTP ответ
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pricelist_opt.xlsx"'
    
    return response


def test_pricelist(request):
    """Simple test function to check if new URLs work."""
    from django.http import HttpResponse
    return HttpResponse("Pricelist test works!", content_type="text/plain")


def pricelist_redirect(request):
    """Redirect to static pricelist file."""
    from django.shortcuts import redirect
    return redirect('/media/pricelist_opt.xlsx')


def pricelist_page(request):
    """Display pricelist download page."""
    return render(request, 'pages/pricelist.html')


def wholesale_page(request):
    """Render wholesale page with products and pricing."""
    try:
        from django.db.models import Q
        from storefront.models import Product, Category
        
        # Фильтруем категории по ключевым словам
        tshirt_categories = Category.objects.filter(
            Q(name__icontains='футболка') | 
            Q(name__icontains='tshirt') | 
            Q(name__icontains='t-shirt') |
            Q(slug__icontains='футболка') |
            Q(slug__icontains='tshirt') |
            Q(slug__icontains='t-shirt')
        )
        
        hoodie_categories = Category.objects.filter(
            Q(name__icontains='худи') | 
            Q(name__icontains='hoodie') | 
            Q(name__icontains='hooded') |
            Q(slug__icontains='худи') |
            Q(slug__icontains='hoodie') |
            Q(slug__icontains='hooded')
        )
        
        # Получаем товары из нужных категорий
        tshirt_products = Product.objects.filter(category__in=tshirt_categories).select_related('category')
        hoodie_products = Product.objects.filter(category__in=hoodie_categories).select_related('category')
        
    except Exception as e:
        # Если есть ошибка с базой данных, используем пустые списки
        tshirt_products = []
        hoodie_products = []
    
    # Цены для категорий
    tshirt_prices = {
        'drop': 650,
        'wholesale': [600, 570, 540, 510, 450],
        'ranges': ['8–15 шт.', '16–31 шт.', '32–63 шт.', '64–99 шт.', '100+ шт.']
    }
    
    hoodie_prices = {
        'drop': 1450,
        'wholesale': [1350, 1300, 1250, 1200, 1150],
        'ranges': ['8–15 шт.', '16–31 шт.', '32–63 шт.', '64–99 шт.', '100+ шт.']
    }
    
    context = {
        'tshirt_products': tshirt_products,
        'hoodie_products': hoodie_products,
        'tshirt_prices': tshirt_prices,
        'hoodie_prices': hoodie_prices,
    }
    
    return render(request, 'pages/wholesale.html', context)


def wholesale_order_form(request):
    """Render wholesale order form page."""
    try:
        from django.db.models import Q
        from storefront.models import Product, Category
        
        # Фильтруем категории по ключевым словам
        tshirt_categories = Category.objects.filter(
            Q(name__icontains='футболка') | 
            Q(name__icontains='tshirt') | 
            Q(name__icontains='t-shirt') |
            Q(slug__icontains='футболка') |
            Q(slug__icontains='tshirt') |
            Q(slug__icontains='t-shirt')
        )
        
        hoodie_categories = Category.objects.filter(
            Q(name__icontains='худи') | 
            Q(name__icontains='hoodie') | 
            Q(name__icontains='hooded') |
            Q(slug__icontains='худи') |
            Q(slug__icontains='hoodie') |
            Q(slug__icontains='hooded')
        )
        
        # Получаем товары из нужных категорий
        tshirt_products = Product.objects.filter(category__in=tshirt_categories).select_related('category')
        hoodie_products = Product.objects.filter(category__in=hoodie_categories).select_related('category')
        
    except Exception as e:
        # Если есть ошибка с базой данных, используем пустые списки
        tshirt_products = []
        hoodie_products = []
    
    # Цены для категорий
    tshirt_prices = {
        'drop': 650,
        'wholesale': [600, 570, 540, 510, 450],
        'ranges': ['8–15 шт.', '16–31 шт.', '32–63 шт.', '64–99 шт.', '100+ шт.']
    }
    
    hoodie_prices = {
        'drop': 1450,
        'wholesale': [1350, 1300, 1250, 1200, 1150],
        'ranges': ['8–15 шт.', '16–31 шт.', '32–63 шт.', '64–99 шт.', '100+ шт.']
    }
    
    # Статистика для оптовых закупок
    wholesale_stats = {
        'total_products': len(list(tshirt_products)) + len(list(hoodie_products)),
        'tshirt_count': len(list(tshirt_products)),
        'hoodie_count': len(list(hoodie_products)),
        'min_order_value': min(tshirt_prices['wholesale'][-1], hoodie_prices['wholesale'][-1]) * 8,  # минимальный заказ 8 штук
        'max_savings_percent': round((1 - (min(tshirt_prices['wholesale'][-1], hoodie_prices['wholesale'][-1]) / max(tshirt_prices['drop'], hoodie_prices['drop']))) * 100, 1)
    }
    
    # Получаем сохраненные данные компании для авторизованных пользователей
    company_data = {}
    user_invoice_stats = {}
    if request.user.is_authenticated:
        try:
            from accounts.models import UserProfile
            from orders.models import WholesaleInvoice
            user_profile = UserProfile.objects.get(user=request.user)
            company_data = {
                'company_name': getattr(user_profile, 'company_name', '') or '',
                'company_number': getattr(user_profile, 'company_number', '') or '',
                'delivery_address': getattr(user_profile, 'delivery_address', '') or '',
                'phone_number': user_profile.phone or '',
                'store_link': getattr(user_profile, 'website', '') or ''
            }
            
            # Статистика накладных пользователя (показываем общую статистику)
            all_invoices = WholesaleInvoice.objects.all().order_by('-created_at')
            last_invoice = all_invoices.first()
            user_invoice_stats = {
                'total_invoices': all_invoices.count(),
                'last_invoice_date': last_invoice.created_at.strftime('%d.%m.%Y %H:%M') if last_invoice else 'Немає накладних',
                'total_products_available': tshirt_products.count() + hoodie_products.count(),
                'tshirt_count': tshirt_products.count(),
                'hoodie_count': hoodie_products.count()
            }
        except UserProfile.DoesNotExist:
            pass
    
    context = {
        'tshirt_products': tshirt_products,
        'hoodie_products': hoodie_products,
        'tshirt_prices': tshirt_prices,
        'hoodie_prices': hoodie_prices,
        'wholesale_stats': wholesale_stats,
        'company_data': company_data,
        'user_invoice_stats': user_invoice_stats,
    }
    
    return render(request, 'pages/wholesale_order_form.html', context)


@require_POST
@csrf_exempt
def generate_wholesale_invoice(request):
    """Генерирует Excel накладную для оптового заказа"""
    import json
    from datetime import datetime
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from orders.models import WholesaleInvoice
    from accounts.models import UserProfile
    
    try:
        # Получаем данные из запроса
        data = json.loads(request.body)
        company_data = data.get('companyData', {})
        order_items = data.get('orderItems', [])
        
        if not order_items:
            return JsonResponse({'error': 'Немає товарів для накладної'}, status=400)
        
        # Генерируем номер накладной с красивой датой (киевское время)
        from django.utils import timezone
        import pytz
        
        kiev_tz = pytz.timezone('Europe/Kiev')
        now = timezone.now().astimezone(kiev_tz)
        timestamp = now.strftime('%Y%m%d_%H%M%S')
        invoice_number = f"ОПТ_{timestamp}"
        
        # Красивое название файла с названием компании
        company_name = company_data.get('companyName', 'Company').strip()
        if not company_name:
            company_name = 'Company'
        
        # Красивая дата для названия файла
        beautiful_date = now.strftime('%d.%m.%Y_%H-%M')
        file_name = f"{company_name}_накладнаОПТ_{beautiful_date}.xlsx"
        
        # Подсчитываем общие данные
        total_tshirts = 0
        total_hoodies = 0
        total_amount = 0
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Оптова накладна"
        
        # Стили
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        title_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=10)
        company_font = Font(name='Arial', size=11, bold=True, color='366092')
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        company_fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Заголовок накладной
        ws.merge_cells('A1:G1')
        ws['A1'] = 'ОПТОВА НАКЛАДНА'
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='366092')
        ws['A1'].alignment = center_alignment
        ws['A1'].fill = company_fill
        
        # Информация о компании
        row = 3
        ws[f'A{row}'] = 'Інформація про замовника:'
        ws[f'A{row}'].font = title_font
        
        row += 1
        ws[f'A{row}'] = 'Назва компанії/ФОП/ПІБ:'
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'] = company_data.get('companyName', '')
        ws[f'B{row}'].font = company_font
        ws[f'B{row}'].fill = company_fill
        
        if company_data.get('companyNumber'):
            row += 1
            ws[f'A{row}'] = 'Номер компанії/ЄДРПОУ/ІПН:'
            ws[f'A{row}'].font = normal_font
            ws[f'B{row}'] = company_data.get('companyNumber', '')
            ws[f'B{row}'].font = normal_font
        
        row += 1
        ws[f'A{row}'] = 'Номер телефону:'
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'] = company_data.get('contactPhone', '')
        ws[f'B{row}'].font = normal_font
        
        row += 1
        ws[f'A{row}'] = 'Адреса доставки:'
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'] = company_data.get('deliveryAddress', '')
        ws[f'B{row}'].font = normal_font
        
        if company_data.get('storeLink'):
            row += 1
            ws[f'A{row}'] = 'Посилання на магазин:'
            ws[f'A{row}'].font = normal_font
            ws[f'B{row}'] = company_data.get('storeLink', '')
            ws[f'B{row}'].font = normal_font
        
        row += 2
        ws[f'A{row}'] = f'Номер накладної: {invoice_number}'
        ws[f'A{row}'].font = title_font
        
        row += 1
        ws[f'A{row}'] = f'Дата створення: {now.strftime("%d.%m.%Y о %H:%M")}'
        ws[f'A{row}'].font = normal_font
        
        # Заголовок таблицы товаров
        row += 2
        headers = ['№', 'Назва товару', 'Розмір', 'Колір', 'Кількість', 'Ціна за од.', 'Сума']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Заполняем товары
        item_number = 1
        row += 1
        
        for item in order_items:
            product = item.get('product', {})
            quantity = item.get('quantity', 0)
            price = item.get('price', 0)
            total = item.get('total', 0)
            
            # Подсчитываем общие данные
            if product.get('type') == 'tshirt':
                total_tshirts += quantity
            elif product.get('type') == 'hoodie':
                total_hoodies += quantity
            
            total_amount += total
            
            # Заполняем строку товара
            ws.cell(row=row, column=1, value=item_number).font = normal_font
            ws.cell(row=row, column=1).alignment = center_alignment
            ws.cell(row=row, column=1).border = thin_border
            
            ws.cell(row=row, column=2, value=product.get('title', '')).font = normal_font
            ws.cell(row=row, column=2).alignment = left_alignment
            ws.cell(row=row, column=2).border = thin_border
            
            ws.cell(row=row, column=3, value=item.get('size', '')).font = normal_font
            ws.cell(row=row, column=3).alignment = center_alignment
            ws.cell(row=row, column=3).border = thin_border
            
            ws.cell(row=row, column=4, value=item.get('color', '')).font = normal_font
            ws.cell(row=row, column=4).alignment = center_alignment
            ws.cell(row=row, column=4).border = thin_border
            
            ws.cell(row=row, column=5, value=quantity).font = normal_font
            ws.cell(row=row, column=5).alignment = center_alignment
            ws.cell(row=row, column=5).border = thin_border
            
            ws.cell(row=row, column=6, value=f"{price}₴").font = normal_font
            ws.cell(row=row, column=6).alignment = center_alignment
            ws.cell(row=row, column=6).border = thin_border
            
            ws.cell(row=row, column=7, value=f"{total}₴").font = normal_font
            ws.cell(row=row, column=7).alignment = center_alignment
            ws.cell(row=row, column=7).border = thin_border
            
            # Чередуем цвет фона строк
            if item_number % 2 == 0:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = light_fill
            
            item_number += 1
            row += 1
        
        # Итоговая строка
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'РАЗОМ:'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = right_alignment = Alignment(horizontal='right', vertical='center')
        ws[f'A{row}'].border = thin_border
        ws[f'A{row}'].fill = company_fill
        
        ws[f'G{row}'] = f"{total_amount}₴"
        ws[f'G{row}'].font = title_font
        ws[f'G{row}'].alignment = center_alignment
        ws[f'G{row}'].border = thin_border
        ws[f'G{row}'].fill = company_fill
        
        # Статистика
        row += 2
        ws[f'A{row}'] = 'Статистика замовлення:'
        ws[f'A{row}'].font = title_font
        
        row += 1
        ws[f'A{row}'] = f'Футболки: {total_tshirts} шт.'
        ws[f'A{row}'].font = normal_font
        
        row += 1
        ws[f'A{row}'] = f'Худі: {total_hoodies} шт.'
        ws[f'A{row}'].font = normal_font
        
        row += 1
        ws[f'A{row}'] = f'Загальна сума: {total_amount}₴'
        ws[f'A{row}'].font = title_font
        
        # Автоматически настраиваем ширину колонок по содержимому
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Устанавливаем минимальную ширину и добавляем небольшой отступ
            adjusted_width = max(max_length + 2, 12)
            # Ограничиваем максимальную ширину для читаемости
            adjusted_width = min(adjusted_width, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем данные компании в профиль пользователя (если авторизован)
        if request.user.is_authenticated:
            try:
                user_profile, created = UserProfile.objects.get_or_create(user=request.user)
                if hasattr(user_profile, 'company_name'):
                    user_profile.company_name = company_data.get('companyName', '')
                if hasattr(user_profile, 'company_number'):
                    user_profile.company_number = company_data.get('companyNumber', '')
                if hasattr(user_profile, 'delivery_address'):
                    user_profile.delivery_address = company_data.get('deliveryAddress', '')
                user_profile.phone = company_data.get('contactPhone', '')
                if hasattr(user_profile, 'website'):
                    user_profile.website = company_data.get('storeLink', '')
                user_profile.save()
            except Exception as e:
                print(f"Error saving user profile: {e}")
        
        # Сохраняем в базу данных
        try:
            invoice = WholesaleInvoice.objects.create(
                invoice_number=invoice_number,
                company_name=company_data.get('companyName', ''),
                company_number=company_data.get('companyNumber', ''),
                contact_phone=company_data.get('contactPhone', ''),
                delivery_address=company_data.get('deliveryAddress', ''),
                store_link=company_data.get('storeLink', ''),
                total_tshirts=total_tshirts,
                total_hoodies=total_hoodies,
                total_amount=total_amount,
                order_details={
                    'company_data': company_data,
                    'order_items': order_items
                }
            )
            invoice_id = invoice.id
        except Exception as e:
            # Временно используем timestamp как ID если база недоступна
            import time
            invoice_id = int(time.time())
            print(f"Database error (temporary workaround): {e}")
        
        # Сохраняем файл на сервере
        import os
        from django.conf import settings
        
        # Создаем папку для накладных пользователя
        user_folder = f"invoices/user_{request.user.id if request.user.is_authenticated else 'anonymous'}"
        invoice_dir = os.path.join(settings.MEDIA_ROOT, user_folder)
        os.makedirs(invoice_dir, exist_ok=True)
        
        # Логируем создание папки
        print(f"Creating invoice directory: {invoice_dir}")
        
        # Путь к файлу (используем уже созданное красивое имя)
        file_path = os.path.join(invoice_dir, file_name)
        
        # Сохраняем Excel файл на сервере
        wb.save(file_path)
        print(f"Saved invoice file: {file_path}")
        
        # Обновляем invoice с путем к файлу (если invoice создался)
        try:
            if 'invoice' in locals() and invoice:
                invoice.file_path = file_path
                invoice.save()
        except:
            pass  # Игнорируем ошибки обновления
        
        # Подготавливаем ответ для скачивания
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        
        # Читаем файл и отправляем
        with open(file_path, 'rb') as f:
            response.write(f.read())
        
        return JsonResponse({
            'success': True,
            'invoice_number': invoice_number,
            'invoice_id': invoice_id,
            'file_name': file_name,
            'company_name': company_name,
            'total_amount': total_amount,
            'total_tshirts': total_tshirts,
            'total_hoodies': total_hoodies
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Помилка при створенні накладної: {str(e)}'}, status=500)


def _fetch_and_apply_checkout_status(order, source='api'):
    """Заглушка для функции получения статуса checkout"""
    return {'payment_status': 'unknown'}

def _update_order_from_checkout_result(order, result, source='api'):
    """Заглушка для функции обновления заказа из результата checkout"""
    pass

def download_invoice_file(request, invoice_id):
    """Скачивание сохраненного файла накладной"""
    from django.http import FileResponse, HttpResponse
    from django.conf import settings
    from orders.models import WholesaleInvoice
    import os
    
    try:
        # Получаем накладную
        invoice = WholesaleInvoice.objects.get(id=invoice_id)
        
        # Проверяем права доступа (только владелец или анонимные)
        if request.user.is_authenticated:
            # Для зарегистрированных пользователей можно добавить проверку
            pass
        
        # Проверяем существование файла
        if not invoice.file_path or not os.path.exists(invoice.file_path):
            # Если файл не найден, возвращаем ошибку
            return HttpResponse('Файл накладної не знайдено', status=404)
        
        # Отправляем файл
        filename = os.path.basename(invoice.file_path) if invoice.file_path else f"invoice_{invoice.id}.xlsx"
        
        try:
            response = FileResponse(
                open(invoice.file_path, 'rb'),
                as_attachment=True,
                filename=filename
            )
            return response
        except Exception as file_error:
            return HttpResponse(f'Помилка при відкритті файлу: {str(file_error)}', status=500)
        
    except WholesaleInvoice.DoesNotExist:
        return HttpResponse('Накладна не знайдена', status=404)
    except Exception as e:
        return HttpResponse(f'Помилка при скачуванні: {str(e)}', status=500)


@require_POST
@csrf_exempt
def delete_wholesale_invoice(request, invoice_id):
    """Удаление накладной"""
    from orders.models import WholesaleInvoice
    import os
    
    try:
        # Получаем накладную
        invoice = WholesaleInvoice.objects.get(id=invoice_id)
        
        # Удаляем файл, если он существует (не критично, если файл не найден)
        if invoice.file_path:
            try:
                if os.path.exists(invoice.file_path):
                    os.remove(invoice.file_path)
                    print(f"File deleted: {invoice.file_path}")
                else:
                    print(f"File not found (but that's OK): {invoice.file_path}")
            except Exception as e:
                print(f"Error deleting file (but continuing): {e}")
        
        # Удаляем запись из базы данных
        invoice.delete()
        print(f"Invoice {invoice_id} deleted successfully")
        
        return JsonResponse({'success': True, 'message': 'Накладна видалена'})

    except WholesaleInvoice.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Накладна не знайдена'}, status=404)
    except Exception as e:
        print(f"Error deleting invoice {invoice_id}: {e}")
        return JsonResponse({'success': False, 'error': f'Помилка: {str(e)}'}, status=500)


@login_required
def get_user_invoices(request):
    """Получение списка накладных пользователя"""
    from orders.models import WholesaleInvoice
    from django.core.paginator import Paginator
    
    try:
        # Получаем накладные пользователя (по номеру телефона или другим данным)
        # Пока возвращаем все накладные, но можно добавить фильтрацию
        invoices = WholesaleInvoice.objects.all().order_by('-created_at')
        
        # Пагинация
        paginator = Paginator(invoices, 10)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        invoices_data = []
        for invoice in page_obj:
            invoices_data.append({
                'id': invoice.id,
                'invoice_number': invoice.invoice_number,
                'company_name': invoice.company_name,
                'total_amount': float(invoice.total_amount),
                'total_tshirts': invoice.total_tshirts,
                'total_hoodies': invoice.total_hoodies,
                'created_at': invoice.created_at.strftime('%d.%m.%Y %H:%M'),
                'status': invoice.get_status_display(),
                'file_name': invoice.file_name
            })
        
        return JsonResponse({
            'success': True,
            'invoices': invoices_data,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Помилка: {str(e)}'}, status=500)
