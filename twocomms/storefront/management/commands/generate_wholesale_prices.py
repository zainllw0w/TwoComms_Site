from django.core.management.base import BaseCommand
from django.db.models import Q
from storefront.models import Product, Category
from productcolors.models import ProductColorVariant, Color
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os


def _translate_color_to_ukrainian(color_name):
    """Translate color names from Russian to Ukrainian."""
    color_translations = {
        'черный': 'чорний',
        'белый': 'білий',
        'красный': 'червоний',
        'синий': 'синій',
        'зеленый': 'зелений',
        'желтый': 'жовтий',
        'оранжевый': 'помаранчевий',
        'фиолетовый': 'фіолетовий',
        'розовый': 'рожевий',
        'серый': 'сірий',
        'коричневый': 'коричневий',
        'голубой': 'блакитний',
        'бордовый': 'бордовий',
        'темно-синий': 'темно-синій',
        'светло-серый': 'світло-сірий',
        'темно-серый': 'темно-сірий',
        'бежевый': 'бежевий',
        'кремовый': 'кремовий',
        'оливковый': 'оливковий',
        'бирюзовый': 'бірюзовий',
        'лайм': 'лайм',
        'лаванда': 'лаванда',
        'мятный': 'м\'ятний',
        'коралловый': 'кораловий',
        'персиковый': 'персиковий',
        'золотой': 'золотий',
        'серебряный': 'срібний',
        'медный': 'мідний',
        'бронзовый': 'бронзовий',
        'хаки': 'хаки',
        'фуксия': 'фуксія',
        'индиго': 'індиго',
        'пурпурный': 'пурпурний',
        'малиновый': 'малиновий',
        'вишневый': 'вишневий',
        'изумрудный': 'смарагдовий',
        'нефрит': 'нефрит',
        'бирюза': 'бірюза',
        'бирюзово-зеленый': 'бірюзово-зелений',
        'темно-зеленый': 'темно-зелений',
        'светло-зеленый': 'світло-зелений',
        'темно-красный': 'темно-червоний',
        'светло-красный': 'світло-червоний',
        'темно-синий': 'темно-синій',
        'светло-синий': 'світло-синій',
        'темно-фиолетовый': 'темно-фіолетовий',
        'светло-фиолетовый': 'світло-фіолетовий',
        'темно-коричневый': 'темно-коричневий',
        'светло-коричневый': 'світло-коричневий',
        'темно-розовый': 'темно-рожевий',
        'светло-розовый': 'світло-рожевий',
        'темно-желтый': 'темно-жовтий',
        'светло-желтый': 'світло-жовтий',
        'темно-оранжевый': 'темно-помаранчевий',
        'светло-оранжевый': 'світло-помаранчевий',
        # Додаткові кольори
        'кайот': 'кайот',
        'navy': 'navy',
        'charcoal': 'charcoal',
        'heather': 'heather',
        'maroon': 'maroon',
        'forest': 'forest',
        'royal': 'royal',
        'sport grey': 'sport grey',
        'ash': 'ash',
        'dark heather': 'dark heather',
        'red': 'red',
        'blue': 'blue',
        'green': 'green',
        'yellow': 'yellow',
        'orange': 'orange',
        'purple': 'purple',
        'pink': 'pink',
        'black': 'чорний',
        'white': 'білий',
        'grey': 'сірий',
        'brown': 'коричневий',
    }
    
    if not color_name:
        return 'чорний'
    
    color_lower = color_name.lower().strip()
    return color_translations.get(color_lower, color_name)


class Command(BaseCommand):
    help = 'Генерирует XLSX-прайс оптовых цен для футболок и худи'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output',
            type=str,
            default='wholesale_prices.xlsx',
            help='Путь к выходному XLSX файлу'
        )

    def handle(self, *args, **options):
        output_path = options['output']
        
        # Фильтруем категории по ключевым словам
        tshirt_categories = Category.objects.filter(
            Q(name__icontains='футболка') | 
            Q(name__icontains='tshirt') | 
            Q(name__icontains='t-shirt') |
            Q(slug__icontains='футболка') |
            Q(slug__icontains='tshirt') |
            Q(slug__icontains='t-shirt')
        )
        
        hoodie_categories = Category.objects.filter(
            Q(name__icontains='худи') | 
            Q(name__icontains='hoodie') | 
            Q(name__icontains='hooded') |
            Q(slug__icontains='худи') |
            Q(slug__icontains='hoodie') |
            Q(slug__icontains='hooded')
        )
        
        # Получаем товары из нужных категорий
        tshirt_products = Product.objects.filter(category__in=tshirt_categories)
        hoodie_products = Product.objects.filter(category__in=hoodie_categories)
        
        # Создаем XLSX файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Прайс (опт)"
        
        # Стили
        header_font = Font(bold=True, size=14)
        table_header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Заголовок (строка 1)
        ws.merge_cells('A1:L1')
        ws['A1'] = "Оптові ціни від кількості. Мінімальне замовлення по моделі — від 8 шт. і кратно 8."
        ws['A1'].font = header_font
        ws['A1'].alignment = center_alignment
        
        # Заголовки таблицы (строка 2)
        headers = [
            'Категорія',
            'Товар (S–XL)',
            'Артикул',
            'Колір',
            '8–15 шт. (за 1 шт.)',
            '16–31 шт.',
            '32–63 шт.',
            '64–99 шт.',
            '100+ шт.',
            'Посилання'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = table_header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Цены для категорий
        tshirt_prices = [750, 700, 650, 600, 550]
        hoodie_prices = [1600, 1500, 1400, 1300, 1200]
        
        row = 3
        
        # Сначала худи
        for product in hoodie_products.order_by('title'):
            # Получаем артикул (если есть)
            sku = getattr(product, 'sku', None) or f"TC-{product.id}"
            
            # Получаем цвета товара
            color_variants = ProductColorVariant.objects.filter(product=product)
            colors = []
            for variant in color_variants:
                if variant.color:
                    colors.append(variant.color.name or str(variant.color))
            
            # Формируем название с "[фліс]" в конце
            product_title = f"{product.title} (S–XL) [фліс]"
            
            # Создаем ссылку на товар
            product_url = f"https://twocomms.shop/product/{product.slug}/"
            
            if colors:
                # Если есть цвета, создаем строку для каждого цвета
                for color in colors:
                    ws.cell(row=row, column=1, value='Худі')
                    ws.cell(row=row, column=2, value=product_title)
                    ws.cell(row=row, column=3, value=sku)
                    ws.cell(row=row, column=4, value=_translate_color_to_ukrainian('чорний'))  # Для худи завжди чорний
                    
                    # Добавляем цены
                    for col, price in enumerate(hoodie_prices, 5):
                        ws.cell(row=row, column=col, value=price)
                        ws.cell(row=row, column=col).alignment = center_alignment
                    
                    # Добавляем ссылку на товар
                    ws.cell(row=row, column=10, value=product_url)
                    
                    row += 1
            else:
                # Если нет цветов, создаем одну строку
                ws.cell(row=row, column=1, value='Худі')
                ws.cell(row=row, column=2, value=product_title)
                ws.cell(row=row, column=3, value=sku)
                ws.cell(row=row, column=4, value=_translate_color_to_ukrainian('чорний'))  # Для худи завжди чорний
                
                # Добавляем цены
                for col, price in enumerate(hoodie_prices, 5):
                    ws.cell(row=row, column=col, value=price)
                    ws.cell(row=row, column=col).alignment = center_alignment
                
                # Добавляем ссылку на товар
                ws.cell(row=row, column=10, value=product_url)
                
                row += 1
        
        # Затем футболки
        for product in tshirt_products.order_by('title'):
            # Получаем артикул (если есть)
            sku = getattr(product, 'sku', None) or f"TC-{product.id}"
            
            # Получаем цвета товара
            color_variants = ProductColorVariant.objects.filter(product=product)
            colors = []
            for variant in color_variants:
                if variant.color:
                    colors.append(variant.color.name or str(variant.color))
            
            # Формируем название без [фліс]
            product_title = f"{product.title} (S–XL)"
            
            # Создаем ссылку на товар
            product_url = f"https://twocomms.shop/product/{product.slug}/"
            
            if colors:
                # Если есть цвета, создаем строку для каждого цвета
                for color in colors:
                    ws.cell(row=row, column=1, value='Футболки')
                    ws.cell(row=row, column=2, value=product_title)
                    ws.cell(row=row, column=3, value=sku)
                    ws.cell(row=row, column=4, value=_translate_color_to_ukrainian(color if color else 'чорний'))
                    
                    # Добавляем цены
                    for col, price in enumerate(tshirt_prices, 5):
                        ws.cell(row=row, column=col, value=price)
                        ws.cell(row=row, column=col).alignment = center_alignment
                    
                    # Добавляем ссылку на товар
                    ws.cell(row=row, column=10, value=product_url)
                    
                    row += 1
            else:
                # Если нет цветов, создаем одну строку с черным цветом
                ws.cell(row=row, column=1, value='Футболки')
                ws.cell(row=row, column=2, value=product_title)
                ws.cell(row=row, column=3, value=sku)
                ws.cell(row=row, column=4, value=_translate_color_to_ukrainian('чорний'))
                
                # Добавляем цены
                for col, price in enumerate(tshirt_prices, 5):
                    ws.cell(row=row, column=col, value=price)
                    ws.cell(row=row, column=col).alignment = center_alignment
                
                # Добавляем ссылку на товар
                ws.cell(row=row, column=10, value=product_url)
                
                row += 1
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(output_path)
        
        self.stdout.write(
            self.style.SUCCESS(f'Прайс успешно создан: {output_path}')
        )
        
        # Статистика
        tshirt_count = tshirt_products.count()
        hoodie_count = hoodie_products.count()
        
        self.stdout.write(f'Найдено товаров:')
        self.stdout.write(f'  - Футболки: {tshirt_count}')
        self.stdout.write(f'  - Худи: {hoodie_count}')
        self.stdout.write(f'  - Всего: {tshirt_count + hoodie_count}')
