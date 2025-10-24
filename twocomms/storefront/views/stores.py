"""
Offline Stores Management Views
Управління оффлайн магазинами

Відповідальність:
- CRUD операції для магазинів (створення, редагування, видалення, активація)
- Управління інвентарем та продажами
- Генерація накладних та рахунків
- Статистика та аналітика роботи магазинів
- Замовлення для поповнення магазинів

Основні функції:
- admin_offline_stores: Список всіх offline магазинів
- admin_offline_store_create/edit/toggle/delete: CRUD операції
- admin_store_management: Головна сторінка управління магазином
- admin_store_add_products_to_store: Додавання товарів до магазину
- admin_store_generate_invoice: Генерація Excel накладної
- admin_store_mark_product_sold: Позначення товару як проданого

Helper функції:
- _get_store_inventory_items: Отримання інвентарю магазину
- _get_store_sales: Отримання продажів магазину
- _calculate_inventory_stats: Розрахунок статистики інвентарю
- _calculate_sales_stats: Розрахунок статистики продажів
- _build_category_stats: Статистика по категоріях
- _serialize_sale: Серіалізація даних продажу
- _compose_store_stats: Збірка повної статистики магазину
"""

import json
import os
from datetime import datetime

from django import forms
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, Alignment

from ..models import (
    OfflineStore,
    Product,
    Category,
    StoreProduct,
    StoreOrder,
    StoreOrderItem,
    StoreInvoice,
    StoreSale,
)
from productcolors.models import Color, ProductColorVariant


# ===== CRUD операції для оффлайн магазинів =====

@login_required
def admin_offline_stores(request):
    """Список оффлайн магазинів"""
    if not request.user.is_staff:
        return redirect('home')
    
    stores = OfflineStore.objects.all()
    
    # Підрахунок статистики
    total_stores = stores.count()
    active_stores = stores.filter(is_active=True).count()
    
    return render(request, 'pages/admin_offline_stores.html', {
        'stores': stores,
        'total_stores': total_stores,
        'active_stores': active_stores,
        'section': 'offline_stores'
    })


class OfflineStoreForm(forms.ModelForm):
    class Meta:
        model = OfflineStore
        fields = ['name', 'address', 'phone', 'email', 'working_hours', 'description', 'is_active', 'order']
        widgets = {
            'name': forms.TextInput(attrs={
                'class': 'form-control bg-elevate',
                'placeholder': 'Назва магазину'
            }),
            'address': forms.Textarea(attrs={
                'class': 'form-control bg-elevate',
                'rows': 3,
                'placeholder': 'Повна адреса магазину'
            }),
            'phone': forms.TextInput(attrs={
                'class': 'form-control bg-elevate',
                'placeholder': 'Телефон магазину'
            }),
            'email': forms.EmailInput(attrs={
                'class': 'form-control bg-elevate',
                'placeholder': 'Email магазину'
            }),
            'working_hours': forms.TextInput(attrs={
                'class': 'form-control bg-elevate',
                'placeholder': 'Наприклад: Пн-Пт 9:00-18:00, Сб 10:00-16:00'
            }),
            'description': forms.Textarea(attrs={
                'class': 'form-control bg-elevate',
                'rows': 4,
                'placeholder': 'Опис магазину, особливості, послуги'
            }),
            'order': forms.NumberInput(attrs={
                'class': 'form-control bg-elevate',
                'min': '0',
                'placeholder': 'Порядок відображення'
            }),
            'is_active': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)


@login_required
def admin_offline_store_create(request):
    """Створення нового оффлайн магазину"""
    if not request.user.is_staff:
        return redirect('home')
    
    if request.method == 'POST':
        form = OfflineStoreForm(request.POST)
        if form.is_valid():
            store = form.save()
            return redirect('admin_store_management', store_id=store.id)
    else:
        form = OfflineStoreForm()
    
    return render(request, 'pages/admin_offline_store_form.html', {
        'form': form,
        'mode': 'create',
        'section': 'offline_stores'
    })


@login_required
def admin_offline_store_edit(request, pk):
    """Редагування оффлайн магазину"""
    if not request.user.is_staff:
        return redirect('home')
    
    store = get_object_or_404(OfflineStore, pk=pk)
    
    if request.method == 'POST':
        form = OfflineStoreForm(request.POST, instance=store)
        if form.is_valid():
            form.save()
            return redirect('admin_offline_stores')
    else:
        form = OfflineStoreForm(instance=store)
    
    return render(request, 'pages/admin_offline_store_form.html', {
        'form': form,
        'mode': 'edit',
        'store': store,
        'section': 'offline_stores'
    })


@login_required
def admin_offline_store_toggle(request, pk):
    """Активація/деактивація оффлайн магазину"""
    if not request.user.is_staff:
        return redirect('home')
    
    store = get_object_or_404(OfflineStore, pk=pk)
    store.is_active = not store.is_active
    store.save()
    
    return redirect('admin_offline_stores')


@login_required
def admin_offline_store_delete(request, pk):
    """Видалення оффлайн магазину"""
    if not request.user.is_staff:
        return redirect('home')
    
    store = get_object_or_404(OfflineStore, pk=pk)
    store.delete()
    
    return redirect('admin_offline_stores')


# ===== Helper функції для статистики =====

def _get_store_inventory_items(store):
    """Отримання інвентарю магазину"""
    return list(
        store.store_products.filter(is_active=True)
        .select_related('product__category', 'product', 'color')
        .order_by('-created_at')
    )


def _get_store_sales(store):
    """Отримання продажів магазину"""
    return list(
        store.store_sales.select_related('product__category', 'product', 'color')
        .order_by('-sold_at')
    )


def _calculate_inventory_stats(inventory_items):
    """Розрахунок статистики інвентарю"""
    stats = {
        'units': 0,
        'cost': 0,
        'revenue': 0,
        'margin': 0,
        'avg_price': 0,
        'avg_margin_per_unit': 0,
        'margin_percent': 0,
        'active_skus': len(inventory_items),
    }

    for item in inventory_items:
        qty = item.quantity or 0
        cost = (item.cost_price or 0) * qty
        revenue = (item.selling_price or 0) * qty
        margin = revenue - cost

        stats['units'] += qty
        stats['cost'] += cost
        stats['revenue'] += revenue
        stats['margin'] += margin

    if stats['units']:
        stats['avg_price'] = stats['revenue'] / stats['units']
        stats['avg_margin_per_unit'] = stats['margin'] / stats['units'] if stats['margin'] else 0

    if stats['cost']:
        stats['margin_percent'] = (stats['margin'] / stats['cost']) * 100

    return stats


def _calculate_sales_stats(sales_items):
    """Розрахунок статистики продажів"""
    stats = {
        'units': 0,
        'cost': 0,
        'revenue': 0,
        'margin': 0,
        'avg_ticket': 0,
        'margin_percent': 0,
        'last_sale_at': None,
    }

    for sale in sales_items:
        qty = sale.quantity or 0
        cost = (sale.cost_price or 0) * qty
        revenue = (sale.selling_price or 0) * qty
        margin = revenue - cost

        stats['units'] += qty
        stats['cost'] += cost
        stats['revenue'] += revenue
        stats['margin'] += margin

    if stats['units']:
        stats['avg_ticket'] = stats['revenue'] / stats['units']

    if stats['cost']:
        stats['margin_percent'] = (stats['margin'] / stats['cost']) * 100

    if sales_items:
        stats['last_sale_at'] = sales_items[0].sold_at

    return stats


def _build_category_stats(inventory_items):
    """Статистика по категоріях"""
    category_data = {}

    for item in inventory_items:
        category_name = (
            item.product.category.name if getattr(item.product, 'category', None) else 'Без категорії'
        )
        entry = category_data.setdefault(
            category_name,
            {'name': category_name, 'units': 0, 'cost': 0, 'revenue': 0, 'margin': 0},
        )

        qty = item.quantity or 0
        cost = (item.cost_price or 0) * qty
        revenue = (item.selling_price or 0) * qty

        entry['units'] += qty
        entry['cost'] += cost
        entry['revenue'] += revenue
        entry['margin'] = entry['revenue'] - entry['cost']

    for entry in category_data.values():
        if entry['cost']:
            entry['margin_percent'] = (entry['margin'] / entry['cost']) * 100
        else:
            entry['margin_percent'] = 0

    return sorted(category_data.values(), key=lambda x: x['revenue'], reverse=True)


def _serialize_sale(sale):
    """Серіалізація даних продажу"""
    image = None
    display_image = getattr(sale.product, 'display_image', None)
    if display_image:
        try:
            image = display_image.url
        except Exception:
            image = None

    return {
        'id': sale.id,
        'product_id': sale.product_id,
        'title': sale.product.title,
        'size': sale.size,
        'color': sale.color.name if sale.color else None,
        'color_hex': sale.color.primary_hex if sale.color else None,
        'quantity': sale.quantity,
        'cost_price': sale.cost_price,
        'selling_price': sale.selling_price,
        'total_revenue': sale.total_revenue,
        'total_margin': sale.margin,
        'sold_at': sale.sold_at.isoformat(),
        'sold_at_human': sale.sold_at.strftime('%d.%m.%Y %H:%M'),
        'image': image,
    }


def _compose_store_stats(inventory_items, sales_items):
    """Збірка повної статистики магазину"""
    inventory_stats = _calculate_inventory_stats(inventory_items)
    sales_stats = _calculate_sales_stats(sales_items)

    total_units = inventory_stats['units'] + sales_stats['units']
    sell_through_rate = (sales_stats['units'] / total_units * 100) if total_units else 0

    if sales_stats['last_sale_at']:
        last_sale_iso = sales_stats['last_sale_at'].isoformat()
        last_sale_human = sales_stats['last_sale_at'].strftime('%d.%m.%Y %H:%M')
    else:
        last_sale_iso = None
        last_sale_human = None

    return {
        'inventory_units': inventory_stats['units'],
        'inventory_cost': inventory_stats['cost'],
        'inventory_revenue': inventory_stats['revenue'],
        'inventory_margin': inventory_stats['margin'],
        'inventory_margin_percent': round(inventory_stats['margin_percent'], 2) if inventory_stats['margin_percent'] else 0,
        'inventory_avg_price': round(inventory_stats['avg_price'], 2) if inventory_stats['avg_price'] else 0,
        'inventory_avg_margin_per_unit': round(inventory_stats['avg_margin_per_unit'], 2) if inventory_stats['avg_margin_per_unit'] else 0,
        'inventory_active_skus': inventory_stats['active_skus'],
        'sales_units': sales_stats['units'],
        'sales_cost': sales_stats['cost'],
        'sales_revenue': sales_stats['revenue'],
        'sales_margin': sales_stats['margin'],
        'sales_margin_percent': round(sales_stats['margin_percent'], 2) if sales_stats['margin_percent'] else 0,
        'sales_avg_ticket': round(sales_stats['avg_ticket'], 2) if sales_stats['avg_ticket'] else 0,
        'sell_through_rate': round(sell_through_rate, 2) if sell_through_rate else 0,
        'unsold_units': inventory_stats['units'],
        'total_committed_cost': inventory_stats['cost'] + sales_stats['cost'],
        'total_realized_revenue': sales_stats['revenue'],
        'total_expected_revenue': inventory_stats['revenue'] + sales_stats['revenue'],
        'lifetime_margin': sales_stats['margin'],
        'last_sale_at': last_sale_iso,
        'last_sale_at_display': last_sale_human,
    }


# ===== Управління магазином =====

@login_required
def admin_store_management(request, store_id):
    """Головна сторінка управління оффлайн магазином"""
    if not request.user.is_staff:
        return redirect('home')
    
    store = get_object_or_404(OfflineStore, pk=store_id)

    products = Product.objects.select_related('category').prefetch_related(
        'color_variants__color',
        'color_variants__images'
    ).all()

    categories = Category.objects.select_related().filter(is_active=True).order_by('order', 'name')

    inventory_items = _get_store_inventory_items(store)
    sales_items = _get_store_sales(store)

    active_orders = StoreOrder.objects.select_related('user').filter(store=store, status='draft').prefetch_related(
        'order_items__product__category', 'order_items__color'
    )

    store_stats = _compose_store_stats(inventory_items, sales_items)
    category_stats = _build_category_stats(inventory_items)
    sold_products = [_serialize_sale(sale) for sale in sales_items]

    return render(request, 'pages/admin_store_management.html', {
        'store': store,
        'products': products,
        'categories': categories,
        'store_products': inventory_items,
        'sold_products': sold_products,
        'active_orders': active_orders,
        'store_stats': store_stats,
        'category_stats': category_stats,
        'section': 'offline_stores'
    })


@login_required
def admin_store_get_order_items(request, store_id, order_id):
    """Отримання товарів замовлення для AJAX"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Доступ заборонено'})
    
    try:
        store = OfflineStore.objects.get(id=store_id)
        order = StoreOrder.objects.get(id=order_id, store=store)
        
        items = []
        for item in order.order_items.all():
            items.append({
                'id': item.id,
                'product_name': item.product.name,
                'size': item.size,
                'color_name': item.color.name if item.color else None,
                'quantity': item.quantity,
                'cost_price': float(item.cost_price),
                'selling_price': float(item.selling_price),
            })
        
        return JsonResponse({'success': True, 'items': items})
    
    except (OfflineStore.DoesNotExist, StoreOrder.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Замовлення не знайдено'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def admin_store_get_product_colors(request, store_id, product_id):
    """Отримання кольорів товару для AJAX"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ заборонено'}, status=403)
    
    try:
        store = get_object_or_404(OfflineStore, pk=store_id)
        product = get_object_or_404(Product, pk=product_id)
        
        # Отримання кольорів товару через color_variants
        colors = []
        for variant in product.color_variants.all():
            # Отримання першого зображення для цього кольору варіанту
            first_image = variant.images.first()
            image_url = first_image.image.url if first_image else None
            
            colors.append({
                'id': variant.color.id,
                'name': variant.color.name or variant.color.primary_hex,
                'hex_code': variant.color.primary_hex,
                'image_url': image_url,
                'variant_id': variant.id
            })
        
        return JsonResponse({
            'success': True,
            'colors': colors
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_add_product_to_order(request, store_id):
    """Додавання товару в замовлення через AJAX"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ заборонено'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не підтримується'}, status=405)
    
    try:
        store = get_object_or_404(OfflineStore, pk=store_id)
        product_id = request.POST.get('product_id')
        size = request.POST.get('size', '').strip()
        color_id = request.POST.get('color_id')
        cost_price = int(request.POST.get('cost_price', 0))
        selling_price = int(request.POST.get('selling_price', 0))
        quantity = int(request.POST.get('quantity', 1))
        
        product = get_object_or_404(Product, pk=product_id)
        color = None
        if color_id and color_id != 'default':
            color = get_object_or_404(Color, pk=color_id)
        elif color_id == 'default':
            # Створюємо або отримуємо дефолтний чорний колір
            color, created = Color.objects.get_or_create(
                name='Чорний',
                defaults={'hex_code': '#000000'}
            )
        
        # Отримуємо або створюємо чернетку замовлення
        order, created = StoreOrder.objects.get_or_create(
            store=store,
            status='draft',
            defaults={'notes': ''}
        )
        
        # Перевіряємо, чи є вже такий товар у замовленні
        existing_item = StoreOrderItem.objects.filter(
            order=order,
            product=product,
            size=size or None,
            color=color
        ).first()
        
        if existing_item:
            # Оновлюємо існуючий товар
            existing_item.cost_price = cost_price
            existing_item.selling_price = selling_price
            existing_item.quantity += quantity
            existing_item.save()
            item = existing_item
        else:
            # Створюємо новий товар у замовленні
            item = StoreOrderItem.objects.create(
                order=order,
                product=product,
                size=size or None,
                color=color,
                cost_price=cost_price,
                selling_price=selling_price,
                quantity=quantity
            )
        
            # Отримуємо зображення кольору, якщо є
            color_image_url = None
            if item.color:
                try:
                    color_variant = ProductColorVariant.objects.filter(
                        product=item.product, 
                        color=item.color
                    ).first()
                    if color_variant:
                        first_image = color_variant.images.first()
                        if first_image:
                            color_image_url = first_image.image.url
                except:
                    pass
            
            return JsonResponse({
                'success': True,
                'order_id': order.id,
                'item': {
                    'id': item.id,
                    'product_name': item.product.title,
                    'size': item.size or '',
                    'color_name': item.color.name if item.color else 'Чорний',
                    'color_hex': item.color.primary_hex if item.color else '#000000',
                    'color_image': color_image_url,
                    'quantity': item.quantity,
                    'cost_price': float(item.cost_price),
                    'selling_price': float(item.selling_price),
                    'product_image': item.product.display_image.url if item.product.display_image else None,
                },
                'message': 'Товар додано до замовлення'
            })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_remove_product_from_order(request, store_id, order_id, item_id):
    """Видалення товару з замовлення"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ заборонено'}, status=403)
    
    try:
        order = get_object_or_404(StoreOrder, pk=order_id, store_id=store_id)
        item = get_object_or_404(StoreOrderItem, pk=item_id, order=order)
        item.delete()
        
        return JsonResponse({'success': True, 'message': 'Товар видалено з замовлення'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_add_products_to_store(request, store_id):
    """Додавання товарів із замовлення до магазину"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ заборонено'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не підтримується'}, status=405)
    
    try:
        store = get_object_or_404(OfflineStore, pk=store_id)
        order_id = request.POST.get('order_id')
        
        order = get_object_or_404(StoreOrder, pk=order_id, store=store)
        
        added_count = 0
        for item in order.order_items.all():
            # Перевіряємо, чи є вже такий товар у магазині
            existing_product = StoreProduct.objects.select_related('category').filter(
                store=store,
                product=item.product,
                size=item.size,
                color=item.color
            ).first()
            
            if existing_product:
                # Оновлюємо існуючий товар
                existing_product.cost_price = item.cost_price
                existing_product.selling_price = item.selling_price
                existing_product.quantity += item.quantity
                existing_product.is_active = True
                existing_product.save()
            else:
                # Створюємо новий товар у магазині
                StoreProduct.objects.create(
                    store=store,
                    product=item.product,
                    size=item.size,
                    color=item.color,
                    cost_price=item.cost_price,
                    selling_price=item.selling_price,
                    quantity=item.quantity
                )
            added_count += 1
        
        # Оновлюємо статус замовлення
        order.status = 'completed'
        order.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Додано {added_count} товарів до магазину'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_generate_invoice(request, store_id):
    """Генерація Excel накладної"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ заборонено'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не підтримується'}, status=405)
    
    try:
        store = get_object_or_404(OfflineStore, pk=store_id)
        order_id = request.POST.get('order_id')
        
        order = get_object_or_404(StoreOrder, pk=order_id, store=store)
        
        # Створюємо Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Накладна"
        
        # Заголовки
        ws['A1'] = f"Накладна для магазину: {store.name} (під реалізацію)"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].font = Font(size=12)
        
        # Заголовки таблиці
        headers = ['Товар', 'Кількість', 'Ціна (грн)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Дані товарів
        row = 5
        category_totals = {}
        
        for item in order.order_items.all():
            # Формуємо назву товару
            product_name = f"TwoComms {item.product.title}"
            if item.size:
                product_name += f" ({item.size})"
            if item.color:
                product_name += f" [{item.color.name}]"
            
            ws.cell(row=row, column=1, value=product_name)
            ws.cell(row=row, column=2, value=item.quantity)
            ws.cell(row=row, column=3, value=item.selling_price)
            
            # Підраховуємо по категоріях
            category = item.product.category.name
            if category not in category_totals:
                category_totals[category] = {'count': 0, 'total': 0}
            category_totals[category]['count'] += item.quantity
            category_totals[category]['total'] += item.selling_price * item.quantity
            
            row += 1
        
        # Підсумки по категоріях
        row += 1
        ws.cell(row=row, column=1, value="Підсумки по категоріях:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        for category, totals in category_totals.items():
            ws.cell(row=row, column=1, value=f"{category}: {totals['count']} шт., сума {totals['total']} грн")
            row += 1
        
        # Загальний підсумок
        row += 1
        total_amount = sum(totals['total'] for totals in category_totals.values())
        ws.cell(row=row, column=1, value=f"ВСЬОГО: {total_amount} грн")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        
        # Налаштування ширини колонок
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        
        # Зберігаємо файл
        filename = f"TwoComms_накладна_{store.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, 'invoices', filename)
        
        # Створюємо директорію якщо не існує
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        wb.save(filepath)
        
        # Зберігаємо інформацію про накладну в БД
        StoreInvoice.objects.create(
            store=store,
            order=order,
            file_name=filename,
            file_path=filepath
        )
        
        # Замовлення залишається активним для продовження роботи
        # order.status = 'completed'  # Прибрано - замовлення залишається draft
        
        # Повертаємо файл для завантаження
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_update_product(request, store_id, product_id):
    """Оновлення товару в магазині"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ заборонено'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не підтримується'}, status=405)
    
    try:
        store_product = get_object_or_404(StoreProduct, pk=product_id, store_id=store_id)

        cost_price = request.POST.get('cost_price')
        selling_price = request.POST.get('selling_price')
        quantity = request.POST.get('quantity')
        
        if cost_price:
            store_product.cost_price = int(cost_price)
        if selling_price:
            store_product.selling_price = int(selling_price)
        if quantity:
            store_product.quantity = int(quantity)

        store_product.save()

        store = store_product.store
        inventory_items = _get_store_inventory_items(store)
        sales_items = _get_store_sales(store)
        stats = _compose_store_stats(inventory_items, sales_items)
        category_stats = _build_category_stats(inventory_items)

        return JsonResponse({
            'success': True,
            'message': 'Товар оновлено',
            'product': {
                'id': store_product.id,
                'cost_price': store_product.cost_price,
                'selling_price': store_product.selling_price,
                'quantity': store_product.quantity,
                'margin_per_unit': store_product.margin,
                'total_cost': store_product.cost_price * store_product.quantity,
                'total_revenue': store_product.selling_price * store_product.quantity,
                'total_margin': (store_product.selling_price - store_product.cost_price) * store_product.quantity,
            },
            'stats': stats,
            'category_stats': category_stats,
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_remove_product(request, store_id, product_id):
    """Видалення товару з магазину"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ заборонено'}, status=403)

    try:
        store_product = get_object_or_404(StoreProduct, pk=product_id, store_id=store_id)
        store = store_product.store
        removed_product_id = store_product.id
        store_product.delete()

        inventory_items = _get_store_inventory_items(store)
        sales_items = _get_store_sales(store)
        stats = _compose_store_stats(inventory_items, sales_items)
        category_stats = _build_category_stats(inventory_items)

        return JsonResponse({
            'success': True,
            'message': 'Товар видалено з магазину',
            'removed_product_id': removed_product_id,
            'stats': stats,
            'category_stats': category_stats,
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def admin_store_mark_product_sold(request, store_id, product_id):
    """Позначає товар як проданий та переносить його в блок статистики продажів."""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ заборонено'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не підтримується'}, status=405)

    try:
        store_product = get_object_or_404(StoreProduct, pk=product_id, store_id=store_id)
        store = store_product.store

        quantity_param = request.POST.get('quantity')

        if quantity_param is None and request.body:
            try:
                payload = json.loads(request.body.decode('utf-8'))
                quantity_param = payload.get('quantity')
            except (ValueError, json.JSONDecodeError):
                quantity_param = None

        if quantity_param is None or quantity_param == '':
            quantity_to_sell = store_product.quantity
        else:
            try:
                quantity_to_sell = int(quantity_param)
            except (TypeError, ValueError):
                return JsonResponse({'error': 'Невірне значення кількості'}, status=400)

        if quantity_to_sell <= 0:
            return JsonResponse({'error': 'Кількість повинна бути більшою за нуль'}, status=400)

        if quantity_to_sell > store_product.quantity:
            return JsonResponse({'error': 'Кількість перевищує залишок товару'}, status=400)

        sale = StoreSale.objects.create(
            store=store,
            product=store_product.product,
            color=store_product.color,
            size=store_product.size,
            quantity=quantity_to_sell,
            cost_price=store_product.cost_price,
            selling_price=store_product.selling_price,
            source_store_product=store_product,
        )

        remaining_product = None
        removed_product_id = None

        if quantity_to_sell >= store_product.quantity:
            removed_product_id = store_product.id
            store_product.delete()
        else:
            store_product.quantity -= quantity_to_sell
            store_product.save(update_fields=['quantity', 'updated_at'])
            remaining_product = {
                'id': store_product.id,
                'quantity': store_product.quantity,
                'cost_price': store_product.cost_price,
                'selling_price': store_product.selling_price,
                'margin_per_unit': store_product.margin,
                'total_cost': store_product.cost_price * store_product.quantity,
                'total_revenue': store_product.selling_price * store_product.quantity,
                'total_margin': (store_product.selling_price - store_product.cost_price) * store_product.quantity,
            }

        inventory_items = _get_store_inventory_items(store)
        sales_items = _get_store_sales(store)
        stats = _compose_store_stats(inventory_items, sales_items)
        category_stats = _build_category_stats(inventory_items)

        return JsonResponse({
            'success': True,
            'message': 'Товар позначено як проданий',
            'sale': _serialize_sale(sale),
            'stats': stats,
            'category_stats': category_stats,
            'remaining_product': remaining_product,
            'removed_product_id': removed_product_id,
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

