"""
Wholesale views - Оптовые продажи.

Содержит views и helper функции для:
- Страницы оптовых продаж
- Формы оптовых заказов
- Генерации и управления накладными (invoices)
- Прайс-листов в XLSX формате
- Платежей через Monobank для оптовых заказов
"""

import logging
import json
import io
import os
from datetime import datetime
from decimal import Decimal

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.contrib.admin.views.decorators import staff_member_required
from django.conf import settings
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Q

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pytz

from ..models import Product, Category
from orders.models import WholesaleInvoice
from accounts.models import UserProfile

# Импорт из других модулей
from .monobank import _monobank_api_request, MonobankAPIError

# Loggers
wholesale_logger = logging.getLogger('storefront.wholesale')
monobank_logger = logging.getLogger('storefront.monobank')


# ==================== HELPER FUNCTIONS ====================

def _translate_color_to_ukrainian(color_name):
    """
    Translate color names from Russian to Ukrainian.
    
    Args:
        color_name: Color name in Russian
        
    Returns:
        str: Color name in Ukrainian
    """
    if not color_name:
        return 'чорний'
    
    color_translations = {
        'черный': 'чорний',
        'белый': 'білий',
        'красный': 'червоний',
        'синий': 'синій',
        'зеленый': 'зелений',
        'желтый': 'жовтий',
        'оранжевый': 'помаранчевий',
        'фиолетовый': 'фіолетовий',
        'розовый': 'рожевий',
        'серый': 'сірий',
        'коричневый': 'коричневий',
        'голубой': 'блакитний',
        'бордовый': 'бордовий',
        'темно-синий': 'темно-синій',
        'светло-серый': 'світло-сірий',
        'темно-серый': 'темно-сірий',
        'бежевый': 'бежевий',
        'кремовый': 'кремовий',
    }
    
    color_lower = color_name.lower()
    return color_translations.get(color_lower, color_name)


# ==================== PRICELIST & XLSX FUNCTIONS ====================

def wholesale_prices_xlsx(request):
    """Generate and serve wholesale prices XLSX file."""
    from productcolors.models import ProductColorVariant, Color
    
    # Фильтруем категории по ключевым словам
    tshirt_categories = Category.objects.select_related().filter(
        Q(name__icontains='футболка') | 
        Q(name__icontains='tshirt') | 
        Q(name__icontains='t-shirt') |
        Q(slug__icontains='футболка') |
        Q(slug__icontains='tshirt') |
        Q(slug__icontains='t-shirt')
    )
    
    hoodie_categories = Category.objects.select_related().filter(
        Q(name__icontains='худи') | 
        Q(name__icontains='hoodie') | 
        Q(name__icontains='hooded') |
        Q(slug__icontains='худи') |
        Q(slug__icontains='hoodie') |
        Q(slug__icontains='hooded')
    )
    
    # Получаем товары из нужных категорий
    tshirt_products = Product.objects.select_related('category').filter(category__in=tshirt_categories)
    hoodie_products = Product.objects.select_related('category').filter(category__in=hoodie_categories)
    
    # Создаем XLSX файл в памяти
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прайс (опт)"
    
    # Стили
    header_font = Font(bold=True, size=14)
    table_header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Заголовок (строка 1)
    ws.merge_cells('A1:L1')
    ws['A1'] = "Оптові ціни від кількості. Мінімальне замовлення по моделі — від 8 шт. і кратно 8."
    ws['A1'].font = header_font
    ws['A1'].alignment = center_alignment
    
    # Заголовки таблицы (строка 2)
    headers = [
        'Категорія',
        'Товар (S–XL)',
        'Колір',
        'Дроп (фікс. ціна)',
        '8–15 шт. (за 1 шт.)',
        '16–31 шт.',
        '32–63 шт.',
        '64–99 шт.',
        '100+ шт.',
        'Посилання'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = table_header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Цены для категорий
    tshirt_prices = [540, 520, 500, 490, 480]
    hoodie_prices = [1300, 1250, 1200, 1175, 1150]
    
    row = 3
    
    # Сначала худи
    for product in hoodie_products.order_by('title'):
        
        # Получаем цвета товара
        from productcolors.models import ProductColorVariant
        color_variants = ProductColorVariant.objects.filter(product=product)
        colors = []
        for variant in color_variants:
            if variant.color:
                colors.append(variant.color.name or str(variant.color))
        
        # Формируем название с "(фліс)" в конце
        product_title = f"{product.title} (S–XL) [фліс]"
        
        # Создаем ссылку на товар
        product_url = request.build_absolute_uri(f'/product/{product.slug}/')
        
        if colors:
            # Если есть цвета, создаем строку для каждого цвета
            for color in colors:
                ws.cell(row=row, column=1, value='Худі')
                ws.cell(row=row, column=2, value=product_title)
                ws.cell(row=row, column=3, value=_translate_color_to_ukrainian('чорний'))  # Для худи завжди чорний
                
                # Добавляем дроп цену (фиксированная 1350)
                ws.cell(row=row, column=4, value=1350)
                ws.cell(row=row, column=4).alignment = center_alignment
                
                # Добавляем оптовые цены
                for col, price in enumerate(hoodie_prices, 5):
                    ws.cell(row=row, column=col, value=price)
                    ws.cell(row=row, column=col).alignment = center_alignment
                
                # Добавляем ссылку на товар
                ws.cell(row=row, column=10, value=product_url)
                
                row += 1
        else:
            # Если нет цветов, создаем одну строку
            ws.cell(row=row, column=1, value='Худі')
            ws.cell(row=row, column=2, value=product_title)
            ws.cell(row=row, column=3, value=_translate_color_to_ukrainian('чорний'))  # Для худи завжди чорний
            
            # Добавляем дроп цену (фиксированная 1350)
            ws.cell(row=row, column=4, value=1350)
            ws.cell(row=row, column=4).alignment = center_alignment
            
            # Добавляем оптовые цены
            for col, price in enumerate(hoodie_prices, 5):
                ws.cell(row=row, column=col, value=price)
                ws.cell(row=row, column=col).alignment = center_alignment
            
            # Добавляем ссылку на товар
            ws.cell(row=row, column=10, value=product_url)
            
            row += 1
    
    # Затем футболки
    for product in tshirt_products.order_by('title'):
        
        # Получаем цвета товара
        from productcolors.models import ProductColorVariant
        color_variants = ProductColorVariant.objects.filter(product=product)
        colors = []
        for variant in color_variants:
            if variant.color:
                colors.append(variant.color.name or str(variant.color))
        
        # Формируем название без [фліс]
        product_title = f"{product.title} (S–XL)"
        
        # Создаем ссылку на товар
        product_url = request.build_absolute_uri(f'/product/{product.slug}/')
        
        if colors:
            # Если есть цвета, создаем строку для каждого цвета
            for color in colors:
                ws.cell(row=row, column=1, value='Футболки')
                ws.cell(row=row, column=2, value=product_title)
                ws.cell(row=row, column=3, value=_translate_color_to_ukrainian(color if color else 'чорний'))
                
                # Добавляем дроп цену (фиксированная 570 для футболок)
                ws.cell(row=row, column=4, value=570)
                ws.cell(row=row, column=4).alignment = center_alignment
                
                # Добавляем оптовые цены
                for col, price in enumerate(tshirt_prices, 5):
                    ws.cell(row=row, column=col, value=price)
                    ws.cell(row=row, column=col).alignment = center_alignment
                
                # Добавляем ссылку на товар
                ws.cell(row=row, column=10, value=product_url)
                
                row += 1
        else:
            # Если нет цветов, создаем одну строку с черным цветом
            ws.cell(row=row, column=1, value='Футболки')
            ws.cell(row=row, column=2, value=product_title)
            ws.cell(row=row, column=3, value=_translate_color_to_ukrainian('чорний'))
            
            # Добавляем дроп цену (фиксированная 570 для футболок)
            ws.cell(row=row, column=4, value=570)
            ws.cell(row=row, column=4).alignment = center_alignment
            
            # Добавляем оптовые цены
            for col, price in enumerate(tshirt_prices, 5):
                ws.cell(row=row, column=col, value=price)
                ws.cell(row=row, column=col).alignment = center_alignment
            
            # Добавляем ссылку на товар
            ws.cell(row=row, column=10, value=product_url)
            
            row += 1
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Создаем HTTP ответ
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pricelist_opt.xlsx"'
    
    return response


def pricelist_redirect(request):
    """Generate and serve wholesale prices XLSX file (redirect)."""
    return wholesale_prices_xlsx(request)


def pricelist_page(request):
    """Display pricelist download page."""
    return render(request, 'pages/pricelist.html')


def test_pricelist(request):
    """Simple test function to check if new URLs work."""
    return HttpResponse("Pricelist test works!", content_type="text/plain")


# ==================== WHOLESALE PAGES ====================

def wholesale_page(request):
    """Render wholesale page with products and pricing."""
    try:
        # Фильтруем категории по ключевым словам
        tshirt_categories = Category.objects.select_related().filter(
            Q(name__icontains='футболка') | 
            Q(name__icontains='tshirt') | 
            Q(name__icontains='t-shirt') |
            Q(slug__icontains='футболка') |
            Q(slug__icontains='tshirt') |
            Q(slug__icontains='t-shirt')
        )
        
        hoodie_categories = Category.objects.select_related().filter(
            Q(name__icontains='худи') | 
            Q(name__icontains='hoodie') | 
            Q(name__icontains='hooded') |
            Q(slug__icontains='худи') |
            Q(slug__icontains='hoodie') |
            Q(slug__icontains='hooded')
        )
        
        # Получаем товары из нужных категорий
        tshirt_products = Product.objects.select_related('category').filter(category__in=tshirt_categories).select_related('category')
        hoodie_products = Product.objects.select_related('category').filter(category__in=hoodie_categories).select_related('category')
        
    except Exception as e:
        # Если есть ошибка с базой данных, используем пустые списки
        tshirt_products = []
        hoodie_products = []
    
    # Цены для категорий
    tshirt_prices = {
        'drop': 570,
        'wholesale': [540, 520, 500, 490, 480],
        'ranges': ['8–15 шт.', '16–31 шт.', '32–63 шт.', '64–99 шт.', '100+ шт.']
    }
    
    hoodie_prices = {
        'drop': 1350,
        'wholesale': [1300, 1250, 1200, 1175, 1150],
        'ranges': ['8–15 шт.', '16–31 шт.', '32–63 шт.', '64–99 шт.', '100+ шт.']
    }
    
    context = {
        'tshirt_products': tshirt_products,
        'hoodie_products': hoodie_products,
        'tshirt_prices': tshirt_prices,
        'hoodie_prices': hoodie_prices,
    }
    
    return render(request, 'pages/wholesale.html', context)


def wholesale_order_form(request):
    """Render wholesale order form page."""
    try:
        # Фильтруем категории по ключевым словам
        tshirt_categories = Category.objects.select_related().filter(
            Q(name__icontains='футболка') | 
            Q(name__icontains='tshirt') | 
            Q(name__icontains='t-shirt') |
            Q(slug__icontains='футболка') |
            Q(slug__icontains='tshirt') |
            Q(slug__icontains='t-shirt')
        )
        
        hoodie_categories = Category.objects.select_related().filter(
            Q(name__icontains='худи') | 
            Q(name__icontains='hoodie') | 
            Q(name__icontains='hooded') |
            Q(slug__icontains='худи') |
            Q(slug__icontains='hoodie') |
            Q(slug__icontains='hooded')
        )
        
        # Получаем товары из нужных категорий
        tshirt_products = Product.objects.select_related('category').filter(category__in=tshirt_categories).select_related('category')
        hoodie_products = Product.objects.select_related('category').filter(category__in=hoodie_categories).select_related('category')
        
    except Exception as e:
        # Если есть ошибка с базой данных, используем пустые списки
        tshirt_products = []
        hoodie_products = []
    
    # Цены для категорий
    tshirt_prices = {
        'drop': 570,
        'wholesale': [540, 520, 500, 490, 480],
        'ranges': ['8–15 шт.', '16–31 шт.', '32–63 шт.', '64–99 шт.', '100+ шт.']
    }
    
    hoodie_prices = {
        'drop': 1350,
        'wholesale': [1300, 1250, 1200, 1175, 1150],
        'ranges': ['8–15 шт.', '16–31 шт.', '32–63 шт.', '64–99 шт.', '100+ шт.']
    }
    
    # Статистика для оптовых закупок
    wholesale_stats = {
        'total_products': len(list(tshirt_products)) + len(list(hoodie_products)),
        'tshirt_count': len(list(tshirt_products)),
        'hoodie_count': len(list(hoodie_products)),
        'min_order_value': min(tshirt_prices['wholesale'][-1], hoodie_prices['wholesale'][-1]) * 8,  # минимальный заказ 8 штук
        'max_savings_percent': round((1 - (min(tshirt_prices['wholesale'][-1], hoodie_prices['wholesale'][-1]) / max(tshirt_prices['drop'], hoodie_prices['drop']))) * 100, 1)
    }
    
    # Получаем сохраненные данные компании для авторизованных пользователей
    company_data = {}
    user_invoice_stats = {}
    if request.user.is_authenticated:
        try:
            user_profile = UserProfile.objects.get(user=request.user)
            company_data = {
                'company_name': getattr(user_profile, 'company_name', '') or '',
                'company_number': getattr(user_profile, 'company_number', '') or '',
                'delivery_address': getattr(user_profile, 'delivery_address', '') or '',
                'phone_number': user_profile.phone or '',
                'store_link': getattr(user_profile, 'website', '') or ''
            }
            
            # Статистика накладных пользователя (показываем общую статистику)
            all_invoices = WholesaleInvoice.objects.all().order_by('-created_at')
            last_invoice = all_invoices.first()
            user_invoice_stats = {
                'total_invoices': all_invoices.count(),
                'last_invoice_date': last_invoice.created_at.strftime('%d.%m.%Y %H:%M') if last_invoice else 'Немає накладних',
                'total_products_available': tshirt_products.count() + hoodie_products.count(),
                'tshirt_count': tshirt_products.count(),
                'hoodie_count': hoodie_products.count()
            }
        except UserProfile.DoesNotExist:
            pass
    
    context = {
        'tshirt_products': tshirt_products,
        'hoodie_products': hoodie_products,
        'tshirt_prices': tshirt_prices,
        'hoodie_prices': hoodie_prices,
        'wholesale_stats': wholesale_stats,
        'company_data': company_data,
        'user_invoice_stats': user_invoice_stats,
    }
    
    return render(request, 'pages/wholesale_order_form.html', context)


# ==================== INVOICE GENERATION & MANAGEMENT ====================

@require_POST
@csrf_exempt
def generate_wholesale_invoice(request):
    """Генерирует Excel накладную для оптового заказа."""
    try:
        # Получаем данные из запроса
        data = json.loads(request.body)
        company_data = data.get('companyData', {})
        order_items = data.get('orderItems', [])
        
        if not order_items:
            return JsonResponse({'error': 'Немає товарів для накладної'}, status=400)
        
        # Генерируем номер накладной с красивой датой (киевское время)
        kiev_tz = pytz.timezone('Europe/Kiev')
        now = timezone.now().astimezone(kiev_tz)
        timestamp = now.strftime('%Y%m%d_%H%M%S')
        invoice_number = f"ОПТ_{timestamp}"
        
        # Красивое название файла с названием компании
        company_name = company_data.get('companyName', 'Company').strip()
        if not company_name:
            company_name = 'Company'
        
        # Красивая дата для названия файла
        beautiful_date = now.strftime('%d.%m.%Y_%H-%M')
        file_name = f"{company_name}_накладнаОПТ_{beautiful_date}.xlsx"
        
        # Подсчитываем общие данные
        total_tshirts = 0
        total_hoodies = 0
        total_amount = 0
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Оптова накладна"
        
        # Стили
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        title_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=10)
        company_font = Font(name='Arial', size=11, bold=True, color='366092')
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        company_fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Заголовок накладной
        ws.merge_cells('A1:G1')
        ws['A1'] = 'ОПТОВА НАКЛАДНА'
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='366092')
        ws['A1'].alignment = center_alignment
        ws['A1'].fill = company_fill
        
        # Информация о компании
        row = 3
        ws[f'A{row}'] = 'Інформація про замовника:'
        ws[f'A{row}'].font = title_font
        
        row += 1
        ws[f'A{row}'] = 'Назва компанії/ФОП/ПІБ:'
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'] = company_data.get('companyName', '')
        ws[f'B{row}'].font = company_font
        ws[f'B{row}'].fill = company_fill
        
        if company_data.get('companyNumber'):
            row += 1
            ws[f'A{row}'] = 'Номер компанії/ЄДРПОУ/ІПН:'
            ws[f'A{row}'].font = normal_font
            ws[f'B{row}'] = company_data.get('companyNumber', '')
            ws[f'B{row}'].font = normal_font
        
        row += 1
        ws[f'A{row}'] = 'Номер телефону:'
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'] = company_data.get('contactPhone', '')
        ws[f'B{row}'].font = normal_font
        
        row += 1
        ws[f'A{row}'] = 'Адреса доставки:'
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'] = company_data.get('deliveryAddress', '')
        ws[f'B{row}'].font = normal_font
        
        if company_data.get('storeLink'):
            row += 1
            ws[f'A{row}'] = 'Посилання на магазин:'
            ws[f'A{row}'].font = normal_font
            ws[f'B{row}'] = company_data.get('storeLink', '')
            ws[f'B{row}'].font = normal_font
        
        row += 2
        ws[f'A{row}'] = f'Номер накладної: {invoice_number}'
        ws[f'A{row}'].font = title_font
        
        row += 1
        ws[f'A{row}'] = f'Дата створення: {now.strftime("%d.%m.%Y о %H:%M")}'
        ws[f'A{row}'].font = normal_font
        
        # Заголовок таблицы товаров
        row += 2
        headers = ['№', 'Назва товару', 'Розмір', 'Колір', 'Кількість', 'Ціна за од.', 'Сума']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Заполняем товары
        item_number = 1
        row += 1
        
        for item in order_items:
            product = item.get('product', {})
            quantity = item.get('quantity', 0)
            price = item.get('price', 0)
            total = item.get('total', 0)
            
            # Подсчитываем общие данные
            if product.get('type') == 'tshirt':
                total_tshirts += quantity
            elif product.get('type') == 'hoodie':
                total_hoodies += quantity
            
            total_amount += total
            
            # Заполняем строку товара
            ws.cell(row=row, column=1, value=item_number).font = normal_font
            ws.cell(row=row, column=1).alignment = center_alignment
            ws.cell(row=row, column=1).border = thin_border
            
            ws.cell(row=row, column=2, value=product.get('title', '')).font = normal_font
            ws.cell(row=row, column=2).alignment = left_alignment
            ws.cell(row=row, column=2).border = thin_border
            
            ws.cell(row=row, column=3, value=item.get('size', '')).font = normal_font
            ws.cell(row=row, column=3).alignment = center_alignment
            ws.cell(row=row, column=3).border = thin_border
            
            ws.cell(row=row, column=4, value=item.get('color', '')).font = normal_font
            ws.cell(row=row, column=4).alignment = center_alignment
            ws.cell(row=row, column=4).border = thin_border
            
            ws.cell(row=row, column=5, value=quantity).font = normal_font
            ws.cell(row=row, column=5).alignment = center_alignment
            ws.cell(row=row, column=5).border = thin_border
            
            ws.cell(row=row, column=6, value=f"{price}₴").font = normal_font
            ws.cell(row=row, column=6).alignment = center_alignment
            ws.cell(row=row, column=6).border = thin_border
            
            ws.cell(row=row, column=7, value=f"{total}₴").font = normal_font
            ws.cell(row=row, column=7).alignment = center_alignment
            ws.cell(row=row, column=7).border = thin_border
            
            # Чередуем цвет фона строк
            if item_number % 2 == 0:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = light_fill
            
            item_number += 1
            row += 1
        
        # Итоговая строка
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'РАЗОМ:'
        ws[f'A{row}'].font = title_font
        right_alignment = Alignment(horizontal='right', vertical='center')
        ws[f'A{row}'].alignment = right_alignment
        ws[f'A{row}'].border = thin_border
        ws[f'A{row}'].fill = company_fill
        
        ws[f'G{row}'] = f"{total_amount}₴"
        ws[f'G{row}'].font = title_font
        ws[f'G{row}'].alignment = center_alignment
        ws[f'G{row}'].border = thin_border
        ws[f'G{row}'].fill = company_fill
        
        # Статистика
        row += 2
        ws[f'A{row}'] = 'Статистика замовлення:'
        ws[f'A{row}'].font = title_font
        
        row += 1
        ws[f'A{row}'] = f'Футболки: {total_tshirts} шт.'
        ws[f'A{row}'].font = normal_font
        
        row += 1
        ws[f'A{row}'] = f'Худі: {total_hoodies} шт.'
        ws[f'A{row}'].font = normal_font
        
        row += 1
        ws[f'A{row}'] = f'Загальна сума: {total_amount}₴'
        ws[f'A{row}'].font = title_font
        
        # Автоматически настраиваем ширину колонок по содержимому
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Устанавливаем минимальную ширину и добавляем небольшой отступ
            adjusted_width = max(max_length + 2, 12)
            # Ограничиваем максимальную ширину для читаемости
            adjusted_width = min(adjusted_width, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем данные компании в профиль пользователя (если авторизован)
        if request.user.is_authenticated:
            try:
                user_profile, created = UserProfile.objects.get_or_create(user=request.user)
                if hasattr(user_profile, 'company_name'):
                    user_profile.company_name = company_data.get('companyName', '')
                if hasattr(user_profile, 'company_number'):
                    user_profile.company_number = company_data.get('companyNumber', '')
                if hasattr(user_profile, 'delivery_address'):
                    user_profile.delivery_address = company_data.get('deliveryAddress', '')
                user_profile.phone = company_data.get('contactPhone', '')
                if hasattr(user_profile, 'website'):
                    user_profile.website = company_data.get('storeLink', '')
                user_profile.save()
            except Exception as e:
                wholesale_logger.error(f"Error saving user profile: {e}")
        
        # Сохраняем в базу данных
        try:
            invoice = WholesaleInvoice.objects.create(
                invoice_number=invoice_number,
                company_name=company_data.get('companyName', ''),
                company_number=company_data.get('companyNumber', ''),
                contact_phone=company_data.get('contactPhone', ''),
                delivery_address=company_data.get('deliveryAddress', ''),
                store_link=company_data.get('storeLink', ''),
                total_tshirts=total_tshirts,
                total_hoodies=total_hoodies,
                total_amount=total_amount,
                status='draft',  # Явно указываем статус
                order_details={
                    'company_data': company_data,
                    'order_items': order_items
                }
            )
            invoice_id = invoice.id
        except Exception as e:
            # Временно используем timestamp как ID если база недоступна
            import time
            invoice_id = int(time.time())
            wholesale_logger.error(f"Database error (temporary workaround): {e}")
        
        # Сохраняем файл на сервере
        # Создаем папку для накладных пользователя
        user_folder = f"invoices/user_{request.user.id if request.user.is_authenticated else 'anonymous'}"
        invoice_dir = os.path.join(settings.MEDIA_ROOT, user_folder)
        os.makedirs(invoice_dir, exist_ok=True)
        
        wholesale_logger.info(f"Creating invoice directory: {invoice_dir}")
        
        # Путь к файлу (используем уже созданное красивое имя)
        file_path = os.path.join(invoice_dir, file_name)
        
        # Сохраняем Excel файл на сервере
        wb.save(file_path)
        wholesale_logger.info(f"Saved invoice file: {file_path}")
        
        # Обновляем invoice с путем к файлу (если invoice создался)
        try:
            if 'invoice' in locals() and invoice:
                invoice.file_path = file_path
                invoice.save()
        except:
            pass  # Игнорируем ошибки обновления
        
        # Подготавливаем ответ для скачивания
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        
        # Читаем файл и отправляем
        with open(file_path, 'rb') as f:
            response.write(f.read())
        
        return JsonResponse({
            'success': True,
            'invoice_number': invoice_number,
            'invoice_id': invoice_id,
            'file_name': file_name,
            'company_name': company_name,
            'total_amount': total_amount,
            'total_tshirts': total_tshirts,
            'total_hoodies': total_hoodies,
            'timestamp': int(now.timestamp() * 1000)
        })
        
    except Exception as e:
        wholesale_logger.exception(f'Error generating wholesale invoice: {e}')
        return JsonResponse({'error': f'Помилка при створенні накладної: {str(e)}'}, status=500)


def download_invoice_file(request, invoice_id):
    """Скачивание сохраненного файла накладной."""
    try:
        # Получаем накладную
        invoice = WholesaleInvoice.objects.get(id=invoice_id)
        
        # Проверяем права доступа (только владелец или анонимные)
        if request.user.is_authenticated:
            # Для зарегистрированных пользователей можно добавить проверку
            pass
        
        # Проверяем существование файла
        if not invoice.file_path or not os.path.exists(invoice.file_path):
            # Если файл не найден, возвращаем ошибку
            return HttpResponse('Файл накладної не знайдено', status=404)
        
        # Отправляем файл
        filename = os.path.basename(invoice.file_path) if invoice.file_path else f"invoice_{invoice.id}.xlsx"
        
        try:
            response = FileResponse(
                open(invoice.file_path, 'rb'),
                as_attachment=True,
                filename=filename
            )
            return response
        except Exception as file_error:
            return HttpResponse(f'Помилка при відкритті файлу: {str(file_error)}', status=500)
        
    except WholesaleInvoice.DoesNotExist:
        return HttpResponse('Накладна не знайдена', status=404)
    except Exception as e:
        return HttpResponse(f'Помилка при скачуванні: {str(e)}', status=500)


@require_POST
@csrf_exempt
def delete_wholesale_invoice(request, invoice_id):
    """Удаление накладной."""
    try:
        # Получаем накладную
        invoice = WholesaleInvoice.objects.get(id=invoice_id)
        
        # Удаляем файл(ы) накладной на диске
        try:
            deleted_any = False
            # 1) Если знаем точный путь — удалить его
            if invoice.file_path and os.path.isabs(invoice.file_path):
                if os.path.exists(invoice.file_path):
                    os.remove(invoice.file_path)
                    deleted_any = True
                    wholesale_logger.info(f"File deleted: {invoice.file_path}")
            # 2) На случай если путь не сохранён или файл перемещён:
            # Пытаемся найти все файлы в media/invoices/, содержащие номер накладной или часть имени компании
            base_dir = os.path.join(settings.MEDIA_ROOT, 'invoices')
            if os.path.isdir(base_dir):
                safe_company = (invoice.company_name or '').strip().replace(' ', '_')
                # Поиск файлов по шаблону
                import glob
                pattern_files = glob.glob(os.path.join(base_dir, '**', f'*{safe_company}*.xlsx'), recursive=True)
                for f in pattern_files:
                    try:
                        os.remove(f)
                        deleted_any = True
                        wholesale_logger.info(f"Pattern-matched file deleted: {f}")
                    except:
                        pass
            
            if deleted_any:
                wholesale_logger.info(f"Successfully deleted invoice files for ID {invoice_id}")
            else:
                wholesale_logger.warning(f"No files found to delete for invoice ID {invoice_id}")
        except Exception as file_error:
            wholesale_logger.error(f"Error deleting files for invoice {invoice_id}: {file_error}")
        
        # Удаляем запись из БД
        invoice.delete()
        
        return JsonResponse({'success': True})
        
    except WholesaleInvoice.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Накладна не знайдена'}, status=404)
    except Exception as e:
        wholesale_logger.exception(f'Error deleting invoice: {e}')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def get_user_invoices(request):
    """Получение списка накладных пользователя."""
    try:
        # Получаем накладные пользователя (по номеру телефона или другим данным)
        # Пока возвращаем все накладные, но можно добавить фильтрацию
        invoices = WholesaleInvoice.objects.all().order_by('-created_at')
        
        # Пагинация
        paginator = Paginator(invoices, 10)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        invoices_data = []
        for invoice in page_obj:
            invoices_data.append({
                'id': invoice.id,
                'invoice_number': invoice.invoice_number,
                'company_name': invoice.company_name,
                'total_amount': float(invoice.total_amount),
                'total_tshirts': invoice.total_tshirts,
                'total_hoodies': invoice.total_hoodies,
                'created_at': invoice.created_at.strftime('%d.%m.%Y %H:%M'),
                'status': invoice.get_status_display(),
                'file_name': invoice.file_name
            })
        
        return JsonResponse({
            'success': True,
            'invoices': invoices_data,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Помилка: {str(e)}'}, status=500)


@require_POST
@staff_member_required
def update_invoice_status(request, invoice_id):
    """Обновление статуса накладной (только для staff)."""
    try:
        invoice = WholesaleInvoice.objects.get(id=invoice_id)
        data = json.loads(request.body or '{}')
        new_status = data.get('status')
        allowed = {'processing','approved','shipped','delivered','cancelled','pending'}
        if new_status not in allowed:
            return JsonResponse({'success': False, 'error': 'Невірний статус'}, status=400)
        
        old_status = invoice.status
        invoice.status = new_status
        invoice.save(update_fields=['status'])
        
        # Отправляем уведомление в Telegram при отправке накладной на проверку
        if new_status == 'pending' and old_status == 'draft':
            try:
                from orders.telegram_notifications import telegram_notifier
                # Отправляем уведомление о новой накладной
                telegram_notifier.send_invoice_notification(invoice)
                
                # Если есть файл накладной, отправляем его тоже
                if invoice.file_path:
                    file_path = os.path.join(settings.MEDIA_ROOT, invoice.file_path)
                    if os.path.exists(file_path):
                        telegram_notifier.send_invoice_document(invoice, file_path)
            except Exception as e:
                # Логируем ошибку, но не прерываем выполнение
                wholesale_logger.error(f"Ошибка отправки Telegram уведомления: {e}")
        
        return JsonResponse({'success': True})
    except WholesaleInvoice.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Накладна не знайдена'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def reset_all_invoices_status(request):
    """Reset all invoices to 'draft' status."""
    try:
        # Reset all invoices to 'draft' status
        updated_count = WholesaleInvoice.objects.filter(
            status__in=['pending', 'processing', 'shipped', 'delivered']
        ).update(status='draft')
        
        return JsonResponse({
            'success': True, 
            'message': f'Статус {updated_count} накладних скинуто до "Чернетка"',
            'updated_count': updated_count
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_POST
@staff_member_required
def toggle_invoice_approval(request, invoice_id):
    """Переключение одобрения накладной для оплаты."""
    try:
        invoice = WholesaleInvoice.objects.get(id=invoice_id)
        data = json.loads(request.body or '{}')
        is_approved = data.get('is_approved', False)
        
        invoice.is_approved = is_approved
        invoice.save(update_fields=['is_approved'])
        
        status_text = "одобрена" if is_approved else "отклонена"
        return JsonResponse({
            'success': True, 
            'message': f'Накладна {status_text} для оплаты',
            'is_approved': is_approved
        })
    except WholesaleInvoice.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Накладна не знайдена'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def check_invoice_approval_status(request, invoice_id):
    """Проверка статуса одобрения накладной."""
    try:
        invoice = WholesaleInvoice.objects.get(id=invoice_id)
        return JsonResponse({
            'success': True,
            'is_approved': invoice.is_approved,
            'invoice_number': invoice.invoice_number
        })
    except WholesaleInvoice.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Накладна не знайдена'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_POST
@staff_member_required
def toggle_invoice_payment_status(request, invoice_id):
    """Переключение статуса оплаты накладной администратором."""
    try:
        invoice = WholesaleInvoice.objects.get(id=invoice_id)
        data = json.loads(request.body or '{}')
        new_status = data.get('payment_status', 'not_paid')
        
        # Проверяем валидность статуса
        valid_statuses = ['not_paid', 'paid', 'pending', 'failed']
        if new_status not in valid_statuses:
            return JsonResponse({'success': False, 'error': 'Невірний статус оплати'}, status=400)
        
        invoice.payment_status = new_status
        invoice.save(update_fields=['payment_status'])
        
        status_text = {
            'not_paid': 'не оплачена',
            'paid': 'оплачена',
            'pending': 'очікує оплати',
            'failed': 'помилка оплати'
        }.get(new_status, new_status)
        
        return JsonResponse({
            'success': True,
            'message': f'Статус оплати змінено на "{status_text}"',
            'payment_status': new_status
        })
        
    except WholesaleInvoice.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Накладна не знайдена'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def check_payment_status(request, invoice_id):
    """Проверка статуса оплаты накладной."""
    try:
        invoice = WholesaleInvoice.objects.get(id=invoice_id)
        return JsonResponse({
            'success': True,
            'payment_status': invoice.payment_status,
            'invoice_number': invoice.invoice_number
        })
    except WholesaleInvoice.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Накладна не знайдена'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ==================== PAYMENT INTEGRATION ====================

@csrf_exempt
def create_wholesale_payment(request):
    """Создание платежа для накладной через Monobank."""
    # Для POST запросов
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST method allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        invoice_id = data.get('invoice_id')
        amount = data.get('amount')
        description = data.get('description', '')
        invoice_number = data.get('invoice_number', '')
        
        # Проверяем invoice_id
        if not invoice_id:
            return JsonResponse({'success': False, 'error': 'Відсутній ID накладної'}, status=400)
        
        # Проверяем amount
        if amount is None or amount == '' or amount == 0:
            return JsonResponse({'success': False, 'error': 'Невірна сума накладної'}, status=400)
        
        # Получаем накладную
        try:
            invoice = WholesaleInvoice.objects.get(id=invoice_id)
        except WholesaleInvoice.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Накладна не знайдена'}, status=404)
        
        # Проверяем, что накладная одобрена
        if not invoice.is_approved:
            return JsonResponse({'success': False, 'error': 'Накладна не одобрена для оплаты. Обратитесь к администратору.'}, status=400)
        
        # Проверяем, что накладная еще не оплачена
        if invoice.payment_status == 'paid':
            return JsonResponse({'success': False, 'error': 'Накладна уже оплачена'}, status=400)
        
        # Проверяем, что накладная не в процессе оплаты
        if invoice.payment_status == 'pending':
            return JsonResponse({'success': False, 'error': 'Накладна уже в процессі оплати'}, status=400)
        
        # Конвертируем сумму в копейки для Monobank
        amount_decimal = Decimal(str(amount))
        amount_kopecks = int(amount_decimal * 100)
        
        # Создаем payload для Monobank
        payload = {
            'amount': amount_kopecks,
            'ccy': 980,  # UAH
            'merchantPaymInfo': {
                'reference': f'INV-{invoice_id}',
                'destination': description or f'Оплата накладної {invoice_number}',
                'basketOrder': [
                    {
                        'name': f'Накладна {invoice_number}',
                        'qty': 1,
                        'sum': amount_kopecks,
                        'icon': '',
                        'unit': 'шт'
                    }
                ]
            },
            'redirectUrl': request.build_absolute_uri('/payments/monobank/return/'),
            'webHookUrl': request.build_absolute_uri('/payments/monobank/webhook/'),
        }
        
        monobank_logger.info('Monobank payload: %s', payload)
        
        # Отправляем запрос в Monobank
        try:
            creation_data = _monobank_api_request('POST', '/api/merchant/invoice/create', json_payload=payload)
            monobank_logger.info('Monobank response: %s', creation_data)
        except MonobankAPIError as exc:
            monobank_logger.warning('Monobank invoice creation failed for wholesale invoice %s: %s', invoice_id, exc)
            return JsonResponse({'success': False, 'error': f'Помилка створення платежу: {str(exc)}'}, status=500)
        
        # Получаем URL для оплаты (Monobank может возвращать pageUrl или invoiceUrl)
        invoice_url = creation_data.get('invoiceUrl') or creation_data.get('pageUrl')
        if not invoice_url:
            monobank_logger.error('No payment URL in Monobank response: %s', creation_data)
            return JsonResponse({'success': False, 'error': 'Не отримано URL для оплати'}, status=500)
        
        # Устанавливаем статус "ожидает оплаты"
        invoice.payment_status = 'pending'
        invoice.save(update_fields=['payment_status'])
        
        # Сохраняем информацию о платеже в сессии
        request.session['wholesale_payment_invoice_id'] = invoice_id
        request.session['wholesale_payment_amount'] = str(amount)
        
        return JsonResponse({
            'success': True,
            'payment_url': invoice_url,
            'invoice_id': invoice_id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Невірний формат даних'}, status=400)
    except Exception as e:
        monobank_logger.exception('Failed to create wholesale payment: %s', e)
        return JsonResponse({'success': False, 'error': 'Помилка створення платежу'}, status=500)


@csrf_exempt
def wholesale_payment_webhook(request):
    """Webhook для обработки уведомлений о платежах накладных от Monobank."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        invoice_id = request.session.get('wholesale_payment_invoice_id')
        
        if not invoice_id:
            return JsonResponse({'error': 'No invoice ID in session'}, status=400)
        
        # Получаем накладную
        try:
            invoice = WholesaleInvoice.objects.get(id=invoice_id)
        except WholesaleInvoice.DoesNotExist:
            return JsonResponse({'error': 'Invoice not found'}, status=404)
        
        # Проверяем статус платежа
        status = data.get('status', '').lower()
        
        if status in ['success', 'hold']:
            # Платеж успешен
            invoice.payment_status = 'paid'
            invoice.save(update_fields=['payment_status'])
            
            # Отправляем уведомление в Telegram
            try:
                from orders.telegram_notifications import telegram_notifier
                if telegram_notifier.is_configured():
                    telegram_notifier.send_invoice_payment_notification(invoice)
            except Exception as e:
                pass  # Игнорируем ошибки Telegram
            
            return JsonResponse({'success': True, 'status': 'paid'})
        
        elif status in ['failure', 'expired', 'rejected', 'canceled', 'cancelled', 'reversed']:
            # Платеж неуспешен
            invoice.payment_status = 'failed'
            invoice.save(update_fields=['payment_status'])
            return JsonResponse({'success': True, 'status': 'failed'})
        
        else:
            # Неизвестный статус
            return JsonResponse({'success': True, 'status': 'unknown'})
            
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ==================== DEBUG FUNCTIONS ====================

def debug_invoices(request):
    """Debug endpoint для проверки накладных."""
    try:
        invoices = WholesaleInvoice.objects.all()[:10]
        invoice_data = []
        for inv in invoices:
            invoice_data.append({
                'id': inv.id,
                'invoice_number': inv.invoice_number,
                'is_approved': inv.is_approved,
                'payment_status': inv.payment_status,
                'total_amount': inv.total_amount
            })
        
        return JsonResponse({
            'success': True,
            'count': invoices.count(),
            'invoices': invoice_data
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

