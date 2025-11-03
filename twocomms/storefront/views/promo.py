"""
Promo codes views - Управление промокодами.

Содержит views для:
- Админ-панели промокодов
- Управления группами промокодов
- Создания/редактирования/удаления промокодов
- Статистики использования
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from decimal import Decimal

from django.db.models import Count, Q, Avg, Sum
from django.http import JsonResponse
from django import forms

from ..models import PromoCode, PromoCodeGroup, PromoCodeUsage


# ==================== FORMS ====================

class PromoCodeGroupForm(forms.ModelForm):
    """Форма для создания/редактирования группы промокодов"""
    class Meta:
        model = PromoCodeGroup
        fields = ['name', 'description', 'one_per_account', 'is_active']
        widgets = {
            'name': forms.TextInput(attrs={
                'class': 'form-control bg-elevate',
                'placeholder': 'Назва групи (наприклад, "Instagram реклама")'
            }),
            'description': forms.Textarea(attrs={
                'class': 'form-control bg-elevate',
                'rows': 3,
                'placeholder': 'Опис групи промокодів'
            }),
            'one_per_account': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'is_active': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }


class PromoCodeForm(forms.ModelForm):
    """Форма для создания/редактирования промокода"""
    class Meta:
        model = PromoCode
        fields = [
            'code', 'promo_type', 'discount_type', 'discount_value', 
            'description', 'group', 'max_uses', 'one_time_per_user',
            'min_order_amount', 'valid_from', 'valid_until', 'is_active'
        ]
        widgets = {
            'code': forms.TextInput(attrs={
                'class': 'form-control bg-elevate',
                'placeholder': 'Введіть код або залиште порожнім для автогенерації'
            }),
            'promo_type': forms.Select(attrs={'class': 'form-select bg-elevate'}),
            'discount_type': forms.Select(attrs={'class': 'form-select bg-elevate'}),
            'discount_value': forms.NumberInput(attrs={
                'class': 'form-control bg-elevate',
                'step': '0.01',
                'min': '0'
            }),
            'description': forms.Textarea(attrs={
                'class': 'form-control bg-elevate',
                'rows': 3,
                'placeholder': 'Опис промокоду (необов\'язково)'
            }),
            'group': forms.Select(attrs={'class': 'form-select bg-elevate'}),
            'max_uses': forms.NumberInput(attrs={
                'class': 'form-control bg-elevate',
                'min': '0',
                'placeholder': '0 = безліміт'
            }),
            'one_time_per_user': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'min_order_amount': forms.NumberInput(attrs={
                'class': 'form-control bg-elevate',
                'step': '0.01',
                'min': '0',
                'placeholder': 'Мінімальна сума (необов\'язково)'
            }),
            'valid_from': forms.DateTimeInput(attrs={
                'class': 'form-control bg-elevate',
                'type': 'datetime-local'
            }),
            'valid_until': forms.DateTimeInput(attrs={
                'class': 'form-control bg-elevate',
                'type': 'datetime-local'
            }),
            'is_active': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }
    
    def clean_code(self):
        """Позволяет оставить поле code пустым для автоматической генерации"""
        code = self.cleaned_data.get('code', '').strip()
        
        # Если код пустой, это нормально - он будет сгенерирован автоматически
        if not code:
            return ''
        
        # Если код указан, проверяем его уникальность
        if PromoCode.objects.filter(code=code).exists():
            if self.instance and self.instance.pk:
                # При редактировании исключаем текущий промокод
                if not PromoCode.objects.filter(code=code).exclude(pk=self.instance.pk).exists():
                    return code
            raise forms.ValidationError("Промокод з таким кодом вже існує")
        
        return code


# ==================== HELPER FUNCTIONS ====================

def get_promo_admin_context(request):
    """Функція для отримання контексту промокодів для адмін-панелі"""
    try:
        # Получаем текущий таб и параметры фильтрации
        promo_tab = request.GET.get('tab', 'promocodes')
        view_type = request.GET.get('view', 'all')
        group_id = request.GET.get('group')

        # ===== ТАБ 1: ПРОМОКОДЫ =====
        promocodes = PromoCode.objects.select_related('group').prefetch_related('usages')

        if view_type == 'vouchers':
            promocodes = promocodes.filter(promo_type='voucher')
        elif view_type == 'grouped':
            promocodes = promocodes.filter(promo_type='grouped', group__isnull=False)
        elif view_type == 'regular':
            promocodes = promocodes.filter(promo_type='regular')

        if group_id:
            promocodes = promocodes.filter(group_id=group_id)

        # ===== ТАБ 2: ГРУППЫ =====
        groups = PromoCodeGroup.objects.prefetch_related('promo_codes').annotate(
            codes_count=Count('promo_codes'),
            active_codes_count=Count('promo_codes', filter=Q(promo_codes__is_active=True)),
            total_usages=Count('usages'),
        )

        # ===== ТАБ 3: СТАТИСТИКА =====
        recent_usages = PromoCodeUsage.objects.select_related(
            'user', 'promo_code', 'group', 'order'
        ).order_by('-used_at')[:50]

        top_promos = PromoCode.objects.annotate(
            usage_count=Count('usages')
        ).filter(usage_count__gt=0).order_by('-usage_count')[:10]

        top_groups = PromoCodeGroup.objects.annotate(
            usage_count=Count('usages')
        ).filter(usage_count__gt=0).order_by('-usage_count')[:10]

        # ===== ОБЩАЯ СТАТИСТИКА =====
        total_promocodes = PromoCode.objects.count()
        active_promocodes = PromoCode.objects.filter(is_active=True).count()
        total_vouchers = PromoCode.objects.filter(promo_type='voucher').count()
        total_groups = groups.count()
        total_usages = PromoCodeUsage.objects.count()
        unique_users = PromoCodeUsage.objects.values('user').distinct().count()

        # ===== РОЗШИРЕНА СТАТИСТИКА =====
        avg_discount_percent = PromoCode.objects.filter(
            discount_type='percentage'
        ).aggregate(avg=Avg('discount_value'))['avg'] or Decimal('0')

        avg_discount_fixed = PromoCode.objects.filter(
            discount_type='fixed'
        ).aggregate(avg=Avg('discount_value'))['avg'] or Decimal('0')

        used_promocodes_count = PromoCode.objects.annotate(
            usages_count=Count('usages')
        ).filter(usages_count__gt=0).count()

        conversion_rate = (
            used_promocodes_count / total_promocodes * 100 if total_promocodes > 0 else 0
        )

        avg_usages_per_promo = total_usages / total_promocodes if total_promocodes > 0 else 0

        total_savings = Decimal('0')

        fixed_savings = PromoCodeUsage.objects.filter(
            promo_code__discount_type='fixed'
        ).aggregate(total=Sum('promo_code__discount_value'))['total'] or Decimal('0')
        total_savings += fixed_savings

        percent_usages_count = PromoCodeUsage.objects.filter(
            promo_code__discount_type='percentage'
        ).count()

        most_popular_promo = PromoCode.objects.annotate(
            usage_count=Count('usages')
        ).filter(usage_count__gt=0).order_by('-usage_count').first()

        most_successful_group = PromoCodeGroup.objects.annotate(
            usage_count=Count('usages')
        ).filter(usage_count__gt=0).order_by('-usage_count').first()

        return {
            'promo_tab': promo_tab,
            'view_type': view_type,
            'current_group_id': int(group_id) if group_id else None,
            'promocodes': promocodes,
            'groups': groups,
            'recent_usages': recent_usages,
            'top_promos': list(top_promos),
            'top_groups': list(top_groups),
            'total_promocodes': total_promocodes,
            'active_promocodes': active_promocodes,
            'total_vouchers': total_vouchers,
            'total_groups': total_groups,
            'total_usages': total_usages,
            'unique_users': unique_users,
            'avg_discount_percent': round(float(avg_discount_percent), 2),
            'avg_discount_fixed': round(float(avg_discount_fixed), 2),
            'conversion_rate': round(conversion_rate, 2),
            'avg_usages_per_promo': round(avg_usages_per_promo, 2),
            'total_savings': float(total_savings),
            'percent_usages_count': percent_usages_count,
            'used_promocodes_count': used_promocodes_count,
            'most_popular_promo': most_popular_promo,
            'most_successful_group': most_successful_group,
        }
    except Exception as e:
        import logging

        logger = logging.getLogger(__name__)
        logger.error('Ошибка в get_promo_admin_context: %s', e, exc_info=True)
        return {
            'promo_tab': request.GET.get('tab', 'promocodes'),
            'view_type': request.GET.get('view', 'all'),
            'current_group_id': None,
            'promocodes': [],
            'groups': [],
            'recent_usages': [],
            'top_promos': [],
            'top_groups': [],
            'total_promocodes': 0,
            'active_promocodes': 0,
            'total_vouchers': 0,
            'total_groups': 0,
            'total_usages': 0,
            'unique_users': 0,
            'avg_discount_percent': 0,
            'avg_discount_fixed': 0,
            'conversion_rate': 0,
            'avg_usages_per_promo': 0,
            'total_savings': 0,
            'percent_usages_count': 0,
            'used_promocodes_count': 0,
            'most_popular_promo': None,
            'most_successful_group': None,
        }


# ==================== ADMIN VIEWS ====================

@login_required
def admin_promocodes(request):
    """
    DEPRECATED: Ця view більше не використовується.
    Промокоди тепер в головній адмін-панелі через ?section=promocodes
    Редирект для backward compatibility
    """
    return redirect('/admin-panel/?section=promocodes')


@login_required
def admin_promocode_create(request):
    """Создание нового промокода (поддержка AJAX)"""
    if not request.user.is_staff:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Доступ заборонено'}, status=403)
        return redirect('home')
    
    if request.method == 'POST':
        form = PromoCodeForm(request.POST)
        if form.is_valid():
            promocode = form.save(commit=False)
            
            # Если код не указан или пустой, генерируем автоматически
            if not promocode.code or promocode.code.strip() == '':
                promocode.code = PromoCode.generate_code()
            
            promocode.save()
            
            # AJAX ответ
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Промокод успішно створено',
                    'promo_id': promocode.id,
                    'promo_code': promocode.code
                })
            
            return redirect('/admin-panel/?section=promocodes&tab=promocodes')
        else:
            # AJAX ответ с ошибками
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': 'Помилка валідації',
                    'errors': form.errors
                }, status=400)
    else:
        form = PromoCodeForm()
    
    return render(request, 'pages/admin_promocode_form.html', {
        'form': form,
        'mode': 'create',
        'section': 'promocodes'
    })


@login_required
def admin_promocode_edit(request, pk):
    """Редактирование промокода"""
    if not request.user.is_staff:
        return redirect('home')
    
    promocode = get_object_or_404(PromoCode, pk=pk)
    
    if request.method == 'POST':
        form = PromoCodeForm(request.POST, instance=promocode)
        if form.is_valid():
            edited_promocode = form.save(commit=False)
            
            # Если код не указан или пустой, генерируем автоматически
            if not edited_promocode.code or edited_promocode.code.strip() == '':
                edited_promocode.code = PromoCode.generate_code()
            
            edited_promocode.save()
            return redirect('/admin-panel/?section=promocodes&tab=promocodes')
    else:
        form = PromoCodeForm(instance=promocode)
    
    return render(request, 'pages/admin_promocode_form.html', {
        'form': form,
        'mode': 'edit',
        'promocode': promocode,
        'section': 'promocodes'
    })


@login_required
def admin_promocode_toggle(request, pk):
    """Активация/деактивация промокода"""
    if not request.user.is_staff:
        return redirect('home')
    
    promocode = get_object_or_404(PromoCode, pk=pk)
    promocode.is_active = not promocode.is_active
    promocode.save()
    
    return redirect('/admin-panel/?section=promocodes&tab=promocodes')


@login_required
def admin_promocode_delete(request, pk):
    """Удаление промокода"""
    if not request.user.is_staff:
        return redirect('home')
    
    promocode = get_object_or_404(PromoCode, pk=pk)
    promocode.delete()
    
    return redirect('/admin-panel/?section=promocodes&tab=promocodes')


# ==================== PROMO CODE GROUPS ====================

@login_required
def admin_promo_groups(request):
    """Редирект на единую панель промокодов (таб Групи)"""
    return redirect('/admin-panel/?section=promocodes&tab=groups')


@login_required
def admin_promo_group_create(request):
    """Создание новой группы промокодов (поддержка AJAX)"""
    if not request.user.is_staff:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Доступ заборонено'}, status=403)
        return redirect('home')
    
    if request.method == 'POST':
        form = PromoCodeGroupForm(request.POST)
        if form.is_valid():
            group = form.save()
            
            # AJAX ответ
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Групу успішно створено',
                    'group_id': group.id,
                    'group_name': group.name
                })
            
            return redirect('/admin-panel/?section=promocodes&tab=groups')
        else:
            # AJAX ответ с ошибками
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': 'Помилка валідації',
                    'errors': form.errors
                }, status=400)
    else:
        form = PromoCodeGroupForm()
    
    return render(request, 'pages/admin_promo_group_form.html', {
        'form': form,
        'mode': 'create',
        'section': 'promocodes'
    })


@login_required
def admin_promo_group_edit(request, pk):
    """Редактирование группы промокодов"""
    if not request.user.is_staff:
        return redirect('home')
    
    group = get_object_or_404(PromoCodeGroup, pk=pk)
    
    if request.method == 'POST':
        form = PromoCodeGroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            return redirect('/admin-panel/?section=promocodes&tab=groups')
    else:
        form = PromoCodeGroupForm(instance=group)
    
    # Получаем промокоды этой группы
    promocodes = PromoCode.objects.filter(group=group)
    
    return render(request, 'pages/admin_promo_group_form.html', {
        'form': form,
        'mode': 'edit',
        'group': group,
        'promocodes': promocodes,
        'section': 'promocodes'
    })


@login_required
def admin_promo_group_delete(request, pk):
    """Удаление группы промокодов"""
    if not request.user.is_staff:
        return redirect('home')
    
    group = get_object_or_404(PromoCodeGroup, pk=pk)
    
    # При удалении группы, промокоды остаются, но теряют связь с группой
    PromoCode.objects.filter(group=group).update(group=None)
    
    group.delete()
    
    return redirect('/admin-panel/?section=promocodes&tab=groups')


# ==================== STATISTICS ====================

@login_required
def admin_promo_stats(request):
    """Редирект на единую панель промокодов (таб Статистика)"""
    return redirect('/admin-panel/?section=promocodes&tab=stats')


# ==================== AJAX ENDPOINTS ====================

@login_required
def admin_promocode_get_form(request, pk):
    """
    AJAX endpoint: Получить данные промокода для формы редактирования
    GET /admin-panel/promocode/<id>/get-form/
    """
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Доступ заборонено'}, status=403)
    
    if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Тільки AJAX запити'}, status=400)
    
    promocode = get_object_or_404(PromoCode, pk=pk)
    groups = PromoCodeGroup.objects.filter(is_active=True)
    
    # Преобразуем даты в ISO формат для datetime-local input
    valid_from = promocode.valid_from.strftime('%Y-%m-%dT%H:%M') if promocode.valid_from else ''
    valid_until = promocode.valid_until.strftime('%Y-%m-%dT%H:%M') if promocode.valid_until else ''
    
    return JsonResponse({
        'success': True,
        'promo': {
            'id': promocode.id,
            'code': promocode.code,
            'promo_type': promocode.promo_type,
            'discount': float(promocode.discount) if promocode.discount else None,
            'fixed_amount': float(promocode.fixed_amount) if promocode.fixed_amount else None,
            'description': promocode.description or '',
            'group_id': promocode.group.id if promocode.group else None,
            'min_order_amount': float(promocode.min_order_amount) if promocode.min_order_amount else None,
            'usage_limit': promocode.usage_limit,
            'one_time_per_user': promocode.one_time_per_user,
            'valid_from': valid_from,
            'valid_until': valid_until,
            'is_active': promocode.is_active,
        },
        'groups': [{'id': g.id, 'name': g.name} for g in groups]
    })


@login_required
def admin_promocode_edit_ajax(request, pk):
    """
    AJAX endpoint: Обновить промокод
    POST /admin-panel/promocode/<id>/edit-ajax/
    """
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Доступ заборонено'}, status=403)
    
    if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Тільки AJAX запити'}, status=400)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Тільки POST запити'}, status=405)
    
    promocode = get_object_or_404(PromoCode, pk=pk)
    form = PromoCodeForm(request.POST, instance=promocode)
    
    if form.is_valid():
        updated_promo = form.save(commit=False)
        
        # Если код пустой, генерируем
        if not updated_promo.code or updated_promo.code.strip() == '':
            updated_promo.code = PromoCode.generate_code()
        
        updated_promo.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Промокод успішно оновлено',
            'promo_id': updated_promo.id,
            'promo_code': updated_promo.code
        })
    else:
        return JsonResponse({
            'success': False,
            'error': 'Помилка валідації',
            'errors': form.errors
        }, status=400)


@login_required
def admin_promo_group_get_form(request, pk):
    """
    AJAX endpoint: Получить данные группы для формы редактирования
    GET /admin-panel/promo-group/<id>/get-form/
    """
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Доступ заборонено'}, status=403)
    
    if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Тільки AJAX запити'}, status=400)
    
    group = get_object_or_404(PromoCodeGroup, pk=pk)
    promocodes = PromoCode.objects.filter(group=group).order_by('-created_at')
    
    return JsonResponse({
        'success': True,
        'group': {
            'id': group.id,
            'name': group.name,
            'description': group.description or '',
            'one_per_account': group.one_per_account,
            'is_active': group.is_active,
        },
        'promocodes': [{
            'id': p.id,
            'code': p.code,
            'is_active': p.is_active,
            'times_used': getattr(p, 'current_uses', 0),
            'usage_limit': getattr(p, 'max_uses', 0),
            'promo_type': p.promo_type,
        } for p in promocodes]
    })


@login_required
def admin_promo_group_edit_ajax(request, pk):
    """
    AJAX endpoint: Обновить группу промокодов
    POST /admin-panel/promo-group/<id>/edit-ajax/
    """
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Доступ заборонено'}, status=403)
    
    if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Тільки AJAX запити'}, status=400)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Тільки POST запити'}, status=405)
    
    group = get_object_or_404(PromoCodeGroup, pk=pk)
    form = PromoCodeGroupForm(request.POST, instance=group)
    
    if form.is_valid():
        updated_group = form.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Групу успішно оновлено',
            'group_id': updated_group.id,
            'group_name': updated_group.name
        })
    else:
        return JsonResponse({
            'success': False,
            'error': 'Помилка валідації',
            'errors': form.errors
        }, status=400)


@login_required
def admin_promo_export(request):
    """
    Експорт статистики промокодів (CSV/Excel/PDF)
    GET /admin-panel/promo-export/?format=csv&period=all
    """
    if not request.user.is_staff:
        return JsonResponse({'error': 'Доступ заборонено'}, status=403)
    
    import csv
    from django.http import HttpResponse
    from datetime import datetime, timedelta
    
    export_format = request.GET.get('format', 'csv')
    period = request.GET.get('period', 'all')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Визначити діапазон дат
    queryset = PromoCode.objects.all()
    
    if period == 'today':
        today = datetime.now().date()
        queryset = queryset.filter(created_at__date=today)
    elif period == 'week':
        week_ago = datetime.now() - timedelta(days=7)
        queryset = queryset.filter(created_at__gte=week_ago)
    elif period == 'month':
        month_ago = datetime.now() - timedelta(days=30)
        queryset = queryset.filter(created_at__gte=month_ago)
    elif period == 'custom' and date_from and date_to:
        queryset = queryset.filter(created_at__date__range=[date_from, date_to])
    
    # CSV експорт
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="promo_stats_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        # Додати BOM для правильного відображення UTF-8 в Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow(['Код', 'Тип', 'Знижка', 'Група', 'Використань', 'Макс. використань', 'Активний', 'Створено'])
        
        for promo in queryset:
            writer.writerow([
                promo.code,
                promo.get_promo_type_display(),
                promo.get_discount_display(),
                promo.group.name if promo.group else '-',
                promo.current_uses,
                promo.max_uses if promo.max_uses > 0 else '∞',
                'Так' if promo.is_active else 'Ні',
                promo.created_at.strftime('%d.%m.%Y %H:%M'),
            ])
        
        return response
    
    # Excel експорт (потребує openpyxl)
    elif export_format == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Промокоди"
            
            # Заголовки
            headers = ['Код', 'Тип', 'Знижка', 'Група', 'Використань', 'Макс. використань', 'Активний', 'Створено']
            ws.append(headers)
            
            # Стилізація заголовків
            header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Дані
            for promo in queryset:
                ws.append([
                    promo.code,
                    promo.get_promo_type_display(),
                    promo.get_discount_display(),
                    promo.group.name if promo.group else '-',
                    promo.current_uses,
                    promo.max_uses if promo.max_uses > 0 else '∞',
                    'Так' if promo.is_active else 'Ні',
                    promo.created_at.strftime('%d.%m.%Y %H:%M'),
                ])
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="promo_stats_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except ImportError:
            return HttpResponse("Excel export requires 'openpyxl' library. Please install it.", status=500)
    
    # PDF експорт (потребує reportlab)
    elif export_format == 'pdf':
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            
            # Дані для таблиці
            data = [['Код', 'Тип', 'Знижка', 'Група', 'Використань', 'Активний']]
            
            for promo in queryset[:50]:  # Обмежуємо до 50 для PDF
                data.append([
                    promo.code,
                    promo.get_promo_type_display(),
                    promo.get_discount_display(),
                    promo.group.name if promo.group else '-',
                    f"{promo.current_uses}/{promo.max_uses if promo.max_uses > 0 else '∞'}",
                    'Так' if promo.is_active else 'Ні',
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            pdf = buffer.getvalue()
            buffer.close()
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="promo_stats_{datetime.now().strftime("%Y%m%d")}.pdf"'
            response.write(pdf)
            
            return response
            
        except ImportError:
            return HttpResponse("PDF export requires 'reportlab' library. Please install it.", status=500)
    
    return HttpResponse("Invalid format", status=400)


@login_required
def admin_promocode_change_group(request, pk):
    """
    AJAX endpoint: Змінити групу промокода (Drag & Drop)
    POST /admin-panel/promocode/<id>/change-group/
    """
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Доступ заборонено'}, status=403)
    
    if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Тільки AJAX запити'}, status=400)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Тільки POST запити'}, status=405)
    
    promocode = get_object_or_404(PromoCode, pk=pk)
    
    try:
        import json
        data = json.loads(request.body)
        group_id = data.get('group_id')
        
        if group_id:
            group = get_object_or_404(PromoCodeGroup, pk=group_id)
            promocode.group = group
            promocode.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Промокод переміщено до групи "{group.name}"',
                'group_id': group.id,
                'group_name': group.name
            })
        else:
            # Прибрати з групи
            promocode.group = None
            promocode.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Промокод прибрано з групи',
                'group_id': None,
                'group_name': None
            })
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Невірний JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
