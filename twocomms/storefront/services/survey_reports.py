from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Optional

from django.conf import settings
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from accounts.models import UserProfile
from .survey_engine import SurveyEngine


def build_survey_report(
    session,
    definition: Dict[str, Any],
    status: str,
    file_path: Path,
) -> Path:
    """Build or update Excel report for a survey session.

    The workbook is structured for a non-technical admin reading it on a
    phone or laptop: a clean "Огляд" cover sheet with the key facts, a
    "Відповіді" sheet with every answer in plain language, and a "Сигнали"
    sheet for quick aggregate reads. All labels are in Ukrainian.
    """
    engine = SurveyEngine(definition)
    file_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Огляд"

    # Brand-aligned palette (TwoComms orange + neutral greys).
    brand_orange = "FF8C42"
    dark_fill = PatternFill(start_color="1C2533", end_color="1C2533", fill_type="solid")
    orange_fill = PatternFill(start_color=brand_orange, end_color=brand_orange, fill_type="solid")
    soft_orange_fill = PatternFill(start_color="FFF1E6", end_color="FFF1E6", fill_type="solid")
    soft_gray_fill = PatternFill(start_color="F4F5F7", end_color="F4F5F7", fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    blue_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    title_font = Font(name="Calibri", bold=True, size=18, color="FFFFFF")
    subtitle_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    section_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    label_font = Font(name="Calibri", bold=True, color="374151")
    value_font = Font(name="Calibri", color="111827")
    note_font = Font(name="Calibri", italic=True, color="4B5563")

    thin_gray = Side(style="thin", color="E5E7EB")
    panel_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    status_labels = {
        "FINAL": "✅ Завершено",
        "PARTIAL": "⏳ Без активності",
        "UPDATED": "♻️ Оновлено",
    }
    status_text = status_labels.get(status, status)
    status_fill = green_fill if status == "FINAL" else (amber_fill if status == "PARTIAL" else blue_fill)

    def style_header_row(sheet, row: int, max_col: int) -> None:
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = dark_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = panel_border
        sheet.row_dimensions[row].height = 26

    def section_bar(sheet, cell_range: str, text: str) -> None:
        first = cell_range.split(":")[0]
        sheet.merge_cells(cell_range)
        sheet[first] = text
        sheet[first].font = section_font
        sheet[first].fill = orange_fill
        sheet[first].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        row_no = int("".join(ch for ch in first if ch.isdigit()))
        sheet.row_dimensions[row_no].height = 22

    def write_pair(sheet, row: int, label: str, value: Any, *, col: int = 1, fill=None) -> None:
        label_cell = sheet.cell(row=row, column=col, value=label)
        value_cell = sheet.cell(row=row, column=col + 1, value=value)
        label_cell.font = label_font
        label_cell.fill = fill or soft_gray_fill
        value_cell.font = value_font
        value_cell.fill = fill or PatternFill(fill_type=None)
        for cell in (label_cell, value_cell):
            cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
            cell.border = panel_border
        sheet.row_dimensions[row].height = 20

    profile = None
    try:
        profile = UserProfile.objects.filter(user=session.user).first()
    except Exception:
        profile = None

    created = timezone.localtime(session.started_at).strftime("%d.%m.%Y %H:%M")
    last_activity = timezone.localtime(session.last_activity_at).strftime("%d.%m.%Y %H:%M")
    completed = (
        timezone.localtime(session.completed_at).strftime("%d.%m.%Y %H:%M")
        if session.completed_at
        else "—"
    )
    promo_expires = (
        timezone.localtime(session.awarded_promocode.valid_until).strftime("%d.%m.%Y %H:%M")
        if session.awarded_promocode and session.awarded_promocode.valid_until
        else "—"
    )

    title = get_survey_title(definition)
    answered_count = len(session.history or [])
    text_count = 0
    for qid in session.history or []:
        question = engine.get_question(qid)
        if question and question.get("type") in ("text_short", "text_long"):
            text_count += 1
    promo_code = session.awarded_promocode.code if session.awarded_promocode else "—"
    contact_value = " / ".join(
        filter(None, [session.user.email or (profile.email if profile else ""), profile.phone if profile else ""])
    ) or "—"

    # --- Cover header -----------------------------------------------------
    ws.merge_cells("A1:F1")
    ws["A1"] = "TWOCOMMS · Звіт опитування"
    ws["A1"].font = title_font
    ws["A1"].fill = dark_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:F2")
    ws["A2"] = title
    ws["A2"].font = subtitle_font
    ws["A2"].fill = orange_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 26

    # --- Key facts --------------------------------------------------------
    section_bar(ws, "A4:F4", "Основне")
    write_pair(ws, 5, "Статус", status_text, fill=status_fill)
    write_pair(ws, 6, "Користувач", f"{session.user.username} (ID {session.user_id})")
    write_pair(ws, 7, "Контакт", contact_value)
    write_pair(ws, 5, "Відповідей", answered_count, col=4, fill=soft_orange_fill)
    write_pair(ws, 6, "Промокод", promo_code, col=4, fill=soft_orange_fill)
    write_pair(ws, 7, "Пристрій", session.device_type or "—", col=4)

    # --- Timeline + reward ------------------------------------------------
    section_bar(ws, "A9:C9", "Хронологія")
    section_bar(ws, "D9:F9", "Винагорода")
    write_pair(ws, 10, "Початок", created)
    write_pair(ws, 11, "Остання активність", last_activity)
    write_pair(ws, 12, "Завершено", completed)
    write_pair(ws, 10, "Промокод діє до", promo_expires, col=4)
    write_pair(ws, 11, "Текстових відповідей", text_count, col=4)
    write_pair(ws, 12, "Звіт сформовано", timezone.localtime(timezone.now()).strftime("%d.%m.%Y %H:%M"), col=4)

    # --- Notes ------------------------------------------------------------
    section_bar(ws, "A14:F14", "Як читати звіт")
    ws.merge_cells("A15:F16")
    ws["A15"] = (
        "Лист «Відповіді» — точні відповіді респондента по кожному питанню. "
        "Лист «Сигнали» — швидке зведення по модулях, типах питань і стану завершення."
    )
    ws["A15"].font = note_font
    ws["A15"].alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    for row in ws.iter_rows(min_row=15, max_row=16, min_col=1, max_col=6):
        for cell in row:
            cell.fill = soft_gray_fill
            cell.border = panel_border

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 6

    # --- Answers sheet ----------------------------------------------------
    ws_answers = wb.create_sheet("Відповіді")
    ws_answers.sheet_view.showGridLines = False
    ws_answers.append(["№", "Розділ", "ID питання", "Питання", "Відповідь", "Тип"])
    style_header_row(ws_answers, 1, 6)

    answer_row = 2
    section_counts: Dict[str, int] = {}
    type_counts: Dict[str, int] = {}
    for index, qid in enumerate(session.history or [], start=1):
        question = engine.get_question(qid)
        if not question:
            continue
        answer = (session.answers or {}).get(qid)
        section = question.get("section") or ""
        qtype = question.get("type") or ""
        section_counts[section] = section_counts.get(section, 0) + 1
        type_counts[qtype] = type_counts.get(qtype, 0) + 1
        ws_answers.append([
            index,
            section,
            qid,
            question.get("prompt_uk") or question.get("prompt") or "",
            engine.format_answer(question, answer),
            qtype,
        ])
        fill = soft_gray_fill if answer_row % 2 == 0 else PatternFill(fill_type=None)
        for col in range(1, 7):
            cell = ws_answers.cell(row=answer_row, column=col)
            cell.fill = fill
            cell.border = panel_border
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal="center" if col in (1, 6) else "left",
            )
        answer_row += 1

    if answer_row == 2:
        ws_answers.append(["—", "—", "—", "Респондент ще не відповів на жодне питання", "—", "—"])
        for col in range(1, 7):
            cell = ws_answers.cell(row=2, column=col)
            cell.border = panel_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        answer_row = 3

    ws_answers.freeze_panes = "A2"
    ws_answers.auto_filter.ref = f"A1:F{max(answer_row - 1, 1)}"
    for column, width in {
        1: 6,
        2: 22,
        3: 28,
        4: 64,
        5: 58,
        6: 22,
    }.items():
        ws_answers.column_dimensions[get_column_letter(column)].width = width

    # --- Signals sheet ----------------------------------------------------
    ws_signals = wb.create_sheet("Сигнали")
    ws_signals.sheet_view.showGridLines = False
    ws_signals.append(["Сигнал", "Значення", "Що це означає"])
    style_header_row(ws_signals, 1, 3)

    signal_rows = [
        ("Ключ опитування", session.survey_key, "Яке визначення опитування сформувало звіт"),
        ("Статус", status_text, "«Завершено» означає, що респондент дійшов до екрана промокоду"),
        ("Відповідей", answered_count, "Усього збережених відповідей у сесії"),
        ("Текстових відповідей", text_count, "Вільні відповіді, які варто переглянути вручну"),
        ("Порядок модулів", ", ".join(session.module_order or []) or "—", "Динамічні модулі, підібрані для респондента"),
        ("Промокод", promo_code, "Код, показаний респонденту"),
    ]
    for section, count in sorted(section_counts.items()):
        signal_rows.append((f"Розділ: {section or 'н/д'}", count, "Кількість питань у цьому розділі"))
    for qtype, count in sorted(type_counts.items()):
        signal_rows.append((f"Тип: {qtype}", count, "Кількість питань цього типу у відповіді"))

    for row_index, row_values in enumerate(signal_rows, start=2):
        ws_signals.append(list(row_values))
        fill = soft_gray_fill if row_index % 2 == 0 else PatternFill(fill_type=None)
        for col in range(1, 4):
            cell = ws_signals.cell(row=row_index, column=col)
            cell.fill = fill
            cell.border = panel_border
            cell.font = value_font if col != 1 else label_font
            cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)

    ws_signals.freeze_panes = "A2"
    ws_signals.auto_filter.ref = f"A1:C{max(len(signal_rows) + 1, 1)}"
    ws_signals.column_dimensions["A"].width = 34
    ws_signals.column_dimensions["B"].width = 34
    ws_signals.column_dimensions["C"].width = 70

    ws.sheet_properties.tabColor = brand_orange
    ws_answers.sheet_properties.tabColor = "1C2533"
    ws_signals.sheet_properties.tabColor = "2563EB"

    wb.save(file_path)
    return file_path


def _safe_filename_component(value: Optional[str], fallback: str, allow_unicode: bool = False) -> str:
    if not value:
        return fallback
    slug = slugify(str(value), allow_unicode=allow_unicode)
    return slug or fallback


def get_survey_title(definition: Optional[Dict[str, Any]]) -> str:
    """Resolve a human-friendly survey title for notifications/files."""
    definition = definition or {}
    ui = definition.get("ui_copy", {}) or {}
    modal_title = (ui.get("modal") or {}).get("title_uk")
    home_title = (ui.get("homepage_block") or {}).get("title_uk")
    return modal_title or home_title or definition.get("survey_key", "Survey")


def resolve_report_path(session, definition: Optional[Dict[str, Any]] = None) -> Path:
    """Resolve report path for a session, using a stable human-friendly filename."""
    base_dir = Path(getattr(settings, "SURVEY_REPORTS_DIR", settings.MEDIA_ROOT / "survey_reports"))
    title = get_survey_title(definition)
    title_slug = _safe_filename_component(title, "survey", allow_unicode=True)
    username = _safe_filename_component(session.user.username, f"user{session.user_id}", allow_unicode=True)
    email = _safe_filename_component(session.user.email, "", allow_unicode=False)
    parts = [title_slug, username]
    if email:
        parts.append(email)
    parts.append(f"s{session.id}")
    filename = f"opituvannia_{'_'.join(parts)}.xlsx"
    return base_dir / filename


class AnonymousSurveySession:
    """Mock/wrapper class that mimics SurveySession for anonymous users.
    
    Allows using the same Excel builder and path resolution logic without DB persistence.
    """
    def __init__(self, state: Dict[str, Any], awarded_promocode=None):
        self.id = "anon"
        self.survey_key = state.get("survey_key", "print_feedback_v1")
        self.status = state.get("status", "completed")
        self.answers = state.get("answers", {})
        self.history = state.get("history", [])
        self.current_question_id = state.get("current_question_id")
        self.back_used = state.get("back_used", False)
        self.version = state.get("version", 1)
        self.device_type = state.get("device_type")
        self.user_agent = state.get("user_agent")
        self.module_order = state.get("module_order", [])

        # Datetime parsing
        def parse_dt(val):
            if not val:
                return timezone.now()
            try:
                from django.utils.dateparse import parse_datetime
                dt = parse_datetime(val)
                if not dt:
                    return timezone.now()
                return timezone.is_aware(dt) and dt or timezone.make_aware(dt)
            except Exception:
                return timezone.now()

        self.started_at = parse_dt(state.get("started_at"))
        self.last_activity_at = parse_dt(state.get("last_activity_at"))
        self.completed_at = parse_dt(state.get("completed_at")) if state.get("completed_at") else None

        class MockUser:
            id = 0
            username = "Анонім"
            email = ""

        self.user = MockUser()
        self.user_id = 0
        self.awarded_promocode = awarded_promocode


def send_anonymous_survey_report(state: Dict[str, Any], definition: Dict[str, Any], status: str) -> bool:
    """Generate and send survey report to Telegram for anonymous users synchronously."""
    import html
    from storefront.models import PromoCode
    from orders.telegram_notifications import telegram_notifier

    # 1. Resolve awarded promo code
    promo_code = None
    promo_id = state.get("awarded_promocode_id")
    if promo_id:
        try:
            promo_code = PromoCode.objects.get(id=promo_id)
        except PromoCode.DoesNotExist:
            pass

    # 2. Build mock session
    session = AnonymousSurveySession(state, promo_code)

    # 3. Build report
    report_path = resolve_report_path(session, definition)
    build_survey_report(session, definition, status, report_path)

    # 4. Prepare caption
    def fmt_dt(value):
        if not value:
            return "—"
        return timezone.localtime(value).strftime("%d.%m.%Y %H:%M")

    title = html.escape(get_survey_title(definition))
    status_text = "✅ Завершено (Анонімно)"

    caption_lines = [
        f"<b>{title}</b>",
        f"Статус: {status_text}",
        f"Користувач: Анонімний відвідувач",
    ]

    code = promo_code.code if promo_code else ""
    expires = fmt_dt(promo_code.valid_until) if promo_code and promo_code.valid_until else ""
    if code:
        if expires:
            caption_lines.append(f"Промокод: {code} (до {expires})")
        else:
            caption_lines.append(f"Промокод: {code}")

    caption_lines.extend([
        f"Початок: {fmt_dt(session.started_at)}",
        f"Остання активність: {fmt_dt(session.last_activity_at)}",
    ])
    if session.completed_at:
        caption_lines.append(f"Завершено: {fmt_dt(session.completed_at)}")

    caption = "\n".join(caption_lines)

    # 5. Send report
    return telegram_notifier.send_admin_document(str(report_path), caption, filename=Path(report_path).name)

