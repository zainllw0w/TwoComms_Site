"""
Tests for survey engine, promo award idempotency, and access rules.
"""
import json
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from django.contrib.auth.models import User
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from storefront.models import PromoCode, SurveySession, UserAction, UserPromoCode
from storefront.services.survey_engine import (
    SurveyEngine,
    award_survey_promocode,
    clear_survey_definition_cache,
    load_survey_definition,
)
from storefront.services.survey_reports import build_survey_report
from storefront.tasks import queue_survey_report


TEST_DEFINITION = {
    "survey_key": "test_survey",
    "reward": {"amount_uah": 200, "expires_in_days": 5, "title_uk": "Reward"},
    "caps": {
        "total_max_questions": 3,
        "max_text_questions_total": 1,
        "max_followups_per_multi_choice": 1,
    },
    "end_conditions": {"required_questions": ["q1"], "min_answered_questions": 2},
    "flow": {"core": ["q1", "q2"], "closing": ["q3"]},
    "questions": [
        {
            "id": "q1",
            "type": "slider_1_10",
            "prompt": "Overall",
            "required": True,
            "scale": {"min": 1, "max": 10},
        },
        {
            "id": "q2",
            "type": "single_choice",
            "prompt": "Pick",
            "options": [
                {"value": "a", "label": "A"},
                {"value": "b", "label": "B"},
            ],
            "show_if": {"var": "answers.q1", "op": "gte", "value": 5},
        },
        {
            "id": "q3",
            "type": "text_short",
            "prompt": "Why",
            "required": False,
        },
    ],
}

V34_DEFINITION_PATH = (
    Path(__file__).resolve().parents[2]
    / "surveys"
    / "twocomms_survey_v3_4_adaptive_research.json"
)


class SurveyEngineTests(TestCase):
    def test_next_question_respects_show_if(self):
        engine = SurveyEngine(TEST_DEFINITION)
        first = engine.get_next_question({}, [])
        self.assertEqual(first["id"], "q1")

        answers = {"q1": 4}
        history = ["q1"]
        next_question = engine.get_next_question(answers, history)
        self.assertEqual(next_question["id"], "q3")

    def test_validate_slider_range(self):
        engine = SurveyEngine(TEST_DEFINITION)
        question = engine.get_question("q1")
        valid, _, reason = engine.validate_answer(question, 11)
        self.assertFalse(valid)
        self.assertEqual(reason, "range")

    def test_v34_definition_uses_screens_modules_and_compile_safe_price_ladder(self):
        definition = load_survey_definition(V34_DEFINITION_PATH)
        engine = SurveyEngine(definition)

        self.assertEqual(definition["survey_key"], "twocomms_survey_v3_4_adaptive_research")
        self.assertEqual(engine.flow_core[:3], ["q010_entry_goal", "q020_purchase_stage", "q030_product_interest"])
        self.assertIn("q930_promo_delivery", engine.flow_closing)
        self.assertEqual(engine.validate_definition(), [])

        selected = engine.select_modules(
            {
                "q010_entry_goal": "partner",
                "q020_purchase_stage": "partner_discussion",
                "q070_main_barrier": "quality_trust",
                "q080_priority_tradeoff": "custom",
            }
        )
        self.assertEqual(selected[0], "partner_lab")
        self.assertLessEqual(len(selected), 2)

    def test_v34_smart_skip_autofills_single_product_without_counting_history(self):
        definition = load_survey_definition(V34_DEFINITION_PATH)
        engine = SurveyEngine(definition)
        answers = {
            "q010_entry_goal": "ready_for_self",
            "q020_purchase_stage": "considering",
            "q030_product_interest": ["hoodie"],
        }

        question = engine.get_next_question(answers, ["q010_entry_goal", "q020_purchase_stage", "q030_product_interest"])

        self.assertEqual(question["id"], "q040_military_code_interest")
        self.assertEqual(answers["q035_primary_product_focus"], "hoodie")
        self.assertTrue(answers["q035_auto_answered"])

    def test_v34_b2b_path_skips_personal_style_questions(self):
        definition = load_survey_definition(V34_DEFINITION_PATH)
        engine = SurveyEngine(definition)
        answers = {
            "q010_entry_goal": "partner",
            "q020_purchase_stage": "partner_discussion",
            "q030_product_interest": ["not_sure"],
            "q035_primary_product_focus": "not_sure",
        }

        question = engine.get_next_question(
            answers,
            ["q010_entry_goal", "q020_purchase_stage", "q030_product_interest", "q035_primary_product_focus"],
        )

        self.assertEqual(question["id"], "q050_overall")

    def test_v34_new_question_types_validate_and_serialize(self):
        definition = load_survey_definition(V34_DEFINITION_PATH)
        engine = SurveyEngine(definition)

        valid, cleaned, reason = engine.validate_answer(
            engine.get_question("q060_purchase_driver_maxdiff"),
            {"best": "quality", "worst": "price"},
        )
        self.assertTrue(valid, reason)
        self.assertEqual(cleaned, {"best": "quality", "worst": "price"})

        valid, _, reason = engine.validate_answer(
            engine.get_question("q060_purchase_driver_maxdiff"),
            {"best": "quality", "worst": "quality"},
        )
        self.assertFalse(valid)
        self.assertEqual(reason, "distinct_required")

        valid, cleaned, reason = engine.validate_answer(
            engine.get_question("q600_assortment_purchase_matrix"),
            {"women_line": "купив/ла б", "caps": "не потрібно"},
        )
        self.assertTrue(valid, reason)
        self.assertEqual(cleaned["women_line"], "купив/ла б")

        valid, cleaned, reason = engine.validate_answer(
            engine.get_question("q603_assortment_budget_allocation"),
            {"new_prints": 40, "better_base": 30, "custom": 30},
        )
        self.assertTrue(valid, reason)
        self.assertEqual(sum(cleaned.values()), 100)

        valid, _, reason = engine.validate_answer(
            engine.get_question("q603_assortment_budget_allocation"),
            {"new_prints": 40, "better_base": 30},
        )
        self.assertFalse(valid)
        self.assertEqual(reason, "total_must_be_100")

        valid, cleaned, reason = engine.validate_answer(
            engine.get_question("q921_contact_bundle"),
            {"channel": "telegram", "value": "@twocomms_user"},
        )
        self.assertTrue(valid, reason)
        self.assertEqual(cleaned["channel"], "telegram")

        price_question = engine.serialize_question(
            engine.get_question("q403_price_ladder"),
            answers={
                "q010_entry_goal": "custom_team_unit",
                "q020_purchase_stage": "custom_request",
                "q035_primary_product_focus": "hoodie",
            },
        )
        self.assertEqual(price_question["price_context"], "team_unit_per_unit")
        self.assertEqual(price_question["range_key"], "team_unit_bulk_fallback")
        self.assertIn("team_low", {option["value"] for option in price_question["options"]})

        project_price_question = engine.serialize_question(
            engine.get_question("q403_price_ladder"),
            answers={
                "q010_entry_goal": "custom_team_unit",
                "q020_purchase_stage": "custom_request",
                "q305_custom_quantity": "21_50",
            },
        )
        self.assertEqual(project_price_question["price_context"], "team_unit_project_budget")
        self.assertEqual(project_price_question["range_key"], "team_unit_project_budget_fallback")


class SurveyPromoCodeTests(TestCase):
    def test_award_promocode_idempotent(self):
        user = User.objects.create_user(username="surveyuser", password="pass1234")
        reward = {"amount_uah": 200, "expires_in_days": 5, "title_uk": "Reward"}

        promo_first, _ = award_survey_promocode(user, "test_survey", reward)
        promo_second, _ = award_survey_promocode(user, "test_survey", reward)

        self.assertEqual(promo_first.id, promo_second.id)
        self.assertEqual(user.promo_grants.count(), 1)


class SurveyReportTests(TestCase):
    def test_report_workbook_has_admin_friendly_overview_answers_and_signals(self):
        user = User.objects.create_user(
            username="surveyuser",
            email="admin-readable@example.com",
            password="pass1234",
        )
        promo, _ = award_survey_promocode(user, "test_survey", TEST_DEFINITION["reward"])
        session = SurveySession.objects.create(
            user=user,
            survey_key="test_survey",
            status="completed",
            answers={"q1": 7, "q2": "a", "q3": "Better checkout"},
            history=["q1", "q2", "q3"],
            awarded_promocode=promo,
        )

        with TemporaryDirectory() as temp_dir:
            report_path = Path(temp_dir) / "survey-report.xlsx"
            build_survey_report(session, TEST_DEFINITION, "FINAL", report_path)

            workbook = load_workbook(report_path)
            self.assertEqual(workbook.sheetnames[:3], ["Overview", "Answers", "Signals"])

            overview = workbook["Overview"]
            self.assertEqual(overview["A1"].value, "TWOCOMMS Survey Report")
            self.assertEqual(overview["A4"].value, "Status")
            self.assertEqual(overview["B4"].value, "FINAL")
            self.assertEqual(overview.freeze_panes, "A8")

            answers = workbook["Answers"]
            self.assertEqual(
                [answers.cell(row=1, column=col).value for col in range(1, 7)],
                ["#", "Section", "Question ID", "Question", "Answer", "Type"],
            )
            self.assertEqual(answers.freeze_panes, "A2")
            self.assertEqual(answers.auto_filter.ref, "A1:F4")

            signals = workbook["Signals"]
            self.assertEqual(signals["A1"].value, "Signal")
            self.assertEqual(signals["B1"].value, "Value")
            self.assertIn("Answered questions", [signals.cell(row=row, column=1).value for row in range(2, 12)])


class SurveyTaskQueueTests(TestCase):
    def test_background_survey_report_queue_returns_without_inline_report_generation(self):
        with patch("storefront.tasks.send_survey_report_task") as report_task, patch("storefront.tasks.Thread") as thread_cls:
            queued = queue_survey_report(123, "FINAL", background=True)

        self.assertTrue(queued)
        report_task.delay.assert_not_called()
        report_task.assert_not_called()
        thread_cls.assert_called_once()
        thread_cls.return_value.start.assert_called_once()


class SurveyViewTests(TestCase):
    def setUp(self):
        self.temp_dir = TemporaryDirectory()
        self.definition_path = Path(self.temp_dir.name) / "survey.json"
        self.definition_path.write_text(json.dumps(TEST_DEFINITION), encoding="utf-8")
        self.override = override_settings(SURVEY_DEFINITION_PATH=self.definition_path)
        self.override.enable()
        clear_survey_definition_cache()
        self.client = Client(
            HTTP_HOST="twocomms.shop",
            SERVER_PORT="443",
            **{"wsgi.url_scheme": "https"},
        )

    def tearDown(self):
        clear_survey_definition_cache()
        self.override.disable()
        self.temp_dir.cleanup()

    def test_anonymous_start_returns_first_question(self):
        response = self.client.post(reverse("survey_start_or_resume"), data="{}", content_type="application/json", secure=True)
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload.get("success"))
        self.assertEqual(payload["question"]["id"], "q1")
        self.assertEqual(SurveySession.objects.count(), 0)
        self.assertEqual(UserAction.objects.filter(action_type="survey_start").count(), 1)

    def test_anonymous_can_complete_and_receive_single_use_promo(self):
        start_response = self.client.post(reverse("survey_start_or_resume"), data="{}", content_type="application/json", secure=True)
        start_payload = start_response.json()

        first_answer = self.client.post(
            reverse("survey_submit_answer"),
            data=json.dumps({
                "question_id": "q1",
                "answer": 3,
                "version": start_payload["version"],
            }),
            content_type="application/json",
            secure=True,
        )
        first_payload = first_answer.json()

        complete_response = self.client.post(
            reverse("survey_submit_answer"),
            data=json.dumps({
                "question_id": "q3",
                "answer": "Better checkout",
                "version": first_payload["version"],
            }),
            content_type="application/json",
            secure=True,
        )

        self.assertEqual(complete_response.status_code, 200)
        complete_payload = complete_response.json()
        self.assertEqual(complete_payload["status"], "completed")
        self.assertIn("promo", complete_payload)
        promo = PromoCode.objects.get(code=complete_payload["promo"]["code"])
        self.assertEqual(promo.discount_value, 200)
        self.assertEqual(promo.max_uses, 1)
        self.assertTrue(promo.one_time_per_user)
        self.assertEqual(SurveySession.objects.count(), 0)
        self.assertEqual(UserPromoCode.objects.count(), 0)

        resume_response = self.client.post(reverse("survey_start_or_resume"), data="{}", content_type="application/json", secure=True)
        self.assertEqual(resume_response.json()["promo"]["code"], promo.code)

    def test_authenticated_start_and_answer(self):
        user = User.objects.create_user(username="surveyuser", password="pass1234")
        self.client.force_login(user)

        start_response = self.client.post(reverse("survey_start_or_resume"), data="{}", content_type="application/json", secure=True)
        self.assertEqual(start_response.status_code, 200)
        payload = start_response.json()
        self.assertTrue(payload.get("success"))
        self.assertEqual(payload["question"]["id"], "q1")
        self.assertEqual(UserAction.objects.filter(action_type="survey_start").count(), 1)

        answer_response = self.client.post(
            reverse("survey_submit_answer"),
            data=json.dumps({
                "question_id": "q1",
                "answer": 3,
                "version": payload["version"],
            }),
            content_type="application/json",
            secure=True,
        )
        self.assertEqual(answer_response.status_code, 200)
        answer_payload = answer_response.json()
        self.assertEqual(answer_payload["question"]["id"], "q3")
        self.assertEqual(UserAction.objects.filter(action_type="survey_answer").count(), 1)
        answer_action = UserAction.objects.get(action_type="survey_answer")
        self.assertEqual(answer_action.metadata.get("question_id"), "q1")

    def test_authenticated_completion_returns_promo_before_report_generation(self):
        user = User.objects.create_user(username="surveyuser", password="pass1234")
        self.client.force_login(user)

        start_response = self.client.post(reverse("survey_start_or_resume"), data="{}", content_type="application/json", secure=True)
        start_payload = start_response.json()
        first_answer = self.client.post(
            reverse("survey_submit_answer"),
            data=json.dumps({
                "question_id": "q1",
                "answer": 3,
                "version": start_payload["version"],
            }),
            content_type="application/json",
            secure=True,
        )
        first_payload = first_answer.json()

        with patch("storefront.views.survey.queue_survey_report") as queue_mock:
            with self.captureOnCommitCallbacks(execute=False) as callbacks:
                complete_response = self.client.post(
                    reverse("survey_submit_answer"),
                    data=json.dumps({
                        "question_id": "q3",
                        "answer": "Better checkout",
                        "version": first_payload["version"],
                    }),
                    content_type="application/json",
                    secure=True,
                )

            self.assertEqual(complete_response.status_code, 200)
            complete_payload = complete_response.json()
            self.assertEqual(complete_payload["status"], "completed")
            self.assertIn("promo", complete_payload)
            queue_mock.assert_not_called()
            self.assertEqual(len(callbacks), 1)

            callbacks[0]()
            session = SurveySession.objects.get(user=user, survey_key="test_survey")
            queue_mock.assert_called_once_with(session.id, "FINAL", background=True)

    @override_settings(SURVEY_DEFINITION_PATH=V34_DEFINITION_PATH)
    def test_v34_anonymous_partner_path_skips_style_questions(self):
        clear_survey_definition_cache()
        start_response = self.client.post(reverse("survey_start_or_resume"), data="{}", content_type="application/json", secure=True)
        start_payload = start_response.json()
        self.assertEqual(start_payload["question"]["id"], "q010_entry_goal")

        steps = [
            ("q010_entry_goal", "partner"),
            ("q020_purchase_stage", "partner_discussion"),
            ("q030_product_interest", ["not_sure"]),
            ("q035_primary_product_focus", "not_sure"),
        ]
        payload = start_payload
        for question_id, answer in steps:
            response = self.client.post(
                reverse("survey_submit_answer"),
                data=json.dumps({
                    "question_id": question_id,
                    "answer": answer,
                    "version": payload["version"],
                }),
                content_type="application/json",
                secure=True,
            )
            self.assertEqual(response.status_code, 200)
            payload = response.json()

        self.assertEqual(payload["question"]["id"], "q050_overall")
        bucket = self.client.session["anonymous_survey_sessions"]
        answers = bucket["twocomms_survey_v3_4_adaptive_research"]["answers"]
        self.assertNotIn("q040_military_code_interest", answers)
        self.assertNotIn("q041_streetwear_fit_interest", answers)
