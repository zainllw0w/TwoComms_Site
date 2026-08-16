from __future__ import annotations

import io
import warnings

from django.test import TestCase
from django.urls import reverse
from django.utils.deprecation import RemovedInDjango70Warning
from openpyxl import load_workbook

from productcolors.models import Color, ProductColorVariant
from storefront.models import Category, Product


class WholesalePricelistRouteTests(TestCase):
    def setUp(self):
        tshirt_category = Category.objects.create(
            name="Футболки",
            slug="tshirts-pricelist-test",
        )
        hoodie_category = Category.objects.create(
            name="Худи",
            slug="hoodies-pricelist-test",
        )
        tshirt = Product.objects.create(
            title="Test T-shirt",
            slug="test-tshirt-pricelist",
            category=tshirt_category,
            price=1000,
        )
        hoodie = Product.objects.create(
            title="Test Hoodie",
            slug="test-hoodie-pricelist",
            category=hoodie_category,
            price=2000,
        )
        black = Color.objects.create(name="black", primary_hex="#000000")
        white = Color.objects.create(name="white", primary_hex="#ffffff")
        for product in (tshirt, hoodie):
            ProductColorVariant.objects.create(product=product, color=black)
            ProductColorVariant.objects.create(product=product, color=white)

    def test_download_is_warning_free_and_preserves_expected_output(self):
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            response = self.client.get(
                reverse("wholesale_prices_xlsx"),
                HTTP_HOST="twocomms.shop",
            )

        django70_warnings = [
            item
            for item in caught
            if issubclass(item.category, RemovedInDjango70Warning)
        ]
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(
            response["Content-Disposition"],
            'attachment; filename="pricelist_opt.xlsx"',
        )
        self.assertEqual(django70_warnings, [])

        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        flattened = {value for row in rows for value in row if value is not None}
        self.assertIn("Test T-shirt (S–XL)", flattened)
        self.assertIn("Test Hoodie (S–XL) [фліс]", flattened)
        self.assertIn("чорний", flattened)
        self.assertIn("білий", flattened)
        self.assertIn(
            "http://twocomms.shop/product/test-tshirt-pricelist/",
            flattened,
        )
