"""Імпорт банківської виписки CSV/XLSX з антидублями (ТЗ 05 §8, 09 §7).

Формат за замовчуванням (universal): колонки date, amount, comment[, external_id].
amount > 0 → дохід, amount < 0 → витрата. Дублі визначаються за
provider+external_id або (date+amount+account+comment).
"""
from __future__ import annotations

import csv
import datetime as dt
import io
from decimal import Decimal, InvalidOperation

from django.utils import timezone

from ..models import Transaction, get_default_company
from . import transactions as txn_service


def _parse_amount(raw):
    if raw in (None, ''):
        return None
    s = str(raw).replace(' ', '').replace(' ', '').replace(',', '.')
    s = s.replace('+', '')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _parse_date(raw):
    if not raw:
        return None
    raw = str(raw).strip()
    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S',
                '%d.%m.%Y %H:%M', '%d.%m.%Y %H:%M:%S'):
        try:
            return dt.datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def parse_file(file_obj, filename=''):
    """Повертає список рядків dict(date, amount, comment, external_id, raw)."""
    name = (filename or getattr(file_obj, 'name', '') or '').lower()
    rows = []
    if name.endswith('.xlsx') or name.endswith('.xls'):
        rows = _parse_xlsx(file_obj)
    else:
        rows = _parse_csv(file_obj)
    return rows


def _normalize_header(h):
    return (h or '').strip().lower()


_DATE_KEYS = {'date', 'дата', 'дата операції', 'data', 'дата операции'}
_AMOUNT_KEYS = {'amount', 'сума', 'сумма', 'amount, uah', 'сума у валюті картки'}
_COMMENT_KEYS = {'comment', 'коментар', 'комментарий', 'опис', 'описание', 'призначення', 'description', 'деталі операції'}
_EXTID_KEYS = {'external_id', 'id', 'reference', 'ref'}


def _map_row(headers, values):
    data = {}
    for h, v in zip(headers, values):
        data[_normalize_header(h)] = v
    def pick(keys):
        for k in keys:
            if k in data and data[k] not in (None, ''):
                return data[k]
        return None
    return {
        'date': pick(_DATE_KEYS),
        'amount': pick(_AMOUNT_KEYS),
        'comment': pick(_COMMENT_KEYS) or '',
        'external_id': pick(_EXTID_KEYS) or '',
        'raw': data,
    }


def _parse_csv(file_obj):
    content = file_obj.read()
    if isinstance(content, bytes):
        for enc in ('utf-8-sig', 'utf-8', 'cp1251'):
            try:
                content = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            content = content.decode('utf-8', errors='ignore')
    sample = content[:2048]
    delimiter = ';' if sample.count(';') > sample.count(',') else ','
    reader = csv.reader(io.StringIO(content), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return []
    headers = rows[0]
    return [_map_row(headers, r) for r in rows[1:] if any(c.strip() for c in r)]


def _parse_xlsx(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else '' for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        out.append(_map_row(headers, [('' if c is None else c) for c in r]))
    return out


def preview_rows(parsed_rows, limit=50):
    """Готує рядки для предпрогляду з валідацією."""
    preview = []
    for r in parsed_rows[:limit]:
        amount = _parse_amount(r['amount'])
        date = _parse_date(r['date'])
        preview.append({
            'date': date.strftime('%Y-%m-%d %H:%M') if date else '—',
            'amount': str(amount) if amount is not None else '—',
            'comment': str(r['comment'])[:120],
            'external_id': str(r['external_id']),
            'valid': amount is not None and date is not None,
            'type': 'income' if (amount and amount > 0) else 'expense',
        })
    return preview


def _is_duplicate(company, account, *, external_id, date, amount, comment):
    if external_id:
        if Transaction.objects.filter(company=company, external_id=external_id).exists():
            return True
    # Фолбек: дата+сума+рахунок+коментар.
    return Transaction.objects.filter(
        company=company, account=account, amount=abs(amount),
        date_actual__date=date.date(), comment=comment[:255],
    ).exists()


def import_rows(parsed_rows, *, user, account, apply_rules=True):
    """Створює операції з виписки, пропускаючи дублі. Повертає статистику."""
    company = get_default_company()
    created = 0
    skipped = 0
    errors = 0
    for r in parsed_rows:
        amount = _parse_amount(r['amount'])
        date = _parse_date(r['date'])
        if amount is None or date is None or amount == 0:
            errors += 1
            continue
        comment = str(r['comment'] or '')
        external_id = str(r['external_id'] or '')
        if _is_duplicate(company, account, external_id=external_id, date=date,
                         amount=amount, comment=comment):
            skipped += 1
            continue
        txn_type = Transaction.TYPE_INCOME if amount > 0 else Transaction.TYPE_EXPENSE
        if timezone.is_naive(date):
            date = timezone.make_aware(date, timezone.get_current_timezone())
        txn = txn_service.create_transaction(
            user=user, type=txn_type, amount=abs(amount), account=account,
            currency=account.currency, date_actual=date, comment=comment,
            source='import', external_id=external_id,
        )
        if apply_rules:
            try:
                from . import rules_engine
                rules_engine.apply_rules_to_transaction(txn, user=user, source='import')
            except Exception:
                pass
        created += 1
    return {'created': created, 'skipped': skipped, 'errors': errors}
